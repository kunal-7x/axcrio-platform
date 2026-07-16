"""caller.py — Famit backend API + (legacy) simple Caller page.

Serves JSON /campaigns,/extract,/leads,/run,/status,/calls,/stats (the Famit panel calls these
via nginx /api -> here) plus the legacy HTML page at '/'. Reuses the working LiveKit agent
(agent_name 'capsy') + campaign-adaptive dispatch metadata {campaign_id, lead_name}. Records every
call. Auth: Basic OR header 'X-Auth: <pass>' OR 'Authorization: Bearer <pass>'.
"""
from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import time
import unicodedata
import urllib.parse
import uuid
from datetime import datetime, timedelta, timezone
from pathlib import Path

import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, Request, Response, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse
from google.protobuf.duration_pb2 import Duration
from livekit import api

from prompt import build_system_prompt
try:
    # W1: vendor-script-aware renderer (used for prompt-preview + dry-run). On an
    # older box prompt.py without v2, fall back to v1 so caller.py still imports.
    from prompt import build_system_prompt_v2 as _build_system_prompt_v2
except Exception:  # noqa: BLE001
    _build_system_prompt_v2 = build_system_prompt

try:
    import whatsapp as wa_mod  # WAVE3 Unit5: provider-agnostic WhatsApp sender
except Exception:  # noqa: BLE001
    wa_mod = None

# WA-AUTO: per-person cross-call memory (voice agent's recap store). Import-safe and
# read-only here — used to enrich the WhatsApp reply brain with prior-call context.
# A missing/broken module leaves the WA path fully functional (memory is additive).
try:
    import memory as _mem_mod  # load_memory(phone) + build_recap(mem)
except Exception:  # noqa: BLE001
    _mem_mod = None

# WAVE A: vendor cost adapters (import-safe; each adapter no-ops without keys).
try:
    from vendors import elevenlabs as v_elevenlabs
    from vendors import vobiz as v_vobiz
    from vendors import groq_meter as v_groq
    from vendors import sarvam_meter as v_sarvam
    from vendors import display_name as vendor_display_name, VENDOR_IDS
except Exception:  # noqa: BLE001
    v_elevenlabs = v_vobiz = v_groq = v_sarvam = None
    VENDOR_IDS = ["vobiz", "elevenlabs", "groq", "sarvam", "livekit"]

    def vendor_display_name(vid):  # type: ignore
        return vid

load_dotenv("/opt/famit-agent/.env")

# P0 config resolver (Doppler-if-DOPPLER_TOKEN else .env/os.environ). Import-safe:
# if config.py is missing or errors, fall back to os.getenv so the service still
# starts. cfg_get/cfg_require are exact pass-throughs to os.getenv/os.environ when
# Doppler is disabled (today), so NO current value changes.
try:
    from config import get as cfg_get, require as cfg_require
except Exception:  # noqa: BLE001
    def cfg_get(key, default=None):
        return os.getenv(key, default)

    def cfg_require(key):
        return os.environ[key]

# P0 auth module (JWT access + rotating refresh). Import-safe: a missing module
# or missing pyjwt leaves _auth_mod usable-but-degraded (legacy auth untouched).
# It is wired via _auth_mod.init(...) AFTER the tenant store + _role_of exist.
try:
    import auth as _auth_mod
except Exception:  # noqa: BLE001
    _auth_mod = None

# P0 append-only audit log (best-effort; never breaks a request). init() after VAR.
try:
    import audit as _audit_mod
except Exception:  # noqa: BLE001
    _audit_mod = None

# System error & event log (white-labeled "System Logs" in the panel). Best-effort + dormant-
# safe; never breaks a request. init() after VAR. See logging_service.py for the design laws.
try:
    import logging_service as _log_mod
except Exception:  # noqa: BLE001
    _log_mod = None

# Observability analytics (read-only ClickHouse queries for the native traces/APM dashboards).
# Import-guarded; every query degrades to a clean error shape. See obs_query.py.
try:
    import obs_query as _obs_q
except Exception:  # noqa: BLE001
    _obs_q = None

# AI campaign-script drafting (Claude Sonnet 3.5). Import-guarded; dormant-safe (no key -> clean
# error). Powers the Script Studio "Generate with AI" button. See script_gen.py.
try:
    import script_gen as _script_gen
except Exception:  # noqa: BLE001
    _script_gen = None

# Post-call transcript CONTENT-quality analysis (LLM via OpenRouter). Import-guarded; dormant-safe.
# Powers the "Transcript quality" card in Voice Analytics. See transcript_quality.py.
try:
    import transcript_quality as _tq
except Exception:  # noqa: BLE001
    _tq = None

# P7 Script Studio 2.0 block compiler. Import-guarded; only engaged when a campaign opts into
# script_studio_v2 (else a pure no-op). Compiles typed blocks DOWN to the consumed fields so the
# live agent path is unchanged. See script_compiler.py.
try:
    import script_compiler as _script_compiler
except Exception:  # noqa: BLE001
    _script_compiler = None

# F2 Business Brain + Knowledge Base (RAG) substrate (import-safe; dormant-degrade).
# brain = structured per-org identity store (JSON mode); kb = pgvector+FTS corpus.
# Both no-op cleanly when their deps are absent. The live voice path imports NEITHER.
try:
    import brain as _brain_mod
except Exception:  # noqa: BLE001
    _brain_mod = None
try:
    import kb as _kb_mod
except Exception:  # noqa: BLE001
    _kb_mod = None

# F4 Credit/Wallet ACID ledger + Action Firewall (import-safe; both degrade cleanly).
# wallet = the money-custody transactional core (reserve/settle/release/topup/balance) on Postgres;
# firewall = PIN/OTP step-up gate. The live RUN-PATH imports NEITHER for spend gating yet (that wiring
# is a later flag-gated unit) — these only back the additive /wallet* and /firewall/* endpoints here.
try:
    import wallet as _wallet_mod
except Exception:  # noqa: BLE001
    _wallet_mod = None
try:
    import firewall as _firewall_mod
except Exception:  # noqa: BLE001
    _firewall_mod = None

# ════════════════════════════════════════════════════════════════════════════════════════════════
# W-WIRE-OPS: real-time ops backbone singletons. ALL import-guarded + flag-gated (default OFF).
# EARNER LAW: these only back the additive seams below; with every flag OFF nothing engages and the
# resting build is byte-identical. voice_kernel/ + voice_ops/ must be on PYTHONPATH (/opt/famit-agent).
# ════════════════════════════════════════════════════════════════════════════════════════════════

# ── W8 EventBus singleton (EVENTBUS_ENABLED default OFF — systemd drop-in only) ─────────────────
try:
    from voice_kernel.events import EventBusConfig as _EvCfg, RedisEventBus as _RedisEvBus  # type: ignore
    _EVCFG = _EvCfg.from_env()
    _EVBUS = _RedisEvBus(_EVCFG) if (_EVCFG and _EVCFG.enabled) else None
except Exception:  # noqa: BLE001
    _EVCFG = None
    _EVBUS = None

async def _ev(make_event):          # noqa: ANN001
    """Fire-and-forget event emit. No-ops when EVENTBUS_ENABLED is OFF. NEVER raises into the loop."""
    if _EVBUS is None:
        return
    try:
        await _EVBUS.emit(make_event)
    except Exception:  # noqa: BLE001
        pass

def _get_event_bus():
    """Accessor so the W9 StagedPipeline + W10 cadence reuse the SAME singleton bus."""
    return _EVBUS

# ── W9 Recording finalize config + providers (RECORDING_FINALIZE_ENABLED default OFF) ───────────
try:
    from voice_ops.recording import (    # type: ignore
        RecordingConfig as _RecCfg, StagedPipeline as _StagedPipeline,
        EgressClient as _EgressClient, ObjectStorage as _ObjStorage,
    )
    from voice_ops.recording.poller import FinalizePoller as _FinalizePoller  # type: ignore
    _RECCFG = _RecCfg.from_env()
except Exception:  # noqa: BLE001
    _RECCFG = None

def _w9_transcript_provider(tenant_id, call_id):   # noqa: ANN001
    """Bind the W9 pipeline to the transcript file caller already reads (room == call_id outbound)."""
    try:
        tr = _read(TRANSCRIPT_DIR / f"{call_id}.json", {})
    except Exception:  # noqa: BLE001
        return None
    if not tr:
        return None
    return {"turns": tr.get("turns", []), "text": tr.get("summary", "")}

def _w9_summary_provider(tenant_id, call_id, transcript):  # noqa: ANN001
    try:
        tr = _read(TRANSCRIPT_DIR / f"{call_id}.json", {})
    except Exception:  # noqa: BLE001
        return None
    if not tr.get("summary") and not tr.get("outcome"):
        return None
    return {
        "summary": tr.get("summary", ""),
        "lifecycle": tr.get("outcome", ""),
        "conversion_prob": tr.get("interest"),
    }

# ── W14 Reporting service (REPORTING_ENABLED default OFF) ────────────────────────────────────────
# ReportingConfig reads NO env, so the enabled flag is read here from REPORTING_ENABLED. The query
# API works against an empty in-mem store (dashboard sees zeros, never an error). A separate consumer
# worker (NOT in this process) fills the store from the W8 stream.
try:
    from voice_ops.reporting import (    # type: ignore
        ReportingStore as _RepStore, ReportingConfig as _RepCfg, ReportingService as _RepSvc,
    )
    _REPORTING_ON = (cfg_get("REPORTING_ENABLED", "0") or "0").strip().lower() in ("1", "true", "yes", "on")
    if _REPORTING_ON:
        _REPCFG = _RepCfg(enabled=True)
        _REPSTORE = _RepStore()
        _REPSVC = _RepSvc(_REPSTORE, _REPCFG)
    else:
        _REPCFG = None; _REPSTORE = None; _REPSVC = None
except Exception:  # noqa: BLE001
    _REPCFG = None; _REPSTORE = None; _REPSVC = None

# ── W14-WIRE: read-model bridge (caller.py /report* <- the SAME W8 stream the worker reads) ───────
# ROOT-CAUSE FIX (dashboard zeros): the reporting WORKER fills a ReportingStore in a SEPARATE process;
# this API process had its OWN empty in-mem store -> /report returned zeros. We bridge by REPLAYING the
# per-tenant Redis stream `vk:events:{tenant}` (read-only XRANGE, no consumer group => zero interference
# with the worker) into THIS process's _REPSTORE via the SAME reducer the worker uses, on each query.
# Idempotent (the reducer is latest-wins on call_id) + per-tenant last-id cached so a repeat query only
# replays new entries. Fail-safe: on any error we leave the store as-is (the old empty-store behavior,
# never a 500). Bounded scan so a huge backlog can't stall a request.
_REP_LAST_ID: dict = {}          # tenant_id -> last XRANGE stream id replayed
_REP_REDIS = None                # cached sync redis client (lazy)
_REP_HYDRATE_MAX = 5000          # max entries replayed per query (bounded; reducer is latest-wins)
try:
    from voice_kernel.events import serde as _vk_serde            # type: ignore
    from voice_ops.reporting import build_consumer_handler as _rep_build_handler  # type: ignore
    _REP_HANDLER = _rep_build_handler(_REPSTORE) if _REPSTORE is not None else None
except Exception:  # noqa: BLE001
    _vk_serde = None
    _REP_HANDLER = None


def _rep_redis():
    """Lazy SYNC redis client on the SAME bus url as the W8 emit. None if unavailable."""
    global _REP_REDIS
    if _REP_REDIS is not None:
        return _REP_REDIS
    if _EVCFG is None:
        return None
    try:
        import redis as _r  # type: ignore  # noqa: PLC0415
        _REP_REDIS = _r.from_url(_EVCFG.url, socket_timeout=2.0, socket_connect_timeout=2.0)
        return _REP_REDIS
    except Exception:  # noqa: BLE001
        return None


def _rep_xrange(stream: str, start_id: str):
    """Blocking XRANGE (start exclusive) -> list[(id, fields)]. Empty on any error."""
    rc = _rep_redis()
    if rc is None:
        return []
    try:
        return rc.xrange(stream, min=start_id, max="+", count=_REP_HYDRATE_MAX)
    except Exception:  # noqa: BLE001
        return []


async def _w14_hydrate(tenant_id: str) -> None:
    """Replay any NEW `vk:events:{tenant}` entries into THIS process's _REPSTORE so /report* serves
    LIVE numbers. Read-only, no consumer group, fail-safe, bounded. No-op when reporting is off."""
    if _REPSTORE is None or _REP_HANDLER is None or _vk_serde is None or _EVCFG is None or not tenant_id:
        return
    try:
        stream = _EVCFG.stream_for(tenant_id)
    except Exception:  # noqa: BLE001
        return
    last = _REP_LAST_ID.get(tenant_id, "0-0")
    # XRANGE min is INCLUSIVE; use "(" to make it exclusive of the last id we already replayed.
    start = "(" + last if last != "0-0" else "0-0"
    try:
        entries = await asyncio.to_thread(_rep_xrange, stream, start)
    except Exception:  # noqa: BLE001
        return
    for _id, fields in entries or []:
        try:
            ev = _vk_serde.decode(fields)
            await _REP_HANDLER(ev)          # reduce+upsert onto the in-proc read-model (never raises)
            _REP_LAST_ID[tenant_id] = _id.decode() if isinstance(_id, (bytes, bytearray)) else str(_id)
        except Exception:  # noqa: BLE001 — one bad entry must never fail the query
            continue

# ── W14b AI-Manager live-data service (rides the SAME reporting store) ───────────────────────────
try:
    from voice_ops.ai_manager_live import AIManagerLiveService as _AIMLiveSvc  # type: ignore
    _AIM_LIVE = _AIMLiveSvc(_REPSVC) if _REPSVC is not None else None
except Exception:  # noqa: BLE001
    _AIM_LIVE = None

# ── W7 Lead lifecycle + AI summary (LEAD_LIFECYCLE_ENABLED default OFF) ──────────────────────────
_LEAD_LIFECYCLE_ON = (cfg_get("LEAD_LIFECYCLE_ENABLED", "0") or "0").strip().lower() in ("1", "true", "yes", "on")
try:
    from voice_kernel.memory.lifecycle import classify_lifecycle as _vk_classify_lifecycle  # type: ignore
    from voice_kernel.packet import Lifecycle as _VKLifecycle  # type: ignore
except Exception:  # noqa: BLE001
    _vk_classify_lifecycle = None
    _VKLifecycle = None

# ── W10 Smart callback cadence (CALLBACK_CADENCE_ENABLED default OFF; anti-runaway) ─────────────
try:
    from voice_ops.callback import (    # type: ignore
        CallbackConfig as _CbCfg, InMemoryCallbackStore as _CbStore,
        enqueue_smart as _cb_enqueue_smart, fire_due as _cb_fire_due, release as _cb_release,
    )
    _CB_CFG = _CbCfg.from_env()
    _CB_STORE = _CbStore() if (_CB_CFG and _CB_CFG.enabled) else None
except Exception:  # noqa: BLE001
    _CB_CFG = None
    _CB_STORE = None
    _cb_enqueue_smart = None
    _cb_fire_due = None
    _cb_release = None

# ── NCPR / DND scrub-before-dial (W26 compliance; R4 A2) ─────────────────────────────────────────
# An ADDITIONAL pre-dial compliance gate on the cadence dial path: scrub every due number against the
# NCPR national DND register before dialing (on top of the always-on local opt-out / suppression
# check already in the fire_due seam). DndScrubber is FAIL-CLOSED (a cache miss BLOCKS) — so with an
# empty NCPR cache it would block EVERY dial. Therefore it is gated behind CALLBACK_NCPR_SCRUB_ENABLED
# (default OFF): only the founder flips it ON once the NCPR cache is being populated by the scrub job.
# The local-suppression check stays always-on regardless of this flag.
_NCPR_SCRUB_ENABLED = (cfg_get("CALLBACK_NCPR_SCRUB_ENABLED", "0") or "0").strip().lower() \
    in ("1", "true", "yes", "on")
try:
    from voice_ops.compliance.dnd import DndScrubber as _DndScrubber  # type: ignore
    _NCPR_SCRUBBER = _DndScrubber(
        salt=(cfg_get("DND_HASH_SALT", "") or ""),
        refresh_days=int(cfg_get("DND_REFRESH_DAYS", "30") or "30"),
    ) if _NCPR_SCRUB_ENABLED else None
except Exception:  # noqa: BLE001
    _DndScrubber = None
    _NCPR_SCRUBBER = None
# ════════════════════════════════════════════════════════════════════════════════════════════════

# CRM CORE: contact spine + unified timeline + next-best-action (import-safe; PG-native projection).
# A read-model/intent layer over the existing leads/calls/wa/suppression/events stores — NEVER a second
# writer of core records, NEVER on the voice run-path. Degrades to empty shapes when PG is absent.
try:
    import crm as _crm_mod
except Exception:  # noqa: BLE001
    _crm_mod = None

# CONTROL LAYER: Foundation entitlement engine (CL-B1). Import-safe; PG-native projection like crm.
# Provides resolve_modes/mode_for/assert_access/feature_key_for_path + the /me/entitlements payload.
# The enforcement choke-point (CL-B2, below) reads through it. Degrades to all-default 'on' when PG is
# absent → live site untouched. NOTHING is enforced unless CONTROL_ENABLED is on (default off).
try:
    import entitlements as _ent_mod
except Exception:  # noqa: BLE001
    _ent_mod = None

# P0 per-tenant rate limiter (Redis-if-reachable else in-proc; FAIL-OPEN always).
try:
    import ratelimit as _rl_mod
except Exception:  # noqa: BLE001
    _rl_mod = None

# P0 observability: Prometheus /metrics + structured request logs (best-effort).
try:
    import obs as _obs_mod
except Exception:  # noqa: BLE001
    _obs_mod = None

# Legacy-token kill switch. Default ON so NOTHING breaks today. Set
# LEGACY_TOKEN_ENABLED=false ONLY after every client uses JWT (post-cutover).
LEGACY_TOKEN_ENABLED = (cfg_get("LEGACY_TOKEN_ENABLED", "true") or "true").strip().lower() \
    not in ("0", "false", "no", "off")

# CONTROL LAYER master gate. DEFAULT OFF -> the entitlement enforcement middleware is a pure no-op
# passthrough (resting state byte-identical to today, F2/F4 discipline). Flip to true ONLY after the
# T1-T18 isolation/impersonation probes pass (CONTROL_LAYER_EXECUTION_PLAN §5 C12).
CONTROL_ENABLED = (cfg_get("CONTROL_ENABLED", "false") or "false").strip().lower() \
    in ("1", "true", "yes", "on")

# ---------- W2: full-context cache invalidation bus (flag CTX_CACHE, default OFF) ----------
# context_store.py is a tenant-scoped LRU + Redis version-stamp cache the INBOUND voice brain reads at
# connect. caller.py's only job here is to PUBLISH a version bump when a campaign is created/edited so the
# inbound process reloads the new context immediately (not after the 300s TTL). DEFAULT OFF -> a no-op:
# when CTX_CACHE is off, _publish_ctx_invalidate does nothing, so the save path is byte-identical to today.
# Import is best-effort + side-effect-free; a missing/broken module NEVER affects campaign save.
try:
    import context_store as _ctx_store  # noqa: F401
except Exception:  # noqa: BLE001
    _ctx_store = None  # type: ignore


def _publish_ctx_invalidate(tenant_id: str, cid: str) -> None:
    """Bump the campaign's version stamp on the cache bus so inbound reloads it. Flag-gated (CTX_CACHE);
    no-op + swallows on any error — must NEVER break a campaign create/edit (earner-safe, additive)."""
    try:
        if _ctx_store is None or not _ctx_store.is_enabled():
            return
        _ctx_store.bump_version(tenant_id, cid)
    except Exception:  # noqa: BLE001
        pass

TRUNK = cfg_get("LIVEKIT_SIP_TRUNK_ID", "ST_fmtVmNJmpzKa")
AGENT = cfg_get("LIVEKIT_AGENT_NAME", "capsy")
LK_URL = cfg_get("LIVEKIT_URL", "ws://127.0.0.1:7880")
LK_KEY = cfg_require("LIVEKIT_API_KEY")
LK_SECRET = cfg_require("LIVEKIT_API_SECRET")


# ---------- REC-B: server-side auto-egress for OUTBOUND calls ----------
# Every outbound campaign call records server-side via LiveKit room-composite AUTO-egress
# (attached to CreateRoomRequest.egress at room create — NOT agent.py-side). Audio-only OGG ->
# DO Spaces (S3Upload). DORMANT-UNTIL-CREDS, identical posture to the inbound AIM recorder:
# arms only when AIM_RECORDING_ENABLED is truthy AND the Spaces creds are complete; otherwise a
# no-op (egress=None) so the dial path is byte-identical to pre-REC-B. NEVER raises — an egress
# build failure must never block a paid outbound call (earner-safety: the call always dials).
# Reuses the SAME AIM_SPACES_* creds the inbound recorder already uses (bucket capsy-recordings).
def _outbound_recording_enabled() -> bool:
    if (cfg_get("AIM_RECORDING_ENABLED", "") or "").strip().lower() not in ("1", "true", "yes", "on"):
        return False
    need = ("AIM_SPACES_BUCKET", "AIM_SPACES_KEY", "AIM_SPACES_SECRET", "AIM_SPACES_ENDPOINT")
    return all((cfg_get(k, "") or "").strip() for k in need)


# REC-B-AZURE (flag-gated, ADDITIVE): an alternative recording backend that uploads the same
# room-composite MP3 egress to Azure Blob Storage instead of DO Spaces. FLAG-OFF == byte-identical
# to today: RECORDING_BACKEND defaults to "spaces" so every existing deploy keeps using DO Spaces,
# which stays BOTH the default AND the fallback (if Azure creds are incomplete, Azure is simply not
# armed and the Spaces branch runs unchanged). NEVER changes the dial path when disabled.
_AZURE_RECORDINGS_CONTAINER_DEFAULT = "recordings"


def _azure_recording_enabled() -> bool:
    """True ONLY when RECORDING_BACKEND=='azure' AND the Azure account+key+container are all set.
    Anything missing -> False -> the caller falls back to the DO Spaces path. NEVER raises."""
    if (cfg_get("RECORDING_BACKEND", "spaces") or "spaces").strip().lower() != "azure":
        return False
    account = (cfg_get("AZURE_STORAGE_ACCOUNT", "") or "").strip()
    akey = (cfg_get("AZURE_STORAGE_KEY", "") or "").strip()
    container = (cfg_get("AZURE_RECORDINGS_CONTAINER", _AZURE_RECORDINGS_CONTAINER_DEFAULT) or "").strip()
    return bool(account and akey and container)


def _outbound_recording_key(call_id: str) -> str:
    """Deterministic Spaces object key: outbound-recordings/YYYY/MM/DD/<call_id>.mp3.
    Chosen BEFORE egress confirms so the call row can store it immediately; the object lands here.
    MP3 (was OGG/Opus): plays natively in EVERY browser incl. Safari AND has a real duration header,
    so the panel player gets sound + a moving bar + correct sync (OGG/Opus played silent in Safari)."""
    day = time.strftime("%Y/%m/%d")
    cid = (call_id or uuid.uuid4().hex)[:48]
    return f"outbound-recordings/{day}/{cid}.mp3"


def _build_outbound_egress(call_id: str):
    """Build a RoomEgress(room=RoomCompositeEgressRequest(audio-only MP3 -> Spaces OR Azure)) for
    embedding in CreateRoomRequest.egress. Returns (egress_obj_or_None, recording_key, bucket).
    Never raises; returns (None, "", "") when disabled/dormant or on any build error (call dials
    unrecorded). REC-B-AZURE: when RECORDING_BACKEND=='azure' (and Azure creds are complete), the
    EncodedFileOutput uploads to an Azure Blob container instead of DO Spaces — the returned
    "bucket" is then the Azure container name. Flag-off keeps the S3/Spaces branch BYTE-IDENTICAL."""
    try:
        if not _outbound_recording_enabled():
            return None, "", ""
        key = _outbound_recording_key(call_id)
        # REC-B-AZURE branch: same MP3 audio-only egress, uploaded to Azure Blob Storage. The DO
        # Spaces path below stays the default + the fallback (this only runs when Azure is fully
        # configured AND selected). No s3= field is set on the EncodedFileOutput in this branch.
        if _azure_recording_enabled():
            az_account = (cfg_get("AZURE_STORAGE_ACCOUNT", "") or "").strip()
            az_key = (cfg_get("AZURE_STORAGE_KEY", "") or "").strip()
            az_container = (cfg_get("AZURE_RECORDINGS_CONTAINER",
                                    _AZURE_RECORDINGS_CONTAINER_DEFAULT) or "").strip()
            file_out_az = api.EncodedFileOutput(
                file_type=api.EncodedFileType.MP3,
                filepath=key,
                azure=api.AzureBlobUpload(
                    account_name=az_account, account_key=az_key, container_name=az_container,
                ),
            )
            egress_az = api.RoomEgress(
                room=api.RoomCompositeEgressRequest(
                    audio_only=True,
                    file_outputs=[file_out_az],
                )
            )
            return egress_az, key, az_container
        bucket = (cfg_get("AIM_SPACES_BUCKET", "") or "").strip()
        region = (cfg_get("AIM_SPACES_REGION", "") or "us-east-1").strip()
        endpoint = (cfg_get("AIM_SPACES_ENDPOINT", "") or "").strip()
        s3key = (cfg_get("AIM_SPACES_KEY", "") or "").strip()
        s3secret = (cfg_get("AIM_SPACES_SECRET", "") or "").strip()
        file_out = api.EncodedFileOutput(
            file_type=api.EncodedFileType.MP3,   # audio-only; MP3 plays in EVERY browser (OGG/Opus was silent in Safari)
            filepath=key,
            s3=api.S3Upload(
                access_key=s3key, secret=s3secret, bucket=bucket,
                region=region, endpoint=endpoint, force_path_style=True,
            ),
        )
        egress = api.RoomEgress(
            room=api.RoomCompositeEgressRequest(
                audio_only=True,
                file_outputs=[file_out],
            )
        )
        return egress, key, bucket
    except Exception as exc:  # noqa: BLE001
        try:
            import logging as _lg_rec
            _lg_rec.getLogger("famit-caller").warning(
                "REC-B outbound egress build failed (call dials unrecorded): %r", exc)
        except Exception:  # noqa: BLE001
            pass
        return None, "", ""


USER = cfg_get("CALLER_USER", "famit")
PW = cfg_get("CALLER_PASS", "Famit@2026")
GROQ_KEY = cfg_require("GROQ_API_KEY")
GROQ_MODEL = cfg_get("CALLER_EXTRACT_MODEL", "meta-llama/llama-4-scout-17b-16e-instruct")
VAR = Path(os.getenv("FAMIT_VAR", "/opt/famit-agent/var"))
CAMPAIGN_DIR = VAR / "campaigns"
TRANSCRIPT_DIR = VAR / "transcripts"
LEADS_FILE = VAR / "leads.json"
CALLS_FILE = VAR / "calls.json"
TENANTS_FILE = VAR / "tenants.json"
SECRET_FILE = VAR / "secret"
SUPPRESSION_FILE = VAR / "suppression.json"   # P0.2 DND/suppression store
RETRY_FILE = VAR / "retry_queue.json"         # P0.5 retry + callback queue
# 🚨 KILL-SWITCH (callback/retry SPAM hotfix 2026-06-16): the auto-retry+callback scheduler
# was redialing leads ~every 2h NON-STOP (no-answer reconciliation reset attempts→1 /
# backoff→120min each tick; callbacks enqueued on ANY LLM-extracted callback_at — even on
# answered/completed calls — with attempts never incrementing → infinite redial loop).
# Until the retry engine is REBUILT with correct policy (≤2 retries, next-day cadence, NO
# callback on a completed pickup, busy→short reschedule, "call me at X"→that-time), the
# scheduler's DIALING is DISABLED by default. Set RETRY_SCHEDULER_ENABLED=1 ONLY after the
# rebuild lands. First-calls (run_job) are UNAFFECTED — only auto-retry/callback dialing is gated.
RETRY_SCHEDULER_ENABLED = (cfg_get("RETRY_SCHEDULER_ENABLED", "0") or "0").strip().lower() \
    in ("1", "true", "yes", "on")
WA_LOG_FILE = VAR / "wa_log.json"             # P1.A whatsapp send log
WA_THREADS_DIR = VAR / "wa_threads"           # WAVE A2 per-contact WhatsApp conversation state
WA_UNROUTED_TENANT = "_unrouted"              # P0-LEAK: quarantine bucket for unknown inbound numbers (never ADMIN_ID)
WEBHOOK_FILE = VAR / "webhooks.json"          # P1.C registered webhooks
WEBHOOK_LOG_FILE = VAR / "webhook_log.json"   # P1.C delivery log
BILLING_FILE = VAR / "billing.json"           # WAVE3 Unit4 per-tenant billing plans
LEDGER_DIR = VAR / "ledger"                   # WAVE3 Unit4 per-tenant charge ledgers
# ---------- WAVE A: real vendor-cost metering stores ----------
USAGE_EVENTS_FILE = VAR / "usage_events.json"     # append: per-call per-vendor internal metering rows
COST_LEDGER_FILE = VAR / "cost_ledger.json"       # normalized joined cost rows (usage + vendor CDR)
DAILY_ROLLUPS_FILE = VAR / "daily_rollups.json"   # precomputed daily/vendor/tenant rollups
VENDOR_SNAPSHOTS_FILE = VAR / "vendor_snapshots.json"  # per-vendor sync status/timestamps for Audit tab
MIN_CALL_FLOOR = 22

# ---------- IST time helpers (P0.1 / P0.5) ----------
IST = timezone(timedelta(hours=5, minutes=30))


def now_ist() -> datetime:
    return datetime.now(IST)


def _utc_iso() -> str:
    """W14-WIRE: UTC wall-clock ISO, tz-LABELLED (+00:00) so the panel parses call times correctly.
    The box runs UTC, so the instant is unchanged vs the old naive datetime.now(); we only ADD the
    offset label that was missing (old `started_at`/`ended_at` were unlabelled -> mis-parsed as local)."""
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


app = FastAPI()
JOBS: dict = {}


# LOGGING (best-effort, defined FIRST so it is the INNERMOST user middleware — it wraps the
# route + Starlette's ExceptionMiddleware most closely). It records into "System Logs" BOTH:
#   (1) any UNHANDLED route exception (caught here, logged, then RE-RAISED so Starlette's normal
#       500 handling stays byte-identical), AND
#   (2) any response that comes back with status >= 500 — this is the BIG one: the codebase is
#       built on a "best-effort, never raise" law, so most errors are caught inside the route and
#       returned as JSONResponse(status_code=500) (or `raise HTTPException(5xx)`, which Starlette's
#       INNER ExceptionMiddleware converts to a 5xx response before it reaches us). Those never
#       raise, so the old exception-only capture missed them — which is exactly why operators saw
#       errors that never reached System Logs. We now log the 5xx response too.
# It NEVER alters a response (read-only on .status_code), never swallows an error, and is a no-op
# when logging_service is unconfigured — so it cannot break a request. A per-request flag
# (request.state._sys_logged) de-dupes so an unhandled exception isn't logged twice, and a handler
# that already logged its own event can set the flag to opt out of the generic 5xx line.
_5XX_LOG_EXEMPT = {"/health", "/metrics", "/favicon.ico"}  # high-freq infra pings -> would flood

# Self-hosted HTTP request telemetry for the Performance page (replaces the SigNoz APM source).
# Import-guarded + flag-gated (HTTP_METRICS_ENABLED) + best-effort: a missing/broken module or
# ClickHouse can NEVER affect a request. Recording is a cheap in-memory append; the write batches on
# a background task. /health, /metrics etc. are skipped (infra pings would dominate the dashboard).
try:
    import http_metrics as _http_metrics  # noqa: E402
except Exception:  # noqa: BLE001
    _http_metrics = None

_HTTP_METRICS_EXEMPT = {"/health", "/metrics", "/favicon.ico"}


def _record_http_metric(request: "Request", status_code: int, t0: float) -> None:
    """Record one request to the HTTP-metrics buffer. Never raises into the request path."""
    if _http_metrics is None:
        return
    try:
        path = request.url.path
        if path in _HTTP_METRICS_EXEMPT:
            return
        import uuid as _uuid
        try:                                   # the matched route template (low cardinality), else the raw path
            _rt = request.scope.get("route")
            route = getattr(_rt, "path", "") or path
        except Exception:  # noqa: BLE001
            route = path
        _http_metrics.record(method=request.method, route=route, status_code=int(status_code),
                             duration_ms=(time.perf_counter() - t0) * 1000.0,
                             trace_id=_uuid.uuid4().hex[:16])
    except Exception:  # noqa: BLE001
        pass


@app.middleware("http")
async def _capture_errors_mw(request: Request, call_next):
    _http_t0 = time.perf_counter()
    if _http_metrics is not None:
        try:
            _http_metrics.ensure_started()   # idempotent; starts the flush loop on the running loop
        except Exception:  # noqa: BLE001
            pass
    try:
        resp = await call_next(request)
    except Exception as exc:  # noqa: BLE001
        try:
            request.state._sys_logged = True
            if _log_mod is not None:
                try:
                    _t = resolve_tenant(request)
                except Exception:  # noqa: BLE001
                    _t = None
                _log_mod.record(
                    "error", "backend",
                    f"{request.method} {request.url.path}: {exc!r}",
                    tenant_id=(_t or {}).get("tenant_id", "") if isinstance(_t, dict) else "",
                    error_type=type(exc).__name__,
                    context={"path": request.url.path, "method": request.method},
                )
        except Exception:  # noqa: BLE001
            pass
        _record_http_metric(request, 500, _http_t0)
        raise
    # (2) capture swallowed/converted SERVER-FAULT responses that did NOT raise. We log 500/502/504
    # (real faults) but DELIBERATELY SKIP 503 — "Service Unavailable" is the conventional degrade for a
    # dependency/feature that isn't configured yet (a dormant managed-provider, reporting, brain/kb/crm
    # subsystem). Those endpoints already return a clean shape AND are polled by the panel, so
    # auto-logging every 503 floods the store with non-actionable noise. A genuine crash still surfaces
    # via the unhandled-exception branch above (any status) — only the swallowed-503 polling noise is cut.
    try:
        _sc = getattr(resp, "status_code", 200)
        if (_log_mod is not None
                and _sc >= 500 and _sc != 503
                and not getattr(request.state, "_sys_logged", False)
                and request.url.path not in _5XX_LOG_EXEMPT):
            try:
                _t = resolve_tenant(request)
            except Exception:  # noqa: BLE001
                _t = None
            _log_mod.record(
                "error", "backend",
                f"{request.method} {request.url.path} -> {resp.status_code}",
                tenant_id=(_t or {}).get("tenant_id", "") if isinstance(_t, dict) else "",
                error_type=f"http_{resp.status_code}",
                context={"path": request.url.path, "method": request.method,
                         "status": resp.status_code},
            )
    except Exception:  # noqa: BLE001
        pass
    _record_http_metric(request, getattr(resp, "status_code", 200), _http_t0)
    return resp


# ── Phase 3 observability — OpenTelemetry traces -> SigNoz (DORMANT-by-default) ────────────
# NOTE: a Prometheus /metrics endpoint ALREADY exists (the `obs` module, see `@app.get("/metrics")`
# below) — the obs droplet's Prometheus scrapes THAT. Here we only ADD distributed tracing.
# Guarded + gated: OTel is set up ONLY when OTEL_EXPORTER_OTLP_ENDPOINT is set, so until the
# observability droplet is wired this is a NO-OP and prod is byte-identical.
if (os.getenv("OTEL_EXPORTER_OTLP_ENDPOINT") or "").strip():
    import logging as _otel_log  # caller.py has no module-level `logging` import (it imports locally)
    try:
        from opentelemetry import trace as _otel_trace  # noqa: E402
        from opentelemetry.sdk.resources import Resource as _OtelResource  # noqa: E402
        from opentelemetry.sdk.trace import TracerProvider as _OtelTP  # noqa: E402
        from opentelemetry.sdk.trace.export import BatchSpanProcessor as _OtelBSP  # noqa: E402
        from opentelemetry.exporter.otlp.proto.grpc.trace_exporter import (  # noqa: E402
            OTLPSpanExporter as _OtelExporter)
        from opentelemetry.instrumentation.fastapi import (  # noqa: E402
            FastAPIInstrumentor as _OtelFastAPI)
        _otel_res = _OtelResource.create(
            {"service.name": os.getenv("OTEL_SERVICE_NAME", "haptica-backend")})
        _otel_tp = _OtelTP(resource=_otel_res)
        _otel_tp.add_span_processor(_OtelBSP(_OtelExporter()))  # reads endpoint from env
        _otel_trace.set_tracer_provider(_otel_tp)
        _OtelFastAPI.instrument_app(app)
        _otel_log.getLogger("famit-caller").info("OTel tracing -> %s",
                                                 os.getenv("OTEL_EXPORTER_OTLP_ENDPOINT"))
    except Exception:  # noqa: BLE001
        _otel_log.getLogger("famit-caller").warning("OTel setup skipped", exc_info=True)


# ---------- P0: per-tenant rate-limit middleware (additive, FAIL-OPEN) ----------
# Pick the backend once at startup (redis on :6380 if reachable, else in-proc, else
# disabled). On ANY problem this middleware allows the request — it never blocks
# real traffic. Health/metrics and the legacy HTML page are exempt.
_RL_BACKEND = "disabled"
if _rl_mod is not None:
    try:
        _RL_BACKEND = _rl_mod.init()
    except Exception:  # noqa: BLE001
        _RL_BACKEND = "disabled"

_RL_EXEMPT_PATHS = {"/", "/health", "/metrics", "/favicon.ico"}
# Auth endpoints get the strict 'auth' class (brute-force guard), keyed by IP.
_RL_AUTH_PATHS = {"/login", "/auth/login", "/auth/refresh"}


def _rl_route_class(method: str, path: str) -> str:
    if path in _RL_AUTH_PATHS:
        return "auth"
    if method in ("POST", "PUT", "PATCH", "DELETE"):
        return "write"
    return "read"


@app.middleware("http")
async def _rate_limit_mw(request: Request, call_next):
    # Disabled / module missing -> straight through.
    if _rl_mod is None or _RL_BACKEND == "disabled":
        return await call_next(request)
    try:
        path = request.url.path
        if path in _RL_EXEMPT_PATHS:
            return await call_next(request)
        route_class = _rl_route_class(request.method, path)
        # Key by tenant when we can resolve one cheaply; else by client IP so
        # anonymous floods (e.g. login brute-force) are still bounded.
        try:
            t = resolve_tenant(request)
        except Exception:  # noqa: BLE001
            t = None
        key = (t or {}).get("tenant_id") or f"ip:{_client_ip(request)}"
        allowed, info = _rl_mod.allow(key, route_class)
        if not allowed:
            retry = int(info.get("reset_in", 1) or 1)
            resp = JSONResponse(
                {"error": "rate limit exceeded", "route_class": route_class,
                 "retry_after": retry}, status_code=429)
            resp.headers["Retry-After"] = str(retry)
            resp.headers["X-RateLimit-Limit"] = str(info.get("limit", ""))
            resp.headers["X-RateLimit-Remaining"] = str(info.get("remaining", 0))
            return resp
    except Exception:  # noqa: BLE001 — limiter must never break the request path
        return await call_next(request)
    return await call_next(request)


# ---------- P0: observability (Prometheus + structured access log) ----------
# Cost gauge REUSES the existing metered cost_ledger (lazy callback; the function
# is defined later in this file but only invoked at /metrics render time).
def _cost_by_currency() -> dict:
    out: dict = {}
    try:
        for r in _read_cost_ledger():
            ccy = r.get("currency") or "INR"
            out[ccy] = out.get(ccy, 0.0) + float(r.get("cost", 0) or 0)
    except Exception:  # noqa: BLE001
        pass
    return out or {"INR": 0.0}


if _obs_mod is not None:
    try:
        _obs_mod.init(cost_provider=_cost_by_currency, component="famit-caller")
    except Exception:  # noqa: BLE001
        pass


@app.middleware("http")
async def _metrics_mw(request: Request, call_next):
    if _obs_mod is None or not _obs_mod.ready():
        return await call_next(request)
    start = time.perf_counter()
    _obs_mod.inprogress_inc()
    status = 500
    try:
        resp = await call_next(request)
        status = resp.status_code
        return resp
    finally:
        _obs_mod.inprogress_dec()
        try:
            dt = time.perf_counter() - start
            # Route TEMPLATE (e.g. /campaigns/{cid}) to bound metric cardinality.
            route = request.url.path
            r = request.scope.get("route")
            if r is not None and getattr(r, "path", None):
                route = r.path
            _obs_mod.observe(request.method, route, status, dt)
            # one-line structured access log (journald)
            tid = ""
            try:
                _t = resolve_tenant(request)
                tid = (_t or {}).get("tenant_id", "")
            except Exception:  # noqa: BLE001
                pass
            _obs_mod.log_request(request.method, request.url.path, route, status,
                                 dt * 1000.0, tenant=tid, ip=_client_ip(request))
        except Exception:  # noqa: BLE001
            pass


# ════════════════════════════════════════════════════════════════════════════
# CONTROL LAYER — entitlement ENFORCEMENT middleware (CL-B2 / plan C3)
# ════════════════════════════════════════════════════════════════════════════
# THE REAL SECURITY BOUNDARY (control-security.md §1.3, spec-control-layer §3):
#   path -> feature_key (via entitlements.feature_key_for_path, longest-prefix +
#   explicit shared map) -> mode -> HIDDEN=404 (no existence leak), LOCKED=402
#   (+upsell JSON), CORE=always pass, unmapped-legacy-path=pass, ON=pass.
#
# It is implemented as an @app.middleware("http") that RETURNS a JSONResponse for a
# block (it NEVER raises HTTPException inside the middleware — the researched
# Starlette gotcha is that raising inside a custom middleware runs OUTSIDE the
# ExceptionMiddleware and leaks a 500; the existing _rate_limit_mw already follows
# this return-don't-raise pattern, so we match it). Frontend HIDE/LOCK is cosmetic;
# this layer is what a saved token / curl / devtools actually hits.
#
# GATED behind CONTROL_ENABLED (default OFF) -> pure no-op passthrough = resting
# byte-identical (T17). On the off path the function returns before any work.
#
# Paths this layer NEVER governs (entitlement is N/A — orthogonal gates own them):
#   * infra/docs (/, /health, /metrics, /favicon.ico, /docs, /openapi.json, /redoc)
#   * /admin/*  -> role-gated by require_super_admin (403 for a vendor); NOT a
#                  hidden FEATURE, so it must not 404/402 here (control-security §1.4).
_CONTROL_EXEMPT_EXACT = {"/", "/health", "/metrics", "/favicon.ico",
                         "/docs", "/openapi.json", "/redoc"}


def _control_path_exempt(path: str) -> bool:
    if path in _CONTROL_EXEMPT_EXACT:
        return True
    # admin plane (role-gated, not entitlement-gated) + swagger assets.
    if path.startswith("/admin") or path.startswith("/docs") or path.startswith("/openapi"):
        return True
    return False


# CONTROL LAYER (CL-B3): act-as READ-ONLY write block. ALWAYS active (independent of CONTROL_ENABLED —
# impersonation is a live capability, not gated by the entitlement master flag). A read_only act-as token
# may ONLY make safe (GET/HEAD/OPTIONS) requests; any POST/PUT/DELETE/PATCH is refused (T10). The act-as
# exit + the firewall verify-pin (to elevate) are the only mutating exceptions. Costs nothing on a normal
# request: we only decode when the cred is present AND carries the act_as marker.
_ACT_AS_WRITE_OK_PATHS = {"/admin/act-as/exit", "/firewall/verify-pin"}


def _act_as_readonly_block(request: Request):
    """Return a 403 JSONResponse if this is a read_only act-as token attempting a mutation; else None."""
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    if _auth_mod is None:
        return None
    try:
        cred = _extract_cred(request)
        if not cred or cred.count(".") != 2:
            return None
        claims = _auth_mod.act_as_claims(cred)
    except Exception:  # noqa: BLE001
        return None
    if not claims:
        return None
    if claims.get("scope") == "read_only" and request.url.path not in _ACT_AS_WRITE_OK_PATHS:
        return JSONResponse(
            {"error": "act-as session is read-only", "act_as": claims.get("act_as")},
            status_code=403)
    return None


@app.middleware("http")
async def _enforce_entitlement_mw(request: Request, call_next):
    # (-1) ACT-AS READ-ONLY GUARD — always on (impersonation safety, not entitlement enforcement).
    blocked = _act_as_readonly_block(request)
    if blocked is not None:
        return blocked

    # (0) MASTER GATE: off -> byte-identical passthrough (no resolve, no lookup).
    if not CONTROL_ENABLED or _ent_mod is None:
        return await call_next(request)

    path = request.url.path
    if _control_path_exempt(path):
        return await call_next(request)

    feature_key = None
    try:
        # (1) path -> governing feature_key (longest-prefix + shared map). None for an
        #     ungoverned legacy route -> pass through (CI registry-drift guard closes
        #     that gap later; spec §3). A resolver error -> treat as ungoverned (pass).
        feature_key = _ent_mod.feature_key_for_path(path)
        if not feature_key:
            return await call_next(request)

        # (2) CORE floor: un-hideable keys (login/me/settings/health/wallet-pay) ALWAYS
        #     pass — anti-lockout (a misconfig must never brick the way back in).
        reg = _ent_mod.load_registry().get(feature_key) or {}
        if reg.get("is_core"):
            return await call_next(request)

        # (3) resolve the calling tenant from the TOKEN (never the body). If we can't,
        #     do NOT 404/402 here — let the route's own auth return 401 (don't mask the
        #     auth failure as a not-found). Anonymous traffic on a public route (e.g.
        #     /f/ public forms) is governed by that route, not by an entitlement.
        try:
            tenant = resolve_tenant(request)
        except Exception:  # noqa: BLE001
            tenant = None
        if tenant is None:
            return await call_next(request)
        tid = tenant.get("tenant_id") or ""

        # (4) effective mode for THIS tenant+feature. The engine is fail-closed for an
        #     unknown/garbage mode (-> hidden); it degrades to 'on' only when PG is down
        #     (resting safety, by design). Admin tenants are NOT entitlement-gated here
        #     (they manage entitlements; their own plane is role-gated) -> always pass.
        if tenant.get("is_admin"):
            return await call_next(request)
        mode = _ent_mod.evaluate(tid, feature_key)
    except Exception as exc:  # noqa: BLE001
        # FAIL-CLOSED: an unexpected error AFTER a governed feature_key was resolved =
        # deny (404, no existence leak). If no key was resolved we already passed above.
        if feature_key:
            return JSONResponse({"error": "not_found"}, status_code=404)
        return await call_next(request)

    if mode == "hidden":
        return JSONResponse({"error": "not_found"}, status_code=404)
    if mode == "locked":
        return JSONResponse(
            {"error": "locked", "feature": feature_key, "upgrade": True},
            status_code=402)
    # mode == "on" (or any unexpected-but-non-blocking value already normalized by the
    # engine to hidden above) -> proceed.
    return await call_next(request)


# In-memory live concurrency per tenant (P0.7). Source of truth for active SIP calls.
ACTIVE_CALLS: dict = {}

# Serialize read-modify-write on shared JSON stores. run_job (dial loop) AND
# scheduler_loop both write leads/calls/suppression/retry concurrently -> without
# this lock concurrent writers lose updates. ALL async writers MUST use _awrite.
_STORE_LOCK = asyncio.Lock()

# P1 strangler: per-store MODE router (store.py). Declared = None HERE (B4) so the rewritten
# _read/_write shims are safe at IMPORT TIME (_migrate_to_admin() @504 and CALLS=_read() @683 run
# before store.init below). Assigned only AFTER store.init() returns; shims guard `_store is not None`.
_store = None

# WAVE A Unit3: timestamp (monotonic-ish epoch) of the last full vendor-sync pass.
# 0 -> forces a sync on the first scheduler tick after startup.
_LAST_VENDOR_SYNC = 0.0


# ---------- secret (HMAC signing key) ----------
def _load_secret() -> str:
    try:
        if SECRET_FILE.exists():
            s = SECRET_FILE.read_text(encoding="utf-8").strip()
            if s:
                return s
    except Exception:  # noqa: BLE001
        pass
    s = secrets.token_hex(32)
    try:
        VAR.mkdir(parents=True, exist_ok=True)
        SECRET_FILE.write_text(s, encoding="utf-8")
        try:
            os.chmod(SECRET_FILE, 0o600)
        except Exception:  # noqa: BLE001
            pass
    except Exception:  # noqa: BLE001
        pass
    return s


SECRET = _load_secret()


# ---------- tenant store ----------
def _hash_pw(password: str, salt: str) -> str:
    return hashlib.sha256((salt + ":" + (password or "")).encode("utf-8")).hexdigest()


def _read_tenants() -> list[dict]:
    return _read(TENANTS_FILE, [])


def _write_tenants(tenants: list[dict]):
    _write(TENANTS_FILE, tenants)


def _seed_admin() -> dict:
    """Ensure exactly one admin tenant exists; its password == legacy PW so panel login keeps working."""
    tenants = _read_tenants()
    admin = next((t for t in tenants if t.get("is_admin")), None)
    if admin:
        return admin
    salt = secrets.token_hex(8)
    admin = {
        "tenant_id": "admin",
        "email": "admin@famit.in",
        "salt": salt,
        "pass_hash": _hash_pw(PW, salt),
        "name": "Famit Admin",
        "is_admin": True,
        "role": "admin",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    tenants.insert(0, admin)
    _write_tenants(tenants)
    return admin


def _make_token(tenant_id: str) -> str:
    sig = hmac.new(SECRET.encode("utf-8"), tenant_id.encode("utf-8"), hashlib.sha256).hexdigest()
    return f"{tenant_id}.{sig}"


def _tenant_by_id(tenant_id: str) -> dict | None:
    return next((t for t in _read_tenants() if t.get("tenant_id") == tenant_id), None)


def _verify_token(token: str) -> dict | None:
    """token == tenant_id.hmac(tenant_id, SECRET). Returns the tenant dict or None."""
    if not token or "." not in token:
        return None
    tid, _, sig = token.partition(".")
    expect = hmac.new(SECRET.encode("utf-8"), tid.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(sig, expect):
        return None
    return _tenant_by_id(tid)


# NOTE: ADMIN_TENANT / ADMIN_ID are seeded after _read/_write are defined (see below).
ADMIN_TENANT: dict = {}
ADMIN_ID = "admin"


# ---------- auth ----------
def _extract_cred(request: Request) -> str:
    """Pull the raw credential string from Basic / Bearer / X-Auth."""
    h = request.headers.get("authorization", "")
    if h.startswith("Basic "):
        try:
            _, p = base64.b64decode(h[6:]).decode().split(":", 1)
            return p
        except Exception:  # noqa: BLE001
            return ""
    if h.startswith("Bearer "):
        return h[7:].strip()
    return request.headers.get("x-auth", "")


# ---------- Client lifecycle: status (active/suspended) + demo TTL (file-based) ----------
# Self-contained on tenants.json so it works on deployments WITHOUT the Postgres
# control layer. Admins are NEVER blocked (anti-lockout).
def _demo_remaining_s(tenant: dict | None) -> int | None:
    """Seconds left on a demo account's clock; None when not a demo. The clock starts
    at demo_started_at (falls back to created_at); demo_minutes is the budget."""
    if not tenant or not tenant.get("demo"):
        return None
    try:
        mins = int(tenant.get("demo_minutes") or 0)
    except (TypeError, ValueError):
        mins = 0
    started = tenant.get("demo_started_at") or tenant.get("created_at") or ""
    if mins <= 0 or not started:
        return 0
    try:
        start_dt = datetime.fromisoformat(started)
    except ValueError:
        return 0
    return max(0, int(mins * 60 - (datetime.now() - start_dt).total_seconds()))


def _client_blocked(tenant: dict | None) -> bool:
    """A non-admin client is locked out when explicitly suspended OR its demo TTL ran out."""
    if not tenant or tenant.get("is_admin"):
        return False
    if (tenant.get("status") or "active").strip().lower() == "suspended":
        return True
    rem = _demo_remaining_s(tenant)
    return rem is not None and rem <= 0


def resolve_tenant(request: Request) -> dict | None:
    """Resolve the calling tenant. Accepts (in order, all ADDITIVE):
       - a P0 JWT access token (Bearer/X-Auth) -> that tenant   [new, optional]
       - legacy bare password (== PW)  -> admin tenant (keeps panel login working)
       - signed token  tenant_id.hmac  -> that tenant
    Returns tenant dict or None. The legacy/hmac branch is gated by
    LEGACY_TOKEN_ENABLED (default True) so it can be turned off AFTER cutover
    without code changes; JWT is never gated.
    """
    cred = _extract_cred(request)
    if not cred:
        return None
    # 1) JWT access token (no-op/None if pyjwt missing or cred isn't a JWT).
    if _auth_mod is not None:
        try:
            t = _auth_mod.resolve_token(cred)
            if t:
                return None if _client_blocked(t) else t
        except Exception:  # noqa: BLE001 — never let auth module break the request
            pass
    # 2) legacy paths (unchanged), gated by the flag.
    if not LEGACY_TOKEN_ENABLED:
        return None
    if cred == PW:
        return _tenant_by_id(ADMIN_ID) or ADMIN_TENANT
    t = _verify_token(cred)
    return None if _client_blocked(t) else t


def authed(request: Request) -> bool:
    return resolve_tenant(request) is not None


def need_auth() -> Response:
    return Response(status_code=401)


# ════════════════════════════════════════════════════════════════════════════
# CONTROL LAYER — admin-plane gate (CL-B2 / plan C2 gate)
# ════════════════════════════════════════════════════════════════════════════
# THE #1 SECURITY FINDING (control-security.md §1.1): the legacy static password
# (CALLER_PASS / "FamitCall2026") is a permanent, un-revocable, un-audited admin
# BEARER token. It MUST NOT reach the /admin/* control plane. require_super_admin
# = is_admin AND auth_method != legacy_pw. This mirrors the existing /tenants gate
# (`if not t.get("is_admin"): 403`) but ADDS the legacy-password exclusion.
def _auth_method(request: Request) -> str:
    """Re-derive HOW the caller authenticated, using the SAME precedence as
    resolve_tenant (JWT first, then legacy bare-password, then signed hmac token).
    Returns 'jwt' | 'legacy_pw' | 'hmac' | 'none'. Non-mutating: resolve_tenant's
    return contract is unchanged — this is a separate, cheap re-classification so
    the admin gate can exclude the static password without touching every reader."""
    cred = _extract_cred(request)
    if not cred:
        return "none"
    # 1) a valid Famit ACCESS JWT (revocable, short-TTL) — the strong path.
    if _auth_mod is not None:
        try:
            if _auth_mod.resolve_token(cred):
                return "jwt"
        except Exception:  # noqa: BLE001
            pass
    if not LEGACY_TOKEN_ENABLED:
        return "none"
    # 2) the legacy bare password -> admin tenant. THE excluded path.
    if cred == PW:
        return "legacy_pw"
    # 3) a signed tenant_id.hmac token (per-tenant, derived from var/secret).
    if _verify_token(cred) is not None:
        return "hmac"
    return "none"


def _is_super_admin(tenant: dict | None, request: Request) -> bool:
    """Phase 1 predicate (control-security.md §1.2): is_admin AND non-legacy auth.
    Phase 2 (when Logto lands): admin-org membership + 'manage_tenants' scope.
    The bare static password is REJECTED here even though it still authenticates
    vendor-grade routes during the transition (residual risk #1, flagged)."""
    if not tenant or not tenant.get("is_admin"):
        return False
    return _auth_method(request) != "legacy_pw"


def require_super_admin(request: Request) -> dict | JSONResponse:
    """The ONE centralized /admin/* gate. Returns the resolved admin tenant dict
    on success, or a Response (401/403) to return directly. Usage in a handler:
        t = require_super_admin(request)
        if isinstance(t, JSONResponse): return t
    Returns 403 (NOT 404) for a non-admin: the EXISTENCE of an admin plane is not a
    secret (control-security.md §1.4); only hidden FEATURES return 404. A vendor
    token, an unauthenticated caller, or the legacy password all fail here."""
    t = resolve_tenant(request)
    if t is None:
        return JSONResponse({"error": "unauthenticated"}, status_code=401)
    if not _is_super_admin(t, request):
        return _forbidden("super-admin required")
    return t


# ---------- helpers / stores ----------
def norm(n: str) -> str:
    d = re.sub(r"\D", "", n or "")
    if d.startswith("0"):
        d = d[1:]
    if len(d) == 10:
        d = "91" + d
    return "+" + d if len(d) >= 11 else ""


def _read_raw(path: Path, default):
    try:
        if path.exists():
            return json.loads(path.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        pass
    return default


def _write_raw(path: Path, data):
    try:
        VAR.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass


async def _awrite_raw(path: Path, data):
    """Lock-guarded write for shared stores written from concurrent async tasks
    (run_job + scheduler_loop). Prevents lost updates on leads/calls/suppression/retry."""
    async with _STORE_LOCK:
        _write_raw(path, data)


def _atomic_write_json(path: Path, data) -> None:
    """Atomic durable write of a JSON record: write to a temp sibling, fsync, then
    os.replace() (atomic rename on the same filesystem). A crash mid-write can never
    leave a torn/partial campaign file — the old file stays intact until the rename.
    Used for the campaign JSON mirror, which is AUTHORITATIVE + write-first for the
    earner (VOICE-BRAIN-MASTER-PLAN red-team fix #4)."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.tmp.{os.getpid()}.{uuid.uuid4().hex[:8]}")
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    try:
        with open(tmp, "w", encoding="utf-8") as fh:
            fh.write(payload)
            fh.flush()
            try:
                os.fsync(fh.fileno())
            except OSError:
                pass
        os.replace(tmp, path)
    finally:
        try:
            if tmp.exists():
                tmp.unlink()
        except OSError:
            pass


# ---- P1 strangler shims: route registered non-json stores through store.py; else byte-identical
#      pass-through to the *_raw bodies above (R3: json mode NEVER reserializes). _store guarded
#      `is not None` (B4) so import-time calls (504/683) are safe before store.init runs. ----
def _read(path: Path, default):
    if _store is not None and _store.mode_of(path) != "json":
        return _store.read(path, default)
    return _read_raw(path, default)


def _write(path: Path, data):
    if _store is not None and _store.mode_of(path) != "json":
        _store.write(path, data)
        return
    _write_raw(path, data)


async def _awrite(path: Path, data):
    if _store is not None and _store.mode_of(path) != "json":
        await _store.awrite(path, data)
        return
    await _awrite_raw(path, data)


# ---- P1: initialize the store router ONCE, here, BEFORE _migrate_to_admin() (@504) and
#      CALLS=_read() (@683) run at import. With STORE_MODES empty (default) every store stays json,
#      so the shims above are a transparent pass-through and behavior is byte-identical to pre-P1.
#      Import-safe: any failure leaves _store=None -> shims call the raw funcs (degrade-to-json). ----
class _StoreConfigShim:
    """Adapts caller's cfg_get into the .get(key, default) surface store.py/db.engine expect,
    so it works whether config.py imported or fell back to os.getenv."""
    @staticmethod
    def get(key, default=""):
        return cfg_get(key, default)

try:
    import store as _store_mod
    _store_mod.init(_read_raw, _write_raw, _awrite_raw, _STORE_LOCK, _StoreConfigShim)
    _store = _store_mod
except Exception:  # noqa: BLE001
    _store = None


# ---------- calling-window helpers (P0.1 / P0.5) ----------
def _in_window(fields: dict) -> tuple[bool, str]:
    """True if 'now' (IST) is within the campaign's calling window. Default 09:00-21:00."""
    fields = fields or {}
    s = fields.get("call_window_start") or "09:00"
    e = fields.get("call_window_end") or "21:00"
    now = now_ist().strftime("%H:%M")
    ok = s <= now <= e
    return ok, f"{s}-{e} IST"


def _clamp_to_window(dt: datetime, fields: dict) -> datetime:
    """Move dt into the calling window: before start -> start same day; after end -> start next day."""
    fields = fields or {}
    s = (fields.get("call_window_start") or "09:00").split(":")
    e = (fields.get("call_window_end") or "21:00").split(":")
    try:
        sh, sm = int(s[0]), int(s[1]); eh, em = int(e[0]), int(e[1])
    except Exception:  # noqa: BLE001
        sh, sm, eh, em = 9, 0, 21, 0
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=IST)
    start = dt.replace(hour=sh, minute=sm, second=0, microsecond=0)
    end = dt.replace(hour=eh, minute=em, second=0, microsecond=0)
    if dt < start:
        return start
    if dt > end:
        return (start + timedelta(days=1))
    return dt


# ---- seed admin + migrate legacy records to admin tenant (now that _read/_write exist) ----
ADMIN_TENANT = _seed_admin()
ADMIN_ID = ADMIN_TENANT["tenant_id"]


def _migrate_to_admin():
    """Backwards-compatible: any existing record with no tenant_id belongs to admin.
    Never loses data; only fills the missing field in place."""
    # campaigns
    try:
        for p in CAMPAIGN_DIR.glob("*.json"):
            d = json.loads(p.read_text(encoding="utf-8"))
            if not d.get("tenant_id"):
                d["tenant_id"] = ADMIN_ID
                p.write_text(json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass
    # leads
    leads = _read(LEADS_FILE, [])
    changed = False
    for x in leads:
        if not x.get("tenant_id"):
            x["tenant_id"] = ADMIN_ID
            changed = True
    if changed:
        _write(LEADS_FILE, leads)
    # calls
    calls = _read(CALLS_FILE, [])
    changed = False
    for c in calls:
        if not c.get("tenant_id"):
            c["tenant_id"] = ADMIN_ID
            changed = True
    if changed:
        _write(CALLS_FILE, calls)


_migrate_to_admin()


def _migrate_tenant_limits():
    """P0.7: backfill plan limits on existing tenants with generous defaults (admin = high)."""
    tenants = _read_tenants()
    changed = False
    for x in tenants:
        is_admin = bool(x.get("is_admin"))
        if "max_concurrency" not in x:
            x["max_concurrency"] = 20 if is_admin else 3; changed = True
        if "daily_call_cap" not in x:
            x["daily_call_cap"] = 100000 if is_admin else 500; changed = True
        if "monthly_minutes_cap" not in x:
            x["monthly_minutes_cap"] = 1000000 if is_admin else 5000; changed = True
    if changed:
        _write_tenants(tenants)


_migrate_tenant_limits()


# ---------- WAVE3 Unit 1: RBAC ----------
# Roles: admin | manager | agent.
#   admin   = full control + tenant management + limits + billing config.
#   manager = run campaigns, manage campaigns/leads/webhooks within own tenant.
#   agent   = read-only (view calls/leads/analytics); cannot /run, create or delete.
ROLES = ("admin", "manager", "agent")


def _role_of(tenant: dict) -> str:
    """Resolve a tenant/user's role. Backwards-compatible migration:
    explicit `role` wins; else is_admin -> admin, otherwise -> manager."""
    if not tenant:
        return "agent"
    r = (tenant.get("role") or "").strip().lower()
    if r in ROLES:
        return r
    return "admin" if tenant.get("is_admin") else "manager"


def _migrate_tenant_roles():
    """Backfill `role` on existing tenants so nothing breaks. admin tenants -> admin,
    all other existing users -> manager (they could already run campaigns)."""
    tenants = _read_tenants()
    changed = False
    for x in tenants:
        if not (x.get("role") or "").strip().lower() in ROLES:
            x["role"] = "admin" if x.get("is_admin") else "manager"
            changed = True
    if changed:
        _write_tenants(tenants)


_migrate_tenant_roles()


def can(tenant: dict, action: str) -> bool:
    """Lightweight permission check.
    action: 'manage_tenants' (admin only), 'write' (admin/manager: run/create/delete
    campaigns+leads+webhooks+whatsapp), 'read' (everyone authed)."""
    role = _role_of(tenant)
    if action == "manage_tenants":
        return role == "admin"
    if action == "write":
        return role in ("admin", "manager")
    return True  # read


def _forbidden(msg: str = "insufficient permissions") -> JSONResponse:
    return JSONResponse({"error": msg}, status_code=403)


# ---------- P0: BOLA ownership guard (defensive depth) ----------
def _owns(tenant: dict | None, obj: dict | None) -> bool:
    """True if `tenant` may access `obj`. Admin sees all; otherwise obj.tenant_id
    must equal the tenant's id. Legacy objects without tenant_id default to the
    admin tenant (ADMIN_ID), matching how the rest of the code scopes data."""
    if not tenant or obj is None:
        return False
    if tenant.get("is_admin"):
        return True
    return obj.get("tenant_id", ADMIN_ID) == tenant.get("tenant_id")


def require_object(tenant: dict | None, obj: dict | None,
                   not_found: bool = True) -> JSONResponse | None:
    """Central BOLA assert for per-ID handlers. Returns an error Response if the
    tenant does NOT own `obj` (or it's missing), else None so the caller proceeds.

    Defense-in-depth: existing handlers already fetch via tenant-scoped readers
    (get_campaign_for / calls_for / _leads_for) that return None cross-tenant.
    This makes the ownership check explicit and uniform wherever a raw object is
    in hand, so a future handler that forgets to scope its read still can't leak
    another tenant's object. By default returns 404 (don't reveal existence of
    other tenants' objects); pass not_found=False for an explicit 403.
    """
    if obj is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    if _owns(tenant, obj):
        return None
    if not_found:
        return JSONResponse({"error": "not found"}, status_code=404)
    return JSONResponse({"error": "forbidden: not your object"}, status_code=403)


# ---------- P0: wire JWT auth module to the existing tenant store ----------
def _verify_password_for_auth(email: str, password: str) -> dict | None:
    """Password check used by the JWT /auth/login. Mirrors the legacy /login
    semantics EXACTLY so both stay in lock-step:
      - bare password == PW (no email) -> admin tenant
      - email + matching salted hash    -> that tenant
      - admin email + legacy PW         -> admin tenant
    Returns the tenant dict or None.
    """
    email = (email or "").strip().lower()
    if not email and password == PW:
        return _tenant_by_id(ADMIN_ID) or ADMIN_TENANT
    if email:
        t = next((x for x in _read_tenants() if (x.get("email") or "").lower() == email), None)
        if t and t.get("pass_hash") == _hash_pw(password, t.get("salt", "")):
            return t
        if t and t.get("is_admin") and password == PW:
            return t
    return None


# Initialise the JWT module (reuses the SAME var/secret as the hmac tokens, so no
# new secret to provision and tokens survive restart). Degrades to NO-OP if pyjwt
# is unavailable — legacy auth is untouched either way.
AUTH_JWT_READY = False
if _auth_mod is not None:
    try:
        AUTH_JWT_READY = _auth_mod.init(
            secret=SECRET,
            refresh_file=VAR / "refresh_tokens.json",
            tenant_by_id=_tenant_by_id,
            verify_password=_verify_password_for_auth,
            role_of=_role_of,
        )
    except Exception:  # noqa: BLE001
        AUTH_JWT_READY = False

# Initialise the append-only audit log.
if _audit_mod is not None:
    try:
        _audit_mod.init(VAR / "audit_log.jsonl")
    except Exception:  # noqa: BLE001
        pass

# Initialise the system error/event log (shared /data volume; the voice agent records to the
# SAME file so the panel surfaces backend + call failures together). Best-effort, but we now
# CAPTURE the readiness result: a False here means the volume was unwritable, so every record()
# would be a silent no-op (the dormant-logging trap). We surface it on stderr + /admin/logs/health
# so the operator isn't left wondering why nothing shows up.
LOG_READY = False
if _log_mod is not None:
    try:
        LOG_READY = bool(_log_mod.init(VAR / "system_events.jsonl"))
        if not LOG_READY:
            import sys as _sys
            print(f"[logging_service] init FAILED for {VAR / 'system_events.jsonl'} — "
                  f"System Logs will be a silent no-op (check FAMIT_VAR volume is writable)",
                  file=_sys.stderr, flush=True)
    except Exception:  # noqa: BLE001
        LOG_READY = False

# F4 Action Firewall: reuse the SAME var/secret (SECRET) the JWT/hmac path uses, + a pins store.
# Degrades to pass-through (no gating) if pyjwt absent or no secret. wallet needs NO init (it rides
# db.engine, which store.init() already wires at startup).
FIREWALL_READY = False
if _firewall_mod is not None:
    try:
        FIREWALL_READY = _firewall_mod.init(secret=SECRET, pin_file=VAR / "pins.json")
    except Exception:  # noqa: BLE001
        FIREWALL_READY = False

# CONTROL LAYER: load the entitlement catalog/plans seed + best-effort apply the additive control
# schema. NEVER raises (degrades to all-default 'on' when PG/seed absent -> resting byte-identical).
# This only loads the engine; NOTHING is enforced unless CONTROL_ENABLED is on.
CONTROL_READY = False
if _ent_mod is not None:
    try:
        CONTROL_READY = _ent_mod.init()
    except Exception:  # noqa: BLE001
        CONTROL_READY = False


def _client_ip(request: Request) -> str:
    """Best-effort client IP (honours X-Forwarded-For from nginx)."""
    try:
        xff = request.headers.get("x-forwarded-for", "")
        if xff:
            return xff.split(",")[0].strip()
        return request.client.host if request.client else ""
    except Exception:  # noqa: BLE001
        return ""


def _audit(request: Request, tenant: dict | None, action: str,
           object_type: str = "", object_id: str = "",
           channel: str = "api", meta: dict | None = None) -> None:
    """Tiny wrapper so mutating endpoints can log in one line. Best-effort."""
    if _audit_mod is None:
        return
    try:
        tid = (tenant or {}).get("tenant_id", "")
        _audit_mod.record(actor=tid, action=action, object_type=object_type,
                          object_id=object_id, ip=_client_ip(request),
                          channel=channel, tenant_id=tid,
                          actor_role=_role_of(tenant) if tenant else "",
                          meta=meta)
    except Exception:  # noqa: BLE001
        pass


def _log_event(level: str, source: str, message: str, *, request: "Request | None" = None,
               tenant: dict | None = None, call_id: str = "", error_type: str = "",
               context: dict | None = None) -> None:
    """One-line structured event recorder for the System Logs panel (best-effort, like
    _audit). No-op when the logging module is unavailable; never raises."""
    if _log_mod is None:
        return
    try:
        tid = (tenant or {}).get("tenant_id", "") if isinstance(tenant, dict) else ""
        ctx = dict(context or {})
        if request is not None:
            try:
                ctx.setdefault("ip", _client_ip(request))
            except Exception:  # noqa: BLE001
                pass
        _log_mod.record(level, source, message, tenant_id=tid, call_id=call_id,
                        error_type=error_type, context=ctx)
    except Exception:  # noqa: BLE001
        pass


CALLS: list = _read(CALLS_FILE, [])


def parse_leads(text: str, csv_bytes: bytes | None) -> list[dict]:
    leads: list[dict] = []

    def add(name: str, raw: str):
        nn = norm(raw)
        if nn:
            leads.append({"name": (name or "").strip(), "num": nn})

    for line in (text or "").splitlines():
        if not line.strip():
            continue
        parts = re.split(r"[,\t]", line)
        phone, name = "", ""
        for p in parts:
            if re.search(r"\d{8,}", p):
                phone = p
            elif p.strip():
                name = name or p.strip()
        add(name, phone or line)
    if csv_bytes:
        try:
            for r in csv.reader(io.StringIO(csv_bytes.decode("utf-8", errors="ignore"))):
                phone, name = "", ""
                for cell in r:
                    if re.search(r"\d{8,}", cell or ""):
                        phone = cell
                    elif (cell or "").strip() and not name and not re.fullmatch(r"[\d\s\-+]+", cell or ""):
                        name = cell.strip()
                if phone:
                    add(name, phone)
        except Exception:  # noqa: BLE001
            pass
    seen, out = set(), []
    for x in leads:
        if x["num"] not in seen:
            seen.add(x["num"])
            out.append(x)
    return out


def parse_xlsx(xlsx_bytes: bytes | None) -> list[dict]:
    """Parse an Excel (.xlsx/.xls) workbook into [{name,num}] using openpyxl
    (read-only, pure-python). Sniffs the phone column (>=8 digits) and the first
    non-numeric text cell as the name — mirrors parse_leads's column logic so the
    behaviour is identical whether a vendor uploads CSV or Excel. Never raises:
    on any failure (missing dep, corrupt file) returns [] so the caller degrades
    to the existing csv/text path. RC2."""
    out: list[dict] = []
    if not xlsx_bytes:
        return out
    try:
        import openpyxl  # lazy: only loaded when an xlsx is actually uploaded
    except Exception:  # noqa: BLE001 — dep absent -> behave like no xlsx given
        return out
    try:
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return out
    seen: set[str] = set()
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                phone, name = "", ""
                for cell in row:
                    if cell is None:
                        continue
                    # openpyxl gives ints/floats for numeric cells -> stringify.
                    s = str(cell).strip()
                    if not s:
                        continue
                    if re.search(r"\d{8,}", s):
                        if not phone:
                            phone = s
                    elif not name and not re.fullmatch(r"[\d\s\-+.eE]+", s):
                        name = s
                if not phone:
                    continue
                nn = norm(phone)
                if nn and nn not in seen:
                    seen.add(nn)
                    out.append({"name": name, "num": nn})
    except Exception:  # noqa: BLE001
        pass
    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass
    return out


def _is_xlsx_name(filename: str | None) -> bool:
    """True when an uploaded filename looks like an Excel workbook."""
    fn = (filename or "").lower().strip()
    return fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xls")


def parse_upload(text: str, upload: "UploadFile | None", file_bytes: bytes | None) -> list[dict]:
    """Unified parse: text lines + an uploaded file routed by filename
    (.xlsx/.xls -> openpyxl, else stdlib csv). Returns the de-duplicated
    [{name,num}] list. Back-compat: with no file this is exactly parse_leads(text).
    RC2."""
    if file_bytes and upload is not None and _is_xlsx_name(getattr(upload, "filename", "")):
        rows = parse_leads(text, None) + parse_xlsx(file_bytes)
    else:
        rows = parse_leads(text, file_bytes)
    seen, out = set(), []
    for x in rows:
        if x["num"] and x["num"] not in seen:
            seen.add(x["num"])
            out.append(x)
    return out


# ============================================================================
# W1 — VENDOR SCRIPT (lossless raw_script + sanitized script_meta)
# VOICE-BRAIN-MASTER-PLAN §3-A. Flag-gated, earner-safe, injection-guarded.
#   - raw_script is stored VERBATIM (no truncation/summarization) inside fields.
#   - script_meta holds only PARSED HINTS (sanitized); it is NEVER lossy over raw.
#   - when raw_script is present AND the inject flag is on, the lossy derived
#     projections (summary/usps/price/...) are gated OFF so they cannot reach the
#     live turn (red-team fix #5) — leaving raw_script as the single source.
# Legacy campaigns (no raw_script) are byte-identical → the golden oracle stays
# green with the flag OFF or ON.
# ============================================================================

# Default OFF -> byte-identical earner render. A per-campaign override
# (fields["vendor_script_inject"]) can opt a single campaign in without a global flip.
_VENDOR_SCRIPT_INJECT_ENV = (
    (os.getenv("VENDOR_SCRIPT_INJECT", "0") or "0").strip().lower() in ("1", "true", "yes", "on"))
# Hard ceiling on the stored verbatim script (DoS guard, NOT a lossy cap — generous
# vs the ~3-8K-token typical vendor script; full text round-trips well under this).
_RAW_SCRIPT_MAX_CHARS = int(os.getenv("RAW_SCRIPT_MAX_CHARS", "60000") or "60000")
_SCRIPT_META_STR_MAX = 600   # per sanitized persona-hint string
_SCRIPT_META_LIST_MAX = 24   # max items in a do/dont list


def _strip_zero_width(s: str) -> str:
    """Remove zero-width / BOM / directional-override chars used to smuggle
    instructions past a denylist or to hide a forged close-tag."""
    return "".join(ch for ch in s if ch not in (
        "​", "‌", "‍", "⁠", "﻿",
        "‪", "‫", "‬", "‭", "‮",
        "⁦", "⁧", "⁨", "⁩"))


def _clean_text(s, *, max_chars: int) -> str:
    """NFKC-normalize, strip zero-width/control chars, clamp length. Used for any
    free-form string that originates from a vendor or an LLM extraction (untrusted)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = _strip_zero_width(s)
    # drop control chars except tab/newline/carriage-return (keep script formatting)
    s = "".join(ch for ch in s if ch in ("\t", "\n", "\r") or ord(ch) >= 0x20)
    if max_chars and len(s) > max_chars:
        s = s[:max_chars]
    return s


def _escape_vendor_script(text: str) -> str:
    """Neutralize any forged close-tag inside a vendor script BEFORE it is fenced
    in <vendor_script>…</vendor_script> at render time. Without this a vendor could
    inject `</vendor_script> SYSTEM: ...` and break out of the data fence (red-team
    fix: escape the close-tag). Case/space-insensitive; also covers <vendor_data.
    Idempotent enough for round-trip (we escape '<' of the tag to a fullwidth '＜')."""
    if not text:
        return text
    # match the OPENING of any vendor_* tag (open or close form), e.g.
    # </vendor_script>, < vendor_script, </vendor_data ...  -> defang the leading '<'
    return re.sub(r"<(\s*/?\s*vendor_(?:script|data)\b)",
                  lambda m: "＜" + m.group(1), text, flags=re.IGNORECASE)


def _sanitize_script_meta(meta) -> dict:
    """Clamp the optional parsed persona/tone hints. NEVER lossy over raw_script —
    this is a convenience projection only; the verbatim truth is fields['raw_script'].
    Unknown keys are dropped; strings cleaned+clamped; do/dont coerced to short lists."""
    if not isinstance(meta, dict):
        return {}
    out: dict = {}
    for k in ("tone", "greeting", "persona", "language", "style"):
        v = meta.get(k)
        if v not in (None, ""):
            out[k] = _clean_text(v, max_chars=_SCRIPT_META_STR_MAX)
    for k in ("do", "dont", "do_list", "dont_list"):
        v = meta.get(k)
        if isinstance(v, str):
            v = [s for s in re.split(r"[\n;]+", v) if s.strip()]
        if isinstance(v, list):
            items = [_clean_text(s, max_chars=_SCRIPT_META_STR_MAX)
                     for s in v if str(s).strip()][:_SCRIPT_META_LIST_MAX]
            if items:
                out[k] = items
    return out


def _coerce_vendor_script(out: dict) -> None:
    """Mutate `out` in place: store raw_script VERBATIM (escaped close-tags only,
    no truncation/summarization beyond a generous DoS ceiling) + a sanitized
    script_meta + a clamped trust_tier + the per-campaign inject override. When the
    script is authoritative (present AND inject on), blank the lossy derived
    projections so they cannot reach the live turn. Idempotent: re-coercing an
    already-coerced record round-trips byte-equal on raw_script."""
    raw = out.get("raw_script")
    # raw_script: clean (NFKC + zero-width strip + control strip) + escape vendor
    # close-tags + clamp to the DoS ceiling. This is the ONLY processing — the full
    # text is preserved verbatim (no lossy summarization).
    if raw in (None, ""):
        out["raw_script"] = ""
    else:
        raw = _clean_text(raw, max_chars=_RAW_SCRIPT_MAX_CHARS)
        out["raw_script"] = _escape_vendor_script(raw)
    # sanitized parsed hints (never authoritative over raw)
    out["script_meta"] = _sanitize_script_meta(out.get("script_meta"))
    # trust tier: sandbox (default) | trusted. Sandbox scripts are INBOUND-ONLY until
    # a super-admin promotes to trusted (hard precondition for any earner exposure).
    tt = str(out.get("trust_tier", "") or "").strip().lower()
    out["trust_tier"] = tt if tt in ("sandbox", "trusted") else "sandbox"
    # per-campaign opt-in (effective = global env OR this flag)
    out["vendor_script_inject"] = bool(out.get("vendor_script_inject", False))
    # red-team fix #5: when the script is authoritative, gate OFF the lossy derived
    # projections (from the truncated extract_fields) so the agent has ONE source —
    # the verbatim script — not script-plus-stale-compression. We blank them at the
    # data layer (so build_system_prompt renders the same empty-list path that legacy
    # empty campaigns already render, proven byte-stable by the golden oracle).
    script_present = bool(out.get("raw_script"))
    inject_on = _VENDOR_SCRIPT_INJECT_ENV or bool(out.get("vendor_script_inject"))
    if script_present and inject_on:
        for k in ("product_summary", "location", "price_offer"):
            out[k] = ""
        for k in ("usps", "talking_points", "qualifying_questions"):
            out[k] = []
        out["objections"] = []
        out["_derived_suppressed"] = True
    else:
        out["_derived_suppressed"] = False


def _sanitize_extracted(out: dict) -> dict:
    """Schema-validate + value-clamp the LLM-returned extract_fields output — the
    OPEN INJECTION SINK (a malicious brief can make the model emit hostile field
    values that later render into the live prompt). Keep only known keys; clean all
    strings; coerce list shapes. Defensive: never raises. (raw_script is NOT produced
    by extract_fields; it is authored separately and coerced in _coerce_vendor_script.)"""
    if not isinstance(out, dict):
        return {"agent_name": "Riya", "company_name": "", "product_name": "",
                "product_summary": "", "language": "Hinglish"}
    safe: dict = {}
    for k in ("company_name", "agent_name", "product_name", "product_summary",
              "location", "price_offer", "language"):
        v = out.get(k)
        if v is not None:
            safe[k] = _clean_text(v, max_chars=4000 if k == "product_summary" else 400)
    for k in ("usps", "talking_points", "qualifying_questions"):
        v = out.get(k)
        if isinstance(v, str):
            v = [s for s in re.split(r"[\n;]+", v) if s.strip()]
        if isinstance(v, list):
            safe[k] = [_clean_text(s, max_chars=400) for s in v if str(s).strip()][:40]
    obj = out.get("objections")
    if isinstance(obj, list):
        norm = []
        for o in obj:
            if isinstance(o, dict):
                norm.append({"q": _clean_text(o.get("q", ""), max_chars=400),
                             "a": _clean_text(o.get("a", ""), max_chars=600)})
        safe["objections"] = norm[:40]
    # preserve a benign error marker if extract_fields fell back
    if out.get("_error"):
        safe["_error"] = _clean_text(out.get("_error"), max_chars=200)
    safe.setdefault("agent_name", "Riya")
    safe.setdefault("language", "Hinglish")
    return safe


def extract_fields(brief: str) -> dict:
    # Script generation uses the BEST-VALUE model via OpenRouter (cheap, strong) when
    # OPENROUTER_API_KEY is set, falling back to Groq. The system prompt is kept SHORT
    # (low input tokens / cost) but quality-directed: talking_points + objection answers
    # must read like a warm, natural HUMAN telecaller — light emotion + proper punctuation.
    sysmsg = (
        "You are a senior tele-sales script writer. Convert the brief into ONE JSON object with EXACTLY these keys: "
        "company_name, agent_name, product_name, product_summary, location, price_offer, "
        "usps (array of short strings), talking_points (array of short strings), "
        "objections (array of {q,a}), qualifying_questions (array of short strings), language. "
        "Write talking_points and objection answers as warm, natural SPOKEN Hinglish that a real human "
        "telecaller would say — short sentences, a little emotion, proper punctuation (commas, a dash '—' "
        "for a pause, '?' for questions). No exclamation marks. No heavy/bookish words. Keep every string short. "
        "agent_name default 'Riya', language default 'Hinglish'. Use sensible short defaults if unknown."
    )
    or_key = (cfg_get("OPENROUTER_API_KEY", "") or "").strip()
    or_model = cfg_get("OPENROUTER_EXTRACT_MODEL", "google/gemini-2.0-flash-001")
    # (endpoint, bearer, model) — try OpenRouter first when configured, else Groq.
    routes = []
    if or_key:
        routes.append(("https://openrouter.ai/api/v1/chat/completions", or_key, or_model))
    routes.append(("https://api.groq.com/openai/v1/chat/completions", GROQ_KEY, GROQ_MODEL))
    last_exc: Exception | None = None
    for url, key, model in routes:
        try:
            r = httpx.post(
                url,
                headers={"Authorization": "Bearer " + key,
                         "HTTP-Referer": "https://haptica.famit.in", "X-Title": "Haptica AI"},
                json={"model": model, "temperature": 0.4, "max_tokens": 1100,
                      "response_format": {"type": "json_object"},
                      "messages": [{"role": "system", "content": sysmsg},
                                   {"role": "user", "content": brief[:8000]}]},
                timeout=45,
            )
            content = r.json()["choices"][0]["message"]["content"]
            m = re.search(r"\{.*\}", content, re.DOTALL)
            # SANDBOX the LLM output: schema-validate + value-clamp the open injection
            # sink before it can render into the live prompt (red-team, in-wave).
            return _sanitize_extracted(json.loads(m.group(0) if m else content))
        except Exception as exc:  # noqa: BLE001 — try the next route
            last_exc = exc
            continue
    if last_exc is not None:  # every LLM route failed -> surface it (was silently swallowed)
        _log_event("warning", "llm", f"extract_call_meta: all LLM routes failed: {last_exc!r}"[:300],
                   error_type=type(last_exc).__name__, context={"stage": "extract_fields"})
    return _sanitize_extracted(
        {"_error": repr(last_exc)[:200], "agent_name": "Riya", "company_name": "",
         "product_name": "", "product_summary": brief[:400], "language": "Hinglish"})


def _groq_chat(messages: list, max_tokens: int = 300, temperature: float = 0.5,
               timeout: int = 20) -> str:
    """Reuse the Groq client/key for a one-shot chat completion. Returns the
    assistant text, or "" on any failure (never raises). Used by WhatsApp AI drafts."""
    try:
        r = httpx.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": "Bearer " + GROQ_KEY},
            json={"model": GROQ_MODEL, "temperature": temperature,
                  "max_tokens": max_tokens, "messages": messages},
            timeout=timeout,
        )
        return (r.json()["choices"][0]["message"]["content"] or "").strip()
    except Exception:  # noqa: BLE001
        return ""


def save_campaign(fields: dict, tenant_id: str) -> dict:
    CAMPAIGN_DIR.mkdir(parents=True, exist_ok=True)
    cid = uuid.uuid4().hex[:10]
    rec = {"id": cid, "tenant_id": tenant_id,
           "name": (fields.get("product_name") or fields.get("company_name") or cid),
           "company": fields.get("company_name", ""), "product": fields.get("product_name", ""),
           "status": "ready", "created_at": datetime.now().isoformat(timespec="seconds"),
           "fields": fields, "system_prompt": build_system_prompt(fields)}
    # ATOMIC write (red-team fix #4): the JSON file mirror is AUTHORITATIVE + write-first
    # for the earner; temp+rename so a crash mid-write never leaves a torn campaign file.
    _atomic_write_json(CAMPAIGN_DIR / f"{cid}.json", rec)
    # P1 DUAL MIRROR (best-effort, additive): campaigns are written per-id (bypassing _write), so the
    # store seam can't see them — mirror this one record to PG explicitly. No-op unless campaigns is
    # flipped to dual in STORE_MODES; off-loop; swallows all errors (must NOT break campaign create).
    try:
        if _store is not None:
            _store.mirror_campaign_upsert(rec)
    except Exception:  # noqa: BLE001
        pass
    # W2: publish a cache-invalidate (version bump) so the inbound context cache reloads this new
    # campaign immediately. Flag-gated (CTX_CACHE); no-op when off; never breaks the create.
    _publish_ctx_invalidate(tenant_id, cid)
    return rec


def get_campaign(cid: str) -> dict | None:
    cid = "".join(ch for ch in (cid or "") if ch.isalnum() or ch in "-_")
    return _read(CAMPAIGN_DIR / f"{cid}.json", None) if cid else None


def get_campaign_for(cid: str, tenant: dict) -> dict | None:
    """Load a campaign only if the tenant owns it (admin sees all)."""
    d = get_campaign(cid)
    if not d:
        return None
    if tenant.get("is_admin") or d.get("tenant_id", ADMIN_ID) == tenant["tenant_id"]:
        return d
    return None


def list_campaigns(tenant: dict | None = None) -> list[dict]:
    out = []
    try:
        for p in sorted(CAMPAIGN_DIR.glob("*.json"), key=lambda x: x.stat().st_mtime, reverse=True):
            d = json.loads(p.read_text(encoding="utf-8"))
            tid = d.get("tenant_id", ADMIN_ID)
            if tenant is not None and not tenant.get("is_admin") and tid != tenant["tenant_id"]:
                continue
            out.append({"id": d["id"], "name": d.get("name", d["id"]),
                        "company": d.get("company", ""), "product": d.get("product", ""),
                        "status": d.get("status", "ready"), "created_at": d.get("created_at", ""),
                        "voice_id": (d.get("fields") or {}).get("voice_id", ""),
                        "tenant_id": tid})
    except Exception:  # noqa: BLE001
        pass
    return out


def record_call(rec: dict):
    CALLS.insert(0, rec)
    del CALLS[2000:]
    _write(CALLS_FILE, CALLS)


def calls_for(tenant: dict) -> list[dict]:
    if tenant.get("is_admin"):
        return CALLS
    return [c for c in CALLS if c.get("tenant_id", ADMIN_ID) == tenant["tenant_id"]]


# ---------- P0.2 suppression / DND ----------
def _suppressed_set(tenant_id: str) -> set:
    return {x["phone"] for x in _read(SUPPRESSION_FILE, [])
            if x.get("tenant_id") == tenant_id and x.get("phone")}


async def _add_suppression(tenant_id: str, phone: str, reason: str, source: str = ""):
    """Add a number to a tenant's DND list (idempotent). Lock-guarded shared-store write."""
    phone = norm(phone)
    if not phone:
        return
    async with _STORE_LOCK:
        store = _read(SUPPRESSION_FILE, [])
        if not any(x.get("phone") == phone and x.get("tenant_id") == tenant_id for x in store):
            store.append({"tenant_id": tenant_id, "phone": phone, "reason": reason,
                          "source": source, "added_at": datetime.now().isoformat(timespec="seconds")})
            _write(SUPPRESSION_FILE, store)


async def _flip_lead_status(tenant_id: str, phone: str, status: str):
    """Mark a tenant's lead (matched by normalized phone) with a status. Lock-guarded."""
    phone = norm(phone)
    if not phone:
        return
    async with _STORE_LOCK:
        store = _read(LEADS_FILE, [])
        changed = False
        for x in store:
            if x.get("tenant_id", ADMIN_ID) == tenant_id and norm(x.get("phone", "")) == phone:
                if x.get("status") != status:
                    x["status"] = status; changed = True
        if changed:
            _write(LEADS_FILE, store)


# ---------- P0.4 outcome classification (answering machine / no-answer) ----------
def _classify_outcome(rec: dict, tr: dict) -> str:
    turns = tr.get("turns") or []
    user_turns = [x for x in turns if x.get("role") == "user" and (x.get("content") or "").strip()]
    dur = rec.get("duration_s", 0)
    if tr.get("amd_hint") == "no_user_audio" and not user_turns:
        return "no_answer" if dur < 8 else "voicemail"
    if not turns and dur < 8:
        return "no_answer"            # never connected / ring-out
    if len(user_turns) == 0 and dur < 25:
        return "voicemail"            # agent spoke, human never did -> machine/VM
    if len(user_turns) == 0:
        return "no_human"             # connected, no human turns (likely VM/IVR)
    return tr.get("outcome") or "answered"


# DND-GUARD: the LLM `opt_out` flag (from agent._summarize) can MIS-FIRE on short/near-empty calls
# (a 24s/1-turn "answered" call wrongly DND-ing a fresh lead — exactly the bug seen live). We only
# auto-suppress when the CALLER literally said an explicit removal phrase, so a hallucinated flag can
# never blacklist a good number. Compliance-safe: a real "remove me / number hata do" still suppresses.
# Used as an AND-gate WITH the LLM flag — the LLM handles nuance ("baad mein call karna" = callback,
# not opt-out), the keywords guard against hallucination. Phrases cover EN + Hinglish(latin) + Devanagari.
_OPTOUT_PHRASES = (
    "remove my number", "remove me", "do not call", "don't call", "dont call", "stop calling",
    "take me off", "unsubscribe", "never call", "do not contact", "don't contact",
    "number hata", "hata do", "hata dijiye", "hata den", "dobara call mat", "dubara call mat",
    "dobara phone mat", "mat karo call", "call mat kar", "phone mat kar", "mat bulao",
    "list se hata", "block kar", "pareshan mat",
    "नंबर हटा", "हटा दो", "हटा दीजिए", "हटा दें", "दोबारा call मत", "दोबारा कॉल मत", "दुबारा कॉल मत",
    "call मत कर", "कॉल मत कर", "फोन मत कर", "मत बुलाओ", "list से हटा", "परेशान मत", "block कर",
)


def _caller_opted_out(tr: dict) -> bool:
    """True ONLY if a caller (role='user') turn literally contains an explicit opt-out/removal phrase.
    Deterministic gate over the LLM opt_out flag so a mis-classification can't DND a fresh lead."""
    try:
        turns = tr.get("turns") or []
        text = " ".join((x.get("content") or "") for x in turns if x.get("role") == "user").lower()
        return bool(text.strip()) and any(p in text for p in _OPTOUT_PHRASES)
    except Exception:  # noqa: BLE001 — a guard error must never crash finalize; treat as "no opt-out"
        return False


_REAL_CONVO = ("answered", "interested", "not_interested", "callback", "opt_out")


# ---------- P0.6 lead scoring ----------
async def _update_lead_after_call(tenant_id: str, phone: str, score, outcome: str,
                                  call_at: str = ""):
    """Update a lead from a finalized call. Never regresses: keeps the MOST RECENT call's
    outcome and the HIGHEST interest score seen (so a later voicemail can't wipe an earlier
    'interested/80'). call_at = the call's started_at (chronology); falls back to now."""
    phone = norm(phone)
    if not phone:
        return
    call_at = call_at or datetime.now().isoformat(timespec="seconds")
    try:
        sc = int(score or 0)
    except Exception:  # noqa: BLE001
        sc = 0
    async with _STORE_LOCK:
        store = _read(LEADS_FILE, [])
        changed = False
        for x in store:
            if x.get("tenant_id", ADMIN_ID) == tenant_id and norm(x.get("phone", "")) == phone:
                prev_at = x.get("last_call_at", "")
                if call_at >= prev_at:                       # only the most recent call sets outcome
                    x["last_outcome"] = outcome or ""
                    x["last_call_at"] = call_at
                best = max(int(x.get("score", 0) or 0), sc)  # keep best interest ever seen
                x["score"] = best
                x["hot"] = best >= 70
                changed = True
        if changed:
            _write(LEADS_FILE, store)


async def _w7_lifecycle_after_call(tenant_id: str, rec: dict, tr: dict, outcome: str, score: int):
    """W7 (LEAD_LIFECYCLE_ENABLED): after each call, classify the lead lifecycle
    (hot/warm/cold/dead) via the deterministic FSM + stamp an AI summary + next action
    onto the lead record so the CRM updates after every call. Pure-Python classify (no PG
    dependency); NEVER raises into finalize. Additive: only WRITES new fields
    (lifecycle/ai_summary/next_action/conversion_prob), never regresses existing outcome/score."""
    if _vk_classify_lifecycle is None or _VKLifecycle is None:
        return
    phone = norm(rec.get("phone", ""))
    if not phone:
        return
    # Map this call's observed signals onto the FSM inputs.
    booked = bool(tr.get("booked") or tr.get("appointment") or outcome in ("booked", "converted"))
    handoff = score >= 70
    dead = bool(tr.get("opt_out") or outcome == "opt_out")
    had_commitment = bool(tr.get("callback_at") or tr.get("commitment") or outcome == "callback")
    had_objection = bool(tr.get("objection") or tr.get("objections"))
    engaged = outcome in _REAL_CONVO or bool(tr.get("summary"))
    try:
        prior_val = _VKLifecycle.NEW
    except Exception:  # noqa: BLE001
        prior_val = None
    async with _STORE_LOCK:
        store = _read(LEADS_FILE, [])
        changed = False
        for x in store:
            if x.get("tenant_id", ADMIN_ID) == tenant_id and norm(x.get("phone", "")) == phone:
                try:
                    prior = _VKLifecycle(x.get("lifecycle", "new")) if x.get("lifecycle") else prior_val
                except Exception:  # noqa: BLE001
                    prior = prior_val
                new_lc = _vk_classify_lifecycle(
                    prior=prior, booked=booked, handoff=handoff, dead=dead,
                    had_objection=had_objection, had_commitment=had_commitment, engaged=engaged)
                x["lifecycle"] = getattr(new_lc, "value", str(new_lc))
                summ = (tr.get("summary", "") or "")[:300]
                if summ:
                    x["ai_summary"] = summ
                na = (tr.get("next_action", "") or "")
                if na:
                    x["next_action"] = na
                x["conversion_prob"] = round((int(score or 0)) / 100.0, 3)
                x["lifecycle_at"] = datetime.now().isoformat(timespec="seconds")
                changed = True
        if changed:
            _write(LEADS_FILE, store)


# ---------- P0.7 usage metering ----------
def _tenant_usage(tenant_id: str, since_iso: str) -> dict:
    rows = [c for c in CALLS if c.get("tenant_id") == tenant_id and c.get("started_at", "") >= since_iso]
    return {"calls": len(rows),
            "minutes": round(sum(c.get("duration_s", 0) for c in rows) / 60, 1)}


def _today_iso() -> str:
    return now_ist().date().isoformat()


def _month_iso() -> str:
    return now_ist().strftime("%Y-%m")


# ---------- P0.5 retry / callback queue ----------
async def _enqueue_retry(tenant_id, campaign_id, name, phone, attempts, max_attempts,
                         next_at, reason):
    phone = norm(phone)
    if not phone:
        return
    async with _STORE_LOCK:
        store = _read(RETRY_FILE, [])
        existing = next((r for r in store if r.get("phone") == phone
                         and r.get("campaign_id") == campaign_id
                         and r.get("tenant_id") == tenant_id), None)
        if existing:
            existing["attempts"] = attempts
            existing["next_attempt_at"] = next_at
            existing["reason"] = reason
            existing["max_attempts"] = max_attempts
        else:
            store.append({"id": uuid.uuid4().hex[:10], "tenant_id": tenant_id,
                          "campaign_id": campaign_id, "name": name or "", "phone": phone,
                          "attempts": attempts, "max_attempts": max_attempts,
                          "next_attempt_at": next_at, "reason": reason,
                          "created_at": datetime.now().isoformat(timespec="seconds")})
        _write(RETRY_FILE, store)


async def _remove_retry(retry_id: str):
    async with _STORE_LOCK:
        store = _read(RETRY_FILE, [])
        store = [r for r in store if r.get("id") != retry_id]
        _write(RETRY_FILE, store)


# ---------- P1.A WhatsApp follow-up (fire-and-forget stub) ----------
async def _send_whatsapp(tenant_id, rec, outcome, camp_fields):
    """Optional per-campaign WhatsApp follow-up via a BSP. Never blocks/raises into the loop."""
    try:
        if not (camp_fields or {}).get("wa_enabled"):
            return
        api_url = os.getenv("WA_API_URL", "")
        api_key = os.getenv("WA_API_KEY", "")
        score = rec.get("interest", 0) or 0
        if outcome == "interested" or score >= 70:
            template = camp_fields.get("wa_template_qualified", "")
        elif outcome in ("no_answer", "voicemail", "no_human"):
            template = camp_fields.get("wa_template_noanswer", "")
        else:
            template = ""
        if not template:
            return
        status = "skipped_no_bsp"
        if api_url and api_key:
            try:
                async with httpx.AsyncClient(timeout=10) as cli:
                    resp = await cli.post(api_url, headers={"Authorization": "Bearer " + api_key},
                                          json={"to": rec.get("phone"), "template": template})
                    status = f"sent:{resp.status_code}"
            except Exception as exc:  # noqa: BLE001
                status = f"error:{repr(exc)[:60]}"
                _log_event("warning", "whatsapp",
                           f"WhatsApp template send failed: {exc!r}"[:300],
                           tenant={"tenant_id": tenant_id}, error_type=type(exc).__name__,
                           context={"phone": rec.get("phone"), "template": template})
        async with _STORE_LOCK:
            log = _read(WA_LOG_FILE, [])
            log.insert(0, {"tenant_id": tenant_id, "phone": rec.get("phone"), "template": template,
                           "status": status, "at": datetime.now().isoformat(timespec="seconds")})
            del log[2000:]
            _write(WA_LOG_FILE, log)
    except Exception:  # noqa: BLE001
        pass


# ---------- WAVE3 Unit5: WhatsApp send via clean module ----------
async def _wa_log(tenant_id: str, to: str, template: str, result: dict, kind: str = "manual"):
    async with _STORE_LOCK:
        log = _read(WA_LOG_FILE, [])
        log.insert(0, {"tenant_id": tenant_id, "phone": to, "template": template,
                       "kind": kind, "status": result.get("status", ""),
                       "ok": bool(result.get("ok")),
                       "at": datetime.now().isoformat(timespec="seconds")})
        del log[2000:]
        _write(WA_LOG_FILE, log)


# ============================================================================
# WAFX — approved Meta WhatsApp TEMPLATE registry + name resolver.
# Meta rejects (#132001 / "template name does not exist") any send whose template
# name is not registered+approved on the WABA. We keep a small allow-list of the
# names that ARE registered, map common internal aliases onto them, and refuse
# unregistered names (hot_lead_alert / benefit_focus / special_offer) with a CLEAR
# error instead of a generic failure. Override via WAFX_APPROVED_TEMPLATES (CSV) and
# WAFX_TEMPLATE_ALIASES ("alias=approved,alias2=approved2").
# ============================================================================
WAFX_APPROVED_TEMPLATES = {
    s.strip() for s in (os.getenv("WAFX_APPROVED_TEMPLATES",
                                  "post_call_followup,hello_world") or "").split(",")
    if s.strip()
}
def _wafx_alias_map() -> dict:
    m = {}
    raw = (os.getenv("WAFX_TEMPLATE_ALIASES", "") or "").strip()
    if raw:
        for pair in raw.split(","):
            if "=" in pair:
                a, _, b = pair.partition("=")
                if a.strip() and b.strip():
                    m[a.strip()] = b.strip()
    return m
def _wafx_resolve_template(name: str) -> dict:
    """Resolve a requested template name to an APPROVED one.
    Returns {"ok":True,"name":<approved>} or {"ok":False,"status":"template_not_registered",
    "requested":<name>,"approved":[...]}. Empty/whitespace name -> ok (caller validates)."""
    n = (name or "").strip()
    if not n:
        return {"ok": True, "name": n}
    if n in WAFX_APPROVED_TEMPLATES:
        return {"ok": True, "name": n}
    mapped = _wafx_alias_map().get(n)
    if mapped and mapped in WAFX_APPROVED_TEMPLATES:
        return {"ok": True, "name": mapped, "mapped_from": n}
    return {"ok": False, "status": "template_not_registered", "requested": n,
            "approved": sorted(WAFX_APPROVED_TEMPLATES)}


async def _wa_send(tenant_id: str, to: str, template: str, params, kind: str = "manual",
                   is_text: bool = False) -> dict:
    """Send via whatsapp.py (no-ops gracefully if unconfigured). Logs every attempt.

    is_text=True => `template` carries a RAW free-form text body (valid only inside the 24h
    customer-service window) and is routed to the Meta TEXT path, NOT the template path —
    else a free-form string is mis-sent as a template name (Graph #132001). Default False
    preserves the template-name behaviour for every existing caller (auto-followup etc.)."""
    to = norm(to) or (to or "")
    if wa_mod is None:
        result = {"ok": False, "status": "module_unavailable"}
    elif is_text and hasattr(wa_mod, "send_whatsapp_text_async"):
        result = await wa_mod.send_whatsapp_text_async(to, template)
    else:
        # WAFX: template sends must use a REGISTERED+APPROVED Meta template name.
        # Map known aliases; refuse unregistered names with a clear error (not a
        # generic failure / "not connected").
        _res = _wafx_resolve_template(template)
        if not _res.get("ok"):
            result = {"ok": False, "status": "template_not_registered",
                      "to": to, "provider": "meta",
                      "meta_error": {"error_user_title": "Template not registered",
                                     "error_user_msg": ("WhatsApp template '" + (template or "")
                                                        + "' is not registered/approved on Meta. "
                                                        "Approved: " + ", ".join(_res.get("approved", []))),
                                     "requested": _res.get("requested", template),
                                     "approved": _res.get("approved", [])}}
            await _wa_log(tenant_id, to, template, result, kind=kind)
            return result
        template = _res.get("name", template)
        result = await wa_mod.send_whatsapp_async(to, template, params)
    await _wa_log(tenant_id, to, template, result, kind=kind)
    return result


# ============================================================================
# HUMAN HANDOFF — per-vendor handoff team list (Business Brain `handoff` block) +
# hot-lead WhatsApp notify. ADDITIVE + ISOLATED: lives on the Brain JSON (no new
# table, no new auth), reuses the existing /brain auth + whatsapp.py sender.
# ============================================================================
# Approved Meta template for cold hot-lead alerts to the team (no open 24h window
# with team members). Registering it is a Meta-onboarding step (GAP-C1); until then
# the send no-ops gracefully (dormant) like every other WA path.
HOT_LEAD_ALERT_TEMPLATE = (os.getenv("HOT_LEAD_ALERT_TEMPLATE", "hot_lead_alert")
                           or "hot_lead_alert").strip()

import logging as _lg_handoff_mod
_lg_handoff = _lg_handoff_mod.getLogger("famit-caller.handoff")


def _handoff_get(tenant_id: str) -> list[dict]:
    """The vendor's handoff team list: [{phone, whatsapp, role, hours, priority}, ...].
    Read from the Business Brain `handoff` block (var/brain/<tenant>.json). [] when none.
    NEVER raises (a broken brain must not break a call)."""
    if not tenant_id or _brain_mod is None:
        return []
    try:
        prof = _brain_mod.get_profile(tenant_id) or {}
        hl = prof.get("handoff")
        if isinstance(hl, dict):                 # tolerate {"team":[...]} shape
            hl = hl.get("team") or hl.get("numbers") or []
        if not isinstance(hl, list):
            return []
        out = []
        for h in hl:
            if not isinstance(h, dict):
                continue
            ph = norm(str(h.get("phone", ""))) or str(h.get("phone", "")).strip()
            wa = norm(str(h.get("whatsapp", ""))) or str(h.get("whatsapp", "")).strip() or ph
            if not ph and not wa:
                continue
            out.append({"phone": ph, "whatsapp": wa,
                        "name": str(h.get("name", "") or ""),
                        "role": str(h.get("role", "") or ""),
                        "hours": str(h.get("hours", "") or ""),
                        "priority": int(h.get("priority", 99) or 99),
                        # additive: default-True so already-seeded entries (no `enabled`) keep working.
                        "enabled": (False if str(h.get("enabled", True)).strip().lower()
                                    in ("false", "0", "no", "off") else True)})
        out.sort(key=lambda x: x.get("priority", 99))
        return out
    except Exception as exc:  # noqa: BLE001
        _lg_handoff.warning("handoff_get failed tenant=%s: %r", tenant_id, exc)
        _log_event("warning", "handoff", f"handoff team lookup failed: {exc!r}"[:300],
                   tenant={"tenant_id": tenant_id}, error_type=type(exc).__name__)
        return []


def _handoff_set(tenant_id: str, team: list, actor: str = "system") -> list[dict]:
    """Replace the vendor's handoff team list on the Brain. Returns the stored list.
    Validates+normalises each entry; NEVER raises (returns prior list on failure)."""
    if not tenant_id or _brain_mod is None:
        return _handoff_get(tenant_id)
    clean = []
    for h in (team or []):
        if not isinstance(h, dict):
            continue
        ph = str(h.get("phone", "")).strip()
        wa = str(h.get("whatsapp", "")).strip()
        if not ph and not wa:
            continue
        clean.append({"phone": norm(ph) or ph,
                      "whatsapp": norm(wa) or wa or (norm(ph) or ph),
                      "name": str(h.get("name", "") or ""),
                      "role": str(h.get("role", "") or ""),
                      "hours": str(h.get("hours", "") or ""),
                      "priority": int(h.get("priority", 99) or 99),
                      "enabled": (False if str(h.get("enabled", True)).strip().lower()
                                  in ("false", "0", "no", "off") else True)})
    try:
        _brain_mod.upsert_profile(tenant_id, {"handoff": clean}, actor=actor)
    except Exception as exc:  # noqa: BLE001
        _lg_handoff.warning("handoff_set failed tenant=%s: %r", tenant_id, exc)
        return _handoff_get(tenant_id)
    return _handoff_get(tenant_id)


def _handoff_valid_phone(phone: str) -> str:
    """Validate + canonicalise a handoff number to +91XXXXXXXXXX (Indian E.164).
    Returns "" when it is not a valid +91 mobile (so callers reject it). Reuses norm()
    (digits-only -> +91…, drops leading 0, prefixes 91 for a bare 10-digit). We additionally
    REQUIRE a 91 country code + 10 national digits so a tenant can't seed a malformed target."""
    n = norm(phone or "")
    # norm() yields "+<cc><number>"; require India (+91) + exactly 10 national digits.
    if n.startswith("+91") and len(n) == 13 and n[1:].isdigit():
        return n
    return ""


def _handoff_add_one(tenant_id: str, entry: dict, actor: str = "system") -> tuple[list, str]:
    """ADD or UPDATE a single handoff entry (keyed by canonical phone). Returns (stored_list, err).
    err="" on success; non-empty spoken-friendly reason on validation failure. Re-uses the
    existing replace path (_handoff_set) so versioning/audit/history stay in ONE place. Token-
    scoped by the caller (tenant_id from resolve_tenant). NEVER raises."""
    if not tenant_id or _brain_mod is None:
        return (_handoff_get(tenant_id), "brain module unavailable")
    entry = entry or {}
    ph = _handoff_valid_phone(str(entry.get("phone", "") or entry.get("whatsapp", "")))
    if not ph:
        return (_handoff_get(tenant_id),
                "invalid phone — give a valid Indian mobile in +91XXXXXXXXXX form")
    wa = _handoff_valid_phone(str(entry.get("whatsapp", ""))) or ph
    cur = _handoff_get(tenant_id)
    # de-dup by canonical phone: replace an existing entry for the same number (idempotent add).
    cur = [h for h in cur if (h.get("phone") or "") != ph]
    try:
        prio = int(entry.get("priority", 0) or 0)
    except Exception:  # noqa: BLE001
        prio = 0
    if prio <= 0:
        prio = (max([int(h.get("priority", 0) or 0) for h in cur], default=0) + 1) if cur else 1
    cur.append({"phone": ph, "whatsapp": wa,
                "name": str(entry.get("name", "") or ""),
                "role": str(entry.get("role", "") or ""),
                "hours": str(entry.get("hours", "") or ""),
                "priority": prio,
                "enabled": (False if str(entry.get("enabled", "true")).strip().lower()
                            in ("false", "0", "no", "off") else True)})
    return (_handoff_set(tenant_id, cur, actor=actor), "")


def _handoff_remove_one(tenant_id: str, phone: str, actor: str = "system") -> tuple[list, str, bool]:
    """REMOVE the single handoff entry matching `phone` (canonicalised). Returns
    (stored_list, err, removed_bool). Token-scoped. NEVER raises."""
    if not tenant_id or _brain_mod is None:
        return (_handoff_get(tenant_id), "brain module unavailable", False)
    target = _handoff_valid_phone(phone) or norm(phone or "") or str(phone or "").strip()
    if not target:
        return (_handoff_get(tenant_id), "no phone given to remove", False)
    cur = _handoff_get(tenant_id)
    kept = [h for h in cur
            if (h.get("phone") or "") != target and (h.get("whatsapp") or "") != target]
    removed = len(kept) != len(cur)
    if not removed:
        return (cur, "", False)
    return (_handoff_set(tenant_id, kept, actor=actor), "", True)


async def notify_handoff_team(tenant_id: str, lead: dict, summary: str = "",
                              score: int = 0) -> dict:
    """HOT-LEAD -> TEAM WHATSAPP. Send the lead phone + call summary to EVERY handoff-list
    WhatsApp number so a human can take over. Reuses whatsapp.py via _wa_send (logs each).
    Cold path uses the approved `hot_lead_alert` template (the team rarely has an open 24h
    window); if Meta isn't configured the send no-ops gracefully (dormant). NEVER raises into
    the call loop. Returns {ok, sent, attempts, results:[...]}.
    Triggered from the post-call hot branch (interest>=70) AND as the warm-transfer fallback."""
    out = {"ok": False, "sent": 0, "attempts": 0, "results": []}
    try:
        team = _handoff_get(tenant_id)
        if not team:
            out["note"] = "no_handoff_team"
            return out
        name = (lead.get("name") or "").strip() or "Lead"
        phone = (lead.get("phone") or "").strip() or "—"
        summ = (summary or lead.get("summary") or "").strip()[:300] or "Hot lead from a call."
        sc = int(score or lead.get("interest", 0) or lead.get("score", 0) or 0)
        # free-form body (used when a team member's 24h window is open / generic BSP).
        text_body = (f"🔥 Hot lead: {name} ({phone}). Score {sc}/100. "
                     f"Summary: {summ}. Reply to take over.")
        params = [name, phone, summ, str(sc)]   # hot_lead_alert template body vars
        meta = bool(wa_mod and getattr(wa_mod, "meta_configured", lambda: False)())
        for h in team:
            wa = (h.get("whatsapp") or h.get("phone") or "").strip()
            if not wa:
                continue
            out["attempts"] += 1
            try:
                if meta:
                    # cold-safe: approved template (no 24h window assumed with the team).
                    res = await _wa_send(tenant_id, wa, HOT_LEAD_ALERT_TEMPLATE, params,
                                         kind="hot_lead_alert")
                else:
                    # generic BSP / dormant -> free-form text path (no-ops if unconfigured).
                    res = await _wa_send(tenant_id, wa, text_body, None,
                                         kind="hot_lead_alert", is_text=True)
            except Exception as exc:  # noqa: BLE001
                res = {"ok": False, "status": f"error:{type(exc).__name__}"}
            out["results"].append({"to": wa, "ok": bool(res.get("ok")),
                                   "status": res.get("status", ""),
                                   "wamid": _wa_mid(res)})
            if res.get("ok"):
                out["sent"] += 1
        out["ok"] = out["sent"] > 0
        _lg_handoff.info("notify_handoff_team tenant=%s team=%d sent=%d", tenant_id, len(team), out["sent"])
        return out
    except Exception as exc:  # noqa: BLE001
        _lg_handoff.warning("notify_handoff_team failed tenant=%s: %r", tenant_id, exc)
        out["note"] = f"error:{type(exc).__name__}"
        return out


def _wa_mid(result: dict) -> str:
    """Best-effort extract the Meta message id (wamid) from a send result for the audit/proof."""
    try:
        import json as _json
        raw = result.get("response") or ""
        if raw and "wamid" in raw:
            d = _json.loads(raw)
            msgs = d.get("messages") or []
            if msgs and isinstance(msgs, list):
                return str(msgs[0].get("id", ""))
    except Exception:  # noqa: BLE001
        pass
    return ""


async def _wa_followup(tenant_id: str, rec: dict, outcome: str, camp_fields: dict):
    """WAVE3 Unit5: per-campaign auto follow-up after interested/callback calls.
    Gated behind campaign `wa_followup` flag (default OFF). No-op if no creds/template."""
    try:
        if not (camp_fields or {}).get("wa_followup"):
            return
        if outcome == "interested" or (rec.get("interest", 0) or 0) >= 70:
            template = (camp_fields.get("wa_template_interested")
                        or camp_fields.get("wa_template_qualified") or "")
        elif outcome == "callback":
            template = camp_fields.get("wa_template_callback", "")
        else:
            return
        if not template:
            return
        await _wa_send(tenant_id, rec.get("phone", ""), template,
                       [rec.get("name", "")], kind="auto_followup")
    except Exception:  # noqa: BLE001
        pass


# ---------- WAVE A2: per-contact WhatsApp conversation threads ----------
WA_MAX_TURNS = int(os.getenv("WA_MAX_TURNS", "12") or "12")  # human turns before handoff
# WA-AUTO post-call template follow-up — DEFAULT OFF (safe). When ON, a completed call
# instantly sends the APPROVED post-call template (cold-safe vehicle) to the lead.
# Gate = global WA_AUTO_FOLLOWUP=1 OR per-campaign fields.wa_followup=True.
WA_AUTO_FOLLOWUP = (os.getenv("WA_AUTO_FOLLOWUP", "0") or "0").strip().lower() in ("1", "true", "yes", "on")
# Approved template name + its var mapping. Body: "Hi {{1}}, thanks for taking our call
# about {{2}}. ..."  -> {{1}}=lead name, {{2}}=product/enquiry. Lang from WA_LANG (en).
WA_FOLLOWUP_TEMPLATE = (os.getenv("WA_FOLLOWUP_TEMPLATE", "post_call_followup") or "post_call_followup").strip()
WA_FOLLOWUP_ENQUIRY_FALLBACK = (os.getenv("WA_FOLLOWUP_ENQUIRY_FALLBACK", "your enquiry") or "your enquiry").strip()
# #10(c) SCORE-GATE: auto-send the post-call WhatsApp follow-up ONLY to INTERESTED leads
# whose lead score is STRICTLY GREATER than this threshold (default 50). Tunable via env;
# a campaign may also override with fields.wa_followup_min_score. Lowered from the old 70.
try:
    WA_FOLLOWUP_MIN_SCORE = int(os.getenv("WA_FOLLOWUP_MIN_SCORE", "50") or "50")
except Exception:  # noqa: BLE001
    WA_FOLLOWUP_MIN_SCORE = 50
_WA_OPTOUT_WORDS = ("stop", "unsubscribe", "opt out", "optout", "band karo",
                    "band karein", "mat bhejo", "remove me", "do not", "dont contact",
                    "don't contact", "block")
_WA_HANDOFF_WORDS = ("talk to human", "human agent", "real person", "call me",
                     "agent se baat", "complaint", "manager")


def _wa_safe_tenant(tenant_id: str | None) -> str:
    """Filesystem-safe single path segment for a tenant id (no '/', '..', dots).
    Empty/None -> '' (caller falls back to the legacy flat path)."""
    return re.sub(r"[^A-Za-z0-9_-]", "", str(tenant_id or "")).strip("-_")


def _wa_thread_path(phone: str, tenant_id: str | None = None) -> Path:
    """WhatsApp thread file path. P0-LEAK: when a tenant_id is supplied the file
    is namespaced under a per-tenant subdir ``{tenant}/{phone}.json`` so tenant A
    cannot read tenant B's thread. With no tenant_id it stays the LEGACY flat
    ``{phone}.json`` (additive — the path the un-restarted earner still writes)."""
    safe = re.sub(r"[^0-9]", "", phone or "")
    tdir = _wa_safe_tenant(tenant_id)
    if tdir:
        return WA_THREADS_DIR / tdir / f"{safe}.json"
    return WA_THREADS_DIR / f"{safe}.json"


def _wa_thread_read(phone: str, tenant_id: str | None = None) -> dict:
    """Read a WhatsApp thread, tenant-scoped with a TENANT-CHECKED legacy fallback
    + migrate-on-read. Prefer ``{tenant}/{phone}.json``; if absent fall back to the
    legacy flat ``{phone}.json`` (which the un-restarted earner still writes) ONLY
    IF it is attributable to THIS tenant — its stored ``tenant_id`` matches OR is
    empty (we claim it). A legacy thread owned by a DIFFERENT tenant is NEVER
    returned (that would re-open the cross-tenant leak). With no tenant_id this is
    the legacy flat read, unchanged."""
    tdir = _wa_safe_tenant(tenant_id)
    if not tdir:
        return _read(_wa_thread_path(phone), {}) or {}
    tp = _wa_thread_path(phone, tenant_id)
    th = _read(tp, {}) or {}
    if th:
        return th
    legacy = _read(_wa_thread_path(phone), {}) or {}
    if not legacy:
        return {}
    owner = _wa_safe_tenant(legacy.get("tenant_id"))
    if owner and owner != tdir:
        return {}  # legacy thread belongs to a different tenant -> not returned
    # same tenant or unowned -> claim + migrate into the tenant path (best-effort)
    legacy["tenant_id"] = tenant_id
    try:
        tp.parent.mkdir(parents=True, exist_ok=True)
        _write(tp, legacy)
    except Exception:  # noqa: BLE001
        pass
    return legacy


def _wa_thread_find_any(phone: str) -> dict:
    """ADMIN-ONLY: locate a contact's thread regardless of which tenant subdir it
    lives in (or the legacy flat path). Returns the most-recently-updated match, or
    {}. Callers MUST gate this on is_admin — it deliberately ignores tenant scope."""
    safe = re.sub(r"[^0-9]", "", phone or "")
    if not safe:
        return {}
    # Legacy flat path first (cheap), then any per-tenant subdir copy.
    flat = _read(WA_THREADS_DIR / f"{safe}.json", {}) or {}
    best, best_mt = flat, 0.0
    try:
        for p in WA_THREADS_DIR.glob(f"**/{safe}.json"):
            if p.name.startswith("."):
                continue
            try:
                mt = p.stat().st_mtime
            except OSError:
                continue
            if mt >= best_mt:
                th = _read(p, {}) or {}
                if th:
                    best, best_mt = th, mt
    except Exception:  # noqa: BLE001
        pass
    return best


async def _wa_thread_write(phone: str, thread: dict, tenant_id: str | None = None):
    tid = tenant_id or (thread.get("tenant_id") if isinstance(thread, dict) else None)
    async with _STORE_LOCK:
        p = _wa_thread_path(phone, tid)
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
        except Exception:  # noqa: BLE001
            pass
        _write(p, thread)


def _resolve_contact_by_phone(phone: str) -> dict:
    """Best-effort link an inbound WhatsApp number to a known lead/campaign/tenant.
    Returns {tenant_id, name, campaign_id, campaign_name} (any may be empty)."""
    p = norm(phone)
    out = {"tenant_id": "", "name": "", "campaign_id": "", "campaign_name": ""}
    if not p:
        return out
    # 1) most-recent call to this number carries tenant + campaign
    for c in CALLS:
        if norm(c.get("phone", "")) == p:
            out["tenant_id"] = c.get("tenant_id", "") or out["tenant_id"]
            out["name"] = c.get("name", "") or out["name"]
            out["campaign_id"] = c.get("campaign_id", "") or out["campaign_id"]
            out["campaign_name"] = c.get("campaign_name", "") or out["campaign_name"]
            break
    # 2) fall back to a stored lead
    if not out["tenant_id"] or not out["name"]:
        for x in _read(LEADS_FILE, []):
            if norm(x.get("phone", "")) == p:
                # Only adopt the lead's tenant if the lead actually carries one;
                # never default an unknown number to ADMIN_ID (that poisons the
                # admin tenant's thread + memory — P0-LEAK red-team break #1).
                out["tenant_id"] = out["tenant_id"] or x.get("tenant_id", "")
                out["name"] = out["name"] or x.get("name", "")
                break
    if not out["tenant_id"]:
        # Unknown inbound number with no attributable tenant -> route to a quarantine
        # bucket, NEVER the ADMIN_ID default. Keeps unrouted traffic out of every
        # real tenant's data until a human/link resolves it.
        out["tenant_id"] = WA_UNROUTED_TENANT
    return out


def _wa_draft_followup_text(rec: dict, outcome: str, camp_fields: dict, tr: dict) -> str:
    """ONE Groq call -> a short Hinglish WhatsApp follow-up tailored to the call.
    Returns "" on failure (caller then falls back to a template send)."""
    name = rec.get("name", "") or "ji"
    company = (camp_fields or {}).get("company_name", "")
    product = (camp_fields or {}).get("product_name", "")
    agent = (camp_fields or {}).get("agent_name", "Riya")
    summary = (tr or {}).get("summary", "")
    next_action = (tr or {}).get("next_action", "")
    interest = rec.get("interest", 0)
    sysmsg = (
        "You are " + agent + ", a friendly Indian sales assistant writing a SHORT WhatsApp "
        "follow-up in natural Hinglish (Roman script). 1-3 short sentences, warm, no emojis spam "
        "(max one), no markdown. Recap the call briefly and give ONE clear next step "
        "(site visit / share details / answer a pending question / schedule a callback). "
        "Do not invent facts beyond the context. Output ONLY the message text."
    )
    ctx = (
        f"Company: {company}\nProduct: {product}\nLead name: {name}\n"
        f"Call outcome: {outcome}\nInterest score: {interest}\n"
        f"Call summary: {summary}\nSuggested next action: {next_action}\n"
    )
    return _groq_chat([{"role": "system", "content": sysmsg},
                       {"role": "user", "content": ctx}], max_tokens=220, temperature=0.6)


def _wa_memory_recap(phone: str, tenant_id: str | None = None,
                     agent_name: str | None = None) -> str:
    """Per-person prior-call recap from memory.py (the voice agent's cross-call store).
    Read-only, import-safe; returns "" when memory is absent/unreadable. Never raises.

    P0-LEAK: the memory read is TENANT-SCOPED — without the tenant_id this would read
    the legacy flat ``{phone}.json`` which is shared across tenants (a live
    cross-tenant read). load_memory() now namespaces by tenant and only falls back to
    a legacy file attributable to the SAME tenant, so passing tenant_id closes the leak
    while still finding the earner-written legacy file for that tenant."""
    try:
        if _mem_mod is None or not phone:
            return ""
        rec = _mem_mod.load_memory(re.sub(r"[^0-9]", "", phone or ""), tenant_id)
        return (_mem_mod.build_recap(rec, agent_name) or "")[:500]
    except Exception:  # noqa: BLE001
        return ""


def _wa_reply_text(thread: dict, camp_fields: dict, incoming: str) -> str:
    """Multi-turn, context-rich reply: ONE Groq call grounded in the CAMPAIGN brain + what
    happened on the CALL (summary / next step / interest, persisted on the thread by
    _wa_ai_followup) + the per-person MEMORY recap + the last 10 thread turns. "" on failure."""
    agent = (camp_fields or {}).get("agent_name", "Riya")
    company = (camp_fields or {}).get("company_name", "")
    product = (camp_fields or {}).get("product_name", "") or thread.get("product", "")
    summary = (camp_fields or {}).get("product_summary", "")
    name = thread.get("name", "") or "ji"
    # Call grounding (written at follow-up/seed time).
    call_summary = (thread.get("call_summary") or "").strip()
    next_action = (thread.get("next_action") or "").strip()
    call_outcome = (thread.get("call_outcome") or "").strip()
    interest = thread.get("interest", 0)
    # P0-LEAK: scope the cross-call memory read to THIS thread's tenant (else it reads
    # the shared legacy file across tenants). Label assistant turns with the campaign
    # agent name, not a hardcoded "Riya".
    mem_recap = _wa_memory_recap(thread.get("phone", ""),
                                 thread.get("tenant_id"), agent)
    grounding = ""
    if call_summary:
        grounding += f"What happened on the phone call: {call_summary[:400]}. "
    if next_action:
        grounding += f"Agreed/suggested next step from the call: {next_action[:160]}. "
    if call_outcome:
        grounding += f"Call outcome: {call_outcome} (interest {interest}). "
    if mem_recap:
        grounding += f"Earlier history with this person: {mem_recap}. "
    sysmsg = (
        "You are " + agent + (f", a sales assistant for {company}" if company else "") +
        ". You are continuing a conversation with " + name + " on WhatsApp AFTER a phone call. "
        "Reply to the customer's message in SHORT natural Hinglish (Roman script), "
        "1-3 sentences, warm and helpful, at most one emoji, no markdown. "
        + (f"You are following up about: {product}. " if product else "")
        + (f"Product/offer context: {summary[:300]}. " if summary else "")
        + (grounding if grounding else "")
        + "Use the call context above so you don't repeat yourself or contradict the call. "
        "Move the conversation toward a clear next step (site visit / share details / schedule "
        "a callback / booking). If they ask something you don't know, offer to have a human call "
        "them. Do not invent facts beyond the context. Output ONLY the reply text."
    )
    msgs = [{"role": "system", "content": sysmsg}]
    for t in (thread.get("turns") or [])[-10:]:
        role = "assistant" if t.get("role") == "assistant" else "user"
        msgs.append({"role": role, "content": t.get("text", "")})
    msgs.append({"role": "user", "content": incoming})
    return _groq_chat(msgs, max_tokens=220, temperature=0.6)


async def _wa_handle_inbound(phone: str, text: str) -> dict:
    """Core inbound handler (provider-agnostic). Maintains thread state, applies opt-out /
    handoff / max-turn guards, generates+sends the next reply when WA is configured.
    Returns a small dict for logging. Safe + dormant when WA env missing."""
    phone_n = norm(phone) or (phone or "")
    link = _resolve_contact_by_phone(phone_n)
    # P0-LEAK: never default an unknown inbound number to ADMIN_ID; _resolve_contact_by_phone
    # already returns WA_UNROUTED_TENANT when nothing attributes the number to a real tenant.
    tenant_id = link["tenant_id"] or WA_UNROUTED_TENANT
    thread = _wa_thread_read(phone_n, tenant_id)
    if not thread:
        thread = {"phone": phone_n, "tenant_id": tenant_id, "name": link["name"],
                  "campaign_id": link["campaign_id"], "campaign_name": link["campaign_name"],
                  "status": "active", "turns": [],
                  "created_at": datetime.now().isoformat(timespec="seconds")}
    tenant_id = thread.get("tenant_id") or tenant_id
    thread.setdefault("turns", []).append(
        {"role": "user", "text": text, "at": datetime.now().isoformat(timespec="seconds")})
    low = (text or "").strip().lower()
    action = "noted"
    # Opt-out / STOP keywords -> end thread + suppress (reuse existing suppression logic).
    if any(w in low for w in _WA_OPTOUT_WORDS):
        thread["status"] = "opted_out"
        await _add_suppression(tenant_id, phone_n, "wa_opt_out", source="whatsapp")
        await _flip_lead_status(tenant_id, phone_n, "opted_out")
        await _emit_webhook(tenant_id, "lead.opted_out",
                            {"phone": phone_n, "campaign_id": thread.get("campaign_id", "")})
        action = "opted_out"
    elif any(w in low for w in _WA_HANDOFF_WORDS):
        thread["status"] = "needs_human"
        action = "needs_human"
    else:
        human_turns = sum(1 for t in thread["turns"] if t.get("role") == "user")
        if human_turns >= WA_MAX_TURNS:
            thread["status"] = "needs_human"
            action = "max_turns_handoff"
        elif wa_mod and wa_mod.is_configured():
            camp = get_campaign(thread.get("campaign_id", "")) or {}
            reply = _wa_reply_text(thread, camp.get("fields") or {}, text)
            if reply:
                result = (await wa_mod.send_whatsapp_text_async(phone_n, reply)
                          if getattr(wa_mod, "meta_configured", lambda: False)()
                          else await wa_mod.send_whatsapp_async(phone_n, reply, None))
                await _wa_log(tenant_id, phone_n, "ai_reply", result, kind="inbound_reply")
                thread["turns"].append({"role": "assistant", "text": reply,
                                        "at": datetime.now().isoformat(timespec="seconds")})
                action = "replied"
            else:
                action = "draft_failed"
        else:
            # Dormant: store the inbound, nothing to call out to. ("WA not configured")
            action = "stored_dormant"
    thread["updated_at"] = datetime.now().isoformat(timespec="seconds")
    if len(thread["turns"]) > 200:        # keep the thread file bounded
        thread["turns"] = thread["turns"][-200:]
    await _wa_thread_write(phone_n, thread, tenant_id)
    return {"phone": phone_n, "action": action, "status": thread["status"]}


def _wa_followup_product(camp_fields: dict, tr: dict) -> str:
    """Resolve {{2}} for the post-call template: the product/enquiry the call was about.
    Prefers the campaign product, then the transcript's stated enquiry, else a safe
    generic fallback. Trimmed to a WhatsApp-template-safe length. Never raises."""
    cand = ((camp_fields or {}).get("product_name", "")
            or (tr or {}).get("enquiry", "")
            or (tr or {}).get("next_action", "")
            or WA_FOLLOWUP_ENQUIRY_FALLBACK)
    cand = " ".join(str(cand).split())[:60].strip()
    return cand or WA_FOLLOWUP_ENQUIRY_FALLBACK


def _wa_brochure_link(camp_fields: dict) -> str:
    """Resolve a publicly-fetchable URL for a campaign's brochure PDF, for a WhatsApp
    DOCUMENT send. Prefers minting a short-lived presigned GET URL from the stored Spaces
    KEY (the live ``capsy-recordings`` bucket keeps objects PRIVATE — a presigned URL is
    what Meta can actually fetch at send time). Falls back to a stored public URL only if
    no key is present. Returns "" when no brochure is configured. Never raises."""
    cf = camp_fields or {}
    key = (cf.get("brochure_pdf_key") or "").strip()
    if key:
        try:
            from media_gen import spaces as _spaces
            cli = _spaces._client()
            if cli is not None:
                bucket = (os.getenv("SPACES_BUCKET") or "").strip()
                if bucket:
                    return cli.generate_presigned_url(
                        "get_object", Params={"Bucket": bucket, "Key": key},
                        ExpiresIn=3600)
        except Exception:  # noqa: BLE001
            pass
    # Fallback: a directly-public URL stored on the campaign (other bucket / CDN).
    return (cf.get("brochure_pdf_url") or "").strip()


async def _wa_send_brochure(tenant_id: str, phone_n: str, camp_fields: dict,
                            product: str) -> str:
    """#10(b) BROCHURE: after a qualifying call, send the per-campaign brochure PDF as a
    WhatsApp DOCUMENT message (native Meta path). Best-effort, never raises into the call
    loop. Returns the wamid/status string on send, "" when no brochure / not configured."""
    try:
        link = _wa_brochure_link(camp_fields)
        if not link:
            return ""
        if not (wa_mod and getattr(wa_mod, "meta_configured", lambda: False)()):
            return ""  # documents are a native-Meta-only feature here
        fname = (camp_fields or {}).get("brochure_pdf_name", "") or "brochure.pdf"
        if not fname.lower().endswith(".pdf"):
            fname = fname + ".pdf"
        caption = f"Here's more about {product}." if product else "Here's our brochure."
        result = await wa_mod.send_whatsapp_document_async(phone_n, link, fname, caption)
        await _wa_log(tenant_id, phone_n, "brochure_pdf", result, kind="auto_brochure")
        return result.get("status", "") if isinstance(result, dict) else ""
    except Exception:  # noqa: BLE001
        return ""


async def _wa_ai_followup(tenant_id: str, rec: dict, outcome: str, camp_fields: dict, tr: dict):
    """WA-AUTO post-call WhatsApp follow-up.

    Gate = global WA_AUTO_FOLLOWUP=1 OR per-campaign fields.wa_followup=True (default OFF).
    On a fresh completed call there is NO open 24h customer-service window, so a cold
    free-form text would be REJECTED by Meta. The cold-safe vehicle is the APPROVED
    TEMPLATE (`WA_FOLLOWUP_TEMPLATE`, lang from WA_LANG=en) populated with [name, product].
    Idempotent (one send per call), consent-checked (skips suppressed numbers), and seeds
    the conversation thread WITH the call context so an inbound reply is grounded.
    Never blocks / raises into the call loop (fire-and-forget, best-effort)."""
    try:
        # GATE: global flag OR per-campaign flag.
        if not (WA_AUTO_FOLLOWUP or (camp_fields or {}).get("wa_followup")):
            return
        # #10(c) SCORE-GATE: send ONLY to INTERESTED leads — lead score STRICTLY > threshold
        # (default 50, was 70), OR an explicitly interested/callback outcome. The per-campaign
        # fields.wa_followup_min_score overrides the global WA_FOLLOWUP_MIN_SCORE when set.
        try:
            min_score = int((camp_fields or {}).get("wa_followup_min_score")
                            or WA_FOLLOWUP_MIN_SCORE)
        except Exception:  # noqa: BLE001
            min_score = WA_FOLLOWUP_MIN_SCORE
        score = rec.get("interest", 0) or 0
        try:
            score = int(score)
        except Exception:  # noqa: BLE001
            score = 0
        if not (outcome in ("interested", "callback") or score > min_score):
            return
        configured = bool(wa_mod and wa_mod.is_configured())
        phone = rec.get("phone", "")
        phone_n = norm(phone) or phone
        if not phone_n:
            return
        # IDEMPOTENCY: never send the post-call follow-up twice for the same call.
        if rec.get("wa_followup_sent"):
            return
        # CONSENT / opt-out: do not message a suppressed (DND) number.
        try:
            if phone_n in _suppressed_set(tenant_id):
                rec["wa_followup_sent"] = "skipped_suppressed"
                return
        except Exception:  # noqa: BLE001
            pass
        if not configured:
            # Dormant (no WA env): keep the legacy template path (logs skipped_no_config).
            await _wa_followup(tenant_id, rec, outcome, camp_fields)
            rec["wa_followup_sent"] = "dormant"
            return

        name = (rec.get("name", "") or "").strip() or "there"
        product = _wa_followup_product(camp_fields, tr)
        meta = bool(wa_mod and getattr(wa_mod, "meta_configured", lambda: False)())

        # Is the 24h CS window already open? Only true if the lead has sent us an inbound
        # WhatsApp recently (their inbound seeds a 'user' turn). A fresh OUTBOUND call does
        # NOT open it -> cold path MUST use the approved template, never free-form text.
        existing = _wa_thread_read(phone_n, tenant_id)
        window_open = any(t.get("role") == "user" for t in (existing.get("turns") or []))

        sent_text = ""
        if meta and window_open:
            # Window open -> a personalised free-form follow-up is allowed and nicer.
            draft = _wa_draft_followup_text(rec, outcome, camp_fields, tr)
            if draft:
                result = await wa_mod.send_whatsapp_text_async(phone_n, draft)
                await _wa_log(tenant_id, phone_n, "ai_followup", result, kind="auto_followup")
                sent_text = draft
        if not sent_text and meta:
            # COLD post-call (no open window) -> APPROVED TEMPLATE, params [name, product].
            # Lang is taken from WA_LANG (=en) inside whatsapp.send_whatsapp_async.
            result = await wa_mod.send_whatsapp_async(phone_n, WA_FOLLOWUP_TEMPLATE,
                                                      [name, product])
            await _wa_log(tenant_id, phone_n, WA_FOLLOWUP_TEMPLATE, result,
                          kind="auto_followup_template")
            # Mirror the template's body so the reply brain has the opener verbatim.
            sent_text = (f"Hi {name}, thanks for taking our call about {product}. "
                         "Reply here if you have questions or want to take the next step.")
        elif not sent_text:
            # No Meta path (generic/other BSP) -> legacy template fallback.
            await _wa_followup(tenant_id, rec, outcome, camp_fields)
            sent_text = "[template:" + ((camp_fields or {}).get("wa_template_interested")
                                        or (camp_fields or {}).get("wa_template_qualified")
                                        or (camp_fields or {}).get("wa_template_callback")
                                        or WA_FOLLOWUP_TEMPLATE) + "]"

        # Mark idempotency (one send per call) regardless of which path fired.
        rec["wa_followup_sent"] = datetime.now().isoformat(timespec="seconds")

        # #10(b) BROCHURE: alongside the template, send the per-campaign brochure PDF as a
        # WhatsApp DOCUMENT (only the qualifying leads that already passed the score-gate
        # above reach here). Best-effort + idempotent (one brochure per call). No brochure
        # configured -> no-op. Never blocks / raises into the call loop.
        try:
            if not rec.get("wa_brochure_sent"):
                br_status = await _wa_send_brochure(tenant_id, phone_n, camp_fields, product)
                if br_status:
                    rec["wa_brochure_sent"] = br_status
        except Exception:  # noqa: BLE001
            pass

        # Seed / update the conversation thread, persisting the CALL CONTEXT so the inbound
        # reply brain (_wa_reply_text) can ground every later turn in what happened on the call.
        if sent_text:
            thread = existing or _wa_thread_read(phone_n, tenant_id)
            if not thread:
                thread = {"phone": phone_n, "tenant_id": tenant_id,
                          "name": rec.get("name", ""), "campaign_id": rec.get("campaign_id", ""),
                          "campaign_name": rec.get("campaign_name", ""),
                          "status": "active", "turns": [],
                          "created_at": datetime.now().isoformat(timespec="seconds")}
            # Call-context grounding (read back in _wa_reply_text on every inbound turn).
            thread["call_summary"] = (tr or {}).get("summary", "") or thread.get("call_summary", "")
            thread["next_action"] = (tr or {}).get("next_action", "") or thread.get("next_action", "")
            thread["call_outcome"] = outcome or thread.get("call_outcome", "")
            thread["interest"] = score if score else thread.get("interest", 0)
            thread["product"] = product or thread.get("product", "")
            thread["name"] = thread.get("name") or rec.get("name", "")
            thread["campaign_id"] = thread.get("campaign_id") or rec.get("campaign_id", "")
            thread["campaign_name"] = thread.get("campaign_name") or rec.get("campaign_name", "")
            thread.setdefault("turns", []).append(
                {"role": "assistant", "text": sent_text,
                 "at": datetime.now().isoformat(timespec="seconds")})
            thread["updated_at"] = datetime.now().isoformat(timespec="seconds")
            thread["status"] = thread.get("status") or "active"
            await _wa_thread_write(phone_n, thread, tenant_id)
    except Exception:  # noqa: BLE001
        pass


# ---------- P1.C CRM webhook out (fire-and-forget stub) ----------
async def _emit_webhook(tenant_id, event, payload):
    """Deliver an event to a tenant's registered webhook(s), HMAC-signed. Never blocks the loop."""
    try:
        hooks = [w for w in _read(WEBHOOK_FILE, [])
                 if w.get("tenant_id") == tenant_id and w.get("active", True)
                 and (not w.get("events") or event in w.get("events", []))]
        if not hooks:
            return
        body = json.dumps({"event": event, "data": payload,
                           "at": datetime.now().isoformat(timespec="seconds")}, ensure_ascii=False)
        for w in hooks:
            sig = hmac.new((w.get("secret") or "").encode("utf-8"), body.encode("utf-8"),
                           hashlib.sha256).hexdigest()
            status = "error"
            for attempt in range(3):
                try:
                    async with httpx.AsyncClient(timeout=10) as cli:
                        resp = await cli.post(w["url"], content=body,
                                              headers={"Content-Type": "application/json",
                                                       "X-Famit-Signature": sig,
                                                       "X-Famit-Event": event})
                        status = f"sent:{resp.status_code}"
                        if resp.status_code < 500:
                            break
                except Exception as exc:  # noqa: BLE001
                    status = f"error:{repr(exc)[:60]}"
                await asyncio.sleep(2 ** attempt)
            if status.startswith("error") or status.startswith("sent:5"):
                _log_event("warning", "webhook",
                           f"webhook delivery failed ({status}) for '{event}'"[:300],
                           tenant={"tenant_id": tenant_id}, error_type="webhook_delivery",
                           context={"url": w.get("url", ""), "event": event, "status": status})
            async with _STORE_LOCK:
                log = _read(WEBHOOK_LOG_FILE, [])
                log.insert(0, {"tenant_id": tenant_id, "url": w["url"], "event": event,
                               "status": status, "at": datetime.now().isoformat(timespec="seconds")})
                del log[2000:]
                _write(WEBHOOK_LOG_FILE, log)
    except Exception:  # noqa: BLE001
        pass


# ---------- WAVE3 Unit4: billing / metering ----------
# Default plan is GENEROUS + postpaid so nothing is blocked for existing tenants today.
def _default_billing(tenant: dict | None = None) -> dict:
    is_admin = bool((tenant or {}).get("is_admin"))
    return {"plan": "postpaid", "rate_per_min": 0.0, "rate_per_call": 0.0,
            "currency": "INR", "balance": 0.0,
            "included_minutes": 1000000 if is_admin else 100000}


def _read_billing() -> dict:
    """Map of tenant_id -> billing record."""
    b = _read(BILLING_FILE, {})
    return b if isinstance(b, dict) else {}


def _billing_for(tenant_id: str) -> dict:
    """Billing record for a tenant; lazily seeds a generous default (never blocks today)."""
    store = _read_billing()
    rec = store.get(tenant_id)
    if not rec:
        rec = _default_billing(_tenant_by_id(tenant_id))
        store[tenant_id] = rec
        _write(BILLING_FILE, store)
    # backfill any missing keys on legacy/partial records
    base = _default_billing(_tenant_by_id(tenant_id))
    for k, v in base.items():
        rec.setdefault(k, v)
    return rec


def _ledger_path(tenant_id: str) -> Path:
    safe = "".join(ch for ch in tenant_id if ch.isalnum() or ch in "-_") or "unknown"
    return LEDGER_DIR / f"{safe}.json"


def _read_ledger(tenant_id: str) -> list[dict]:
    return _read(_ledger_path(tenant_id), [])


def _call_cost(billing: dict, duration_s: int) -> float:
    mins = (duration_s or 0) / 60.0
    cost = mins * float(billing.get("rate_per_min", 0) or 0)
    cost += float(billing.get("rate_per_call", 0) or 0)
    return round(cost, 4)


async def _charge_call(tenant_id: str, rec: dict):
    """Append an itemized charge to the tenant ledger + decrement prepaid balance.
    Best-effort, lock-guarded, never raises into the call loop. Postpaid plans just
    accrue cost (no balance decrement)."""
    try:
        dur = int(rec.get("duration_s", 0) or 0)
        if dur <= 0 and rec.get("status") in ("suppressed", "failed"):
            return  # nothing billable
        async with _STORE_LOCK:
            billing_store = _read_billing()
            billing = billing_store.get(tenant_id) or _default_billing(_tenant_by_id(tenant_id))
            base = _default_billing(_tenant_by_id(tenant_id))
            for k, v in base.items():
                billing.setdefault(k, v)
            cost = _call_cost(billing, dur)
            entry = {"id": uuid.uuid4().hex[:10], "call_id": rec.get("id"),
                     "phone": rec.get("phone", ""), "campaign_id": rec.get("campaign_id", ""),
                     "duration_s": dur, "cost": cost, "currency": billing.get("currency", "INR"),
                     "outcome": rec.get("outcome", ""),
                     "at": datetime.now().isoformat(timespec="seconds")}
            LEDGER_DIR.mkdir(parents=True, exist_ok=True)
            ledger = _read_ledger(tenant_id)
            ledger.insert(0, entry)
            del ledger[5000:]
            _write(_ledger_path(tenant_id), ledger)
            # prepaid: decrement balance (may go negative; /run gate stops the next batch)
            if billing.get("plan") == "prepaid":
                billing["balance"] = round(float(billing.get("balance", 0) or 0) - cost, 4)
            billing_store[tenant_id] = billing
            _write(BILLING_FILE, billing_store)
    except Exception:  # noqa: BLE001
        pass


# ---------- WAVE A Unit1: internal per-call vendor metering ----------
async def record_usage_event(ev: dict) -> None:
    """Append one usage event to var/usage_events.json (lock-guarded, best-effort).
    Shape: {ts,call_id,room,tenant_id,campaign_id,vendor,service_type,qty,unit,
            est_cost_inr,actual_or_estimated}. A metering failure NEVER breaks a call."""
    try:
        row = {
            "ts": ev.get("ts") or datetime.now().isoformat(timespec="seconds"),
            "call_id": ev.get("call_id", ""),
            "room": ev.get("room", ""),
            "tenant_id": ev.get("tenant_id", ""),
            "campaign_id": ev.get("campaign_id", ""),
            "vendor": ev.get("vendor", ""),
            "service_type": ev.get("service_type", ""),
            "qty": ev.get("qty", 0),
            "unit": ev.get("unit", ""),
            "est_cost_inr": round(float(ev.get("est_cost_inr", 0) or 0), 6),
            "actual_or_estimated": ev.get("actual_or_estimated", "estimated"),
            "sip_call_id": ev.get("sip_call_id", ""),
        }
        async with _STORE_LOCK:
            rows = _read(USAGE_EVENTS_FILE, [])
            if not isinstance(rows, list):
                rows = []
            rows.append(row)
            del rows[:-50000]   # cap growth (keep newest 50k)
            _write(USAGE_EVENTS_FILE, rows)
    except Exception:  # noqa: BLE001
        pass


def _read_usage_events() -> list[dict]:
    rows = _read(USAGE_EVENTS_FILE, [])
    return rows if isinstance(rows, list) else []


USAGE_RAW_DIR = VAR / "usage_events_raw"   # agent drops one file per room here


def _call_by_room(room: str) -> dict | None:
    if not room:
        return None
    return next((c for c in CALLS if c.get("room") == room), None)


async def _drain_usage_raw() -> int:
    """Fold the agent's per-room usage files (usage_events_raw/<room>.json) into the
    shared usage_events.json, stamping tenant_id+call_id by joining on `room`. The agent
    runs in a separate process and writes one file per call to avoid clobbering. We only
    ingest a room once its call record exists (so we can attribute the tenant); files for
    unknown rooms are left for a later tick. Lock-guarded, best-effort."""
    drained = 0
    try:
        if not USAGE_RAW_DIR.exists():
            return 0
        files = list(USAGE_RAW_DIR.glob("*.json"))
        if not files:
            return 0
        async with _STORE_LOCK:
            rows = _read(USAGE_EVENTS_FILE, [])
            if not isinstance(rows, list):
                rows = []
            seen_rooms = {r.get("room") for r in rows}
            for f in files:
                try:
                    room = f.stem
                    if room in seen_rooms:
                        f.unlink(missing_ok=True)   # already ingested; clean up
                        continue
                    call = _call_by_room(room)
                    if not call:
                        continue   # call rec not landed yet; ingest on a later tick
                    evs = _read(f, [])
                    if not isinstance(evs, list):
                        f.unlink(missing_ok=True); continue
                    tid = call.get("tenant_id", "")
                    cid = call.get("campaign_id", "")
                    for ev in evs:
                        ev["tenant_id"] = ev.get("tenant_id") or tid
                        ev["campaign_id"] = ev.get("campaign_id") or cid
                        ev["call_id"] = ev.get("call_id") or call.get("id", "")
                        rows.append(ev)
                        drained += 1
                    f.unlink(missing_ok=True)
                except Exception:  # noqa: BLE001
                    continue
            del rows[:-50000]
            _write(USAGE_EVENTS_FILE, rows)
    except Exception:  # noqa: BLE001
        pass
    return drained


# ---------- dial loop ----------
async def _phone_present(lk, room: str) -> bool:
    try:
        r = await lk.room.list_participants(api.ListParticipantsRequest(room=room))
        return any(p.identity.startswith("phone-") for p in r.participants)
    except Exception:  # noqa: BLE001
        return False


async def _finalize_call(it: dict, now_t: float, tenant_id: str, cid: str, camp_fields: dict):
    """Single finalize touch-point (P0.3/4/5/6 + P1.A/C). Runs once per completed call.
    Order: classify -> update lead score -> enqueue retry/callback -> opt-out suppress -> WA -> webhook."""
    rec = it.get("_rec")
    if not rec:
        return
    rec["status"] = "done"
    rec["ended_at"] = _utc_iso()  # W14-WIRE: tz-labelled UTC so the panel renders the right time
    rec["duration_s"] = int(now_t - it["launched_at"])
    room = rec.get("room", "")
    tr = _read(TRANSCRIPT_DIR / f"{room}.json", {}) if room else {}
    # If the transcript exists with real data, mark reconciled so the scheduler sweep skips it.
    # If it's empty (agent shutdown lags hangup), leave it UNMARKED so the sweep re-reconciles.
    if tr:
        rec["_reconciled"] = True
    # P0.4 classify
    outcome = _classify_outcome(rec, tr)
    rec["outcome"] = outcome
    rec["answered"] = outcome in _REAL_CONVO
    rec["interest"] = tr.get("interest", 0)               # P0.6 surface score on call rec
    async with _STORE_LOCK:
        _write(CALLS_FILE, CALLS)
    # ── W9 recording finalize-poll (RECORDING_FINALIZE_ENABLED) -> recording appears in SECONDS ──
    # DETACHED task: polls LiveKit ListEgress to completion, flips recording_status -> "completed",
    # emits recording_ready/transcript_ready/summary_ready IN ORDER. Never blocks/raises into the loop.
    if _RECCFG is not None and _RECCFG.enabled:
        try:
            _bus_w9 = _get_event_bus()
            _storage_w9 = _ObjStorage(_RECCFG)
            _poller_w9 = _FinalizePoller(_RECCFG, bus=_bus_w9,
                                         egress=_EgressClient(), storage=_storage_w9)
            _pipe_w9 = _StagedPipeline(
                _RECCFG, bus=_bus_w9, poller=_poller_w9,
                transcript_provider=_w9_transcript_provider,
                summary_provider=_w9_summary_provider,
            )
            asyncio.create_task(_pipe_w9.run(
                call_id=cid, tenant_id=tenant_id,
                room_name=rec.get("room", cid) or cid, direction="outbound",
            ))
        except Exception as _w9_exc:  # noqa: BLE001
            try:
                import logging as _lg_w9
                _lg_w9.getLogger("w9.finalize").warning(
                    "W9 finalize schedule failed (non-fatal): %r", _w9_exc)
            except Exception:  # noqa: BLE001
                pass
    # WAVE3 Unit4: bill this completed call (ledger + prepaid balance). Best-effort.
    await _charge_call(tenant_id, rec)
    # P0.6 lead scoring
    await _update_lead_after_call(tenant_id, rec.get("phone", ""), tr.get("interest", 0), outcome,
                                  call_at=rec.get("started_at", ""))
    # ── Haptica Grow (FEATURE_GROW): feed this call outcome into the Revenue-Truth Signal
    #    Loop — score the lead (L5: hot/warm/investor/end-user/junk + why) AND dispatch the
    #    CAPI Lead/QualifiedLead conversion signal (L7, value=lead_score, SHADOW-safe until
    #    Meta creds + GROW_SIGNALS_LIVE=1). ADDITIVE + FLAG-GATED + best-effort, off the
    #    event loop: grow.on_call_outcome never raises, and a live POST must not block the
    #    loop. _grow_mod/FEATURE_GROW are module globals defined in the mount block below.
    if globals().get("_grow_mod") is not None and globals().get("FEATURE_GROW"):
        try:
            _g_phone = rec.get("phone", "") or ""
            _g_lead = norm(_g_phone) or (rec.get("id", "") or "")
            _g_tr = tr or {}
            await asyncio.to_thread(
                _grow_mod.on_call_outcome, tenant_id, _g_lead,
                phone=_g_phone, name=(rec.get("name", "") or ""),
                source_platform=(rec.get("source", "") or ""),
                call_answered=bool(rec.get("answered")),
                call_duration_s=int(rec.get("duration_s", 0) or 0),
                interest_score=int(_g_tr.get("interest", 0) or 0),
                booking_made=bool(_g_tr.get("booked") or _g_tr.get("appointment")
                                  or outcome in ("booked", "converted")),
                site_visit_ready=bool(_g_tr.get("site_visit") or _g_tr.get("site_visit_ready")),
                budget_mentioned=bool(_g_tr.get("budget_mentioned") or _g_tr.get("budget")),
                timeline_mentioned=bool(_g_tr.get("timeline") or _g_tr.get("timeline_mentioned")),
                decision_authority=bool(_g_tr.get("decision_authority")),
                investor_intent=bool(_g_tr.get("investor") or _g_tr.get("investor_intent")),
                end_user_intent=bool(_g_tr.get("end_user") or _g_tr.get("end_user_intent")),
                last_outcome=outcome or "")
        except Exception:  # noqa: BLE001 — Grow can NEVER break the call-finalize path
            pass
    # ── Haptica Flywheel (FLYWHEEL_ENABLED): feed this finalized call into the RLHF/RLAIF self-
    #    improvement engine — capture the (state, move, reward) trajectory + the live policy arm +
    #    outcome so the proprietary dataset compounds (every call is fuel). ADDITIVE + FLAG-GATED +
    #    best-effort, OFF the event loop (asyncio.to_thread). The flywheel can NEVER break the call-
    #    finalize path (mirrors the Grow hook above). _flywheel_mod/FLYWHEEL_ENABLED are module
    #    globals defined in the mount block below.
    if globals().get("_flywheel_mod") is not None and globals().get("FLYWHEEL_ENABLED"):
        try:
            await asyncio.to_thread(_flywheel_mod.on_call_finalized_hook, tenant_id, rec, tr)
        except Exception:  # noqa: BLE001 — the flywheel can NEVER break call finalize
            pass
    # R5P4-2 (ADDITIVE): UPSERT this caller into the CRM contacts store so a freshly-called number is
    # NEVER "(unknown)" in the CRM (the finalize path only wrote leads.json before; crm.upsert_contact
    # had ZERO callers). Idempotent on (org, phone); name only overwrites a blank, so a manual rename is
    # preserved. Off the event loop, best-effort — a PG-down / missing-phone case returns None and the
    # call finalize proceeds untouched. NEVER raises into the finalize path.
    try:
        _crm_phone = rec.get("phone", "") or ""
        _crm_name = (rec.get("name", "") or "").strip()
        if _crm_mod is not None and _crm_phone:
            await asyncio.to_thread(
                lambda: _crm_mod.upsert_contact(tenant_id, _crm_phone, name=_crm_name))
    except Exception:  # noqa: BLE001 — a CRM write can NEVER break call finalize
        pass
    # P0.3 opt-out -> auto-suppress + flip lead. DND-GUARD: require BOTH the LLM flag AND an explicit
    # caller opt-out phrase, so a mis-classified short call can't blacklist a fresh lead (live bug).
    if (tr.get("opt_out") or tr.get("outcome") == "opt_out") and _caller_opted_out(tr):
        rec["outcome"] = "opt_out"; rec["answered"] = True
        await _add_suppression(tenant_id, rec.get("phone", ""), "opt_out_call", source=room)
        await _flip_lead_status(tenant_id, rec.get("phone", ""), "opted_out")
        async with _STORE_LOCK:
            _write(CALLS_FILE, CALLS)
        await _emit_webhook(tenant_id, "lead.opted_out",
                            {"phone": rec.get("phone"), "campaign_id": cid})
        # W8: opt-out -> lead dead + call ended events
        try:
            import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
            await _ev(_vke.lead_classified(rec.get("id", ""), tenant_id, "dead"))
            await _ev(_vke.call_ended(rec.get("id", ""), tenant_id, duration_s=rec.get("duration_s", 0)))
        except Exception:  # noqa: BLE001
            pass
    else:
        # ── W10 smart callback cadence (CALLBACK_CADENCE_ENABLED): when ON it OWNS the enqueue
        #    (anti-runaway: outcome-guarded, monotonic attempts, dedup, DND). The legacy flat-file
        #    enqueue below is SKIPPED only when the cadence engine is armed -> reversible by flag.
        _cb_owned = False
        if _CB_STORE is not None and _cb_enqueue_smart is not None:
            try:
                await _cb_enqueue_smart(
                    tenant_id, cid, rec, tr, outcome,
                    int(it.get("attempt", 0)), camp_fields,
                    store=_CB_STORE, config=_CB_CFG, bus=_get_event_bus(),
                )
                _cb_owned = True
            except Exception:  # noqa: BLE001 — an enqueue can NEVER break the call-finalize path
                _cb_owned = False
        if not _cb_owned:
            # P0.5 retry / callback enqueue (only when not opted out) — LEGACY path (flag OFF)
            maxa = int((camp_fields or {}).get("retry_max_attempts", 3))
            backoff = (camp_fields or {}).get("retry_backoff_mins") or [120, 360, 1440]
            attempts = int(it.get("attempt", 0))
            cb = tr.get("callback_at")
            if cb:
                await _enqueue_retry(tenant_id, cid, rec.get("name", ""), rec.get("phone", ""),
                                     attempts, maxa, cb, "callback")
                await _emit_webhook(tenant_id, "callback.scheduled",
                                    {"phone": rec.get("phone"), "campaign_id": cid,
                                     "name": rec.get("name", ""), "when": cb,
                                     "callback_raw": tr.get("callback_raw", "")})
                # W8: callback scheduled event
                try:
                    import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
                    await _ev(_vke.callback_scheduled(rec.get("id", ""), tenant_id, preferred_ts=cb))
                except Exception:  # noqa: BLE001
                    pass
            elif outcome in ("no_answer", "voicemail", "busy") and attempts < maxa:
                delay = backoff[min(attempts, len(backoff) - 1)]
                next_at = _clamp_to_window(now_ist() + timedelta(minutes=delay), camp_fields)
                await _enqueue_retry(tenant_id, cid, rec.get("name", ""), rec.get("phone", ""),
                                     attempts + 1, maxa, next_at.isoformat(), outcome)
    # P1.A WhatsApp + P1.C webhook (fire-and-forget; never block)
    await _send_whatsapp(tenant_id, rec, rec["outcome"], camp_fields)
    # W8: whatsapp_sent event (optimistic; mirrors the webhook pattern — a later wave can gate on
    # the _send_whatsapp return value). Skip on opt-out (no follow-up is sent).
    if rec.get("outcome") not in ("opt_out",):
        try:
            import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
            await _ev(_vke.whatsapp_sent(rec.get("id", ""), tenant_id, template=rec.get("outcome", "")))
        except Exception:  # noqa: BLE001
            pass
    # WAVE A2: AI-drafted context-aware follow-up (gated by per-campaign wa_followup
    # flag + configured WA env). Falls back to the WAVE3 template path when dormant or
    # when no free-form (Meta) text path is available.
    await _wa_ai_followup(tenant_id, rec, rec["outcome"], camp_fields, tr)
    _score = rec.get("interest", 0) or 0
    # WAVE3 Unit2: mark completed-emitted ONLY when transcript was real here; if it was
    # empty (agent shutdown lags), leave unmarked so the reconciliation sweep re-emits
    # ONCE with the real summary/score.
    if rec.get("_reconciled"):
        rec["_wh_completed"] = True
    await _emit_webhook(tenant_id, "call.completed",
                        {"call_id": rec.get("id"), "phone": rec.get("phone"),
                         "name": rec.get("name", ""), "campaign_id": cid,
                         "campaign_name": rec.get("campaign_name", ""),
                         "outcome": rec["outcome"], "interest": _score, "score": _score,
                         "summary": tr.get("summary", ""), "next_action": tr.get("next_action", ""),
                         "duration_s": rec.get("duration_s", 0), "room": room})
    # W8: call_ended + summary_ready (rich payload -> W14 reducer builds the FactCall the dashboard reads)
    try:
        import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
        await _ev(_vke.call_ended(rec.get("id", ""), tenant_id, duration_s=rec.get("duration_s", 0)))
        await _ev(_vke.summary_ready(
            rec.get("id", ""), tenant_id,
            lifecycle=rec.get("outcome", ""),
            conversion_prob=(_score / 100.0),
            summary=tr.get("summary", ""),
            next_action=tr.get("next_action", ""),
            lead_name=rec.get("name", ""),
            campaign_id=cid,
        ))
    except Exception:  # noqa: BLE001
        pass
    # ── W7 post-call lead lifecycle + AI summary (LEAD_LIFECYCLE_ENABLED) -> CRM updates after each call
    if _LEAD_LIFECYCLE_ON and rec.get("outcome") != "opt_out":
        try:
            await _w7_lifecycle_after_call(tenant_id, rec, tr, outcome, _score)
        except Exception:  # noqa: BLE001 — lifecycle enrichment can NEVER break finalize
            pass
    # WAVE3 Unit2: lead.qualified on a high-interest, real conversation
    if rec["outcome"] != "opt_out" and _score >= 70:
        await _emit_webhook(tenant_id, "lead.qualified",
                            {"call_id": rec.get("id"), "phone": rec.get("phone"),
                             "name": rec.get("name", ""), "campaign_id": cid,
                             "score": _score, "outcome": rec["outcome"],
                             "summary": tr.get("summary", "")})
        # W8: lead classified HOT
        try:
            import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
            await _ev(_vke.lead_classified(rec.get("id", ""), tenant_id, "hot",
                                           conversion_prob=(_score / 100.0)))
        except Exception:  # noqa: BLE001
            pass
        # BUILD#6: HOT-LEAD -> TEAM WHATSAPP. Notify the vendor's handoff team (lead phone +
        # call summary) so a human can take over a hot lead. Reuses whatsapp.py; no-ops if no
        # handoff team or WA dormant. Fire-and-forget, never blocks/raises into the call loop.
        try:
            await notify_handoff_team(
                tenant_id,
                {"name": rec.get("name", ""), "phone": rec.get("phone", "")},
                summary=tr.get("summary", ""), score=_score)
            # W8: handoff requested
            try:
                import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
                await _ev(_vke.handoff_requested(rec.get("id", ""), tenant_id, reason="hot_lead"))
            except Exception:  # noqa: BLE001
                pass
        except Exception as exc:  # noqa: BLE001
            _lg_handoff.warning("hot-lead notify_handoff_team failed: %r", exc)

    # ── COMMUNICATION (W1-P3): post-call founder hot-lead alert + contact auto-summary ──
    # COMMUNICATION-MASTER-PLAN §2.3 / §1.1 / §8 WAVE 1. EARNER LAW (the red-team mandate):
    #   * _finalize_call is AWAITED inside the live dial loop (run_job, ~:2845). So this MUST
    #     NEVER await a network send. We take a PURE-SYNCHRONOUS snapshot of ONLY the fields the
    #     send needs (no ref to the live rec/tr/it the loop keeps mutating, no open files, no db),
    #     then asyncio.create_task a DETACHED fire-and-forget send. The dial loop never waits.
    #   * The detached comm.post_call.run owns a HARD per-channel asyncio.wait_for timeout inside
    #     comm.engine.send — a black-holed Telegram can NEVER keep the task alive past the bound.
    #   * Flag-gated COMM_ENABLED (default OFF => this block is a no-op => resting byte-identical).
    #     The founder-alert / auto-summary sub-flags are checked INSIDE run (also default OFF).
    #   * Snapshot/dispatch is wrapped in its OWN try/except so a comm fault can NEVER disrupt the
    #     finalize hot path (the call still finalizes; rec/tr are already persisted above).
    #   * We DUPLICATE the field reads here (summary/next_action/company/product/agent) — we do
    #     NOT refactor _wa_draft_followup_text (additive+isolated beats DRY when it's the earner).
    if (cfg_get("COMM_ENABLED", "0") or "0").strip().lower() in ("1", "true", "yes", "on"):
        try:
            from comm import post_call as _comm_post_call  # import-guarded; never crashes finalize
            _comm_snap = _comm_post_call.snapshot(            # PURE SYNC — no live-object refs
                rec, tr, camp_fields, tenant_id=tenant_id, call_id=rec.get("id", ""))
            asyncio.create_task(_comm_post_call.run(_comm_snap))  # DETACHED — never awaited here
        except Exception as exc:  # noqa: BLE001 — a comm fault must NEVER disrupt the call loop
            try:
                import logging as _lg_comm_pc
                _lg_comm_pc.getLogger("comm.post_call").warning(
                    "comm post-call hook skipped: %r", type(exc).__name__)
                _log_event("warning", "comm",
                           f"post-call comm hook skipped: {type(exc).__name__}",
                           tenant={"tenant_id": tenant_id}, call_id=rec.get("id", ""),
                           error_type=type(exc).__name__)
            except Exception:  # noqa: BLE001
                pass


def _variant_pool(camp_fields: dict) -> list[str]:
    """Build a weighted round-robin pool of variant ids for a campaign. Empty if no variants."""
    variants = (camp_fields or {}).get("variants") or []
    pool: list[str] = []
    for v in variants:
        if isinstance(v, dict) and v.get("id"):
            pool.extend([v["id"]] * max(1, int(v.get("weight", 1) or 1)))
    return pool


def _variant_by_id(camp_fields: dict, vid: str) -> dict | None:
    for v in (camp_fields or {}).get("variants") or []:
        if isinstance(v, dict) and v.get("id") == vid:
            return v
    return None


async def _job_sleep(job: dict, seconds: float) -> None:
    """Sleep up to `seconds`, but wake within ~0.4s if the operator pressed Stop — so a Stop takes
    effect almost immediately even while the dialer is idling (e.g. queued out-of-window for 60s)."""
    end = time.time() + seconds
    while time.time() < end:
        if job.get("stopped"):
            return
        await asyncio.sleep(min(0.4, max(0.05, end - time.time())))


async def run_job(job_id: str) -> None:
    job = JOBS[job_id]
    cid = job["campaign_id"]
    tenant_id = job.get("tenant_id", ADMIN_ID)
    camp = get_campaign(cid) or {}
    cname = camp.get("name", "")
    camp_fields = camp.get("fields", {}) or {}
    # WAVE3 Unit3: A/B variant assignment (weighted round-robin across dialed leads).
    variant_pool = _variant_pool(camp_fields)
    variant_idx = 0
    # Load these ONCE per job (cheap; not per tick) — risk-note discipline.
    supp = _suppressed_set(tenant_id)
    tenant = _tenant_by_id(tenant_id) or {}
    max_conc = int(tenant.get("max_concurrency", 3))
    daily_cap_tenant = int(tenant.get("daily_call_cap", 500))
    lk = api.LiveKitAPI(url=LK_URL, api_key=LK_KEY, api_secret=LK_SECRET)
    started_ts: list[float] = []
    pending = job["leads"]
    idx = 0
    active: list[dict] = []
    job["state"] = "running"
    try:
        while idx < len(pending) or active:
            # STOP: operator pressed Stop — halt NEW dialing at once, mark the rest "stopped", and let
            # the few in-flight calls drain naturally below (their finalize/recording still runs).
            if job.get("stopped") and not job.get("_stop_applied"):
                job["_stop_applied"] = True
                for _it in pending[idx:]:
                    if _it.get("status") == "queued":
                        _it["status"] = "stopped"
                idx = len(pending)
                job["state"] = "stopping" if active else "stopped"
            now = time.time()
            still = []
            for it in active:
                if now - it["launched_at"] < MIN_CALL_FLOOR or await _phone_present(lk, it["room"]):
                    still.append(it)
                else:
                    it["status"] = "done"
                    ACTIVE_CALLS[tenant_id] = max(0, ACTIVE_CALLS.get(tenant_id, 0) - 1)
                    await _finalize_call(it, now, tenant_id, cid, camp_fields)
            active = still
            # P0.1 window gate: out of window + nothing active -> idle and auto-resume.
            # LPR-FORCE-WINDOW: an AIM/panel "dial now" job (force_window) bypasses this idle
            # so "run <campaign>" actually rings even outside 09-21 IST. SIP/trunk/agent.py
            # are UNTOUCHED — only this compliance idle is skipped for an explicitly-forced job.
            in_win, win = _in_window(camp_fields)
            if job.get("force_window"):
                in_win = True
                job.pop("paused_reason", None)
            elif not in_win:
                job["paused_reason"] = "out_of_window"
                if not active and idx < len(pending):
                    await _job_sleep(job, 60)   # wakes instantly on Stop
                    continue
            else:
                job.pop("paused_reason", None)
            started_ts = [t for t in started_ts if now - t < 86400]
            hourly = len([t for t in started_ts if now - t < 3600])
            daily = len(started_ts)
            while (in_win and idx < len(pending) and len(active) < job["concurrency"]
                   and hourly < job["hourly_cap"] and daily < job["daily_cap"]):
                it = pending[idx]
                if it["status"] != "queued":
                    idx += 1
                    continue
                num = it["num"]
                # P0.2 suppression skip
                if num in supp:
                    it["status"] = "suppressed"; idx += 1
                    record_call({"id": uuid.uuid4().hex[:10], "tenant_id": tenant_id,
                                 "name": it.get("name", ""), "phone": num,
                                 "campaign_id": cid, "campaign_name": cname, "status": "suppressed",
                                 "outcome": "suppressed", "answered": False,
                                 "started_at": _utc_iso(),  # W14-WIRE: tz-labelled UTC
                                 "ended_at": _utc_iso(),    # W14-WIRE: tz-labelled UTC
                                 "duration_s": 0, "room": ""})
                    continue
                # P0.7 per-tenant concurrency cap (stacks under per-job concurrency)
                if ACTIVE_CALLS.get(tenant_id, 0) >= max_conc:
                    break   # leave queued; revisit next tick
                # P0.7 daily cap -> pause job
                if _tenant_usage(tenant_id, _today_iso())["calls"] >= daily_cap_tenant:
                    job["paused_reason"] = "daily_cap_reached"
                    break
                idx += 1
                room = f"famit-{num[1:]}-{uuid.uuid4().hex[:6]}"
                # WAVE3 Unit3: assign an A/B variant (round-robin over the weighted pool).
                md_obj = {"campaign_id": cid, "lead_name": it.get("name", "")}
                v_id, v_label = "", ""
                if variant_pool:
                    v_id = variant_pool[variant_idx % len(variant_pool)]
                    variant_idx += 1
                    vdef = _variant_by_id(camp_fields, v_id) or {}
                    v_label = vdef.get("label", v_id)
                    md_obj["variant_id"] = v_id
                    md_obj["variant_label"] = v_label
                    md_obj["fields_override"] = vdef.get("fields_override") or {}
                # Haptica Flywheel (FLYWHEEL_ENABLED + bandit): let the contextual bandit pick the
                # variant arm via Thompson sampling over the worker's precomputed policy SNAPSHOT (a
                # local dict read — NEVER an inference/ClickHouse call on the live dial path). Dormant
                # or any error ⇒ keep the round-robin choice (byte-identical resting behaviour).
                _fw_arm: dict = {}
                _fw_mod = globals().get("_flywheel_mod")
                if _fw_mod is not None and globals().get("FLYWHEEL_ENABLED"):
                    try:
                        _fw_arm = _fw_mod.select_arm_for_dispatch(tenant_id, cid) or {}
                        if _fw_arm.get("variant_id"):
                            v_id = _fw_arm["variant_id"]
                            vdef = _variant_by_id(camp_fields, v_id) or {}
                            v_label = vdef.get("label", v_id)
                            md_obj["variant_id"] = v_id
                            md_obj["variant_label"] = v_label
                            md_obj["fields_override"] = vdef.get("fields_override") or {}
                    except Exception:  # noqa: BLE001 — the bandit can NEVER break dispatch
                        _fw_arm = {}
                md = json.dumps(md_obj)
                # REC-B: call_id chosen BEFORE create_room so the recording object key embeds it
                # (the call row id == the <call_id> in outbound-recordings/.../<call_id>.ogg).
                _call_id = uuid.uuid4().hex[:10]
                _egress, _rec_key, _rec_bucket = _build_outbound_egress(_call_id)
                try:
                    await lk.room.create_room(api.CreateRoomRequest(
                        name=room, empty_timeout=300, departure_timeout=20, egress=_egress))
                    await lk.agent_dispatch.create_dispatch(api.CreateAgentDispatchRequest(
                        room=room, agent_name=AGENT, metadata=md))
                    _sip_resp = await lk.sip.create_sip_participant(api.CreateSIPParticipantRequest(
                        sip_trunk_id=TRUNK, sip_call_to=num, room_name=room,
                        participant_identity=f"phone-{num}", participant_name=it.get("name") or num,
                        wait_until_answered=False, ringing_timeout=Duration(seconds=45)))
                    _sip_call_id = (getattr(_sip_resp, "sip_call_id", "") or "").strip()
                    it["status"] = "calling"; it["room"] = room; it["launched_at"] = time.time()
                    # W8: call_started event (detached; fire-and-forget; no-op when EVENTBUS off)
                    try:
                        import voice_kernel.events as _vke  # type: ignore  # noqa: PLC0415
                        asyncio.create_task(_ev(_vke.call_started(_call_id, tenant_id, campaign_id=cid)))
                    except Exception:  # noqa: BLE001
                        pass
                    rec = {"id": _call_id, "tenant_id": tenant_id,
                           "name": it.get("name", ""), "phone": num,
                           "campaign_id": cid, "campaign_name": cname, "status": "calling",
                           "variant_id": v_id, "variant_label": v_label,
                           "started_at": _utc_iso(),  # W14-WIRE: tz-labelled UTC
                           "ended_at": "", "duration_s": 0, "room": room,
                           "sip_call_id": _sip_call_id,
                           # REC-B server-side auto-egress handle (additive). egress_id is assigned
                           # asynchronously by LiveKit (auto-egress returns none at room-create), so the
                           # AUTHORITATIVE handle is the deterministic recording_key (object lands there).
                           "recording_key": _rec_key,
                           "recording_bucket": _rec_bucket,
                           "recording_status": ("recording" if _egress is not None else "disabled")}
                    # REC-B-AZURE marker (additive): when the Azure backend armed this egress, stamp
                    # recording_backend="azure" so the read/serve path presigns via Azure SAS instead
                    # of DO Spaces (recording_bucket above is then the Azure container). Flag-off ->
                    # this key is simply absent and every reader stays on the default DO Spaces path.
                    if _egress is not None and _azure_recording_enabled():
                        rec["recording_backend"] = "azure"
                    it["_rec"] = rec
                    # Haptica Flywheel: record the FINAL policy arm (model/voice/variant/propensity)
                    # on the call record so the engine can correlate outcome↔arm later. Pure metadata,
                    # flag-gated, never raises. (_fw_mod/_fw_arm set in the dispatch block above.)
                    if _fw_mod is not None and globals().get("FLYWHEEL_ENABLED"):
                        try:
                            _fw_mod.stamp_arm(rec, camp_fields, md_obj, _fw_arm)
                        except Exception:  # noqa: BLE001
                            pass
                    record_call(rec)
                    ACTIVE_CALLS[tenant_id] = ACTIVE_CALLS.get(tenant_id, 0) + 1
                    active.append(it); started_ts.append(time.time()); hourly += 1; daily += 1
                except Exception as exc:  # noqa: BLE001
                    it["status"] = "failed"; it["error"] = repr(exc)[:140]
                    _log_event("error", "dialer",
                               f"call dial failed for {num}: {exc!r}",
                               tenant=({"tenant_id": tenant_id}), error_type=type(exc).__name__,
                               context={"phone": num, "campaign": cname, "campaign_id": cid})
                    record_call({"id": uuid.uuid4().hex[:10], "tenant_id": tenant_id,
                                 "name": it.get("name", ""), "phone": num,
                                 "campaign_id": cid, "campaign_name": cname, "status": "failed",
                                 "started_at": _utc_iso(),  # W14-WIRE: tz-labelled UTC
                                 "ended_at": "", "duration_s": 0})
            await _job_sleep(job, 4)
        job["state"] = "stopped" if job.get("stopped") else "done"
    finally:
        await lk.aclose()


# ---------- JSON API (frontend hits these via nginx /api -> /) ----------
# HRD #9: dependency-aware health probe. Bare {status:ok} kept as the DEFAULT response
# shape (back-compat: callers that just check 200 still pass when healthy), but the route
# now LIVE-checks DB + redis + LiveKit reachability and returns 503 when a hard dependency
# (DB) is down so the watchdog/load-balancer can see "degraded". Every check is bounded and
# fully exception-guarded — /health itself can NEVER hang or 500. `?deep=0` returns the old
# cheap liveness probe (process-up) for high-frequency pings.
def _hc_db() -> tuple[bool, str]:
    """Live bounded SELECT 1 (NOT the cached startup flag). True/err-string."""
    try:
        from db import engine as _eng  # local import (import-safe; mirrors the rest of caller.py)
        if not _eng.available():
            return False, "engine_unavailable"
        from sqlalchemy import text
        with _eng.session(tenant_id="", is_admin=True) as s:
            s.execute(text("SET LOCAL statement_timeout = '1500ms'"))
            s.execute(text("SELECT 1"))
        return True, ""
    except Exception as exc:  # noqa: BLE001
        return False, repr(exc)[:120]


def _hc_redis() -> tuple[bool, str]:
    """Ping the rate-limiter redis (:6380). Soft dependency — degraded, not fatal."""
    try:
        import redis as _r  # type: ignore
        url = os.getenv("RATELIMIT_REDIS_URL", "redis://127.0.0.1:6380/0")
        cli = _r.from_url(url, socket_connect_timeout=1.0, socket_timeout=1.0)
        return (bool(cli.ping()), "")
    except Exception as exc:  # noqa: BLE001
        return False, repr(exc)[:120]


def _hc_livekit() -> tuple[bool, str]:
    """TCP-reachability of the LiveKit signalling host (ws[s]://host[:port]). Soft dependency."""
    try:
        import socket
        from urllib.parse import urlparse
        raw = cfg_get("LIVEKIT_URL", "ws://127.0.0.1:7880") or "ws://127.0.0.1:7880"
        u = urlparse(raw if "://" in raw else "ws://" + raw)
        host = u.hostname or "127.0.0.1"
        port = u.port or (443 if (u.scheme or "").endswith("s") else 7880)
        with socket.create_connection((host, port), timeout=1.5):
            return True, ""
    except Exception as exc:  # noqa: BLE001
        return False, repr(exc)[:120]


@app.get("/health")
async def health(deep: int = 1):
    # cheap liveness (process up) — for high-frequency pings or load-balancer wiring.
    if not deep:
        return {"status": "ok"}
    db_ok, db_err = await asyncio.to_thread(_hc_db)
    redis_ok, redis_err = await asyncio.to_thread(_hc_redis)
    lk_ok, lk_err = await asyncio.to_thread(_hc_livekit)
    # DB is the only HARD dependency (data plane). redis/livekit are soft (the app FAILS-OPEN
    # on redis and the earner runs in its own process) -> they mark 'degraded' but keep 200
    # unless DB is also down. DB down => 503 so the watchdog/LB pulls the node.
    checks = {
        "db": {"ok": db_ok, **({"error": db_err} if db_err else {})},
        "redis": {"ok": redis_ok, **({"error": redis_err} if redis_err else {})},
        "livekit": {"ok": lk_ok, **({"error": lk_err} if lk_err else {})},
    }
    healthy = db_ok and redis_ok and lk_ok
    status = "ok" if healthy else ("unhealthy" if not db_ok else "degraded")
    code = 200 if db_ok else 503
    return JSONResponse({"status": status, "checks": checks}, status_code=code)


@app.get("/metrics")
async def metrics():
    """Prometheus exposition (no auth — standard scrape endpoint; exposes only
    aggregate counters/latency/cost, never secrets or per-tenant data). Returns a
    tiny stub if prometheus_client is unavailable."""
    if _obs_mod is None:
        return Response(content="# obs module unavailable\nfamit_up 1\n",
                        media_type="text/plain; version=0.0.4; charset=utf-8")
    body, ctype = _obs_mod.render()
    return Response(content=body, media_type=ctype)


def _login_blocked_by_status(tenant: dict | None) -> bool:
    """CONTROL LAYER (CL-B3 / control-security §5.1): a suspended/disabled vendor cannot mint a token.
    Admins are NEVER blocked (anti-lockout). Gated behind CONTROL_ENABLED so resting login is unchanged.
    Suspension's INSTANT kill is auth.revoke_all (next call 401); this closes the fresh-login door too."""
    if not tenant or tenant.get("is_admin"):
        return False
    if _client_blocked(tenant):          # file-based status/demo gate (works without PG)
        return True
    if not CONTROL_ENABLED or _ent_mod is None:
        return False
    try:
        st = _ent_mod.load_status(tenant.get("tenant_id", "")).get("status", "active")
    except Exception:  # noqa: BLE001
        return False
    return st in ("suspended", "disabled")


def _blocked_login_response(t: dict | None) -> JSONResponse:
    """Distinct 403 for an EXPIRED DEMO vs a normally suspended/deactivated account."""
    if t and t.get("demo") and (_demo_remaining_s(t) or 0) <= 0:
        return JSONResponse(
            {"error": "Your demo account has expired. Please contact your admin to continue.",
             "code": "demo_expired"}, status_code=403)
    return JSONResponse(
        {"error": "This account has been deactivated. Please contact your admin.",
         "code": "suspended"}, status_code=403)


@app.post("/login")
async def login(request: Request, password: str = Form(""), email: str = Form("")):
    """Accepts BOTH:
       - {password}            legacy panel login (FamitCall2026) -> admin tenant
       - {email, password}     vendor / tenant login
    Returns {token, tenant_id, name, is_admin}. Token = tenant_id.hmac(tenant_id, SECRET).
    """
    email = (email or "").strip().lower()
    # Legacy: bare password with no email == admin login. Keep this working forever.
    if not email and password == PW:
        t = _tenant_by_id(ADMIN_ID) or ADMIN_TENANT
        return JSONResponse({"token": _make_token(t["tenant_id"]), "tenant_id": t["tenant_id"],
                             "name": t.get("name", ""), "is_admin": True, "role": _role_of(t)})
    if email:
        t = next((x for x in _read_tenants() if (x.get("email") or "").lower() == email), None)
        if t and t.get("pass_hash") == _hash_pw(password, t.get("salt", "")):
            if _login_blocked_by_status(t):
                return _blocked_login_response(t)
            return JSONResponse({"token": _make_token(t["tenant_id"]), "tenant_id": t["tenant_id"],
                                 "name": t.get("name", ""), "is_admin": bool(t.get("is_admin")),
                                 "role": _role_of(t)})
        # Admin may also log in via its email + the legacy password.
        if t and t.get("is_admin") and password == PW:
            return JSONResponse({"token": _make_token(t["tenant_id"]), "tenant_id": t["tenant_id"],
                                 "name": t.get("name", ""), "is_admin": True, "role": _role_of(t)})
    return JSONResponse({"error": "invalid credentials"}, status_code=401)


# ---------- P0: JWT access + rotating refresh (ADDITIVE; legacy /login above stays) ----------
@app.post("/auth/login")
async def auth_login(request: Request, email: str = Form(""), password: str = Form("")):
    """JWT login. Returns {access_token, refresh_token, token_type, expires_in,
    tenant_id, role, is_admin, name}. Same credentials as /login (bare PW -> admin,
    or email+password). The legacy /login endpoint is unchanged and still works."""
    if not AUTH_JWT_READY or _auth_mod is None:
        return JSONResponse({"error": "jwt auth unavailable"}, status_code=503)
    pair = _auth_mod.login((email or "").strip().lower(), password or "")
    if not pair:
        return JSONResponse({"error": "invalid credentials"}, status_code=401)
    _bt = _tenant_by_id(pair.get("tenant_id", ""))
    if _login_blocked_by_status(_bt):
        return _blocked_login_response(_bt)
    return JSONResponse(pair)


@app.post("/auth/refresh")
async def auth_refresh(request: Request, refresh_token: str = Form("")):
    """Rotate tokens: revoke the presented refresh token and return a NEW
    access+refresh pair. 401 if the refresh token is unknown/expired/revoked."""
    if not AUTH_JWT_READY or _auth_mod is None:
        return JSONResponse({"error": "jwt auth unavailable"}, status_code=503)
    rt = (refresh_token or "").strip()
    if not rt:
        # also accept JSON body or Authorization: Bearer for convenience
        rt = (request.headers.get("x-refresh-token", "") or "").strip()
    pair = _auth_mod.refresh(rt)
    if not pair:
        return JSONResponse({"error": "invalid or expired refresh token"}, status_code=401)
    return JSONResponse(pair)


@app.post("/auth/logout")
async def auth_logout(request: Request, refresh_token: str = Form("")):
    """Revoke a refresh token (idempotent). Access tokens expire on their own."""
    if _auth_mod is None:
        return JSONResponse({"ok": True})
    rt = (refresh_token or "").strip() or (request.headers.get("x-refresh-token", "") or "").strip()
    revoked = _auth_mod.logout(rt)
    return JSONResponse({"ok": True, "revoked": bool(revoked)})


@app.get("/me")
async def me(request: Request):
    """Current identity + role. Frontend uses role to show/hide actions."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    out = {"tenant_id": t["tenant_id"], "email": t.get("email", ""),
           "name": t.get("name", ""), "role": _role_of(t),
           "is_admin": bool(t.get("is_admin")),
           "status": (t.get("status") or "active"),
           "restricted": list(t.get("restricted") or [])}
    if t.get("demo"):
        out["demo"] = True
        out["demo_minutes"] = int(t.get("demo_minutes") or 0)
        out["demo_remaining_s"] = int(_demo_remaining_s(t) or 0)
    return JSONResponse(out)


# ---------- F2: Business Brain + Knowledge Base (additive; tenant-scoped; org_id from token only) ----------
# org_id is ALWAYS t["tenant_id"] (resolve_tenant), NEVER a body/param (platform-business-brain RT-5).
# All routes degrade cleanly when brain/kb modules are absent. Nothing here touches an existing route.
@app.get("/brain")
async def brain_get(request: Request):
    """Full Business Brain profile for the caller's org. {} when none yet."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _brain_mod is None:
        return JSONResponse({"profile": {}, "note": "brain module unavailable"})
    return JSONResponse({"profile": _brain_mod.get_profile(t["tenant_id"]),
                         "completeness": _brain_mod.completeness(t["tenant_id"])})


@app.put("/brain")
async def brain_put(request: Request):
    """Upsert (merge) the Business Brain profile. Versioned + audited. write-role gated."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    if _brain_mod is None:
        return JSONResponse({"error": "brain module unavailable"}, status_code=503)
    try:
        patch = await request.json()
        if not isinstance(patch, dict):
            return JSONResponse({"error": "body must be a JSON object"}, status_code=400)
    except Exception:  # noqa: BLE001
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    # never honour a body-supplied org_id/id (RT-5): strip before merge.
    patch.pop("org_id", None)
    patch.pop("id", None)
    prof = _brain_mod.upsert_profile(t["tenant_id"], patch, actor=t["tenant_id"])
    return JSONResponse({"profile": prof,
                         "completeness": _brain_mod.completeness(t["tenant_id"])})


@app.get("/brain/handoff")
async def brain_handoff_get(request: Request):
    """The vendor's HUMAN HANDOFF team list (warm-transfer + hot-lead WhatsApp targets).
    [{phone, whatsapp, role, hours, priority}, ...]. Lives on the Business Brain `handoff` block."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    return JSONResponse({"handoff": _handoff_get(t["tenant_id"])})


@app.put("/brain/handoff")
async def brain_handoff_put(request: Request):
    """Replace the vendor's handoff team list. write-role gated. Body = a JSON array of
    {phone, whatsapp, role, hours, priority}, or {"handoff":[...]} / {"team":[...]}."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    if _brain_mod is None:
        return JSONResponse({"error": "brain module unavailable"}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    if isinstance(body, dict):
        team = body.get("handoff") or body.get("team") or body.get("numbers") or []
    elif isinstance(body, list):
        team = body
    else:
        return JSONResponse({"error": "body must be a JSON array or {handoff:[...]}"},
                            status_code=400)
    if not isinstance(team, list):
        return JSONResponse({"error": "handoff must be an array"}, status_code=400)
    stored = _handoff_set(t["tenant_id"], team, actor=t["tenant_id"])
    return JSONResponse({"handoff": stored})


@app.post("/brain/handoff/add")
async def brain_handoff_add(request: Request):
    """ADD/UPDATE a SINGLE handoff team member (vs PUT which replaces the whole list).
    write-role gated; tenant from TOKEN only (never body). Body = {phone, whatsapp?, role?,
    hours?, priority?, enabled?}. Phone MUST be a valid +91 Indian mobile. Idempotent: re-adding
    the same number updates it. Returns the full priority-sorted list."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    if _brain_mod is None:
        return JSONResponse({"error": "brain module unavailable"}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    if not isinstance(body, dict):
        return JSONResponse({"error": "body must be a JSON object {phone,...}"}, status_code=400)
    body.pop("org_id", None)              # RT-5: never honour a body-supplied tenant
    body.pop("tenant_id", None)
    stored, err = _handoff_add_one(t["tenant_id"], body, actor=t["tenant_id"])
    if err:
        return JSONResponse({"error": err, "handoff": stored}, status_code=400)
    return JSONResponse({"handoff": stored, "ok": True})


@app.delete("/brain/handoff/remove")
async def brain_handoff_remove(request: Request):
    """REMOVE a SINGLE handoff member by phone. write-role gated; tenant from TOKEN only.
    Phone via ?phone= or JSON body {phone}. Returns the remaining priority-sorted list +
    removed:bool (idempotent — removing a missing number is a no-op, not an error)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    if _brain_mod is None:
        return JSONResponse({"error": "brain module unavailable"}, status_code=503)
    phone = (request.query_params.get("phone") or "").strip()
    if not phone:
        try:
            body = await request.json()
            if isinstance(body, dict):
                phone = str(body.get("phone", "") or "").strip()
        except Exception:  # noqa: BLE001
            phone = ""
    stored, err, removed = _handoff_remove_one(t["tenant_id"], phone, actor=t["tenant_id"])
    if err:
        return JSONResponse({"error": err, "handoff": stored}, status_code=400)
    return JSONResponse({"handoff": stored, "removed": removed, "ok": True})


@app.post("/handoff/notify")
async def handoff_notify(request: Request):
    """Internal: fire the HOT-LEAD WhatsApp to the vendor's handoff team. Used by the inbound
    voice agent's warm-transfer fallback (and any caller) to drop lead phone+summary into the
    team's chat. write-role gated; reuses notify_handoff_team. Returns the send report."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    lead = {"name": str((body or {}).get("name", "") or ""),
            "phone": str((body or {}).get("phone", "") or "")}
    summary = str((body or {}).get("summary", "") or "")
    try:
        score = int((body or {}).get("score", 0) or 0)
    except Exception:  # noqa: BLE001
        score = 0
    res = await notify_handoff_team(t["tenant_id"], lead, summary=summary, score=score)
    return JSONResponse(res)


@app.get("/brain/completeness")
async def brain_completeness(request: Request):
    """Readiness score + missing fields (onboarding checklist + hallucination guard)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _brain_mod is None:
        return JSONResponse({"score": 0, "missing": [], "note": "brain module unavailable"})
    return JSONResponse(_brain_mod.completeness(t["tenant_id"]))


@app.post("/brain/knowledge")
async def brain_add_knowledge(request: Request, content: str = Form(""), title: str = Form(""),
                              doc_type: str = Form("generic"), kind: str = Form("paste")):
    """Ingest a long-form knowledge doc into the KB corpus (business-scoped) for the caller's org.
    Chunk + FTS + (optional, dormant-until-key) embed -> upsert. Heavy work runs OFF the event loop."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    if _brain_mod is None:
        return JSONResponse({"ok": False, "reason": "brain module unavailable"}, status_code=503)
    if not (content or "").strip():
        return JSONResponse({"ok": False, "reason": "empty content"}, status_code=400)
    # embed() may do a network round-trip -> run the whole ingest off the uvicorn loop (event-loop safety).
    res = await asyncio.to_thread(_brain_mod.add_knowledge, t["tenant_id"], content,
                                  title=title, kind=kind, doc_type=doc_type, actor=t["tenant_id"])
    return JSONResponse(res)


@app.get("/brain/retrieve")
async def brain_retrieve(request: Request, q: str = "", k: int = 4):
    """Hybrid (FTS + dense-when-configured) retrieve over the caller's org KB. RLS-scoped.
    [] when KB/PG absent. Off the voice hot path (voice uses precomputed blob — later unit)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _brain_mod is None or not (q or "").strip():
        return JSONResponse({"results": []})
    res = await asyncio.to_thread(_brain_mod.retrieve, t["tenant_id"], q, max(1, min(int(k or 4), 20)))
    return JSONResponse({"results": res})


@app.post("/kb/seed-telecaller")
async def kb_seed_telecaller(request: Request):
    """SUPER-ADMIN ONLY: (re)seed the shared `_global` telecaller knowledge corpus — universal,
    business-neutral telecaller BEHAVIOUR (objection handlers, rapport/backchannel, close techniques,
    pricing/value framing, vertical explainer scaffolds), read-shared into every tenant's recall.

    IDEMPOTENT: each corpus entry dedups by sha256(content) under `_global` -> re-running (or hitting
    this twice) writes nothing the second time (returns duplicate counts). Lets the founder re-seed /
    refresh from the panel after editing kb/seed_global_corpus.json on the box.

    `_global` WRITE-LOCK: the seeder ingests under is_admin=True — the ONLY path the kb_chunks RLS
    `WITH CHECK` permits to write `_global` (a tenant request can never reach here: this is
    require_super_admin-gated, and the write itself sets the admin GUC). Heavy work (chunk+FTS+upsert,
    one PG round-trip per entry) runs OFF the uvicorn loop via asyncio.to_thread."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _kb_mod is None:
        return JSONResponse({"ok": False, "reason": "kb module unavailable"}, status_code=503)
    res = await asyncio.to_thread(_kb_mod.seed_global_corpus, actor=t.get("tenant_id", "admin"))
    return JSONResponse(res)


# ============================================================================
# RAG W3 — KB MANAGEMENT (the founder-facing /knowledge control surface)
# Token-derived tenant (resolve_tenant -> t["tenant_id"], NEVER a body/param). Every PG touch runs
# under the caller's OWN GUC (engine.session(tenant_id=t["tenant_id"], is_admin=False)) so RLS gates
# every other tenant out; the shared `_global` corpus is read-shared (kb_chunks RLS USING) and surfaced
# explicitly. All heavy work (chunk/FTS/upsert, multi-roundtrip reads) runs OFF the uvicorn loop via
# asyncio.to_thread. These are ADDITIVE + ISOLATED: nothing here imports the voice run-path, touches
# agent.py, or restarts famit-agent. Degrade cleanly to empty shapes when kb/PG is absent.
# ============================================================================

# PDF text extraction — light, dependency-OPTIONAL. pypdf is a pure-Python wheel (no native deps);
# when absent the upload still works for text and returns a clear reason for a PDF (never 500s).
def _kb_extract_pdf_text(data: bytes) -> tuple[str, str]:
    """Best-effort PDF -> plain text. Returns (text, reason). reason='' on success; a short code
    otherwise ('pdf_parser_unavailable' | 'pdf_parse_failed' | 'pdf_empty'). NEVER raises."""
    try:
        from pypdf import PdfReader  # pure-python; optional
    except Exception:  # noqa: BLE001
        return "", "pdf_parser_unavailable"
    try:
        import io as _io
        reader = PdfReader(_io.BytesIO(data))
        parts: list[str] = []
        for page in reader.pages[:300]:  # bound a malicious 10k-page PDF
            try:
                parts.append(page.extract_text() or "")
            except Exception:  # noqa: BLE001
                continue
        text = "\n\n".join(p.strip() for p in parts if p and p.strip()).strip()
        return (text, "") if text else ("", "pdf_empty")
    except Exception:  # noqa: BLE001
        return "", "pdf_parse_failed"


@app.get("/kb/sources")
async def kb_sources_list(request: Request, scope_campaign_id: str = ""):
    """List THIS tenant's KB sources + the shared `_global` sources, each with chunk count + status.
    Token-derived tenant; RLS-scoped (own GUC, is_admin=False) — the only rows returned are the
    caller's own + `_global`. Read-only. {sources:[{id,title,kind,scope,channel_scope,status,
    kb_version,chunks,is_shared,scope_campaign_id,created_at,updated_at}], total, global_count}."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    tid = t["tenant_id"]

    def _q() -> dict:
        try:
            from db import engine as _eng
            from sqlalchemy import text as _sql
            if not _eng.available():
                return {"sources": [], "total": 0, "global_count": 0, "reason": "pg_unavailable"}
            # ensure the KB tables exist (idempotent; no-op when already applied)
            if _kb_mod is not None:
                try:
                    _kb_mod.ensure_schema()
                except Exception:  # noqa: BLE001
                    pass
            params: dict = {"tid": tid}
            camp_sql = ""
            if scope_campaign_id:
                # match sources whose chunks are campaign-scoped (or business-wide)
                camp_sql = (" AND (s.scope = 'business' OR s.scope = :cscope "
                            "OR EXISTS (SELECT 1 FROM kb_chunks kc WHERE kc.source_id = s.id "
                            "AND (kc.scope_campaign_id = :cid OR kc.scope_campaign_id = '')))")
                params["cscope"] = f"campaign:{scope_campaign_id}"
                params["cid"] = scope_campaign_id
            # RLS USING already read-shares own-tenant + `_global`; the explicit predicate keeps the
            # set to exactly those two (defence in depth; never a `%` wildcard, never is_admin=True).
            with _eng.session(tenant_id=tid, is_admin=False) as s:
                rows = s.execute(_sql(
                    "SELECT s.id, s.title, s.kind, s.scope, s.channel_scope, s.status, "
                    "s.kb_version, s.tenant_id, s.created_at, s.updated_at, "
                    "(SELECT count(*) FROM kb_chunks c WHERE c.source_id = s.id) AS chunks "
                    "FROM kb_sources s "
                    "WHERE (s.tenant_id = :tid OR s.tenant_id = '_global')"
                    + camp_sql +
                    " ORDER BY (s.tenant_id = '_global'), s.created_at DESC LIMIT 500"),
                    params).fetchall()
            out = []
            gcount = 0
            for r in rows:
                is_shared = (r[7] == "_global")
                if is_shared:
                    gcount += 1
                out.append({
                    "id": r[0], "title": r[1] or "(untitled)", "kind": r[2], "scope": r[3],
                    "channel_scope": r[4], "status": r[5], "kb_version": int(r[6] or 1),
                    "is_shared": is_shared, "chunks": int(r[10] or 0),
                    "created_at": r[8].isoformat() if r[8] else "",
                    "updated_at": r[9].isoformat() if r[9] else ""})
            return {"sources": out, "total": len(out), "global_count": gcount}
        except Exception as exc:  # noqa: BLE001
            return {"sources": [], "total": 0, "global_count": 0,
                    "reason": f"error:{type(exc).__name__}"}

    return JSONResponse(await asyncio.to_thread(_q))


@app.post("/kb/upload")
async def kb_upload(request: Request,
                    text: str = Form(""),
                    title: str = Form(""),
                    doc_type: str = Form("generic"),
                    channel_scope: str = Form("all"),
                    scope_campaign_id: str = Form(""),
                    pdf: UploadFile | None = File(None)):
    """Ingest collateral into THIS tenant's KB. Accepts a `text` field OR a `pdf` upload (parsed via
    pypdf, graceful-degrade). Chunks + FTS-indexes + upserts via kb.ingest under the tenant's own GUC
    (RLS-scoped). Optionally tagged to a campaign (scope_campaign_id) so retrieval can scope to it.
    Token-derived tenant; write-gated. Heavy work runs OFF the loop. NEVER writes `_global`.

    Returns {ok, source_id, document_id, chunks, embedded, reason, title}."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot add knowledge")
    if _kb_mod is None:
        return JSONResponse({"ok": False, "reason": "kb module unavailable"}, status_code=503)

    body = (text or "").strip()
    kind = "paste"
    src_title = (title or "").strip()[:300]

    # PDF branch (optional) — sniff, size-guard, extract.
    if pdf is not None:
        try:
            data = await pdf.read()
        except Exception as exc:  # noqa: BLE001
            return JSONResponse({"ok": False, "reason": f"read_failed:{type(exc).__name__}"},
                                status_code=400)
        if not data:
            return JSONResponse({"ok": False, "reason": "empty_file"}, status_code=400)
        if len(data) > 20 * 1024 * 1024:  # keep the shared GIN index lean (quota: RAG plan §7-12)
            return JSONResponse({"ok": False, "reason": "file_too_large_max_20mb"}, status_code=413)
        if not (data[:5] == b"%PDF-" or (pdf.content_type or "").lower() == "application/pdf"):
            return JSONResponse({"ok": False, "reason": "not_a_pdf"}, status_code=415)
        ptext, preason = _kb_extract_pdf_text(data)
        if preason:
            code = 503 if preason == "pdf_parser_unavailable" else 422
            return JSONResponse({"ok": False, "reason": preason}, status_code=code)
        # PDF text augments / supplies the body; an explicit text field (if any) prepends.
        body = (body + "\n\n" + ptext).strip() if body else ptext
        kind = "file"
        if not src_title:
            fn = (pdf.filename or "document.pdf")
            src_title = ("".join(ch for ch in fn if ch.isalnum() or ch in "-_. ").strip()
                         or "document.pdf")[:300]

    if not body:
        return JSONResponse({"ok": False, "reason": "empty_content"}, status_code=400)
    if len(body) > 200_000:  # per-doc max-size guard (RAG plan §7-12)
        return JSONResponse({"ok": False, "reason": "content_too_large_max_200k_chars"},
                            status_code=413)
    if not src_title:
        src_title = (body.strip().splitlines()[0] if body.strip() else "Knowledge")[:80]

    scope = f"campaign:{scope_campaign_id}" if scope_campaign_id else "business"
    dt = (doc_type or "generic").strip().lower()[:40] or "generic"
    chan = (channel_scope or "all").strip().lower()
    if chan not in ("all", "voice", "whatsapp", "support", "creative"):
        chan = "all"

    # ingest under the tenant's OWN GUC (is_admin=False) — RLS WITH CHECK pins the write to this
    # tenant; a `_global` write is impossible from here. Off the uvicorn loop (embed may RTT).
    res = await asyncio.to_thread(
        _kb_mod.ingest, t["tenant_id"], body,
        title=src_title, kind=kind, scope=scope, doc_type=dt, channel_scope=chan,
        scope_campaign_id=scope_campaign_id, is_admin=False)
    res = dict(res or {})
    res["title"] = src_title
    _audit(request, t, "kb.upload", "kb_source", res.get("source_id", ""),
           meta={"chunks": res.get("chunks"), "kind": kind, "scope": scope})
    status_code = 200 if res.get("ok") else 422
    return JSONResponse(res, status_code=status_code)


@app.post("/kb/test-retrieve")
async def kb_test_retrieve(request: Request):
    """THE differentiator: the founder types a question + (optional) channel/campaign and SEES exactly
    which chunks ground the answer — the brain lighting up. Runs the SAME FTS-only (dense=False) path
    the live voice `lookup` uses, under the tenant's OWN GUC (own tenant + `_global`, RLS-gated).
    Logs the query to kb_query_log (grounded flag) so the gap loop learns from real probes too.

    Body {query, channel?, campaign?, top_k?}. Returns
    {query, grounded, count, chunks:[{id,source_id,document_id,section,snippet,score,leg,is_shared}]}."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _kb_mod is None:
        return JSONResponse({"grounded": False, "chunks": [], "count": 0,
                             "reason": "kb module unavailable"}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    query = (str(body.get("query") or "")).strip()
    if not query:
        return JSONResponse({"error": "query is required"}, status_code=400)
    channel = (str(body.get("channel") or "all")).strip().lower() or "all"
    if channel not in ("all", "voice", "whatsapp", "support", "creative"):
        channel = "all"
    campaign = (str(body.get("campaign") or body.get("scope_campaign_id") or "")).strip()
    try:
        top_k = max(1, min(int(body.get("top_k") or 6), 12))
    except Exception:  # noqa: BLE001
        top_k = 6
    tid = t["tenant_id"]

    def _run() -> dict:
        # FTS-only forever on this surface (dense=False) — mirrors the live `lookup` contract (C-3):
        # ZERO embed RTT regardless of EMBED_API_KEY. include_global so the `_global` corpus is
        # visible (what the live voice path actually retrieves).
        hits = _kb_mod.retrieve(
            tid, query, top_k=top_k, channel=channel, scope_campaign_id=campaign,
            dense=False, include_global=True, is_admin=False)
        out = []
        for h in (hits or []):
            content = h.get("content") or ""
            out.append({
                "id": h.get("source_id", ""),  # source_id for the UI to link to a source card
                "source_id": h.get("source_id", ""),
                "document_id": h.get("document_id", ""),
                "section": h.get("section", ""),
                "snippet": (content[:280] + ("…" if len(content) > 280 else "")),
                "score": round(float(h.get("score") or 0.0), 5),
                "leg": h.get("leg", ""),
            })
        grounded = bool(out)
        # learn from the probe too (best-effort; off-loop already — we're inside to_thread).
        try:
            _kb_mod.log_query(tid, query, channel=channel, scope_campaign_id=campaign,
                              grounded=grounded, leg=(out[0]["leg"] if out else ""),
                              top_ids=[h.get("document_id", "") for h in (hits or [])],
                              is_admin=False)
        except Exception:  # noqa: BLE001
            pass
        return {"query": query, "grounded": grounded, "count": len(out), "chunks": out}

    return JSONResponse(await asyncio.to_thread(_run))


@app.get("/kb/gaps")
async def kb_gaps(request: Request, days: int = 30, limit: int = 50):
    """Knowledge-gap loop: recent queries from kb_query_log that grounded NOTHING (grounded=false),
    aggregated by normalized query — 'the questions your AI couldn't answer'. The most sellable
    artifact (RAG plan §7-5). Token-derived tenant; STRICTLY per-tenant (kb_query_log has NO
    `_global` read-share). Read-only. {gaps:[{query, count, last_seen, channels}], total, window_days}."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    tid = t["tenant_id"]
    try:
        win = max(1, min(int(days or 30), 365))
    except Exception:  # noqa: BLE001
        win = 30
    try:
        lim = max(1, min(int(limit or 50), 200))
    except Exception:  # noqa: BLE001
        lim = 50

    def _q() -> dict:
        try:
            from db import engine as _eng
            from sqlalchemy import text as _sql
            if not _eng.available():
                return {"gaps": [], "total": 0, "window_days": win, "reason": "pg_unavailable"}
            if _kb_mod is not None:
                try:
                    _kb_mod.ensure_schema()
                except Exception:  # noqa: BLE001
                    pass
            with _eng.session(tenant_id=tid, is_admin=False) as s:
                rows = s.execute(_sql(
                    "SELECT lower(btrim(query)) AS q, count(*) AS n, max(created_at) AS last_seen, "
                    "array_agg(DISTINCT channel) AS channels "
                    "FROM kb_query_log "
                    "WHERE tenant_id = :tid AND grounded = false "
                    "AND created_at >= now() - make_interval(days => :w) "
                    "AND btrim(query) <> '' "
                    "GROUP BY lower(btrim(query)) "
                    "ORDER BY n DESC, last_seen DESC LIMIT :lim"),
                    {"tid": tid, "w": win, "lim": lim}).fetchall()
            gaps = [{
                "query": r[0],
                "count": int(r[1] or 0),
                "last_seen": r[2].isoformat() if r[2] else "",
                "channels": [c for c in (r[3] or []) if c],
            } for r in rows]
            return {"gaps": gaps, "total": len(gaps), "window_days": win}
        except Exception as exc:  # noqa: BLE001
            return {"gaps": [], "total": 0, "window_days": win,
                    "reason": f"error:{type(exc).__name__}"}

    return JSONResponse(await asyncio.to_thread(_q))


# ============================================================================
# CRM CORE — contact spine + unified timeline + next-best-action (additive; read-model).
# All X-Auth, tenant-scoped (org_id == t["tenant_id"], NEVER a body/param), RBAC per can().
# PG work runs OFF the uvicorn loop (asyncio.to_thread) — project/rebuild do several round-trips.
# These NEVER write leads and NEVER touch the run-path. Degrade to empty shapes when crm/PG absent.
# ============================================================================
@app.get("/contacts")
async def contacts_list(request: Request, stage: str = "", hot: str = "", q: str = "",
                        segment: str = "", sort: str = "last_activity_at", limit: int = 100,
                        sort_by: str = "", order: str = "", offset: int = 0):
    """List/filter/segment contacts for the caller's org. {contacts:[...], total, offset,
    limit, next}. `offset` cursor-pages the list (the panel CRM workspace loads ONE page
    then fetches the next) — without it the infinite-scroll re-requested page 0 and showed
    duplicate rows for books larger than one page."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _crm_mod is None:
        return JSONResponse({"contacts": [], "total": 0, "note": "crm module unavailable"})
    hot_f = None
    if str(hot).lower() in ("1", "true", "yes"):
        hot_f = True
    elif str(hot).lower() in ("0", "false", "no"):
        hot_f = False
    # R5P4-3 (ADDITIVE): the panel CRM sends sort_by + order. crm.list_contacts only sorts on
    # score|stage|last_activity_at, so translate the FE column to that vocabulary (temperature/score
    # -> score; stage -> stage; everything else -> last_activity_at), pass it as `sort`, and then apply
    # `order` + the FE-only columns (name/last_outcome) as a stable post-sort on the returned rows.
    _sb = (sort_by or "").strip().lower()
    if _sb:
        sort = {"temperature": "score", "score": "score", "stage": "stage",
                "last_activity_at": "last_activity_at"}.get(_sb, "last_activity_at")
    res = await asyncio.to_thread(
        lambda: _crm_mod.list_contacts(t["tenant_id"], stage=stage, hot=hot_f, q=q, sort=sort,
                                       limit=max(1, min(int(limit or 100), 1000)),
                                       offset=max(0, int(offset or 0)),
                                       is_admin=bool(t.get("is_admin"))))
    if _sb and isinstance(res, dict) and isinstance(res.get("contacts"), list):
        _desc = (order or "").strip().lower() != "asc"

        def _ck(c):  # noqa: ANN001
            if _sb in ("score", "temperature"):
                return int(c.get("score") or 0)
            if _sb == "name":
                return (c.get("name", "") or "").lower()
            if _sb in ("stage",):
                return (c.get("stage", "") or "").lower()
            if _sb in ("campaign",):
                return (c.get("campaign_id") or c.get("campaign") or "").lower()
            if _sb in ("last_outcome",):
                return (c.get("last_outcome", "") or "").lower()
            return (c.get("last_activity_at") or "") or ""
        try:
            res["contacts"] = sorted(res["contacts"], key=_ck, reverse=_desc)
        except Exception:  # noqa: BLE001 — a sort hiccup never drops the list
            pass
    return JSONResponse(res)


@app.get("/contacts/{phone}")
async def contacts_get(request: Request, phone: str):
    """One contact: full profile + the authoritative lead + the rule-based next-best-action.
    `phone` is any phone form (canonicalized) OR a ct_ contact id."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _crm_mod is None:
        return JSONResponse({"contact": None, "note": "crm module unavailable"}, status_code=503)
    org = t["tenant_id"]
    adm = bool(t.get("is_admin"))
    # PERF UNIT-2: project on read (FRESHNESS-GATED rebuild — at most once per contact / TTL, so a
    # repeat open is a fast cached read, no full timeline rebuild + N+1 transcript disk reads). The
    # projection ALSO hands back the timeline it just read (`_timeline`) so we DON'T issue a second
    # get_timeline; next_best_action is pure rules (no DB) -> compute inline. Off the event loop.
    c = await asyncio.to_thread(lambda: _crm_mod.project_contact(org, phone, is_admin=adm))
    if c is None:
        c = await asyncio.to_thread(lambda: _crm_mod.get_contact(org, phone, is_admin=adm))
    if c is None:
        return JSONResponse({"error": "contact not found"}, status_code=404)
    # reuse the timeline from the projection (newest-first, up to 500) when present; else one read.
    tl_full = c.pop("_timeline", None)
    if tl_full is None:
        tl_full = await asyncio.to_thread(
            lambda: _crm_mod.get_timeline(org, c["id"], limit=50, is_admin=adm))
    tl = tl_full[:50]
    nba = await asyncio.to_thread(lambda: _crm_mod.next_best_action(org, c, timeline=tl, is_admin=adm))
    # REC-FIX: attach this lead's recordings (UNIFIED, both directions, newest-first) directly on the
    # profile so the CRM lead page shows the player URL the moment the object exists — no separate lazy
    # "Load recordings" fetch needed. Reuses the proven /contacts/{phone}/recordings shaper (HEAD-verify
    # + presign + stuck-status self-heal). Tenant-pinned (RLS inbound / tenant filter outbound). NEVER
    # raises -> a degraded recordings side returns [] and the profile still renders.
    recs: list[dict] = []
    try:
        phone_n = norm(c.get("phone", "") or phone) or (c.get("phone", "") or phone or "")
        def _lead_recs() -> list[dict]:
            out: list[dict] = []
            try:
                for cc in calls_for(t):
                    cp = cc.get("phone", "") or ""
                    if norm(cp) == phone_n or cp == phone_n:
                        out.append(_outbound_rec_item(cc))
            except Exception:  # noqa: BLE001
                pass
            try:
                out.extend(_inbound_rec_items(org, phone_n))
            except Exception:  # noqa: BLE001
                pass
            out.sort(key=lambda x: (x.get("started_at", "") or ""), reverse=True)
            return out
        recs = await asyncio.to_thread(_lead_recs)
    except Exception:  # noqa: BLE001
        recs = []
    n_play = sum(1 for x in recs if x.get("playable"))
    return JSONResponse({"contact": c, "timeline": tl, "nba": nba,
                         "recordings": recs, "recordings_playable": n_play})


@app.get("/contacts/{phone}/timeline")
async def contacts_timeline(request: Request, phone: str, kinds: str = "", limit: int = 100):
    """The full chronological interaction history (newest-first) for a person."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _crm_mod is None:
        return JSONResponse({"timeline": [], "note": "crm module unavailable"})
    org = t["tenant_id"]
    adm = bool(t.get("is_admin"))
    kinds_l = [k.strip() for k in kinds.split(",") if k.strip()] or None
    cid = phone if str(phone).startswith("ct_") else _crm_mod.contact_id(org, phone)
    tl = await asyncio.to_thread(
        lambda: _crm_mod.get_timeline(org, cid, limit=max(1, min(int(limit or 100), 1000)),
                                      kinds=kinds_l, is_admin=adm))
    return JSONResponse({"timeline": tl, "contact_id": cid})


@app.get("/contacts/{phone}/nba")
async def contacts_nba(request: Request, phone: str):
    """The recommended next action for a contact (deterministic rules; no metered call)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _crm_mod is None:
        return JSONResponse({"action": "none", "reason": "crm module unavailable"})
    org = t["tenant_id"]
    adm = bool(t.get("is_admin"))
    c = await asyncio.to_thread(lambda: _crm_mod.get_contact(org, phone, is_admin=adm))
    if c is None:
        return JSONResponse({"error": "contact not found"}, status_code=404)
    tl = await asyncio.to_thread(lambda: _crm_mod.get_timeline(org, c["id"], limit=50, is_admin=adm))
    nba = await asyncio.to_thread(lambda: _crm_mod.next_best_action(org, c, timeline=tl, is_admin=adm))
    return JSONResponse(nba)


@app.put("/contacts/{phone}")
async def contacts_update(request: Request, phone: str):
    """Update contact name/email/tags/stage (a CONTACT-only manual override — NEVER writes leads).
    Body = JSON {name?,email?,tags?[],stage?,data?{}}. write-role gated."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    if _crm_mod is None:
        return JSONResponse({"error": "crm module unavailable"}, status_code=503)
    try:
        body = await request.json()
        if not isinstance(body, dict):
            return JSONResponse({"error": "body must be a JSON object"}, status_code=400)
    except Exception:  # noqa: BLE001
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    # never honour a body-supplied org_id/id/phone_key (identity is derived from the path).
    for k in ("org_id", "id", "phone_key"):
        body.pop(k, None)
    org = t["tenant_id"]
    adm = bool(t.get("is_admin"))
    c = await asyncio.to_thread(
        lambda: _crm_mod.update_contact(
            org, phone, name=body.get("name"), email=body.get("email"),
            tags=body.get("tags"), stage=body.get("stage"), data=body.get("data"), is_admin=adm))
    if c is None:
        return JSONResponse({"error": "contact not found / crm unavailable"}, status_code=404)
    return JSONResponse({"ok": True, "contact": c})


# ---------- P0: read-only audit log (admin; tenant-scoped; paginated) ----------
@app.get("/audit")
async def get_audit(request: Request, limit: int = 100, offset: int = 0,
                    action: str = "", channel: str = ""):
    """Append-only audit trail. Admin sees ALL events; a non-admin sees only its
    own tenant's events. Read-only + paginated (newest first). `channel` filters by
    event channel (e.g. ?channel=ai for AI-decision rows — F4 §7)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _audit_mod is None:
        return JSONResponse({"events": [], "total": 0, "limit": limit, "offset": offset,
                             "note": "audit module unavailable"})
    scope = None if t.get("is_admin") else t["tenant_id"]
    return JSONResponse(_audit_mod.tail(limit=limit, offset=offset,
                                        tenant_id=scope, action_prefix=action, channel=channel))


# ============================================================================
# SYSTEM LOGS & ERRORS — super-admin observability surface (white-labeled "System Logs").
# Backed by logging_service (shared /data JSONL the voice agent ALSO writes to). All routes
# are super-admin-gated, read-only, and degrade to a clean empty shape (never 500) when the
# module is unavailable. `/admin/logs/summary` is declared BEFORE `/admin/logs/{event_id}` so
# "summary" is not captured as an id.
# ============================================================================
@app.get("/admin/logs")
async def admin_logs(request: Request, limit: int = 100, offset: int = 0, level: str = "",
                     source: str = "", tenant_id: str = "", q: str = "", since: str = ""):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"events": [], "total": 0, "limit": limit, "offset": offset,
                             "note": "logging module unavailable"})
    return JSONResponse(_log_mod.tail(limit=limit, offset=offset, level=level, source=source,
                                      tenant_id=tenant_id, q=q, since=since))


@app.get("/admin/logs/summary")
async def admin_logs_summary(request: Request):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"by_level": {}, "total": 0, "last_24h": 0, "errors_24h": 0,
                             "top_errors": [], "note": "logging module unavailable"})
    return JSONResponse(_log_mod.summary())


@app.get("/admin/notifications")
async def admin_notifications(request: Request, after: int = 0, limit: int = 30):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"events": [], "latest_seq": 0, "unread": 0, "unread_errors": 0})
    try:
        _after = int(after or 0)
    except Exception:  # noqa: BLE001
        _after = 0
    return JSONResponse(_log_mod.notifications(after_seq=_after, limit=limit))


# NOTE: /admin/logs/health and /admin/logs/test are declared BEFORE /admin/logs/{event_id}
# so "health"/"test" are matched as routes, not captured as an event id.
@app.get("/admin/logs/health")
async def admin_logs_health(request: Request):
    """Operator self-test for System Logs: is the store live, where does it write, can it write,
    how many events are buffered, the current seq, and whether Telegram/AI-fix are configured.
    Answers the #1 question 'is logging even on?' in one call. 503 when not ready/writable."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"ready": False, "note": "logging module unavailable",
                             "init_ok": bool(globals().get("LOG_READY", False))}, status_code=503)
    h = _log_mod.health()
    h["init_ok"] = bool(globals().get("LOG_READY", False))
    code = 200 if (h.get("ready") and h.get("writable")) else 503
    return JSONResponse(h, status_code=code)


@app.post("/admin/logs/test")
async def admin_logs_test(request: Request):
    """Emit a synthetic event so an operator can SEE capture working end-to-end (it should appear
    in the Logs tab immediately). Returns the stored event. Super-admin only."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"ok": False, "note": "logging module unavailable"}, status_code=503)
    tid = t.get("tenant_id", "") if isinstance(t, dict) else ""
    ev = _log_mod.record("info", "selftest",
                         "System Logs self-test — if you can see this, capture is working.",
                         tenant_id=tid, error_type="selftest",
                         context={"by": tid or "super-admin", "ip": _client_ip(request)})
    return JSONResponse({"ok": bool(ev), "event": ev or {}})


@app.post("/admin/logs/client")
async def admin_logs_client(request: Request):
    """Ingest a CLIENT-SIDE (panel) error so UI/runtime/fetch failures surface in System Logs
    alongside backend events. Any authenticated user may report (scoped to their tenant; super-
    admin sees all). Source is FORCED to 'frontend'; level is clamped; message/stack are bounded.
    Best-effort: always returns 200 ({ok}) so a failed report can never cascade into the UI."""
    try:
        t = resolve_tenant(request)
    except Exception:  # noqa: BLE001
        t = None
    if not isinstance(t, dict):
        return JSONResponse({"ok": False}, status_code=401)
    if _log_mod is None:
        return JSONResponse({"ok": False})
    try:
        body = await request.json()
        if not isinstance(body, dict):
            body = {}
    except Exception:  # noqa: BLE001
        body = {}
    lvl = str(body.get("level", "error") or "error").strip().lower()
    if lvl not in ("info", "warning", "error", "critical"):
        lvl = "error"
    msg = str(body.get("message", "") or "")[:1000] or "client error"
    etype = str(body.get("error_type", "") or "ClientError")[:120]
    ctx = {
        "where": "frontend",
        "url": str(body.get("url", "") or "")[:500],
        "stack": str(body.get("stack", "") or "")[:1500],
        "kind": str(body.get("kind", "") or "")[:60],   # error | unhandledrejection | render | fetch
        "ua": (request.headers.get("user-agent", "") or "")[:300],
        "ip": _client_ip(request),
    }
    extra = body.get("context")
    if isinstance(extra, dict):
        for k, v in list(extra.items())[:10]:
            ctx[str(k)[:40]] = str(v)[:300]
    _log_mod.record(lvl, "frontend", msg, tenant_id=t.get("tenant_id", ""),
                    error_type=etype, context=ctx)
    return JSONResponse({"ok": True})


@app.get("/admin/logs/{event_id}")
async def admin_log_detail(request: Request, event_id: str):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"error": "logging module unavailable"}, status_code=404)
    ev = _log_mod.get(event_id)
    if not ev:
        return JSONResponse({"error": "not_found"}, status_code=404)
    return JSONResponse(ev)


@app.post("/admin/logs/{event_id}/suggest")
async def admin_log_suggest(request: Request, event_id: str, force: str = ""):
    """Generate (and cache) an AI 'why + how to fix' suggestion for the event's fingerprint.
    `?force=1` (the panel's Regenerate button) bypasses the per-fingerprint cache and re-asks the LLM."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _log_mod is None:
        return JSONResponse({"suggestion": ""})
    _force = str(force).strip().lower() in ("1", "true", "yes", "on")
    # suggest_fix does a blocking Groq call -> offload so the event loop is never stalled.
    suggestion = await asyncio.to_thread(_log_mod.suggest_fix, event_id, force=_force)
    return JSONResponse({"suggestion": suggestion})


@app.get("/bookings")
async def list_bookings(request: Request, limit: int = 200):
    """Site-visit bookings captured by the voice agent (BC1 fast-capture -> /data/bookings.jsonl).
    A tenant sees ONLY its own; admin sees all. Read-only, newest-first. Degrades to an empty
    list (never 500) when nothing has been captured yet. (The full availability/calendar engine
    + a richer panel surface land later; this exposes the captured slots in the meantime.)"""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    scope = None if t.get("is_admin") else t.get("tenant_id", "")

    def _load():
        out = []
        try:
            p = VAR / "bookings.jsonl"
            if p.exists():
                for ln in p.read_text(encoding="utf-8").splitlines():
                    ln = ln.strip()
                    if not ln:
                        continue
                    try:
                        b = json.loads(ln)
                    except Exception:  # noqa: BLE001
                        continue
                    if not isinstance(b, dict):
                        continue
                    if scope is not None and (b.get("tenant_id") or "") != scope:
                        continue
                    out.append(b)
        except Exception:  # noqa: BLE001
            return []
        out.reverse()  # newest first
        return out[:max(1, min(limit, 1000))]

    rows = await asyncio.to_thread(_load)
    return JSONResponse({"bookings": rows, "total": len(rows)})


# ============================================================================
# PERFORMANCE — white-labeled metrics for the super-admin "Performance" page. A thin proxy to
# the observability droplet's Prometheus HTTP API (PROM_URL). Super-admin-gated, read-only.
# The panel renders native charts from the returned series — the vendor (Prometheus/Grafana/
# SigNoz) is never named or exposed. Degrades to 503 (never 500) when PROM_URL is unset.
# ============================================================================
async def _prom_proxy(path: str, params: dict):
    base = (os.getenv("PROM_URL") or "").strip().rstrip("/")
    if not base:
        return JSONResponse({"status": "error", "error": "metrics backend not configured"},
                            status_code=503)
    try:
        import httpx as _httpx
        async with _httpx.AsyncClient(timeout=10.0) as _c:
            r = await _c.get(base + path, params=params)
        try:
            body = r.json()
        except Exception:  # noqa: BLE001
            body = {"status": "error", "error": "bad upstream response"}
        return JSONResponse(body, status_code=r.status_code)
    except Exception:  # noqa: BLE001
        return JSONResponse({"status": "error", "error": "metrics backend unreachable"},
                            status_code=502)


@app.get("/admin/metrics/instant")
async def admin_metrics_instant(request: Request, query: str):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    return await _prom_proxy("/api/v1/query", {"query": query})


@app.get("/admin/metrics/range")
async def admin_metrics_range(request: Request, query: str, minutes: int = 60, step: int = 60):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    import time as _t
    mins = max(1, min(minutes, 10080))   # honor up to 7d (matches the ClickHouse panels)
    end = int(_t.time())
    start = end - mins * 60
    # derive a step that keeps ~60-360 points (step=0 from the client => auto) so a 7d window
    # doesn't ask Prometheus for tens of thousands of points.
    if step and step > 0:
        st = max(15, min(step, 21600))
    else:
        st = 60 if mins <= 60 else 300 if mins <= 360 else 900 if mins <= 1440 else 3600 if mins <= 10080 else 21600
    return await _prom_proxy("/api/v1/query_range",
                             {"query": query, "start": start, "end": end, "step": st})


# ============================================================================
# OBSERVABILITY ANALYTICS — read-only ClickHouse (trace/APM) queries for the native, white-
# labeled System Logs (Traces/Requests) + Performance (APM) dashboards. Super-admin-gated.
# All degrade to {"error":...,"rows":[]} (never 500) when obs is unconfigured. See obs_query.py.
# ============================================================================
def _obs_guard(request: Request):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _obs_q is None:
        return JSONResponse({"error": "observability not configured", "rows": []}, status_code=503)
    return None


@app.get("/admin/obs/services")
async def admin_obs_services(request: Request, minutes: int = 1440):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.services(minutes))


@app.get("/admin/obs/summary")
async def admin_obs_summary(request: Request, minutes: int = 60, service: str = ""):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.summary(minutes, service))


@app.get("/admin/obs/red")
async def admin_obs_red(request: Request, minutes: int = 60, service: str = ""):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.red_timeseries(minutes, service))


@app.get("/admin/obs/routes")
async def admin_obs_routes(request: Request, minutes: int = 60, service: str = "", limit: int = 50):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.top_routes(minutes, service, limit))


@app.get("/admin/obs/status")
async def admin_obs_status(request: Request, minutes: int = 60, service: str = ""):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.status_dist(minutes, service))


@app.get("/admin/obs/service-dist")
async def admin_obs_service_dist(request: Request, minutes: int = 60):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.service_dist(minutes))


@app.get("/admin/obs/errors")
async def admin_obs_errors(request: Request, minutes: int = 60, service: str = "", limit: int = 20):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.error_ops(minutes, service, limit))


@app.get("/admin/obs/traces")
async def admin_obs_traces(request: Request, minutes: int = 60, service: str = "",
                           errors_only: int = 0, q: str = "", limit: int = 60):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(
        await _obs_q.traces(minutes, service, errors_only, q, limit))


@app.get("/admin/obs/trace/{trace_id}")
async def admin_obs_trace(request: Request, trace_id: str):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.trace_detail(trace_id))


# ── P1 Voice Performance Analytics (reads over the agent-written haptica_voice_* CH tables) ──
def _voice_filters(tenant_id: str = "", campaign_id: str = "", agent_name: str = "",
                   phone: str = "", provider: str = "", model: str = "",
                   status: str = "", stage: str = "") -> dict:
    return {"tenant_id": tenant_id, "campaign_id": campaign_id, "agent_name": agent_name,
            "phone": phone, "provider": provider, "model": model, "status": status, "stage": stage}


@app.get("/admin/obs/voice/summary")
async def admin_obs_voice_summary(request: Request, minutes: int = 60, tenant_id: str = "",
                                  campaign_id: str = "", agent_name: str = "", phone: str = "",
                                  provider: str = "", model: str = "", status: str = "", stage: str = ""):
    g = _obs_guard(request)
    if g is not None:
        return g
    f = _voice_filters(tenant_id, campaign_id, agent_name, phone, provider, model, status, stage)
    return JSONResponse(await _obs_q.voice_summary(minutes, f))


@app.get("/admin/obs/voice/red")
async def admin_obs_voice_red(request: Request, minutes: int = 60, tenant_id: str = "",
                              campaign_id: str = "", agent_name: str = "", phone: str = "",
                              provider: str = "", model: str = "", status: str = "", stage: str = ""):
    g = _obs_guard(request)
    if g is not None:
        return g
    f = _voice_filters(tenant_id, campaign_id, agent_name, phone, provider, model, status, stage)
    return JSONResponse(await _obs_q.voice_red_timeseries(minutes, f))


@app.get("/admin/obs/voice/calls")
async def admin_obs_voice_calls(request: Request, minutes: int = 60, limit: int = 100,
                                tenant_id: str = "", campaign_id: str = "", agent_name: str = "",
                                phone: str = "", provider: str = "", model: str = "",
                                status: str = "", stage: str = ""):
    g = _obs_guard(request)
    if g is not None:
        return g
    f = _voice_filters(tenant_id, campaign_id, agent_name, phone, provider, model, status, stage)
    return JSONResponse(await _obs_q.voice_calls(minutes, f, limit))


@app.get("/admin/obs/voice/filters")
async def admin_obs_voice_filters(request: Request, minutes: int = 1440):
    g = _obs_guard(request)
    return g if g is not None else JSONResponse(await _obs_q.voice_filter_options(minutes))


@app.get("/admin/obs/voice/stack")
async def admin_obs_voice_stack(request: Request, minutes: int = 1440, tenant_id: str = "",
                                campaign_id: str = "", agent_name: str = ""):
    """The AI stack + versions actually in use over the window + each stage's metrics."""
    g = _obs_guard(request)
    if g is not None:
        return g
    f = _voice_filters(tenant_id, campaign_id, agent_name)
    return JSONResponse(await _obs_q.voice_stack(minutes, f))


# NOTE: /call/{call_id} is declared as a sub-path so it never collides with the static routes above.
@app.get("/admin/obs/voice/call/{call_id}")
async def admin_obs_voice_call(request: Request, call_id: str):
    g = _obs_guard(request)
    if g is not None:
        return g
    detail = await _obs_q.voice_call_detail(call_id)
    timeline = await _obs_q.voice_turn_timeline(call_id)
    latency = await _obs_q.voice_call_latency(call_id)
    return JSONResponse({"detail": (detail.get("rows") or [{}])[0] if not detail.get("error") else {},
                         "timeline": timeline.get("rows", []),
                         "latency": latency.get("row", {}),
                         **({"error": detail.get("error")} if detail.get("error") else {})})


@app.get("/admin/obs/voice/call/{call_id}/quality")
async def admin_obs_voice_call_quality(request: Request, call_id: str, force: int = 0, cached: int = 0):
    """LLM CONTENT-quality analysis of ONE call's transcript (repetition/hanging/off-script/goal).
    call_id == the LiveKit room, so the transcript lives at transcripts/{call_id}.json. The result is
    CACHED per call (one paid LLM analysis); ?force=1 re-runs; ?cached=1 PEEKS (returns the cached
    result or not_analyzed, never spends). Dormant-safe; never 500s."""
    g = _obs_guard(request)
    if g is not None:
        return g
    if _tq is None:
        return JSONResponse({"ok": False, "error": "unavailable",
                             "message": "Transcript analysis is not available."}, status_code=503)
    cid = (call_id or "").strip()
    safe = re.sub(r"[^A-Za-z0-9_.-]", "_", cid)[:120]
    cache_path = VAR / "transcript_qa" / f"{safe}.json"
    if not force:
        prev = _read(cache_path, None)
        if isinstance(prev, dict) and prev.get("ok"):
            prev["cached"] = True
            return JSONResponse(prev)
    if cached:   # peek-only: do NOT run a paid analysis
        return JSONResponse({"ok": False, "error": "not_analyzed"})
    turns = _outbound_transcript_turns({"room": cid})
    if not turns:
        return JSONResponse({"ok": False, "error": "no_transcript",
                             "message": "No transcript saved for this call yet."})
    res = await _tq.analyze(turns, {"campaign": cid})
    if res.get("ok"):
        try:
            (VAR / "transcript_qa").mkdir(parents=True, exist_ok=True)
            _write(cache_path, res)
        except Exception:  # noqa: BLE001
            pass
    return JSONResponse(res)


# ════════════════════════════════════════════════════════════════════════════
# FAMIT RESEARCH (W-RES) — "instrumented conversation science" READ routes. Purely additive;
# touch no existing route. Tenant is TOKEN-derived (resolve_tenant), NEVER from the body — and
# research_query binds WHERE tenant_id = {tid:String} on every ClickHouse read (ClickHouse has no
# RLS, so the Python scope IS the boundary). The HEAVY pipeline (acoustic extraction + the affect
# filter) runs POST-CALL off the recording egress in a separate process, so these routes never
# touch the live turn loop. FAMIT_RESEARCH_ENABLED gates WRITING (the recorder); READS always work
# and FALL BACK to a clearly-labelled `demo:true` dataset (the real filter over scripted archetype
# calls) so the premium dashboard is alive day-one instead of an all-zeros dead page.
# ════════════════════════════════════════════════════════════════════════════
try:
    import research_query as _research_q  # type: ignore
except Exception:  # noqa: BLE001
    _research_q = None

_RESEARCH_ON = (os.getenv("FAMIT_RESEARCH_ENABLED", "0") or "0").strip().lower() in ("1", "true", "yes", "on")


@app.get("/research/dashboard")
async def research_dashboard(request: Request, minutes: int = 1440):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _research_q is None:
        return JSONResponse({"error": "research module unavailable"}, status_code=503)
    try:
        data = await _research_q.dashboard(t["tenant_id"], minutes)
    except Exception as exc:  # noqa: BLE001 — never 500 the panel on a metrics hiccup
        return JSONResponse({"error": str(exc)[:200], "summary": {}, "calls": []})
    data["enabled"] = _RESEARCH_ON
    return JSONResponse(data)


@app.get("/research/call/{call_id}")
async def research_call(request: Request, call_id: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _research_q is None:
        return JSONResponse({"error": "research module unavailable"}, status_code=503)
    try:
        return JSONResponse(await _research_q.call_detail(t["tenant_id"], call_id))
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": str(exc)[:200], "call": {}, "turns": []})


@app.get("/research/health")
async def research_health(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    return JSONResponse({"enabled": _RESEARCH_ON, "module_loaded": _research_q is not None})


# ============================================================================
# F4: Credit/Wallet ledger + Action Firewall — ADDITIVE endpoints (tenant-scoped).
# org_id is ALWAYS t["tenant_id"] (resolve_tenant), NEVER a body/param. All routes degrade to a clean
# shape (never 500) when wallet/firewall are unavailable. None of these touch the existing run-path.
# Money crosses the API in MAJOR units (rupees); the wallet core stores INTEGER MINOR units (paise).
# ============================================================================
def _minor_to_major(m) -> float:
    try:
        return round(int(m) / 100.0, 2)
    except Exception:  # noqa: BLE001
        return 0.0


def _wallet_unavailable_body() -> dict:
    return {"wallet_available": False, "note": "wallet ledger unavailable (Postgres down or module absent)"}


def _step_up_guard(request: Request, scope: str, tenant: dict):
    """Apply the Action Firewall step-up gate. Returns an error Response to RETURN (deny), or None to
    proceed. Pass-through when firewall is OFF / unavailable / tenant has no PIN (non-breaking)."""
    if _firewall_mod is None:
        return None
    try:
        _firewall_mod.require_step_up(request, scope, tenant)
        return None
    except _firewall_mod.StepUpDenied as d:
        return JSONResponse(d.body, status_code=d.status)
    except Exception:  # noqa: BLE001
        return None  # firewall must never hard-break a request


@app.get("/wallet")
async def wallet_get(request: Request):
    """Wallet balance for the caller's org (major units). Clean shape when unavailable / no account."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _wallet_mod is None or not _wallet_mod.available():
        return JSONResponse(_wallet_unavailable_body())
    bal = await asyncio.to_thread(_wallet_mod.balance, t["tenant_id"], "INR", bool(t.get("is_admin")))
    plan = (_billing_for(t["tenant_id"]) or {}).get("plan", "postpaid") if "_billing_for" in globals() else "postpaid"
    if bal is None:
        return JSONResponse({"tenant_id": t["tenant_id"], "currency": "INR", "available": 0.0,
                             "held": 0.0, "plan": plan, "lifetime_topup": 0.0, "lifetime_spend": 0.0,
                             "wallet_available": True, "note": "no wallet account yet"})
    return JSONResponse({
        "tenant_id": t["tenant_id"], "currency": bal["currency"], "plan": plan,
        "available": _minor_to_major(bal["available_minor"]),
        "held": _minor_to_major(bal["held_minor"]),
        "lifetime_topup": _minor_to_major(bal["lifetime_topup_minor"]),
        "lifetime_spend": _minor_to_major(bal["lifetime_spend_minor"]),
        "wallet_available": True,
    })


@app.get("/wallet/ledger")
async def wallet_ledger(request: Request, limit: int = 100):
    """Wallet transaction ledger (newest first) for the caller's org. Amounts in major units."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _wallet_mod is None or not _wallet_mod.available():
        return JSONResponse({"transactions": [], "total": 0, **_wallet_unavailable_body()})
    rows = await asyncio.to_thread(_wallet_mod.transactions, t["tenant_id"], limit, bool(t.get("is_admin")))
    out = [{
        "id": r["id"], "kind": r["kind"], "amount": _minor_to_major(r["amount_minor"]),
        "held_delta": _minor_to_major(r["held_delta_minor"]),
        "resource_type": r["resource_type"], "resource_id": r["resource_id"],
        "balance_after": _minor_to_major(r["balance_after_minor"]), "hold_id": r["hold_id"], "at": r["at"],
    } for r in rows]
    return JSONResponse({"transactions": out, "total": len(out)})


@app.get("/wallet/holds")
async def wallet_holds(request: Request, state: str = ""):
    """Open/closed reservations for the caller's org. Amounts in major units."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _wallet_mod is None or not _wallet_mod.available():
        return JSONResponse({"holds": [], **_wallet_unavailable_body()})
    rows = await asyncio.to_thread(_wallet_mod.holds, t["tenant_id"], state, bool(t.get("is_admin")))
    out = [{
        "id": r["id"], "amount": _minor_to_major(r["amount_minor"]), "state": r["state"],
        "resource_type": r["resource_type"], "resource_id": r["resource_id"],
        "settled": (_minor_to_major(r["settled_minor"]) if r["settled_minor"] is not None else None),
        "expires_at": r["expires_at"], "at": r["at"],
    } for r in rows]
    return JSONResponse({"holds": out})


@app.post("/wallet/topup/{tenant_id}")
async def wallet_topup(request: Request, tenant_id: str, amount: float = Form(...),
                       payment_ref: str = Form("")):
    """ADMIN credit to a tenant's wallet (major units rupees). Idempotent on payment_ref so a webhook
    retry can't double-credit. Gated by the Action Firewall (spend scope) when FIREWALL_ENABLED + a PIN
    is set on the ACTING admin. Audited."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "manage_tenants"):
        return _forbidden("admin only")
    # Action Firewall: a topup is spend-sensitive. Gate on the ACTING tenant's step-up.
    denied = _step_up_guard(request, "spend", t)
    if denied is not None:
        _audit(request, t, "firewall.stepup.denied", "wallet", tenant_id,
               meta={"scope": "spend", "action": "wallet.topup"})
        return denied
    if _wallet_mod is None or not _wallet_mod.available():
        return JSONResponse(_wallet_unavailable_body(), status_code=503)
    minor = int(round(float(amount) * 100))
    if minor <= 0:
        return JSONResponse({"ok": False, "reason": "amount must be positive"}, status_code=400)
    idem = f"topup:{payment_ref}" if payment_ref else f"topup:manual:{secrets.token_hex(8)}"
    res = await asyncio.to_thread(_wallet_mod.topup, tenant_id, minor, t["tenant_id"], idem, "INR", True, None)
    _audit(request, t, "wallet.topup", "wallet", tenant_id,
           meta={"amount_minor": minor, "payment_ref": payment_ref, "ok": bool(res.get("ok"))})
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    return JSONResponse({
        "ok": True, "tenant_id": tenant_id, "credited": _minor_to_major(minor),
        "available": _minor_to_major(res.get("available_minor", 0)),
        "held": _minor_to_major(res.get("held_minor", 0)),
    })


@app.put("/firewall/pin")
async def firewall_set_pin(request: Request, pin: str = Form(...)):
    """Set/replace the caller's own Action-Firewall PIN (salted-hash stored; never the raw PIN). Audited."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _firewall_mod is None or not _firewall_mod.available():
        return JSONResponse({"ok": False, "reason": "firewall unavailable"}, status_code=503)
    res = _firewall_mod.set_pin(t["tenant_id"], pin)
    if res.get("ok"):
        _audit(request, t, "firewall.pin.set", "firewall", t["tenant_id"])
    return JSONResponse(res, status_code=200 if res.get("ok") else 400)


@app.post("/firewall/pin/change")
async def firewall_change_pin(request: Request,
                              old_pin: str = Form(...), new_pin: str = Form(...)):
    """Rotate the caller's OWN Action-Firewall PIN. Verifies `old_pin` against the EXISTING salted hash
    (firewall.check_pin — unchanged), then sets `new_pin` (firewall.set_pin — unchanged). Tenant/role
    scoped (token-derived; never a body field). Brute-force protected: ~5 wrong old-PIN attempts inside a
    rolling 15-min window -> time-boxed lockout (429). Every outcome is audited into the immutable events
    leg. ADDITIVE — does not alter the existing PIN verify / step-up logic."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    # role scope: writing one's own security PIN requires write capability (same axis as /firewall/pin
    # setters and the AIM numbers writers). Read-only roles cannot rotate the PIN.
    if not can(t, "write"):
        return _forbidden()
    if _firewall_mod is None or not _firewall_mod.available():
        return JSONResponse({"ok": False, "reason": "firewall unavailable"}, status_code=503)
    tid = t["tenant_id"]
    res = _firewall_mod.change_pin(tid, old_pin, new_pin)
    if res.get("ok"):
        _audit(request, t, "firewall.pin.change", "firewall", tid)
        return JSONResponse({"ok": True}, status_code=200)
    reason = res.get("reason", "")
    if reason == "locked":
        _audit(request, t, "firewall.pin.change.locked", "firewall", tid,
               meta={"retry_after_s": res.get("retry_after_s", 0)})
        return JSONResponse(res, status_code=429)
    if reason == "invalid old PIN":
        _audit(request, t, "firewall.pin.change.fail", "firewall", tid,
               meta={"fails": res.get("fails", 0), "locked": res.get("locked", False)})
        return JSONResponse(res, status_code=401)
    # remaining reasons (no PIN set / new == old / pin length) are client validation -> 400
    return JSONResponse(res, status_code=400)


@app.post("/firewall/verify-pin")
async def firewall_verify_pin(request: Request, pin: str = Form(...), scope: str = Form("spend")):
    """Verify the caller's PIN; on match mint a short-TTL step-up token (the client then replays the
    gated action with X-Step-Up: <token>). On mismatch -> 401 + an audited firewall.stepup.fail row."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _firewall_mod is None or not _firewall_mod.available():
        return JSONResponse({"ok": False, "reason": "firewall unavailable"}, status_code=503)
    if not _firewall_mod.has_pin(t["tenant_id"]):
        return JSONResponse({"ok": False, "reason": "no PIN set"}, status_code=400)
    if not _firewall_mod.check_pin(t["tenant_id"], pin):
        _audit(request, t, "firewall.stepup.fail", "firewall", t["tenant_id"], meta={"scope": scope})
        return JSONResponse({"ok": False, "error": "invalid PIN"}, status_code=401)
    tok = _firewall_mod.mint_step_up(t["tenant_id"], scope)
    _audit(request, t, "firewall.stepup.ok", "firewall", t["tenant_id"], meta={"scope": scope})
    return JSONResponse({"ok": True, **(tok or {})})


@app.get("/firewall/status")
async def firewall_status(request: Request):
    """Firewall enrollment + flag state for the caller's org."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _firewall_mod is None:
        return JSONResponse({"pin_set": False, "firewall_enabled": False, "available": False})
    return JSONResponse(_firewall_mod.status(t["tenant_id"]))


@app.post("/extract")
async def extract(request: Request, brief: str = Form("")):
    if not authed(request):
        return need_auth()
    return JSONResponse(extract_fields(brief or ""))


# ════════════════════════════════════════════════════════════════════════════
# BRAND KITS — tenant-scoped persistence for the Creative Studio brand page (R4 A5)
# ════════════════════════════════════════════════════════════════════════════
# The panel brand page (app/assets) binds /api/assets/brand-kits -> this backend /brand-kits.
# Before R4 there was NO backend, so "save brand kit" silently dropped. Persisted as a tenant-scoped
# JSON file (var/brand_kits/<tenant>.json) — same control-plane pattern as ai-manager sessions /
# tenants: NO PG, NO DDL, byte-identical-when-untouched, fully isolated per tenant (the file path IS
# the tenant boundary; tenant_id comes ONLY from the token via resolve_tenant, never a body field).
_BRAND_KITS_DIR = VAR / "brand_kits"


def _brand_kits_path(tenant_id: str) -> Path:
    safe = re.sub(r"[^A-Za-z0-9_.-]", "_", str(tenant_id or "")) or "unknown"
    return _BRAND_KITS_DIR / f"{safe}.json"


def _read_brand_kits(tenant_id: str) -> list:
    data = _read(_brand_kits_path(tenant_id), [])
    return data if isinstance(data, list) else []


@app.get("/brand-kits")
async def brand_kits_list(request: Request):
    """List the calling tenant's saved brand kits. Tenant derived from token ONLY."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    return JSONResponse({"brand_kits": _read_brand_kits(t["tenant_id"])})


@app.post("/brand-kits")
async def brand_kits_save(request: Request):
    """Create or update a brand kit for the calling tenant. Body is a JSON brand-kit object
    (name, colors, fonts, logo_url, voice/tone, etc.). If it carries an `id` that already exists
    the kit is UPDATED in place; otherwise a new one is appended with a fresh id. The org_id is
    SERVER-DERIVED from the token (never trusted from the body)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    if not isinstance(body, dict):
        return JSONResponse({"ok": False, "reason": "bad_body"}, status_code=400)
    tenant_id = t["tenant_id"]
    kits = _read_brand_kits(tenant_id)
    now_iso = now_ist().isoformat()
    kit = {k: v for k, v in body.items() if k != "tenant_id"}  # never let body set tenant
    kid = str(kit.get("id") or "").strip()
    if kid:
        found = False
        for i, k in enumerate(kits):
            if str(k.get("id")) == kid:
                kit["created_at"] = k.get("created_at", now_iso)
                kit["updated_at"] = now_iso
                kits[i] = kit
                found = True
                break
        if not found:
            kit["created_at"] = kit["updated_at"] = now_iso
            kits.append(kit)
    else:
        kit["id"] = f"bk_{uuid.uuid4().hex[:12]}"
        kit["created_at"] = kit["updated_at"] = now_iso
        kits.append(kit)
    async with _STORE_LOCK:
        _BRAND_KITS_DIR.mkdir(parents=True, exist_ok=True)
        _write(_brand_kits_path(tenant_id), kits)
    return JSONResponse({"ok": True, "brand_kit": kit})


@app.delete("/brand-kits/{kit_id}")
async def brand_kits_delete(request: Request, kit_id: str):
    """Delete one of the calling tenant's brand kits by id. Tenant-scoped + write-gated."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden()
    tenant_id = t["tenant_id"]
    kits = _read_brand_kits(tenant_id)
    remaining = [k for k in kits if str(k.get("id")) != str(kit_id)]
    if len(remaining) == len(kits):
        return JSONResponse({"ok": True, "status": "noop", "reason": "not_found"})
    async with _STORE_LOCK:
        _write(_brand_kits_path(tenant_id), remaining)
    return JSONResponse({"ok": True, "deleted": kit_id})


# === PVS PHASE-1 (provider+voice switcher) === ADDITIVE; agent.py/trunks/firewall/SIP untouched.
# Sarvam Bulbul v2 fixed speaker catalogue (NO per-voice preview URL via API -> we pre-host a tiny
# one-time sample set under var/voice_samples/sarvam/<speaker>.mp3, served by /voice-preview).
_SARVAM_VOICES = [
    {"voice_id": "anushka", "name": "Anushka", "gender": "female", "accent": "Indian", "language": "Hindi/multi"},
    {"voice_id": "manisha", "name": "Manisha", "gender": "female", "accent": "Indian", "language": "Hindi/multi"},
    {"voice_id": "vidya",   "name": "Vidya",   "gender": "female", "accent": "Indian", "language": "Hindi/multi"},
    {"voice_id": "arya",    "name": "Arya",    "gender": "female", "accent": "Indian", "language": "Hindi/multi"},
    {"voice_id": "abhilash","name": "Abhilash","gender": "male",   "accent": "Indian", "language": "Hindi/multi"},
    {"voice_id": "karun",   "name": "Karun",   "gender": "male",   "accent": "Indian", "language": "Hindi/multi"},
    {"voice_id": "hitesh",  "name": "Hitesh",  "gender": "male",   "accent": "Indian", "language": "Hindi/multi"},
]
_VOICE_SAMPLE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "var", "voice_samples")


def _sarvam_voice_list():
    out = []
    for v in _SARVAM_VOICES:
        d = dict(v)
        d["preview_url"] = ""  # served via the proxy below (pre-hosted local clip)
        d["sample_url"] = f"/voice-preview?provider=sarvam&id={v['voice_id']}"
        out.append(d)
    return out


@app.get("/voices")
async def voices(request: Request, provider: str = ""):
    """Voice catalogue per provider. ElevenLabs = live /v1/voices WITH the free public preview_url
    (un-stripped) + accent/gender. Sarvam = the fixed Bulbul v2 speaker catalogue + a sample_url that
    points at the pre-hosted clip proxy. FREE — no synthesis here."""
    if not authed(request):
        return need_auth()
    p = (provider or "").strip().lower()
    if p == "sarvam":
        return JSONResponse({"provider": "sarvam", "voices": _sarvam_voice_list()})
    # default + p in ("", "elevenlabs"): ElevenLabs live catalogue
    try:
        r = httpx.get("https://api.elevenlabs.io/v1/voices",
                      headers={"xi-api-key": os.environ["ELEVENLABS_API_KEY"]}, timeout=15)
        vs = []
        for v in r.json().get("voices", []):
            labels = v.get("labels") or {}
            vs.append({
                "voice_id": v.get("voice_id"),
                "name": v.get("name"),
                "preview_url": v.get("preview_url", ""),  # public GCS MP3, FREE (un-stripped)
                "accent": labels.get("accent", ""),
                "gender": labels.get("gender", ""),
                "language": labels.get("language", ""),
                "sample_url": f"/voice-preview?provider=elevenlabs&id={v.get('voice_id')}",
            })
        return JSONResponse({"provider": "elevenlabs", "voices": vs})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"provider": "elevenlabs", "voices": [], "error": repr(exc)[:140]})


@app.get("/voice-preview")
async def voice_preview(request: Request, provider: str = "", id: str = ""):
    """FREE play-preview proxy. ElevenLabs -> full-buffer the voice's preview clip and return it
    same-origin as audio/mpeg (no key for the clip fetch, no synthesis, no burn). Sarvam -> stream
    the pre-hosted one-time sample clip from var/voice_samples/sarvam/<id>.wav. Used by the panel
    <audio> Play button."""
    # Auth: standard header auth OR t= query-param token (needed because <audio src> cannot
    # send headers; the FE already appends ?t=<jwt> for voice preview).
    _t_param = request.query_params.get("t", "")
    if not authed(request):
        if not _t_param:
            return need_auth()
        # Validate the t= param through the same resolve_tenant path by synthesising a
        # Bearer-like fake request object (scoped to this no-spend, no-PII route only).
        class _FakeReq:
            def __init__(self, token):
                self._token = token
                self.headers = {"authorization": "Bearer " + token, "x-auth": ""}
            @property
            def query_params(self): return {}
        if resolve_tenant(_FakeReq(_t_param)) is None:
            return need_auth()
    from fastapi.responses import RedirectResponse, FileResponse
    p = (provider or "").strip().lower()
    vid = (id or "").strip()
    if not vid:
        return JSONResponse({"error": "id required"}, status_code=400)
    if p == "sarvam":
        safe = "".join(ch for ch in vid if ch.isalnum() or ch in "-_")
        fp = os.path.join(_VOICE_SAMPLE_DIR, "sarvam", f"{safe}.wav")
        if os.path.isfile(fp):
            return FileResponse(fp, media_type="audio/wav", filename=f"sarvam-{safe}.wav")
        return JSONResponse({"error": "sample not available", "voice_id": vid}, status_code=404)
    # elevenlabs (default): resolve the voice's preview_url, then FULL-BUFFER the upstream clip and
    # return the bytes SAME-ORIGIN with content-type FORCED to audio/mpeg. We do NOT 307-redirect:
    # the upstream EL preview bytes are served Content-Type: text/plain (both on the public
    # storage.googleapis.com host AND the signed-expiring api.us.elevenlabs.io host), which Safari/
    # iOS refuses for <audio src> -> silent MEDIA_ERR_SRC_NOT_SUPPORTED. Buffering same-origin and
    # forcing audio/mpeg is the load-bearing fix. Both host shapes are handled identically (we GET
    # whatever URL EL returns); no branch on "is it GCS". ≤32 KB cap, 5 s timeout, 502 on empty/fail.
    _EL_PREVIEW_CAP = 32 * 1024  # tiny clip; cap guards against an HTML error page or a huge body
    try:
        r = httpx.get("https://api.elevenlabs.io/v1/voices",
                      headers={"xi-api-key": os.environ["ELEVENLABS_API_KEY"]}, timeout=15)
        pu = ""
        for v in r.json().get("voices", []):
            if v.get("voice_id") == vid:
                pu = (v.get("preview_url") or "").strip()
                break
        if not pu:
            return JSONResponse({"error": "no preview for this voice", "voice_id": vid}, status_code=404)
        # Fetch the clip from whichever EL host the signed/public URL points at (GCS or api.us).
        async with httpx.AsyncClient(timeout=5.0, follow_redirects=True) as _cli:
            up = await _cli.get(pu)
        if up.status_code != 200:
            return JSONResponse({"error": "upstream preview empty", "status": up.status_code},
                                status_code=502)
        body = up.content or b""
        if not body:
            return JSONResponse({"error": "upstream preview empty"}, status_code=502)
        body = body[:_EL_PREVIEW_CAP]
        # FORCE audio/mpeg; never echo the upstream text/plain. No Accept-Ranges (plain Response).
        return Response(content=body, media_type="audio/mpeg",
                        headers={"Cache-Control": "private, max-age=300"})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": repr(exc)[:140]}, status_code=502)


@app.get("/providers")
async def providers_list(request: Request):
    """Usable providers per role (built-in + custom) with kind + available(>=1 live key).
    Built-ins: their available-ness comes from the provider_pool (>=1 non-cooling key) for groq/
    sarvam; elevenlabs is available iff ELEVENLABS_API_KEY is set. Custom providers append from the
    isolated custom-provider store. Reuses _pk_get_pool / key availability — no new pool."""
    if not authed(request):
        return need_auth()

    def _builtin_available(name):
        try:
            if name == "elevenlabs":
                return bool((os.environ.get("ELEVENLABS_API_KEY") or "").strip())
            if _pk_get_pool is not None:
                pool = _pk_get_pool(name)
                if pool is not None:
                    return pool.available_count() > 0
            # fall back to env presence
            return bool((os.environ.get((name or "").upper() + "_API_KEY") or "").strip())
        except Exception:  # noqa: BLE001
            return False

    builtin = [
        {"id": "sarvam",     "name": "Sarvam",      "builtin": True, "kinds": ["stt", "tts"],
         "available": _builtin_available("sarvam")},
        {"id": "groq",       "name": "Groq",        "builtin": True, "kinds": ["llm"],
         "available": _builtin_available("groq")},
        {"id": "elevenlabs", "name": "ElevenLabs",  "builtin": True, "kinds": ["stt", "tts"],
         "available": _builtin_available("elevenlabs")},
        {"id": "sambanova",  "name": "SambaNova",   "builtin": True, "kinds": ["llm"],
         "available": _builtin_available("sambanova")},
        {"id": "openrouter", "name": "OpenRouter",  "builtin": True, "kinds": ["llm"],
         "available": _builtin_available("openrouter")},
    ]
    custom = []
    try:
        from llm_router import custom_providers as _cp
        for c in _cp.list_masked():
            custom.append({
                "id": c["id"], "name": c["name"], "builtin": False, "kinds": [c["kind"]],
                "kind": c["kind"], "model": c["model"], "base_url": c["base_url"],
                "enabled": c["enabled"], "available": c["available"], "masked": c["masked"],
            })
    except Exception:  # noqa: BLE001
        pass
    # group by role for convenience (UI's 3 per-role selects)
    by_role = {"stt": [], "llm": [], "tts": []}
    for prov in builtin + custom:
        for k in prov.get("kinds", []):
            if k in by_role:
                by_role[k].append({"id": prov["id"], "name": prov["name"],
                                   "builtin": prov["builtin"], "available": prov["available"]})
    return JSONResponse({"providers": builtin + custom, "by_role": by_role})


# ── Realtime provider network-health (signal bars + latency) ──────────────────────────────────
# Powers the "Live status" row in the Run-page Providers card. Measures a lightweight TCP-connect
# RTT to each provider's API host — NO auth, NO request body, NO quota burn — and grades it into
# 5 signal bars + a green/yellow/red status. Key-less providers report red/no-key. Cached briefly
# so the panel can poll every few seconds without hammering. Best-effort; never breaks a request.
_PROVIDER_HOSTS = {
    "groq": ("api.groq.com", 443),
    "elevenlabs": ("api.elevenlabs.io", 443),
    "sarvam": ("api.sarvam.ai", 443),
    "sambanova": ("api.sambanova.ai", 443),
    "openrouter": ("openrouter.ai", 443),
}
_PROVIDER_ROLE = {"groq": "llm", "sambanova": "llm", "openrouter": "llm",
                  "elevenlabs": "tts", "sarvam": "stt"}
_PROVIDER_LABEL = {"groq": "Groq", "elevenlabs": "ElevenLabs", "sarvam": "Sarvam",
                   "sambanova": "SambaNova", "openrouter": "OpenRouter"}
# Latency grading (ms). Tunable per deployment region (US APIs from India sit higher). green=4-5
# bars, yellow=2-3, red<=1. _LAT_GREEN_MS is the "healthy realtime" ceiling.
_LAT_GREEN_MS = float(os.getenv("PROVIDER_LAT_GREEN_MS", "220"))
_LAT_YELLOW_MS = float(os.getenv("PROVIDER_LAT_YELLOW_MS", "800"))
_PROVIDER_HEALTH_TTL = float(os.getenv("PROVIDER_HEALTH_TTL_S", "4"))
_PROVIDER_HEALTH_CACHE: dict = {"at": 0.0, "rtt": {}}


def _provider_has_key(name: str) -> bool:
    """True iff this built-in provider has at least one usable key (mirrors /providers)."""
    try:
        if name == "elevenlabs":
            return bool((os.environ.get("ELEVENLABS_API_KEY") or "").strip())
        if _pk_get_pool is not None:
            pool = _pk_get_pool(name)
            if pool is not None:
                return pool.available_count() > 0
        return bool((os.environ.get((name or "").upper() + "_API_KEY") or "").strip())
    except Exception:  # noqa: BLE001
        return False


def _provider_probe_rtt(host: str, port: int, timeout: float = 2.5):
    """Min TCP-connect RTT (ms) over 2 samples — a real network-reachability signal with zero
    auth/quota. Returns None when the host is unreachable/timed out. Never raises."""
    import socket  # local import (module imports socket lazily elsewhere too)
    best = None
    for _ in range(2):
        t0 = time.perf_counter()
        try:
            with socket.create_connection((host, port), timeout=timeout):
                pass
        except Exception:  # noqa: BLE001
            return None
        dt = (time.perf_counter() - t0) * 1000.0
        best = dt if best is None else min(best, dt)
    return round(best, 1) if best is not None else None


def _latency_grade(ms):
    """(bars 0-5, status green|yellow|red) from a latency in ms. None -> unreachable (red, 1 bar)."""
    if ms is None:
        return 1, "red"
    if ms <= _LAT_GREEN_MS * 0.5:
        return 5, "green"
    if ms <= _LAT_GREEN_MS:
        return 4, "green"
    if ms <= (_LAT_GREEN_MS + _LAT_YELLOW_MS) / 2:
        return 3, "yellow"
    if ms <= _LAT_YELLOW_MS:
        return 2, "yellow"
    return 1, "red"


@app.get("/providers/health")
async def providers_health(request: Request, ids: str = ""):
    """Realtime network latency + signal-strength for the AI providers, for the Providers card's
    "Live status" row. `ids` is an optional comma list (defaults to the 3 headline providers).
    Cached ~4s. Degrades to an empty list (never 500)."""
    if not authed(request):
        return need_auth()
    try:
        want = [s.strip().lower() for s in (ids or "").split(",") if s.strip()]
        if not want:
            want = ["groq", "elevenlabs", "sarvam"]
        # de-dupe, keep only known hosts, bound the fan-out
        seen: set = set()
        want = [p for p in want if p in _PROVIDER_HOSTS and not (p in seen or seen.add(p))][:8]
        now = time.time()
        cache = _PROVIDER_HEALTH_CACHE
        fresh = (now - cache["at"]) < _PROVIDER_HEALTH_TTL
        need_probe = (not fresh) or any(p not in cache["rtt"] for p in want)
        if need_probe:
            rtts = await asyncio.gather(*[
                asyncio.to_thread(_provider_probe_rtt, *_PROVIDER_HOSTS[p]) for p in want
            ])
            merged = dict(cache["rtt"]) if fresh else {}
            merged.update(dict(zip(want, rtts)))
            cache["rtt"] = merged
            cache["at"] = now
        rtt = cache["rtt"]
        out = []
        for pid in want:
            has_key = _provider_has_key(pid)
            ms = rtt.get(pid)
            if not has_key:
                bars, status, ms, note = 0, "red", None, "no key"
            elif ms is None:
                bars, status, note = 1, "red", "unreachable"
            else:
                bars, status = _latency_grade(ms)
                note = ""
            out.append({"id": pid, "role": _PROVIDER_ROLE.get(pid, ""),
                        "label": _PROVIDER_LABEL.get(pid, pid.title()),
                        "available": has_key, "reachable": ms is not None,
                        "latency_ms": ms, "bars": bars, "status": status, "note": note})
        return JSONResponse({"providers": out, "at": _utc_iso(), "ttl_s": _PROVIDER_HEALTH_TTL})
    except Exception:  # noqa: BLE001
        return JSONResponse({"providers": [], "at": _utc_iso(), "ttl_s": _PROVIDER_HEALTH_TTL})


@app.get("/campaigns/{cid}/preflight")
async def campaign_preflight(request: Request, cid: str):
    """Pre-launch readiness for a campaign, with REAL signals: TCP-RTT to the AI providers it rides
    (LLM/STT/TTS, no auth/quota), db/redis/livekit voice infra, and this campaign's recent call
    latency (ClickHouse p95). Returns graded checks + an overall verdict so the panel can warn before
    a slow launch (the browser layers its own network-RTT check on top). Best-effort; never 500s."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    f = (d.get("fields") or {}) if isinstance(d, dict) else {}
    checks: list = []
    worst = 0.0

    def _pid(v, default):
        s = str(v or "").strip().lower()
        return s if s in _PROVIDER_HOSTS else default

    # 1) AI providers (LLM · STT · TTS) — real TCP-connect RTT
    prov_ids: list = []
    for pid in (_pid(f.get("llm_provider"), "groq"), _pid(f.get("stt_provider"), "sarvam"),
                _pid(f.get("tts_provider"), "elevenlabs")):
        if pid not in prov_ids:
            prov_ids.append(pid)
    try:
        rtts = await asyncio.gather(*[asyncio.to_thread(_provider_probe_rtt, *_PROVIDER_HOSTS[p])
                                      for p in prov_ids])
    except Exception:  # noqa: BLE001
        rtts = [None] * len(prov_ids)
    prov_parts: list = []
    prov_worst = 0.0
    prov_down = False
    for pid, ms in zip(prov_ids, rtts):
        label = _PROVIDER_LABEL.get(pid, pid.title())
        if ms is None:
            prov_down = True
            prov_parts.append(f"{label} unreachable")
        else:
            prov_parts.append(f"{label} {ms:.0f}ms")
            prov_worst = max(prov_worst, ms)
    prov_status = "red" if (prov_down or prov_worst > _LAT_YELLOW_MS) else (
        "yellow" if prov_worst > _LAT_GREEN_MS else "green")
    checks.append({"id": "providers", "label": "AI providers (LLM · STT · TTS)",
                   "status": prov_status, "latency_ms": round(prov_worst) if prov_worst else None,
                   "detail": " · ".join(prov_parts)})
    worst = max(worst, 9999.0 if prov_down else prov_worst)

    # 2) Voice infrastructure — LiveKit is the call-signalling plane that actually gates a launch.
    # (db/redis are NOT on the live call path in this file-storage deployment, so they don't block.)
    lk_ok, _ = await asyncio.to_thread(_hc_livekit)
    infra_status = "green" if lk_ok else "red"
    checks.append({"id": "voice_infra", "label": "Voice infrastructure", "status": infra_status,
                   "latency_ms": None,
                   "detail": "LiveKit reachable — call routing up" if lk_ok
                             else "LiveKit unreachable — calls can't connect"})

    # 3) Recent call latency for THIS campaign (ClickHouse). Best-effort; neutral if no data.
    recent_status, recent_detail, recent_ms = "green", "no recent calls to measure", None
    try:
        if _obs_q is not None:
            summ = await _obs_q.voice_summary(60, {"campaign_id": cid})
            if isinstance(summ, dict):
                stages = {s.get("stage"): s for s in (summ.get("latency_by_stage") or [])}
                ncalls = int((summ.get("row") or {}).get("calls") or 0)
                p95s = [float(stages.get(s, {}).get("p95") or 0) for s in ("llm", "tts", "eou")]
                recent_ms = max(p95s) if any(p95s) else None
                if ncalls and recent_ms:
                    recent_status = "red" if recent_ms >= 2500 else ("yellow" if recent_ms >= 1500 else "green")
                    recent_detail = f"p95 {recent_ms / 1000:.1f}s over last hour ({ncalls} calls)"
                    worst = max(worst, recent_ms)
    except Exception:  # noqa: BLE001
        pass
    checks.append({"id": "recent", "label": "Recent call latency", "status": recent_status,
                   "latency_ms": round(recent_ms) if recent_ms else None, "detail": recent_detail})

    # verdict: DOWN only when something is genuinely unreachable (can't run calls); SLOW when the
    # worst reachable latency crosses ~1.5s (the warning the operator asked for); else OK.
    hard_down = prov_down or (not lk_ok)
    if hard_down:
        verdict = "down"
    elif worst >= 1500:
        verdict = "slow"
    else:
        verdict = "ok"
    headline = {"ok": "All systems go", "slow": "Networks look slow — not recommended",
                "down": "Some systems are down"}[verdict]
    return JSONResponse({"ok": True, "verdict": verdict, "headline": headline,
                         "worst_latency_ms": (round(worst) if (worst and worst < 9999) else None),
                         "checks": checks, "at": _utc_iso()})


@app.get("/tiers")
async def tiers_route(request: Request):
    """SINGLE SOURCE OF TRUTH for the Lean/Standard/Premium tier system: the 3 preset triples +
    the per-component rate card + the cost-math the frontend cost-meter uses (client-side, zero
    burn). Mirrors llm_router/tiers.py."""
    if not authed(request):
        return need_auth()
    try:
        from llm_router import tiers as _tiers
        return JSONResponse(_tiers.tiers_payload())
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": repr(exc)[:140], "tiers": []}, status_code=500)
# === /PVS PHASE-1 voices/preview/providers/tiers ===


@app.get("/campaigns")
async def campaigns(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    return JSONResponse({"campaigns": list_campaigns(t)})


def _coerce_script_blocks(out: dict) -> None:
    """P7: when Script Studio 2.0 is on for this campaign, compile its typed `script_blocks` down to
    the fields build_system_prompt consumes (so the live agent path is unchanged). The blocks remain
    on `out` as the source of truth; the compiled keys are what the agent reads. Best-effort + a
    pure no-op when off / unavailable / no blocks. Never raises."""
    try:
        if _script_compiler is None:
            return
        if str(out.get("script_studio_v2") or "").strip().lower() not in ("1", "true", "yes", "on"):
            return
        blocks = out.get("script_blocks")
        if not isinstance(blocks, list) or not blocks:
            return
        overrides = _script_compiler.compile_blocks(blocks, out.get("script_variables") or {})
        if isinstance(overrides, dict):
            for k, v in overrides.items():
                out[k] = v
    except Exception:  # noqa: BLE001
        pass


def _coerce_fields(fields: dict) -> dict:
    """Normalise extracted/edited campaign fields so a save never 500s and the
    agent always gets a usable shape. Missing/odd values get sensible defaults."""
    if not isinstance(fields, dict):
        raise ValueError("fields must be a JSON object")
    out = dict(fields)
    # strings
    for k, default in (("company_name", ""), ("agent_name", "Riya"), ("product_name", ""),
                       ("product_summary", ""), ("location", ""), ("price_offer", ""),
                       ("language", "Hinglish"), ("voice_id", ""),
                       # P0.1 calling window (HH:MM, IST gate); call_window_tz informational
                       ("call_window_start", "09:00"), ("call_window_end", "21:00"),
                       ("call_window_tz", "Asia/Kolkata")):
        v = out.get(k)
        out[k] = (str(v).strip() if v is not None else default) or default
    # P0.5 retry policy
    try:
        out["retry_max_attempts"] = max(0, int(out.get("retry_max_attempts", 3)))
    except Exception:  # noqa: BLE001
        out["retry_max_attempts"] = 3
    bo = out.get("retry_backoff_mins")
    if isinstance(bo, str):
        bo = [s.strip() for s in re.split(r"[,\s]+", bo) if s.strip()]
    if isinstance(bo, list):
        try:
            bo = [int(x) for x in bo if str(x).strip()]
        except Exception:  # noqa: BLE001
            bo = []
    out["retry_backoff_mins"] = bo or [120, 360, 1440]
    # P1.A whatsapp follow-up (per-campaign)
    out["wa_enabled"] = bool(out.get("wa_enabled", False))
    # WAVE3 Unit5: per-campaign auto WhatsApp follow-up after interested/callback calls.
    # Defaults OFF so nothing fires until the user enables it AND adds WA_* creds.
    out["wa_followup"] = bool(out.get("wa_followup", False))
    for k in ("wa_template_qualified", "wa_template_noanswer",
              "wa_template_interested", "wa_template_callback",
              # #10(b) BROCHURE: per-campaign brochure PDF (Spaces key + public/CDN url +
              # display filename), populated by POST /campaigns/{cid}/brochure upload route.
              "brochure_pdf_url", "brochure_pdf_key", "brochure_pdf_name"):
        v = out.get(k)
        out[k] = str(v).strip() if v is not None else ""
    # #10(c) per-campaign override for the post-call WhatsApp score-gate (blank -> global).
    msc = out.get("wa_followup_min_score")
    if msc in (None, ""):
        out["wa_followup_min_score"] = ""
    else:
        try:
            out["wa_followup_min_score"] = int(msc)
        except Exception:  # noqa: BLE001
            out["wa_followup_min_score"] = ""
    # list-of-strings
    for k in ("usps", "talking_points", "qualifying_questions"):
        v = out.get(k)
        if isinstance(v, str):
            v = [s.strip() for s in re.split(r"[\n;]+", v) if s.strip()]
        out[k] = [str(s) for s in v] if isinstance(v, list) else []
    # objections: list of {q,a}
    obj = out.get("objections")
    norm_obj = []
    if isinstance(obj, list):
        for o in obj:
            if isinstance(o, dict):
                norm_obj.append({"q": str(o.get("q", "")), "a": str(o.get("a", ""))})
    out["objections"] = norm_obj
    # WAVE3 Unit3: A/B variants. Each = {id,label,fields_override:{...},weight}.
    # A campaign with no variants behaves exactly as before.
    variants = out.get("variants")
    norm_var = []
    if isinstance(variants, list):
        for v in variants:
            if not isinstance(v, dict):
                continue
            vid = str(v.get("id") or uuid.uuid4().hex[:6])
            ov = v.get("fields_override")
            ov = ov if isinstance(ov, dict) else {}
            try:
                w = max(1, int(v.get("weight", 1)))
            except Exception:  # noqa: BLE001
                w = 1
            norm_var.append({"id": vid, "label": str(v.get("label") or vid),
                             "fields_override": ov, "weight": w})
    out["variants"] = norm_var
    # === PVS PHASE-1: per-campaign provider+voice tier persistence (additive) ===
    # tier in {lean,standard,premium,custom}; default lean (== today's pipeline on outbound).
    _tier = str(out.get("tier", "") or "").strip().lower()
    if _tier not in ("lean", "standard", "premium", "custom"):
        _tier = "lean"
    out["tier"] = _tier
    # explicit per-role provider overrides (used when tier == custom, or as the resolved snapshot)
    for k in ("stt_provider", "llm_provider", "tts_provider", "custom_provider_id"):
        v = out.get(k)
        out[k] = str(v).strip() if v is not None else ""
    # est avg call minutes (for projected campaign spend in the UI; clamp sane)
    try:
        _eac = float(out.get("est_avg_call_min", 1.5))
        out["est_avg_call_min"] = round(min(30.0, max(0.1, _eac)), 2)
    except Exception:  # noqa: BLE001
        out["est_avg_call_min"] = 1.5
    # optional per-campaign budget cap in ₹ (blank/0 -> no cap; UI warn/estimate only in Phase 1)
    _cap = out.get("budget_cap_inr")
    if _cap in (None, "", 0, "0"):
        out["budget_cap_inr"] = ""
    else:
        try:
            out["budget_cap_inr"] = max(0, int(float(_cap)))
        except Exception:  # noqa: BLE001
            out["budget_cap_inr"] = ""
    # snapshot the resolved {stt,llm,tts,voice} triple so a later tiers.py edit never silently
    # rewrites an in-flight campaign. For tier==custom we snapshot the explicit overrides.
    try:
        from llm_router import tiers as _tiers_mod
        if _tier == "custom":
            out["tier_resolved"] = {
                "tier": "custom",
                "stt": {"provider": out.get("stt_provider", "")},
                "llm": {"provider": out.get("llm_provider", "")},
                "tts": {"provider": out.get("tts_provider", "")},
                "voice": {"voice_id": out.get("voice_id", "")},
            }
        else:
            _trip = _tiers_mod.resolve_triple(_tier)
            # let an explicitly chosen voice_id override the tier default voice in the snapshot
            if _trip and out.get("voice_id"):
                _trip.setdefault("voice", {})["voice_id"] = out.get("voice_id")
            out["tier_resolved"] = _trip or {}
    except Exception:  # noqa: BLE001
        out["tier_resolved"] = out.get("tier_resolved") or {}
    # === /PVS PHASE-1 ===
    # === W1 VENDOR SCRIPT (lossless raw_script + sanitized script_meta) ===
    # Runs LAST so the derived-projection suppression sees the fully-coerced lists.
    # No-op for legacy campaigns (no raw_script) -> golden render byte-identical.
    _coerce_vendor_script(out)
    # === /W1 VENDOR SCRIPT ===
    # === P7 SCRIPT STUDIO 2.0 — compile typed blocks DOWN to the consumed fields ===
    # Runs LAST. ONLY when fields.script_studio_v2 is on AND script_blocks present: the typed block
    # model compiles to the SAME fields build_system_prompt already reads, so the LIVE agent path is
    # unchanged. Legacy / flag-off campaigns are byte-identical (pure no-op). Never raises.
    _coerce_script_blocks(out)
    # === /P7 ===
    return out


@app.post("/campaigns")
async def create_campaign(request: Request, fields_json: str = Form("")):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot create campaigns")
    # Bulletproof: validate input, never 500, return a clear error instead.
    if not (fields_json or "").strip():
        return JSONResponse({"error": "fields_json is required"}, status_code=400)
    try:
        raw = json.loads(fields_json.lstrip("﻿"))
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"fields_json is not valid JSON: {exc}"}, status_code=400)
    try:
        fields = _coerce_fields(raw)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"invalid fields: {exc}"}, status_code=400)
    if not (fields.get("company_name") or fields.get("product_name")):
        return JSONResponse({"error": "at least one of company_name or product_name is required"},
                            status_code=400)
    try:
        rec = save_campaign(fields, t["tenant_id"])
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"could not save campaign: {repr(exc)[:160]}"}, status_code=500)
    _audit(request, t, "campaign.create", "campaign", rec["id"], meta={"name": rec.get("name")})
    return JSONResponse({"id": rec["id"], "name": rec["name"], "tenant_id": rec["tenant_id"]})


@app.get("/campaigns/{cid}")
async def get_campaign_detail(request: Request, cid: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    return JSONResponse({"campaign": d})


def _log_script_gen(t, cid: str, kind: str, res: dict, elapsed: float) -> None:
    """Surface Script Studio AI-drafting health in System Logs: an ERROR if the draft failed, a
    WARNING if it ran slow (the ~40s Sonnet calls that used to silently time out at the Next proxy —
    invisible to System Logs because the backend handler itself succeeded). Best-effort; never raises."""
    if _log_mod is None:
        return
    try:
        tid = (t or {}).get("tenant_id", "") if isinstance(t, dict) else ""
        res = res if isinstance(res, dict) else {}
        if not res.get("ok"):
            err = res.get("error") or "gen_failed"
            _log_mod.record(
                "error", "script_studio",
                f"AI script {kind} FAILED for campaign {cid}: {err} ({elapsed:.0f}s)",
                tenant_id=tid, error_type=f"script_{err}"[:120],
                context={"cid": cid, "kind": kind, "elapsed_s": round(elapsed, 1),
                         "message": str(res.get("message", ""))[:300]},
            )
        elif elapsed >= 15.0:
            _log_mod.record(
                "warning", "script_studio",
                f"AI script {kind} slow: {elapsed:.0f}s for campaign {cid} (near the proxy timeout)",
                tenant_id=tid, error_type="script_slow",
                context={"cid": cid, "kind": kind, "elapsed_s": round(elapsed, 1)},
            )
    except Exception:  # noqa: BLE001
        pass


@app.post("/campaigns/{cid}/script/generate")
async def generate_campaign_script(request: Request, cid: str):
    """AI-draft a call script for this campaign with Claude Sonnet 3.5 (Script Studio
    "Generate with AI"). Tenant-scoped (the campaign must belong to the caller). Read-only on
    the campaign — returns the drafted text; the operator edits + saves it themselves. Degrades
    to a clean error (never 500) when the drafter or ANTHROPIC_API_KEY is unavailable."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    if _script_gen is None:
        return JSONResponse({"ok": False, "error": "unavailable",
                             "message": "AI script drafting is not available."}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    brief = str((body or {}).get("brief", "") or "")[:2000]
    tone = str((body or {}).get("tone", "") or "")[:24]
    length = str((body or {}).get("length", "") or "")[:24]
    push = str((body or {}).get("push", "") or "")[:24]
    fields = (d.get("fields") or {}) if isinstance(d, dict) else {}
    _t0 = time.perf_counter()
    res = await _script_gen.generate(fields, brief, tone=tone, length=length, push=push)
    _elapsed = time.perf_counter() - _t0
    _log_script_gen(t, cid, "generate", res, _elapsed)
    return JSONResponse(res, status_code=200 if res.get("ok") else 502)


@app.post("/campaigns/{cid}/script/generate-block")
async def generate_campaign_script_block(request: Request, cid: str):
    """P7.3: AI-draft ONE Script Studio 2.0 block (Claude Sonnet 4.6 via OpenRouter). Returns
    {ok, block:{type,...}} shaped for the builder to merge. Tenant-scoped + read-only on the
    campaign; degrades cleanly (never 500) when the drafter / OPENROUTER_API_KEY is unavailable."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    if _script_gen is None or not hasattr(_script_gen, "generate_block"):
        return JSONResponse({"ok": False, "error": "unavailable",
                             "message": "AI script drafting is not available."}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    block_type = str((body or {}).get("block_type", "") or "")[:40]
    brief = str((body or {}).get("brief", "") or "")[:2000]
    tone = str((body or {}).get("tone", "") or "")[:24]
    length = str((body or {}).get("length", "") or "")[:24]
    push = str((body or {}).get("push", "") or "")[:24]
    fields = (d.get("fields") or {}) if isinstance(d, dict) else {}
    _t0 = time.perf_counter()
    res = await _script_gen.generate_block(fields, block_type, brief, tone=tone, length=length, push=push)
    _elapsed = time.perf_counter() - _t0
    _log_script_gen(t, cid, f"block:{block_type}", res, _elapsed)
    return JSONResponse(res, status_code=200 if res.get("ok") else 502)


@app.get("/campaigns/{cid}/ab")
async def campaign_ab(request: Request, cid: str):
    """WAVE3 Unit3: per-variant A/B stats computed from call records.
    Returns {campaign_id, variants:[{id,label,weight,dialed,connected,interested,
    qualified,avg_interest}]}. Calls with no variant_id are bucketed under '(default)'."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    defs = (d.get("fields") or {}).get("variants") or []
    # seed buckets from defined variants (so a 0-call variant still appears)
    buckets: dict = {}
    for v in defs:
        if isinstance(v, dict) and v.get("id"):
            buckets[v["id"]] = {"id": v["id"], "label": v.get("label", v["id"]),
                                "weight": int(v.get("weight", 1) or 1),
                                "dialed": 0, "connected": 0, "interested": 0,
                                "qualified": 0, "_interest_sum": 0, "_interest_n": 0}
    rows = [c for c in calls_for(t) if c.get("campaign_id") == cid]
    for c in rows:
        vid = c.get("variant_id") or "(default)"
        b = buckets.get(vid)
        if b is None:
            b = buckets[vid] = {"id": vid, "label": c.get("variant_label") or vid,
                                "weight": 0, "dialed": 0, "connected": 0, "interested": 0,
                                "qualified": 0, "_interest_sum": 0, "_interest_n": 0}
        if c.get("status") == "suppressed":
            continue
        b["dialed"] += 1
        if c.get("status") not in ("queued", "suppressed"):
            b["connected"] += 1
        if c.get("answered") is True or (c.get("answered") is None and c.get("duration_s", 0) >= 8):
            pass
        if c.get("outcome") == "interested":
            b["interested"] += 1
        score = c.get("interest", 0) or 0
        if score >= 70:
            b["qualified"] += 1
        if score:
            b["_interest_sum"] += score
            b["_interest_n"] += 1
    out = []
    for b in buckets.values():
        n = b.pop("_interest_n"); s = b.pop("_interest_sum")
        b["avg_interest"] = round(s / n, 1) if n else 0
        out.append(b)
    return JSONResponse({"campaign_id": cid, "variants": out})


@app.post("/campaigns/{cid}")
async def update_campaign(request: Request, cid: str, fields_json: str = Form("")):
    """Update a campaign's fields (incl. voice_id). Rebuilds the system prompt. Tenant-scoped."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot edit campaigns")
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    if not (fields_json or "").strip():
        return JSONResponse({"error": "fields_json is required"}, status_code=400)
    try:
        raw = json.loads(fields_json.lstrip("﻿"))
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"fields_json is not valid JSON: {exc}"}, status_code=400)
    try:
        fields = _coerce_fields(raw)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"invalid fields: {exc}"}, status_code=400)
    try:
        d["fields"] = fields
        d["name"] = fields.get("product_name") or fields.get("company_name") or d["id"]
        d["company"] = fields.get("company_name", "")
        d["product"] = fields.get("product_name", "")
        d["system_prompt"] = build_system_prompt(fields)
        d.setdefault("tenant_id", t["tenant_id"])
        safe = "".join(ch for ch in cid if ch.isalnum() or ch in "-_")
        # ATOMIC write (red-team fix #4): file mirror authoritative + write-first.
        _atomic_write_json(CAMPAIGN_DIR / f"{safe}.json", d)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"could not update: {repr(exc)[:160]}"}, status_code=500)
    # P1 DUAL MIRROR (best-effort, additive): mirror the edited campaign to PG (per-id upsert). No-op
    # unless campaigns is dual in STORE_MODES; off-loop; swallows — must NOT break campaign edit.
    try:
        if _store is not None:
            _store.mirror_campaign_upsert(d)
    except Exception:  # noqa: BLE001
        pass
    # W2: publish a cache-invalidate (version bump) — a compliance-line / script edit must reach the
    # inbound voice brain on the NEXT connect, not after the 300s TTL. Flag-gated; no-op when off.
    _publish_ctx_invalidate(d.get("tenant_id", t["tenant_id"]), d["id"])
    _audit(request, t, "campaign.update", "campaign", d["id"], meta={"name": d.get("name")})
    return JSONResponse({"id": d["id"], "name": d["name"]})


@app.post("/campaigns/{cid}/brochure")
async def upload_campaign_brochure(request: Request, cid: str,
                                   pdf: UploadFile | None = File(None)):
    """#10(b) BROCHURE: upload a per-campaign brochure PDF to DO Spaces and store its key +
    url on the campaign (fields.brochure_pdf_url / brochure_pdf_key / brochure_pdf_name). The
    file is stored PRIVATE (the live bucket has object-ACLs disabled); the post-call sender
    mints a short-lived presigned GET URL for Meta to fetch. Tenant-scoped, write-gated."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot edit campaigns")
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    if pdf is None:
        return JSONResponse({"error": "pdf file is required (form field 'pdf')"},
                            status_code=400)
    try:
        data = await pdf.read()
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"could not read upload: {repr(exc)[:120]}"},
                            status_code=400)
    if not data:
        return JSONResponse({"error": "empty file"}, status_code=400)
    if len(data) > 25 * 1024 * 1024:  # WhatsApp document cap is 100MB; keep brochures lean
        return JSONResponse({"error": "brochure too large (max 25MB)"}, status_code=413)
    # Light PDF sniff (don't hard-block, but reject obviously-wrong types).
    if not (data[:5] == b"%PDF-" or (pdf.content_type or "").lower() == "application/pdf"):
        return JSONResponse({"error": "file does not look like a PDF"}, status_code=415)
    fname = "".join(ch for ch in (pdf.filename or "brochure.pdf")
                    if ch.isalnum() or ch in "-_. ").strip() or "brochure.pdf"
    if not fname.lower().endswith(".pdf"):
        fname = fname + ".pdf"
    # Store on Spaces (no ACL -> private object; presigned at send time). Reuse the shared
    # media_gen.spaces boto3 client (READ-ONLY reuse of the proven AI-Asset client).
    try:
        from media_gen import spaces as _spaces
        cli = _spaces._client()
        if cli is None:
            return JSONResponse({"error": "object storage not configured"},
                                status_code=503)
        import uuid as _uuid
        bucket = (os.getenv("SPACES_BUCKET") or "").strip()
        safe_cid = "".join(ch for ch in cid if ch.isalnum() or ch in "-_")
        key = f"wa_brochures/{safe_cid}/{_uuid.uuid4().hex}.pdf"
        cli.put_object(Bucket=bucket, Key=key, Body=data,
                       ContentType="application/pdf")
        public_url = _spaces._public_url(key)  # canonical reference (may be 403 if private)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"upload failed: {type(exc).__name__}"},
                            status_code=502)
    # Persist onto the campaign fields.
    try:
        fields = d.get("fields") or {}
        fields["brochure_pdf_url"] = public_url
        fields["brochure_pdf_key"] = key
        fields["brochure_pdf_name"] = fname
        d["fields"] = _coerce_fields(fields)
        safe = "".join(ch for ch in cid if ch.isalnum() or ch in "-_")
        (CAMPAIGN_DIR / f"{safe}.json").write_text(
            json.dumps(d, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"saved file but could not update campaign: "
                                      f"{repr(exc)[:120]}"}, status_code=500)
    try:
        if _store is not None:
            _store.mirror_campaign_upsert(d)
    except Exception:  # noqa: BLE001
        pass
    _audit(request, t, "campaign.brochure", "campaign", d["id"],
           meta={"key": key, "name": fname, "bytes": len(data)})
    return JSONResponse({"ok": True, "brochure_pdf_key": key,
                         "brochure_pdf_url": public_url, "brochure_pdf_name": fname,
                         "bytes": len(data)})


@app.delete("/campaigns/{cid}")
async def delete_campaign(request: Request, cid: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot delete campaigns")
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    cid = "".join(ch for ch in cid if ch.isalnum() or ch in "-_")
    p = CAMPAIGN_DIR / f"{cid}.json"
    if p.exists():
        p.unlink()
    # P1 DUAL MIRROR (best-effort, additive): drop the campaign row from PG. No-op unless campaigns is
    # dual in STORE_MODES; off-loop; swallows — must NOT break campaign delete.
    try:
        if _store is not None:
            _store.mirror_campaign_delete(cid)
    except Exception:  # noqa: BLE001
        pass
    _audit(request, t, "campaign.delete", "campaign", cid)
    return JSONResponse({"deleted": cid})


# ===========================================================================
# W1 — SCRIPT STUDIO: prompt-preview + dry-run (vendor-script adopt-persona)
# ---------------------------------------------------------------------------
# Two read-only-ish helper endpoints for the Script Studio UI so a vendor can
# AUTHOR a campaign script and SEE the brain it produces + a sample turn BEFORE
# any real call. Both are tenant-scoped + auth'd via the same resolve_tenant
# pattern as every campaign route. Neither mutates the stored campaign, the
# global VENDOR_SCRIPT_INJECT env flag, or places any paid/outbound call.
#
# EARNER-SAFE: these only READ a campaign + render through build_system_prompt_v2
# (which is byte-identical to build_system_prompt when no raw_script is present)
# and, for dry-run, call the SAME free/cheap Groq extract model (_groq_chat) the
# WhatsApp drafts already use — never the outbound dial path, never the DID.
# The preview forces the vendor-script render ON for PREVIEW ONLY by setting the
# per-campaign opt-in flag on an in-memory COPY of fields (vendor_script_inject),
# so the founder can see the adopted persona even while the global env flag is
# still OFF in production. The stored campaign is never written.
# ===========================================================================
def _preview_fields(d: dict) -> dict:
    """A COPY of the campaign fields with the per-campaign vendor-script opt-in
    forced ON so build_system_prompt_v2 splices the persona for PREVIEW. Never
    mutates the stored campaign. No raw_script => identical to the base render."""
    f = dict((d or {}).get("fields") or {})
    if isinstance(f.get("raw_script"), str) and f.get("raw_script").strip():
        f["vendor_script_inject"] = True
    return f


@app.get("/campaigns/{cid}/prompt-preview")
async def campaign_prompt_preview(request: Request, cid: str):
    """Return the FULLY-RENDERED system prompt for this campaign — the exact brain
    the live inbound agent would adopt — with the vendor-script persona FORCED ON
    for preview (per-campaign opt-in on a copy; global env flag untouched). Tenant-
    scoped + auth'd. Read-only: never writes the campaign, never calls an LLM."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    f = _preview_fields(d)
    try:
        rendered = _build_system_prompt_v2(f)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"could not render prompt: {repr(exc)[:160]}"},
                            status_code=500)
    raw = f.get("raw_script")
    return JSONResponse({
        "campaign_id": cid,
        "name": d.get("name", cid),
        "vendor_script_present": bool(isinstance(raw, str) and raw.strip()),
        "vendor_script_active_in_preview": bool(f.get("vendor_script_inject")),
        "system_prompt": rendered,
        "chars": len(rendered),
    })


@app.post("/campaigns/{cid}/dry-run")
async def campaign_dry_run(request: Request, cid: str,
                           message: str = Form(""), as_returning: str = Form(""),
                           history: str = Form("")):
    """Dry-run ONE sample caller line through the inbound brain so the founder can
    SEE how the agent would greet/respond with the vendor's adopted persona —
    WITHOUT placing a real call. Renders the persona via build_system_prompt_v2
    (vendor-script forced on for preview) + reframes it for INBOUND, then runs a
    SINGLE turn through the FREE/cheap Groq extract model (_groq_chat) — the same
    model the WhatsApp drafts use, NEVER a paid outbound call / the DID. Tenant-
    scoped + auth'd. Reads the campaign; never writes it."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    d = get_campaign_for(cid, t)
    if not d:
        return JSONResponse({"error": "not found"}, status_code=404)
    sample = (message or "").strip() or "Hello"
    if len(sample) > 1000:
        sample = sample[:1000]
    f = _preview_fields(d)
    company = (f.get("company_name") or "").strip() or "the company"
    agent = (f.get("agent_name") or "").strip() or "Riya"
    raw_present = bool(isinstance(f.get("raw_script"), str) and f.get("raw_script").strip())
    # Inbound brain: disclosure OFF (mirrors aim_voice_agent _build_sales_instructions),
    # vendor persona ON for preview.
    try:
        inb = dict(f)
        inb["disclose_ai"] = False
        brain = _build_system_prompt_v2(inb)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"could not render brain: {repr(exc)[:160]}"},
                            status_code=500)
    # INBOUND reframe header — same spirit as the live inbound agent: the caller
    # dialled US; short human beats; mirror the caller's language; adopt the vendor
    # persona; never sound robotic; never ask for a PIN.
    inbound_head = (
        f"You are {agent}, a warm, real human salesperson for {company} on a LIVE INBOUND phone "
        "call — the CALLER dialled US because they're interested. Speak in short, natural beats "
        "(one or two sentences, then STOP), in the SAME language/code-mix the caller uses "
        "(Hinglish/Hindi/English). This is INBOUND: do NOT do the outbound permission opener — "
        "get straight to warmly helping them. NEVER sound robotic, NEVER ask for any PIN. "
        "Reply with ONLY what you would SAY out loud on this turn — no stage directions, no labels.\n\n"
    )
    system = inbound_head + (brain or "")
    if (as_returning or "").strip().lower() in ("1", "true", "yes", "on"):
        system += ("\n\n(NOTE: this caller has spoken to us before — continue warmly, don't "
                   "restart introductions.)\n")
    # P7.5 multi-turn simulator: prior turns passed as JSON [{role:'user'|'assistant', content}],
    # inserted between the system brain and the latest caller line so replies are context-aware.
    # Absent => byte-identical single-turn dry-run. Bounded + clipped.
    msgs = [{"role": "system", "content": system}]
    try:
        hist = json.loads(history) if (history or "").strip() else []
        if isinstance(hist, list):
            for h in hist[-12:]:
                if isinstance(h, dict):
                    role = "assistant" if str(h.get("role")) == "assistant" else "user"
                    content = str(h.get("content", "") or "")[:1000]
                    if content:
                        msgs.append({"role": role, "content": content})
    except Exception:  # noqa: BLE001
        pass
    msgs.append({"role": "user", "content": sample})
    reply = _groq_chat(msgs, max_tokens=220, temperature=0.6)
    used_llm = bool(reply)
    if not reply:
        # Never 500 / never silent — make the dry-run still useful if Groq is down.
        reply = ("(LLM unavailable right now — preview the rendered persona via "
                 "GET /campaigns/{cid}/prompt-preview to verify the adopted greeting.)")
    _audit(request, t, "campaign.dry_run", "campaign", cid,
           meta={"name": d.get("name"), "vendor_script": raw_present})
    return JSONResponse({
        "campaign_id": cid,
        "name": d.get("name", cid),
        "vendor_script_present": raw_present,
        "vendor_script_active_in_preview": bool(f.get("vendor_script_inject")),
        "sample_user": sample,
        "agent_reply": reply,
        "used_llm": used_llm,
        "provider": "groq",
        "model": GROQ_MODEL,
        "note": "dry-run only — no real call, no DID, free/cheap Groq turn",
    })


def _leads_for(tenant: dict) -> list[dict]:
    store = _read(LEADS_FILE, [])
    if tenant.get("is_admin"):
        return store
    return [x for x in store if x.get("tenant_id", ADMIN_ID) == tenant["tenant_id"]]


# Temperature — ONE source of truth shared with the panel (app/crm/_ui.tsx `tempOf`)
# and the profile spine (crm.temperature_of). Hot/Warm/Cold/Dead, where `dead` is an
# EXPLICIT opt-out / not-interested (an unscored brand-new lead is COLD, never dead).
# Delegating to crm keeps GET /leads ?status=, the /run audience targeting, and the
# panel badge from EVER drifting again. Falls back to a band-only read if crm is absent.
def _lead_temp(lead: dict) -> str:
    if _crm_mod is not None:
        try:
            return _crm_mod.temperature_of(lead)
        except Exception:  # noqa: BLE001
            pass
    bag = " ".join(str(lead.get(k, "") or "").lower()
                   for k in ("status", "last_outcome", "outcome", "lifecycle"))
    if any(k in bag for k in ("opt_out", "opted_out", "not_interested", "dead", "lost")):
        return "dead"
    s = int(lead.get("score", 0) or 0)
    if s >= 70:
        return "hot"
    if s >= 40:
        return "warm"
    return "cold"


def _csv_set(raw: str) -> set:
    """Split a comma/space/repeated-field selector string into a clean set."""
    if not raw:
        return set()
    return {p.strip() for p in re.split(r"[,\s]+", raw) if p.strip()}


def _resolve_audience(t: dict, *, source_mode: str = "", lead_ids: str = "",
                      temps: str = "", batch_ids: str = "",
                      score_min: str = "", score_max: str = "") -> list[dict]:
    """RC2: resolve the OPTIONAL composable /run selectors to a concrete set of
    this tenant's stored lead rows (BOLA-safe — only ever this tenant's leads).
    Returns full lead dicts (with phone/name/score/batch_id). Layers, in order:

      BASE POOL  = stored leads, narrowed to selected batch_ids (if any).
      TEMPERATURE/SCORE FILTER = temps (hot/warm/cold) and/or a custom score band.
      MANUAL OVERRIDE = if lead_ids given, use EXACTLY those (intersected with the
                        tenant's own leads — never another tenant's).

    Returns [] (no selector resolution) when none of the selectors are present,
    so the legacy csv/use_stored/text path in /run is completely untouched."""
    pool = _leads_for(t)
    by_id = {x.get("id"): x for x in pool}

    # MANUAL: explicit lead_ids win outright (primary UI path = preview==dials).
    ids = _csv_set(lead_ids)
    if ids:
        return [by_id[i] for i in ids if i in by_id]

    # Nothing else selected -> signal "no server-side resolution" to the caller.
    bset = _csv_set(batch_ids)
    tset = {x.lower() for x in _csv_set(temps)}
    has_band = bool(str(score_min).strip() or str(score_max).strip())
    if not (bset or tset or has_band or source_mode in ("all", "temperature", "upload")):
        return []

    rows = pool
    # BASE POOL narrowing by uploaded batch.
    if bset:
        rows = [x for x in rows if (x.get("batch_id") or "") in bset]
    # TEMPERATURE filter (multi-select).
    if tset:
        rows = [x for x in rows if _lead_temp(x) in tset]
    # CUSTOM score band.
    if has_band:
        try:
            lo = int(float(score_min)) if str(score_min).strip() else 0
        except Exception:  # noqa: BLE001
            lo = 0
        try:
            hi = int(float(score_max)) if str(score_max).strip() else 100
        except Exception:  # noqa: BLE001
            hi = 100
        rows = [x for x in rows if lo <= int(x.get("score", 0) or 0) <= hi]
    return rows


@app.get("/leads")
async def get_leads(request: Request, hot: str = "", sort: str = "",
                    limit: int = 0, offset: int = 0, sort_by: str = "", order: str = ""):
    """Leads. Backward-compatible: with NO limit (the current FE) it returns ALL leads as
    `{leads:[...]}` exactly as before. Pagination is opt-in via limit/offset, and the
    response ALWAYS now also carries total/offset/limit/next so the FE can paginate.
    Lead rows are already small (10 flat fields), so no payload trim is needed."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = _leads_for(t)
    if hot:
        rows = [x for x in rows if (x.get("score", 0) or 0) >= 70]
    # W14-WIRE: apply the filters the panel forwards (from/to/campaign_id/status/batch). The route used
    # to DROP these (the panel sent them, caller ignored them) -> the dashboard's filtered lead views
    # silently showed everything. All optional; absent -> unchanged behavior.
    qp = request.query_params
    _lfrm = (qp.get("from_") or qp.get("from") or "").strip()
    _lto = (qp.get("to_") or qp.get("to") or "").strip()
    _lcamp = (qp.get("campaign_id") or qp.get("campaign") or "").strip()
    _lstatus = (qp.get("status") or "").strip().lower()
    _lbatch = (qp.get("batch") or qp.get("batch_id") or "").strip()
    if _lfrm:
        rows = [x for x in rows if (x.get("added_at") or "") >= _lfrm]
    if _lto:
        rows = [x for x in rows if (x.get("added_at") or "") <= (_lto + "T23:59:59+00:00")]
    if _lcamp:
        rows = [x for x in rows if str(x.get("campaign_id", "")) == _lcamp]
    if _lbatch:
        rows = [x for x in rows if str(x.get("batch_id", "")) == _lbatch]
    if _lstatus:
        if _lstatus in ("hot", "warm", "cold", "dead"):   # temperature band (matches the panel badge)
            # Use the SHARED classifier so server-side paging of Warm/Cold/Dead matches the
            # Hot/Warm/Cold/Dead badge the panel renders (no more client-only partial filter).
            rows = [x for x in rows if _lead_temp(x) == _lstatus]
        else:                                              # literal lead status (new/contacted/...)
            rows = [x for x in rows if (x.get("status", "") or "").lower() == _lstatus]
    # Sort selector (additive; default = newest-first by added_at, latest->oldest).
    # Accepts the legacy `sort`: "" / "recent" -> created_at DESC; "oldest" -> created_at ASC;
    # "name" -> A->Z; "status" -> status A->Z then newest; "score" -> high->low.
    # R5P4-3 (ADDITIVE): also accept the panel's `sort_by` + `order` (column + direction). When sort_by
    # is present it wins; we map the column to a sort key and honor order=asc|desc. No FE breakage: the
    # existing `sort`-only callers are untouched.
    _lsb = (sort_by or "").strip().lower()
    if _lsb:
        _ldesc = (order or "").strip().lower() != "asc"
        if _lsb in ("score", "interest"):
            rows = sorted(rows, key=lambda x: x.get("score", 0) or 0, reverse=_ldesc)
        elif _lsb in ("name", "lead"):
            rows = sorted(rows, key=lambda x: (x.get("name", "") or "").lower(), reverse=_ldesc)
        elif _lsb in ("status",):
            rows = sorted(rows, key=lambda x: (x.get("status", "") or "").lower(), reverse=_ldesc)
        elif _lsb in ("added_at", "created_at", "placed", "recent"):
            rows = sorted(rows, key=lambda x: x.get("added_at", "") or "", reverse=_ldesc)
        else:  # unknown column -> newest-first default
            rows = sorted(rows, key=lambda x: x.get("added_at", "") or "", reverse=True)
    else:
        s = (sort or "recent").lower()
        if s == "score":
            rows = sorted(rows, key=lambda x: x.get("score", 0) or 0, reverse=True)
        elif s == "oldest":
            rows = sorted(rows, key=lambda x: x.get("added_at", "") or "")
        elif s == "name":
            rows = sorted(rows, key=lambda x: (x.get("name", "") or "").lower())
        elif s == "status":
            rows = sorted(rows, key=lambda x: ((x.get("status", "") or "").lower(),
                                               x.get("added_at", "") or ""))
        else:  # "recent" / default / unknown -> newest-first
            rows = sorted(rows, key=lambda x: x.get("added_at", "") or "", reverse=True)
    total = len(rows)
    off = max(0, int(offset))
    lim = int(limit)
    if lim and lim > 0:                       # paginated request
        lim = min(lim, 1000)
        page = rows[off:off + lim]
        nxt = off + lim if (off + lim) < total else None
    else:                                     # legacy: all leads, single page
        page = rows[off:] if off else rows
        lim = total
        nxt = None
    return JSONResponse({"leads": page, "total": total, "offset": off,
                         "limit": lim, "next": nxt})


@app.post("/leads")
async def add_leads(request: Request, leads: str = Form(""), csv: UploadFile | None = File(None)):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot add leads")
    csv_bytes = await csv.read() if csv is not None else None
    parsed = parse_upload(leads, csv, csv_bytes)
    # RC2: stamp a batch_id + source_file when leads arrive via an uploaded file,
    # so leads can later be filtered by which upload they came from. Manual
    # text-only adds stay batch-less (back-compat: existing rows keep working).
    src_file = (getattr(csv, "filename", "") or "") if (csv is not None and csv_bytes) else ""
    batch_id = uuid.uuid4().hex[:8] if src_file else ""
    now_iso = datetime.now().isoformat(timespec="seconds")
    store = _read(LEADS_FILE, [])
    # de-dup within THIS tenant's leads (different tenants may share a number)
    have = {x["phone"] for x in store if x.get("tenant_id", ADMIN_ID) == t["tenant_id"]}
    added = 0
    for x in parsed:
        if x["num"] not in have:
            rec = {"id": uuid.uuid4().hex[:8], "tenant_id": t["tenant_id"],
                   "name": x["name"], "phone": x["num"],
                   "status": "new", "added_at": now_iso}
            if batch_id:
                rec["batch_id"] = batch_id
                rec["source_file"] = src_file
            store.append(rec)
            have.add(x["num"]); added += 1
    _write(LEADS_FILE, store)
    _audit(request, t, "leads.add", "lead", "", meta={"added": added, "batch_id": batch_id,
                                                        "source_file": src_file})
    resp = {"added": added, "total": len(_leads_for(t))}
    if batch_id:
        resp["batch_id"] = batch_id
        resp["source_file"] = src_file
    return JSONResponse(resp)


@app.get("/leads/batches")
async def leads_batches(request: Request):
    """RC2: return the distinct upload batches for this tenant, derived by
    grouping the lead store on batch_id (no new storage). Each batch =
    {batch_id, source_file, count, added_at}. Leads with no batch_id (manual /
    legacy adds) are folded into a single synthetic 'manual' bucket so the UI can
    always show a complete picture. Newest-first by added_at."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    groups: dict[str, dict] = {}
    for x in _leads_for(t):
        bid = x.get("batch_id") or ""
        key = bid or "manual"
        g = groups.get(key)
        added = x.get("added_at", "") or ""
        if g is None:
            groups[key] = {
                "batch_id": bid,
                "source_file": x.get("source_file", "") or ("" if bid else "Manual / legacy"),
                "count": 1,
                "added_at": added,
            }
        else:
            g["count"] += 1
            if added and (not g["added_at"] or added > g["added_at"]):
                g["added_at"] = added
            if bid and not g["source_file"]:
                g["source_file"] = x.get("source_file", "")
    batches = sorted(groups.values(), key=lambda b: b.get("added_at", ""), reverse=True)
    return JSONResponse({"batches": batches})


@app.delete("/leads/{lead_id}")
async def delete_lead(request: Request, lead_id: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot delete leads")
    store = _read(LEADS_FILE, [])
    target = next((x for x in store if x.get("id") == lead_id), None)
    guard = require_object(t, target)  # 404 if missing or not owned (BOLA)
    if guard is not None:
        return guard
    store = [x for x in store if x.get("id") != lead_id]
    _write(LEADS_FILE, store)
    _audit(request, t, "leads.delete", "lead", lead_id)
    return JSONResponse({"deleted": lead_id, "total": len(_leads_for(t))})


@app.post("/leads/delete")
async def delete_leads_bulk(request: Request, ids: str = Form("")):
    """Delete a SET of this tenant's leads by id (multi-select). Idempotent:
    unknown / already-deleted / cross-tenant ids are simply skipped (never error,
    never touch another tenant's rows). Tenant-scoped STRICTLY by tenant_id — even
    an admin only deletes rows it actually owns here (no cross-tenant wipe). The
    ids arrive as a comma/space-separated form field (max form-field compat)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot delete leads")
    want = _csv_set(ids)
    if not want:
        return JSONResponse({"deleted": 0, "total": len(_leads_for(t))})
    tid = t["tenant_id"]
    store = _read(LEADS_FILE, [])
    # Owned + requested = the rows we may remove (BOLA: tenant_id must match).
    kill = {x.get("id") for x in store
            if x.get("id") in want and x.get("tenant_id", ADMIN_ID) == tid}
    if not kill:
        return JSONResponse({"deleted": 0, "total": len(_leads_for(t))})
    store = [x for x in store if x.get("id") not in kill]
    _write(LEADS_FILE, store)
    _audit(request, t, "leads.delete_bulk", "lead", "",
           meta={"deleted": len(kill), "ids": sorted(kill)})
    return JSONResponse({"deleted": len(kill), "total": len(_leads_for(t))})


@app.delete("/leads")
async def delete_all_leads(request: Request, confirm: str = ""):
    """Delete ALL of THIS tenant's leads (destructive). STRICTLY tenant-scoped by
    tenant_id — NEVER cross-tenant, even for an admin token (we filter by tenant_id
    explicitly rather than via _leads_for's admin all-view). Confirm-gated: requires
    ?confirm=DELETE so a stray call can't wipe leads. Idempotent (0 leads -> 0).
    Other tenants' rows are preserved byte-for-byte."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot delete leads")
    if (confirm or "").strip().upper() != "DELETE":
        return JSONResponse({"error": "confirm required",
                             "detail": "pass ?confirm=DELETE to wipe all of this tenant's leads"},
                            status_code=400)
    tid = t["tenant_id"]
    store = _read(LEADS_FILE, [])
    mine = [x for x in store if x.get("tenant_id", ADMIN_ID) == tid]
    keep = [x for x in store if x.get("tenant_id", ADMIN_ID) != tid]
    deleted = len(mine)
    if deleted:
        _write(LEADS_FILE, keep)
    _audit(request, t, "leads.delete_all", "lead", "", meta={"deleted": deleted})
    return JSONResponse({"deleted": deleted, "total": 0})


@app.get("/leads/{phone}/memory")
async def lead_memory_get(request: Request, phone: str):
    """Durable, cross-channel relationship memory for one lead (the profile Memory tab).
    Built by the crm spine from the lead's durable fields + call insights. Dormant-safe:
    with the crm module absent we return {memory: null} (200) so the panel shows a calm
    'no memory yet' state, never an error. `phone` is any phone form OR a ct_ contact id."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    phone_n = norm(phone) or (phone or "")
    if _crm_mod is None or not hasattr(_crm_mod, "lead_memory"):
        return JSONResponse({"phone": phone_n, "memory": None})
    org = t["tenant_id"]
    adm = bool(t.get("is_admin"))
    mem = await asyncio.to_thread(lambda: _crm_mod.lead_memory(org, phone, is_admin=adm))
    return JSONResponse({"phone": phone_n, "memory": mem})


@app.get("/leads/{phone}/episodes")
async def lead_episodes_get(request: Request, phone: str, limit: int = 50, offset: int = 0):
    """Conversation history (one summarised episode per call) for one lead. Dormant-safe:
    crm absent -> {episodes: []} (200). Tenant-scoped from the token."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    phone_n = norm(phone) or (phone or "")
    if _crm_mod is None or not hasattr(_crm_mod, "lead_episodes"):
        return JSONResponse({"phone": phone_n, "episodes": [], "total": 0,
                             "offset": max(0, int(offset)), "limit": int(limit or 50), "next": None})
    org = t["tenant_id"]
    adm = bool(t.get("is_admin"))
    res = await asyncio.to_thread(
        lambda: _crm_mod.lead_episodes(org, phone, limit=int(limit or 50),
                                       offset=max(0, int(offset)), is_admin=adm))
    res = dict(res or {})
    res.setdefault("phone", phone_n)
    res["phone"] = phone_n
    return JSONResponse(res)


@app.post("/run/preview")
async def run_preview(request: Request, leads: str = Form(""),
                      use_stored: str = Form(""), source_mode: str = Form(""),
                      lead_ids: str = Form(""), temps: str = Form(""),
                      batch_ids: str = Form(""), score_min: str = Form(""),
                      score_max: str = Form(""), csv: UploadFile | None = File(None)):
    """RC2: resolve the SAME composable audience selectors as /run and return a
    truthful preview { count, suppressed_count, breakdown } WITHOUT creating a job
    or dialling anyone. Lets the UI show 'N leads will be called' before launch.
    No paid call can originate here — there is no run_job dispatch."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    csv_bytes = await csv.read() if csv is not None else None
    parsed = parse_upload(leads, csv, csv_bytes)
    resolved = _resolve_audience(
        t, source_mode=source_mode, lead_ids=lead_ids, temps=temps,
        batch_ids=batch_ids, score_min=score_min, score_max=score_max)
    rc2_used = bool(resolved or _csv_set(lead_ids) or _csv_set(batch_ids)
                    or _csv_set(temps) or str(score_min).strip() or str(score_max).strip()
                    or source_mode in ("all", "temperature", "upload", "manual"))
    if resolved:
        parsed += [{"name": x.get("name", ""), "num": x.get("phone", "")} for x in resolved]
    if use_stored and not rc2_used:
        parsed += [{"name": x.get("name", ""), "num": x["phone"]} for x in _leads_for(t)]
    seen, uniq = set(), []
    for x in parsed:
        if x["num"] and x["num"] not in seen:
            seen.add(x["num"]); uniq.append(x)
    supp = _suppressed_set(t["tenant_id"])
    suppressed_count = sum(1 for x in uniq if x["num"] in supp)
    _phone_temp = {x.get("phone"): _lead_temp(x) for x in _leads_for(t)}
    breakdown = {"hot": 0, "warm": 0, "cold": 0}
    for x in uniq:
        tb = _phone_temp.get(x["num"])
        if tb in breakdown:
            breakdown[tb] += 1
    # callable count = unique candidates minus those already suppressed (DND).
    callable_count = sum(1 for x in uniq if x["num"] not in supp)
    return JSONResponse({"count": len(uniq), "callable_count": callable_count,
                         "suppressed_count": suppressed_count, "breakdown": breakdown})


@app.post("/run")
async def run(request: Request, campaign_id: str = Form(""), leads: str = Form(""),
              use_stored: str = Form(""), concurrency: int = Form(3),
              hourly_cap: int = Form(200), daily_cap: int = Form(1000),
              force: str = Form(""),
              now: str = Form(""),  # LPR-FORCE-WINDOW: AIM "run now" -> bypass the calling-window idle in run_job
              # RC2 composable OPTIONAL audience selectors (all default ""/absent
              # -> legacy behaviour: csv + leads-text + use_stored unchanged).
              source_mode: str = Form(""), lead_ids: str = Form(""),
              temps: str = Form(""), batch_ids: str = Form(""),
              score_min: str = Form(""), score_max: str = Form(""),
              csv: UploadFile | None = File(None)):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot run campaigns")
    cid = (campaign_id or "").strip()
    camp = get_campaign_for(cid, t) if cid else None
    if cid and not camp:
        return JSONResponse({"error": "campaign not found for this account"}, status_code=404)
    camp_fields = (camp or {}).get("fields", {}) or {}
    tenant_id = t["tenant_id"]
    # P0.7 monthly minutes admission cap
    tenant_rec = _tenant_by_id(tenant_id) or {}
    monthly_cap = int(tenant_rec.get("monthly_minutes_cap", 5000))
    used_min = _tenant_usage(tenant_id, _month_iso())["minutes"]
    if used_min >= monthly_cap:
        return JSONResponse({"error": "monthly minutes cap reached",
                             "used_minutes": used_min, "cap": monthly_cap}, status_code=429)
    # WAVE3 Unit4: prepaid balance gate. Prepaid + balance<=0 -> refuse (402-style), no crash.
    billing = _billing_for(tenant_id)
    if billing.get("plan") == "prepaid" and float(billing.get("balance", 0) or 0) <= 0:
        return JSONResponse({"error": "insufficient balance",
                             "message": "Your prepaid balance is exhausted. Please top up to run campaigns.",
                             "balance": float(billing.get("balance", 0) or 0),
                             "currency": billing.get("currency", "INR")}, status_code=402)
    csv_bytes = await csv.read() if csv is not None else None
    parsed = parse_upload(leads, csv, csv_bytes)
    # RC2: composable audience selectors (lead_ids / temps / batch_ids / score band
    # / source_mode). Resolved over THIS tenant's stored leads (BOLA-safe). When any
    # are present they ADD to the ad-hoc parsed set; legacy use_stored is honoured
    # only when no RC2 selector narrows the set (so use_stored stays "all stored").
    resolved = _resolve_audience(
        t, source_mode=source_mode, lead_ids=lead_ids, temps=temps,
        batch_ids=batch_ids, score_min=score_min, score_max=score_max)
    rc2_used = bool(resolved or _csv_set(lead_ids) or _csv_set(batch_ids)
                    or _csv_set(temps) or str(score_min).strip() or str(score_max).strip()
                    or source_mode in ("all", "temperature", "upload", "manual"))
    if resolved:
        parsed += [{"name": x.get("name", ""), "num": x.get("phone", "")} for x in resolved]
    # Legacy "use all stored" — kept working, but suppressed when RC2 selectors are
    # driving the audience (otherwise picking a subset would still dial everyone).
    if use_stored and not rc2_used:
        parsed += [{"name": x.get("name", ""), "num": x["phone"]} for x in _leads_for(t)]
    seen, uniq = set(), []
    for x in parsed:
        if x["num"] and x["num"] not in seen:
            seen.add(x["num"]); uniq.append(x)
    # P0.2 pre-filter suppressed numbers (visibility; dial loop also enforces)
    supp = _suppressed_set(tenant_id)
    suppressed_count = sum(1 for x in uniq if x["num"] in supp)
    # RC2 preview breakdown: temperature counts over the resolved stored leads.
    _phone_temp = {x.get("phone"): _lead_temp(x) for x in _leads_for(t)}
    temp_breakdown = {"hot": 0, "warm": 0, "cold": 0}
    for x in uniq:
        tb = _phone_temp.get(x["num"])
        if tb in temp_breakdown:
            temp_breakdown[tb] += 1
    # P0.7 clamp concurrency to tenant cap
    conc = max(1, min(int(concurrency), 20, int(tenant_rec.get("max_concurrency", 3))))
    job_id = uuid.uuid4().hex[:10]
    JOBS[job_id] = {
        "state": "queued", "campaign_id": cid, "tenant_id": tenant_id,
        "concurrency": conc,
        "hourly_cap": max(1, int(hourly_cap)), "daily_cap": max(1, int(daily_cap)),
        # LPR-FORCE-WINDOW: when AIM (or the panel) explicitly asks to dial NOW, run_job
        # skips the out-of-window idle. Default False -> normal TRAI 09-21 window honoured.
        # LPR-FORCE-WINDOW-FIX: honour BOTH the AIM "now" field AND the panel's
        # "force" field (the "Start anyway" button) so an explicit dashboard
        # override actually dials outside the TRAI 09-21 window. Default False.
        "force_window": bool(str(now).strip()) or str(force).strip().lower() in ("1","true","yes","on"),
        "leads": [{"name": x["name"], "num": x["num"], "status": "queued", "room": "",
                   "launched_at": 0.0, "attempt": 0}
                  for x in uniq],
    }
    asyncio.create_task(run_job(job_id))
    _audit(request, t, "run.dispatch", "job", job_id,
           meta={"campaign_id": cid, "count": len(uniq), "suppressed": suppressed_count})
    # P0.1 calling-window gate: out of window (and not forced) -> 202, job still created + auto-resumes.
    in_win, win = _in_window(camp_fields)
    _dial_now = bool(str(now).strip())  # LPR-FORCE-WINDOW: explicit dial-now bypasses the window
    if not in_win and not force and not _dial_now:
        return JSONResponse({"queued_out_of_window": True, "window": win, "job_id": job_id,
                             "count": len(uniq), "suppressed_count": suppressed_count,
                             "breakdown": temp_breakdown}, status_code=202)
    return JSONResponse({"job_id": job_id, "count": len(uniq),
                         "suppressed_count": suppressed_count, "breakdown": temp_breakdown})


@app.get("/status")
async def status(request: Request, job: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    j = JOBS.get(job)
    if not j:
        return JSONResponse({"state": "unknown", "leads": []})
    # don't leak another tenant's job
    if not t.get("is_admin") and j.get("tenant_id", ADMIN_ID) != t["tenant_id"]:
        return JSONResponse({"state": "unknown", "leads": []})
    return JSONResponse({"state": j["state"],
                         "leads": [{"name": x["name"], "num": x["num"], "status": x["status"]} for x in j["leads"]]})


@app.post("/jobs/{job_id}/stop")
async def stop_job(request: Request, job_id: str):
    """Stop a running campaign: halt NEW dialing immediately; the few in-flight calls drain on their
    own. Tenant-scoped (can't touch another tenant's job), idempotent, never 500s."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    j = JOBS.get(job_id)
    if not j or (not t.get("is_admin") and j.get("tenant_id", ADMIN_ID) != t["tenant_id"]):
        return JSONResponse({"ok": False, "error": "not_found"}, status_code=404)
    j["stopped"] = True
    if j.get("state") in ("queued", "running"):
        j["state"] = "stopping"
    try:
        _log_event("info", "dialer", f"campaign job {job_id} stopped by operator",
                   tenant={"tenant_id": t.get("tenant_id", "")}, context={"job_id": job_id})
    except Exception:  # noqa: BLE001
        pass
    return JSONResponse({"ok": True, "job_id": job_id, "state": j.get("state", "stopping")})


# === PERF UNIT-1 ===
# Slim list-row payload: the call-logs TABLE only renders these fields. Heavy/internal
# bookkeeping (recording_key/bucket, sip_call_id, _reconciled, _wh_completed, room,
# variant_*) is dropped from LIST rows — the full record still ships on GET /calls/{id}.
_CALLS_LIST_FIELDS = (
    "id", "name", "phone", "campaign_id", "campaign_name", "status",
    "started_at", "ended_at", "duration_s", "interest", "outcome", "answered",
)


def _slim_call_row(c: dict) -> dict:
    return {k: c.get(k) for k in _CALLS_LIST_FIELDS if k in c}


def _call_outcome_cached(c: dict) -> str:
    """Outcome for filtering WITHOUT an N+1 per-row transcript read. 254/263 rows already
    carry `outcome`; for the rare row that doesn't, read its transcript ONCE and cache the
    value back onto the in-RAM record so the next request is O(1). Never raises."""
    o = c.get("outcome")
    if o:
        return o
    room = c.get("room")
    if not room:
        return ""
    try:
        tr = _read(TRANSCRIPT_DIR / f"{room}.json", {}) or {}
        o = tr.get("outcome") or ""
    except Exception:
        o = ""
    if o:
        c["outcome"] = o          # memoize onto the record (in-RAM CALLS list) -> no repeat read
    return o


@app.get("/calls")
async def calls(request: Request, limit: int = 200, offset: int = 0,
                campaign_id: str = "", outcome: str = "", order: str = "",
                slim: str = "", sort_by: str = ""):
    """Call logs. Backward-compatible: a bare `/calls` (or `/calls?limit=N`) returns the
    SAME `{calls:[...]}` in storage order it always did. Pagination is opt-in:
      - offset (int>=0): page start; presence implies the caller wants the paginated contract.
      - order=desc: newest-first by started_at (recommended for the paginated list view).
      - slim=1 (default behaviour for paginated callers; legacy bare call gets full rows for
        strict back-compat): trim list rows to display fields.
    Response ALWAYS now also carries total/offset/limit/next so the FE can paginate; the
    legacy `calls` key is preserved. The N+1 transcript read on the `outcome` filter is gone
    (outcome is read from the row, with a one-time cached fallback)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = calls_for(t)
    if campaign_id:
        rows = [c for c in rows if c.get("campaign_id") == campaign_id]
    if outcome:
        rows = [c for c in rows if _call_outcome_cached(c) == outcome]
    # ADDITIVE date-range filter (the shared dashboard/report range). Read from the query params so
    # `from` (a Python keyword) needs no signature alias. Inclusive, compared on the YYYY-MM-DD prefix
    # of started_at (ISO sorts lexicographically) so it works for either a date or a datetime input.
    # Absent => byte-identical to before.
    _frm = (request.query_params.get("from") or "").strip()[:10]
    _to = (request.query_params.get("to") or "").strip()[:10]
    if _frm:
        rows = [c for c in rows if (c.get("started_at") or "")[:10] >= _frm]
    if _to:
        rows = [c for c in rows if (c.get("started_at") or "")[:10] <= _to]

    paginated = bool(str(offset).strip()) and int(offset) > 0
    paginated = paginated or order.lower() in ("desc", "asc") or bool(str(sort_by).strip()) \
        or str(slim).strip().lower() in ("1", "true", "yes")

    # R5P4-3 (ADDITIVE): column sort the panel call-logs sends (sort_by + order). Maps the FE column
    # keys to call-row fields; unknown column -> the legacy started_at sort. Back-compat: with no
    # sort_by the only behavior change is honoring order=asc (was: only desc) -> bare /calls unchanged.
    _sb = (sort_by or "").strip().lower()
    if _sb:
        _desc = order.lower() != "asc"   # default desc for an explicit column sort

        def _calls_key(c):  # noqa: ANN001
            if _sb in ("duration_s", "duration"):
                return int(c.get("duration_s") or 0)
            if _sb in ("interest", "score"):
                return int(c.get("interest") or 0)
            if _sb in ("name", "lead"):
                return (c.get("name", "") or "").lower()
            if _sb in ("campaign_name", "campaign", "campaign_id"):
                return (c.get("campaign_name") or c.get("campaign_id") or "").lower()
            if _sb in ("status", "outcome"):
                return (_call_outcome_cached(c) or "").lower()
            return c.get("started_at") or ""   # started_at / placed / unknown
        rows = sorted(rows, key=_calls_key, reverse=_desc)
    elif order.lower() == "asc":
        rows = sorted(rows, key=lambda c: c.get("started_at") or "")
    elif order.lower() == "desc":
        rows = sorted(rows, key=lambda c: c.get("started_at") or "", reverse=True)

    total = len(rows)
    off = max(0, int(offset))
    lim = max(1, min(int(limit), 1000))
    page = rows[off:off + lim]
    nxt = off + lim if (off + lim) < total else None

    # Trim list rows: always for paginated callers; the bare legacy call keeps full rows
    # unless it explicitly asks slim=1, so the existing FE is byte-compatible.
    want_slim = str(slim).strip().lower() in ("1", "true", "yes") or (paginated and str(slim).strip() == "")
    out = [_slim_call_row(c) for c in page] if want_slim else page

    return JSONResponse({"calls": out, "total": total, "offset": off,
                         "limit": lim, "next": nxt})
# === /PERF UNIT-1 ===


@app.get("/calls/{call_id}")
async def call_detail(request: Request, call_id: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    # Look up across ALL calls, then assert ownership explicitly (BOLA guard).
    # A genuinely missing id -> 404; another tenant's call -> 403 (require_object).
    rec = next((c for c in CALLS if c.get("id") == call_id), None)
    if rec is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    guard = require_object(t, rec, not_found=False)
    if guard is not None:
        return guard
    transcript = _read(TRANSCRIPT_DIR / f"{rec.get('room', '')}.json", {}) if rec.get("room") else {}
    # REC-FIX: enrich the call detail with the SAME unified recording shape used by
    # /calls/{id}/recording, so the Call-Logs modal gets the presigned, range-streamable player URL in
    # the SAME request (no separate fetch, no missed field). The shaper HEAD-verifies the Spaces object,
    # flips a stuck "recording" status -> "completed" when the object is playable, and degrades to an
    # empty url (never a 500) on any Spaces/boto3 hiccup. We also fold the key fields onto the `call`
    # object so any FE that reads call.recording_presigned_url / call.recording_status works too.
    try:
        recv = _outbound_rec_item(rec)
        if isinstance(rec, dict) and recv:
            rec["recording_presigned_url"] = recv.get("recording_presigned_url", "")
            rec["recording_playable"] = bool(recv.get("playable"))
            rec["recording_status"] = recv.get("recording_status", rec.get("recording_status", ""))
            rec["recording_size_bytes"] = int(recv.get("size_bytes", 0) or 0)
        return JSONResponse({"call": rec, "transcript": transcript, "recording": recv})
    except Exception:  # noqa: BLE001
        # any recording-shaping failure must NEVER break the call-detail read (earner-safe).
        return JSONResponse({"call": rec, "transcript": transcript})


# ==============================================================================================
# REC-C: UNIFIED RECORDINGS API (tenant-scoped). One shape over BOTH call directions:
#   - OUTBOUND: the JSON CALLS store (REC-B server-side auto-egress; recording_key embeds call_id).
#   - INBOUND : the ai_manager_sessions PG table (REC-A; recording_egress_id + finalize-on-read).
# Every recording url is a SHORT-LIVED presigned GET (the capsy-recordings bucket is PRIVATE) that is
# range-streamable + downloadable; the bucket is NEVER made public. Tenant is pinned from the TOKEN
# (resolve_tenant) exactly like /contacts -> a tenant can only read its OWN recordings (RLS on the
# inbound PG side; tenant_id filter + BOLA guard on the outbound JSON side). All helpers NEVER raise
# (a Spaces/boto3/PG hiccup degrades to has_recording with an empty url, never a 500).
# additive: famit-caller only; agent.py / earner / trunks / firewall / SIP untouched.
# ----------------------------------------------------------------------------------------------
_REC_TERMINAL = ("uploaded", "failed", "disabled")


# REC-FIX: HEAD/presign the recording DIRECTLY against DO Spaces (where the egress
# actually uploads, via AIM_SPACES_*). The old path used ai_manager.recorder, which
# reads R2/B2 creds that aren't configured here -> it never found the object, so the
# panel was stuck on "Preparing recording…" even though the OGG was in Spaces.
_AIM_S3_CLIENT = None


def _aim_s3():
    global _AIM_S3_CLIENT
    if _AIM_S3_CLIENT is None:
        import boto3
        from botocore.config import Config as _BotoCfg
        _AIM_S3_CLIENT = boto3.client(
            "s3",
            endpoint_url=(cfg_get("AIM_SPACES_ENDPOINT", "") or "").strip(),
            aws_access_key_id=(cfg_get("AIM_SPACES_KEY", "") or "").strip(),
            aws_secret_access_key=(cfg_get("AIM_SPACES_SECRET", "") or "").strip(),
            region_name=(cfg_get("AIM_SPACES_REGION", "") or "us-east-1").strip(),
            config=_BotoCfg(signature_version="s3v4", s3={"addressing_style": "path"}),
        )
    return _AIM_S3_CLIENT


def _rec_presign(bucket: str, key: str, expires_s: int = 3600) -> str:
    """Mint a short-lived presigned GET for a DO Spaces recording object (sigv4 + path-style,
    range-streamable). '' when no object / any error -> panel shows 'recorded, link unavailable'."""
    if not bucket or not key:
        return ""
    try:
        return _aim_s3().generate_presigned_url(
            "get_object", Params={"Bucket": bucket, "Key": key}, ExpiresIn=int(expires_s)) or ""
    except Exception:  # noqa: BLE001
        return ""


# PERF UNIT-2: a recording object must be non-trivially sized + audio before we call it PLAYABLE.
# Outbound auto-egress can leave a near-empty / 486-busy OGG: duration_s is set but the bytes don't
# decode, so the panel's <audio> runs a timer and plays nothing. We HEAD-verify the object FIRST and
# only mark playable (and only presign) when a real audio file is present.
_MIN_PLAYABLE_REC_BYTES = int(os.getenv("REC_MIN_PLAYABLE_BYTES", "2048") or 2048)


def _rec_playable(bucket: str, key: str) -> dict:
    """HEAD-verify a Spaces recording object. Returns {playable:bool, size_bytes:int}. A file counts
    as playable when it EXISTS, is >= _MIN_PLAYABLE_REC_BYTES, and looks like audio (content-type
    audio/* OR an .ogg/.mp3/.m4a/.wav key — Spaces sometimes returns a generic content-type). NEVER
    raises -> {playable:False} on any error so the FE shows 'preparing', never a broken player."""
    if not bucket or not key:
        return {"playable": False, "size_bytes": 0}
    try:
        h = _aim_s3().head_object(Bucket=bucket, Key=key)
    except Exception:  # noqa: BLE001 — 404 / creds / network -> not yet playable
        return {"playable": False, "size_bytes": 0}
    size = int(h.get("ContentLength", 0) or 0)
    ctype = str(h.get("ContentType", "") or "").lower()
    key_l = key.lower()
    looks_audio = ctype.startswith("audio/") or key_l.endswith((".ogg", ".mp3", ".m4a", ".wav", ".webm"))
    playable = bool(size >= _MIN_PLAYABLE_REC_BYTES and looks_audio)
    return {"playable": playable, "size_bytes": size}


# ── REC-B-AZURE serving helpers ──────────────────────────────────────────────────────────────────
# Mirror of _aim_s3 / _rec_presign / _rec_playable, but for Azure Blob Storage. Used ONLY when a
# recording row carries recording_backend=="azure" (or RECORDING_BACKEND=="azure"). The azure SDK is
# imported LAZILY inside the helpers so the module still imports if `azure-storage-blob` is absent
# (then these just return ""/{"playable":False} and the panel shows "preparing"). NEVER raise.
_AZURE_BLOB_SERVICE_CLIENT = None


def _azure_blob_client():
    """Cached BlobServiceClient built from AZURE_STORAGE_ACCOUNT + AZURE_STORAGE_KEY. Lazy-imports
    azure-storage-blob so a missing lib degrades gracefully (returns None). NEVER raises."""
    global _AZURE_BLOB_SERVICE_CLIENT
    if _AZURE_BLOB_SERVICE_CLIENT is None:
        try:
            from azure.storage.blob import BlobServiceClient  # type: ignore  # noqa: PLC0415
            account = (cfg_get("AZURE_STORAGE_ACCOUNT", "") or "").strip()
            akey = (cfg_get("AZURE_STORAGE_KEY", "") or "").strip()
            if not account or not akey:
                return None
            _AZURE_BLOB_SERVICE_CLIENT = BlobServiceClient(
                account_url=f"https://{account}.blob.core.windows.net",
                credential=akey,
            )
        except Exception:  # noqa: BLE001 — missing lib / bad creds -> caller falls back to "preparing"
            return None
    return _AZURE_BLOB_SERVICE_CLIENT


def _azure_sas_url(container: str, key: str, expires_s: int = 3600) -> str:
    """Mint a short-lived read-only SAS GET URL for an Azure blob:
    https://<acct>.blob.core.windows.net/<container>/<key>?<sas>. '' on any error (incl. missing lib)."""
    if not container or not key:
        return ""
    try:
        from datetime import datetime as _dt, timedelta as _td, timezone as _tz  # noqa: PLC0415
        from azure.storage.blob import generate_blob_sas, BlobSasPermissions  # type: ignore  # noqa: PLC0415
        account = (cfg_get("AZURE_STORAGE_ACCOUNT", "") or "").strip()
        akey = (cfg_get("AZURE_STORAGE_KEY", "") or "").strip()
        if not account or not akey:
            return ""
        sas = generate_blob_sas(
            account_name=account,
            container_name=container,
            blob_name=key,
            account_key=akey,
            permission=BlobSasPermissions(read=True),
            expiry=_dt.now(_tz.utc) + _td(seconds=int(expires_s)),
        )
        return (f"https://{account}.blob.core.windows.net/"
                f"{urllib.parse.quote(container)}/{urllib.parse.quote(key)}?{sas}")
    except Exception:  # noqa: BLE001
        return ""


def _azure_blob_playable(container: str, key: str) -> dict:
    """HEAD an Azure blob via get_blob_properties. Returns {playable:bool, size_bytes:int}. Playable
    when the blob exists, is >= _MIN_PLAYABLE_REC_BYTES, and looks like audio (content-type audio/* OR
    an .ogg/.mp3/.m4a/.wav key). NEVER raises -> {playable:False} on any error (missing lib / 404)."""
    if not container or not key:
        return {"playable": False, "size_bytes": 0}
    try:
        svc = _azure_blob_client()
        if svc is None:
            return {"playable": False, "size_bytes": 0}
        bc = svc.get_blob_client(container=container, blob=key)
        props = bc.get_blob_properties()
        size = int(getattr(props, "size", 0) or 0)
        cs = getattr(props, "content_settings", None)
        ctype = str(getattr(cs, "content_type", "") or "").lower()
    except Exception:  # noqa: BLE001 — 404 / creds / network / missing lib -> not yet playable
        return {"playable": False, "size_bytes": 0}
    key_l = key.lower()
    looks_audio = ctype.startswith("audio/") or key_l.endswith((".ogg", ".mp3", ".m4a", ".wav", ".webm"))
    playable = bool(size >= _MIN_PLAYABLE_REC_BYTES and looks_audio)
    return {"playable": playable, "size_bytes": size}


def _rec_backend_is_azure(rec: dict) -> bool:
    """True when this recording row should be served from Azure: explicit row marker, or the global
    RECORDING_BACKEND=='azure' selection. Row marker wins so historic Spaces rows stay on Spaces."""
    rb = str((rec or {}).get("recording_backend", "") or "").strip().lower()
    if rb == "azure":
        return True
    if rb == "spaces":
        return False
    return (cfg_get("RECORDING_BACKEND", "spaces") or "spaces").strip().lower() == "azure"


def _outbound_rec_item(rec: dict, *, presign: bool = True) -> dict:
    """Shape ONE outbound call row (REC-B) into the unified recording item. The deterministic
    recording_key IS the authoritative handle (auto-egress returns no id at room-create)."""
    # ── W9 self-heal (RECORDING_FINALIZE_ENABLED): if a row is stuck at "recording", HEAD the
    #    deterministic key and flip it to completed + presigned on READ (unifies the manual HEAD below).
    if presign and _RECCFG is not None and _RECCFG.enabled:
        try:
            from voice_ops.recording.api import build_recording_view  # type: ignore  # noqa: PLC0415
            view = build_recording_view(
                rec, storage=_ObjStorage(_RECCFG), tenant_id=rec.get("tenant_id", ""))
            if view:
                return view
        except Exception:  # noqa: BLE001
            pass   # fall through to the existing sync HEAD-check logic
    bucket = (rec.get("recording_bucket", "") or "")
    key = (rec.get("recording_key", "") or "")
    rstatus = (rec.get("recording_status", "") or "")
    has_rec = bool(bucket and key and rstatus not in ("", "disabled"))
    # REC-B-AZURE: route the HEAD-verify + presign through Azure when this row is Azure-backed
    # (recording_backend=="azure" on the row, or RECORDING_BACKEND=="azure"); otherwise the default
    # DO Spaces path runs UNCHANGED. recording_bucket holds the Azure container in the Azure case.
    is_azure = _rec_backend_is_azure(rec)
    # PERF UNIT-2: HEAD-verify the object BEFORE presigning. An auto-egress 486-busy/near-empty OGG has
    # has_recording=True (status uploaded) but won't decode -> only mark playable + presign when the
    # object is a real non-trivial audio file. The HEAD is a single cheap call, only when has_rec.
    if presign and has_rec:
        pv = _azure_blob_playable(bucket, key) if is_azure else _rec_playable(bucket, key)
    else:
        pv = {"playable": False, "size_bytes": 0}
    playable = bool(pv.get("playable"))
    # REC-FIX: the row's recording_status is stamped "recording" at room-create and nothing ever flips
    # it (the W9 finalize-poller is mis-keyed / a no-op). So on READ, once the object is HEAD-verified
    # playable, report the HONEST terminal status "completed" — every UI that gates on
    # recording_status == completed/uploaded then shows the player. Read-only rewrite of the SHAPE only;
    # the underlying calls.json row is left untouched here (a separate best-effort persist does that).
    eff_status = rstatus or ("disabled" if not key else "")
    if playable:
        eff_status = "completed"
    item = {
        "call_id": rec.get("id", ""),
        "direction": "outbound",
        "phone": rec.get("phone", ""),
        "name": rec.get("name", ""),
        "campaign_id": rec.get("campaign_id", ""),
        "started_at": rec.get("started_at", ""),
        "duration_s": int(rec.get("duration_s", 0) or 0),
        "status": rec.get("status", ""),
        "recording_status": eff_status,
        "has_recording": has_rec,
        "playable": playable,
        "size_bytes": int(pv.get("size_bytes", 0) or 0),
        # only hand the FE a URL for a VERIFIED playable object -> the FE renders <audio> iff this is set.
        # REC-B-AZURE: Azure rows get a read-only SAS URL; Spaces rows get an S3 sigv4 presign (default).
        "recording_presigned_url": (
            ((_azure_sas_url(bucket, key) if is_azure else _rec_presign(bucket, key)))
            if playable else ""),
    }
    # REC-FIX: best-effort mutate the IN-MEMORY call row's status to the verified terminal state. This
    # makes GET /calls/{id} (which returns this same `rec` object) and the list report "completed"
    # immediately, and it gets durably persisted on the next normal _write(CALLS_FILE, CALLS) (call
    # finalize / scheduler sweep). We do NOT force a disk write from the read path (avoids lock
    # contention + write amplification on every recording open). NEVER raises into the read path.
    if playable and rstatus != "completed":
        try:
            rec["recording_status"] = "completed"
        except Exception:  # noqa: BLE001
            pass
    return item


def _inbound_rec_items(tid: str, phone_n: str, *, presign: bool = True) -> list[dict]:
    """All INBOUND (ai_manager_sessions) recordings for one phone, RLS-scoped to `tid`. Reuses the
    AIM store (token-scoped vendor reads) + the REC-A finalize-on-read self-heal so a stuck
    'recording'/0 row gets its true terminal status + duration from LiveKit ListEgress before we
    presign. NEVER raises -> [] on any failure (PG down / module absent)."""
    out: list[dict] = []
    try:
        from ai_manager import store as _store
        if not _store.available():
            return out
        # list_sessions is RLS-scoped (engine.session(tenant_id=tid)); filter the voice rows for THIS phone.
        rows = _store.list_sessions(tid, limit=200, channel="voice") or []
        sids = [r.get("id") for r in rows
                if (r.get("caller_phone", "") or "") == phone_n and r.get("id")]
        for sid in sids:
            full = _store.get_session(tid, sid)  # RLS-scoped; bucket/key/egress_id
            if not full:
                continue
            bucket = (full.get("recording_bucket", "") or "")
            key = (full.get("recording_key", "") or "")
            rstatus = (full.get("recording_status", "") or "")
            rdur = int(full.get("recording_duration_s", 0) or 0)
            eg = (full.get("recording_egress_id", "") or "").strip()
            # REC-A finalize-on-read self-heal: a fire-and-forget hangup write often loses the race, so
            # reconcile the authoritative terminal state from LiveKit when the row still looks un-final.
            if eg and (rstatus not in _REC_TERMINAL or rdur <= 0):
                try:
                    from ai_manager import recorder as _recorder
                    fin = _recorder.finalize(eg)
                    if fin.get("complete"):
                        new_key = (fin.get("key", "") or "") or key
                        new_dur = int(fin.get("duration_s", 0) or 0)
                        try:
                            _store.set_recording(tid, sid, status="uploaded",
                                                 key=new_key, duration_s=new_dur)
                        except Exception:  # noqa: BLE001
                            pass
                        rstatus = "uploaded"
                        if new_key:
                            key = new_key
                        if new_dur > 0:
                            rdur = new_dur
                    elif fin.get("status") == "failed":
                        try:
                            _store.set_recording(tid, sid, status="failed")
                        except Exception:  # noqa: BLE001
                            pass
                        rstatus = "failed"
                except Exception:  # noqa: BLE001
                    pass
            has_rec = bool(bucket and key and rstatus not in ("", "disabled"))
            # PERF UNIT-2: inbound is already finalize-gated, but HEAD-verify anyway so the playable
            # contract is uniform across both directions (and a finalize race that wrote a 0-byte key
            # still degrades to 'preparing' instead of a dead player).
            pv = _rec_playable(bucket, key) if (presign and has_rec) else {"playable": False, "size_bytes": 0}
            playable = bool(pv.get("playable"))
            out.append({
                "call_id": sid,
                "direction": "inbound",
                "phone": full.get("caller_phone", "") or phone_n,
                "name": "",
                "campaign_id": "",
                "started_at": str(full.get("started_at", "") or ""),
                "duration_s": rdur,
                "status": full.get("status", "") or "",
                "recording_status": rstatus or ("disabled" if not key else ""),
                "has_recording": has_rec,
                "playable": playable,
                "size_bytes": int(pv.get("size_bytes", 0) or 0),
                "recording_presigned_url": (_rec_presign(bucket, key) if playable else ""),
            })
    except Exception:  # noqa: BLE001
        return out
    return out


@app.get("/calls/{call_id}/recording")
async def call_recording(request: Request, call_id: str):
    """The recording for ONE outbound call (REC-B). Returns a freshly-minted presigned, range-
    streamable + downloadable URL + metadata. Tenant pinned from token; BOLA-guarded (another
    tenant's id -> 404). 404 when the call id is unknown; has_recording=False (url '') when the
    call was placed before recording was armed / dialed unrecorded."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rec = next((c for c in CALLS if c.get("id") == call_id), None)
    if rec is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    guard = require_object(t, rec, not_found=True)   # cross-tenant -> 404 (don't reveal existence)
    if guard is not None:
        return guard
    return JSONResponse({"recording": _outbound_rec_item(rec)})


@app.get("/contacts/{phone}/recordings")
async def contact_recordings(request: Request, phone: str):
    """ALL recordings for one lead, UNIFIED across both directions (newest-first):
      outbound calls (REC-B JSON store) + inbound AI-Manager sessions (REC-A PG table),
    joined by the canonicalized phone. Each item: {call_id, direction, phone, started_at,
    duration_s, status, recording_status, has_recording, recording_presigned_url}. Tenant pinned
    from token (RLS on the inbound side, tenant_id filter on the outbound side) -> a tenant can
    only ever see its OWN lead's recordings. NEVER raises -> a degraded side returns []."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    phone_n = norm(phone) or (phone or "")
    items: list[dict] = []
    # OUTBOUND: the tenant-scoped JSON calls store, matched on the canonicalized phone.
    try:
        for c in calls_for(t):
            if norm(c.get("phone", "") or "") == phone_n or (c.get("phone", "") or "") == phone_n:
                items.append(_outbound_rec_item(c))
    except Exception:  # noqa: BLE001
        pass
    # INBOUND: the RLS-scoped ai_manager_sessions, matched on caller_phone (with finalize self-heal).
    try:
        items.extend(await asyncio.to_thread(
            lambda: _inbound_rec_items(t["tenant_id"], phone_n)))
    except Exception:  # noqa: BLE001
        pass
    items.sort(key=lambda x: (x.get("started_at", "") or ""), reverse=True)
    n_rec = sum(1 for x in items if x.get("has_recording"))
    n_play = sum(1 for x in items if x.get("playable"))
    return JSONResponse({"phone": phone_n, "recordings": items, "total": len(items),
                         "with_recording": n_rec, "with_playable": n_play})


# ════════════════════════════════════════════════════════════════════════════════════════════════
# W14 — REAL-TIME REPORTING ROUTES (REPORTING_ENABLED). Purely additive; touch no existing route.
# Tenant is TOKEN-derived (resolve_tenant), NEVER from the body. The read-model store is filled by a
# SEPARATE consumer worker tailing the W8 event stream; with REPORTING_ENABLED off these return 503.
# Every response echoes the resolved range {preset,from,to,tz} so the panel renders the window
# unambiguously and the "1 day ago" off-by-one is fixed (UTC store + vendor-tz day math).
# ════════════════════════════════════════════════════════════════════════════════════════════════
def _w14_filters(campaign, lead_status, source, agent, call_status, booking_status):  # noqa: ANN001
    return {k: v for k, v in {
        "campaign": campaign, "lead_status": lead_status, "source": source,
        "agent": agent, "call_status": call_status, "booking_status": booking_status,
    }.items() if v}


def _enrich_report_temperature(rep, tenant_id, preset, frm, to, filters):  # noqa: ANN001
    """R5P4-1: add `temperature_distribution` + `hot_leads` to a /report payload IN PLACE.

    The panel dashboard (`report.temperature_distribution` -> donut) and Report page read these two
    top-level keys; the base report only carried `totals.{hot,warm,cold,dead}` + `by_status`. We
    DERIVE the distribution from the SAME `totals` already in `rep` (so the donut can never diverge
    from the KPI counts), and pull `hot_leads` from the live reporting service (the identical scan the
    dashboard's hot-lead widget uses). Purely additive; if `rep` is not a dict we leave it untouched.
    """
    if not isinstance(rep, dict):
        return
    totals = rep.get("totals") or {}
    by_status = rep.get("by_status") or {}
    # prefer the explicit by_status breakdown; fall back to the totals counters (same numbers).
    def _band(name):  # noqa: ANN001
        v = by_status.get(name)
        if v is None:
            v = totals.get(name, 0)
        try:
            return int(v or 0)
        except Exception:  # noqa: BLE001
            return 0
    bands = {k: _band(k) for k in ("hot", "warm", "cold", "dead")}
    tot = sum(bands.values())
    # FE TemperatureBucket shape: {tier, count, pct, delta?}. pct = share of the 4 temperature bands.
    rep["temperature_distribution"] = [
        {"tier": tier, "count": bands[tier],
         "pct": round((bands[tier] * 100.0 / tot), 1) if tot else 0.0}
        for tier in ("hot", "warm", "cold", "dead")
    ]
    # mirror onto by_status too so any client reading `report.by_status.{hot,...}` also sees the bands.
    rep.setdefault("by_status", {})
    for k, v in bands.items():
        rep["by_status"].setdefault(k, v)
    # hot_leads (named rows for the panel list). Best-effort; absent service -> []. The service signature
    # is hot_leads(tenant, preset, *, frm, to, limit) -> list[{call_id,name,phone_masked,score?,...}].
    rows = []
    if _REPSVC is not None and "hot_leads" not in rep:
        try:
            rows = _REPSVC.hot_leads(tenant_id, preset, frm=frm, to=to, limit=25) or []
        except TypeError:
            try:
                rows = _REPSVC.hot_leads(tenant_id, preset, frm=frm, to=to) or []
            except Exception:  # noqa: BLE001
                rows = []
        except Exception:  # noqa: BLE001
            rows = []
        # surface a flat `score` (0-100) alongside conversion_prob so the panel can render either.
        # conversion_prob may already be a 0-100 percentage OR a 0-1 fraction — normalize to 0-100.
        for r in rows:
            if isinstance(r, dict) and "score" not in r and r.get("conversion_prob") is not None:
                try:
                    cp = float(r["conversion_prob"])
                    r["score"] = int(round(cp if cp > 1 else cp * 100))
                except Exception:  # noqa: BLE001
                    pass
    rep.setdefault("hot_leads", rows)


@app.get("/report")
async def report_query(
    request: Request,
    preset: str = "today", frm: str = "", to: str = "",
    campaign: str = "", lead_status: str = "", source: str = "",
    agent: str = "", call_status: str = "", booking_status: str = "",
):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    filters = _w14_filters(campaign, lead_status, source, agent, call_status, booking_status)
    try:
        rep = _REPSVC.report(t["tenant_id"], preset, frm=frm, to=to, filters=filters or None)
        # R5P4-1 (ADDITIVE): the panel dashboard + Report read `hot_leads` + `temperature_distribution`
        # off the SAME /report payload (they were rendering empty). Source both from the live read-model
        # already in `rep` — never a second store, never a divergent number. NEVER raises into the route.
        try:
            _enrich_report_temperature(rep, t["tenant_id"], preset, frm, to, filters)
        except Exception:  # noqa: BLE001 — enrichment is best-effort; the base report still returns
            pass
        return JSONResponse(rep)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": repr(exc)}, status_code=500)


@app.get("/report/funnel")
async def report_funnel(request: Request, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.funnel(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/timeline")
async def report_timeline(request: Request, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.timeline(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/agents")
async def report_agents(request: Request, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.agents(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/sources")
async def report_sources(request: Request, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.sources(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/campaigns")
async def report_campaigns(request: Request, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.campaigns(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/followups")
async def report_followups(request: Request, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.followups(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/hot-leads")
async def report_hot_leads(request: Request, preset: str = "today", frm: str = "", to: str = "",
                           limit: int = 25):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    try:
        return JSONResponse(_REPSVC.hot_leads(t["tenant_id"], preset, frm=frm, to=to, limit=limit))
    except TypeError:
        return JSONResponse(_REPSVC.hot_leads(t["tenant_id"], preset, frm=frm, to=to))


@app.get("/report/metric/{key}")
async def report_metric(request: Request, key: str, preset: str = "today", frm: str = "", to: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _REPSVC is None:
        return JSONResponse({"error": "reporting not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: replay the W8 stream into the in-proc read-model
    return JSONResponse(_REPSVC.metric(t["tenant_id"], key, preset, frm=frm, to=to))


@app.post("/ai-manager/report")
async def ai_manager_report(request: Request):
    """W14b: AI-Manager live-data command box. POST {"message":"how many calls today"} -> the SAME
    LIVE reporting numbers the dashboard shows (can never diverge into a stale cache). Token-scoped."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _AIM_LIVE is None:
        return JSONResponse({"error": "ai-manager live not enabled"}, status_code=503)
    await _w14_hydrate(t["tenant_id"])  # W14-WIRE: AI-Manager reads the SAME live read-model as /report
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    msg = (body.get("message") or body.get("command") or "").strip()
    try:
        return JSONResponse(_AIM_LIVE.handle(t["tenant_id"], msg))
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": repr(exc)}, status_code=500)


# ==============================================================================================
# TRANSCRIPT-PER-CALL API (tenant-scoped). One UNIFIED chat-view shape over BOTH call directions:
#   - OUTBOUND: transcripts/{room}.json turns {role: assistant|user, content} (room is on the CALLS row).
#   - INBOUND : ai_manager_sessions turns {seq, role: agent|user, text, created_at} (RLS PG table).
# Returns ordered turns {role, text, ts, seq} where role is NORMALIZED for the chat bubble side:
#   ai|assistant|agent  -> "ai"        (rendered on the LEFT)
#   user|customer       -> "customer"  (rendered on the RIGHT)
# Tenant is pinned from the TOKEN (resolve_tenant) exactly like /calls -> a tenant can only read its
# OWN call transcript (outbound = require_object BOLA guard; inbound = AIM store RLS by vendor_id).
# All helpers NEVER raise (a missing/empty transcript degrades to turns:[] , never a 500).
# ==============================================================================================
def _norm_turn_role(role: str) -> str:
    """Normalize a stored turn role to the chat-bubble side. ai/assistant/agent -> 'ai' (LEFT);
    user/customer -> 'customer' (RIGHT). Anything else falls back to 'ai' (system/tool lines sit
    on the agent side). NEVER raises."""
    r = (role or "").strip().lower()
    if r in ("user", "customer", "human", "caller", "lead"):
        return "customer"
    return "ai"


def _outbound_transcript_turns(rec: dict) -> list[dict]:
    """Ordered chat turns for ONE outbound call from transcripts/{room}.json. Outbound turns carry
    no per-turn timestamp -> ts is ''. Returns [] when there is no room / no transcript / no turns."""
    room = (rec.get("room", "") or "")
    if not room:
        return []
    tr = _read(TRANSCRIPT_DIR / f"{room}.json", {}) or {}
    out: list[dict] = []
    for i, turn in enumerate(tr.get("turns", []) or []):
        if not isinstance(turn, dict):
            continue
        text = (turn.get("content") or turn.get("text") or "").strip()
        if not text:
            continue
        out.append({
            "role": _norm_turn_role(turn.get("role", "")),
            "text": text,
            "ts": str(turn.get("ts", "") or ""),
            "seq": i,
        })
    return out


def _inbound_transcript_turns(tid: str, session_id: str) -> list[dict]:
    """Ordered chat turns for ONE inbound AI-Manager call from ai_manager_session_turns, RLS-scoped
    to `tid`. Reuses the AIM store (token-scoped vendor read). Returns [] when PG is down / the
    session is not this tenant's / has no turns. NEVER raises."""
    try:
        from ai_manager import store as _store
        if not _store.available():
            return []
        full = _store.get_session(tid, session_id)  # RLS-scoped by vendor_id == tid
        if not full:
            return []
        out: list[dict] = []
        for turn in (full.get("turns", []) or []):
            if not isinstance(turn, dict):
                continue
            text = (turn.get("text") or "").strip()
            if not text:
                continue
            out.append({
                "role": _norm_turn_role(turn.get("role", "")),
                "text": text,
                "ts": str(turn.get("created_at", "") or ""),
                "seq": int(turn.get("seq", 0) or 0),
            })
        out.sort(key=lambda x: (x.get("seq", 0), x.get("ts", "")))
        return out
    except Exception:  # noqa: BLE001
        return []


@app.get("/calls/{call_id}/transcript")
async def call_transcript(request: Request, call_id: str):
    """The FULL ordered transcript for ONE call as a chat view, UNIFIED across both directions.
    `call_id` resolves as an OUTBOUND call id first (the JSON CALLS store -> transcripts/{room}.json),
    else as an INBOUND ai_manager_sessions session_id (room/session_id are accepted as call_id too).
    Returns {call_id, direction, phone, name, turns:[{role:"ai"|"customer", text, ts, seq}], total}.
    Tenant pinned from token: outbound is BOLA-guarded (another tenant's id -> 404); inbound is RLS
    by vendor_id (a cross-tenant id returns turns:[] under the policy). 404 only when no call AND no
    inbound session matches the id for this tenant."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    # OUTBOUND first: a CALLS row (by id, or by room == call_id), tenant-owned.
    rec = next((c for c in CALLS
                if c.get("id") == call_id or c.get("room") == call_id), None)
    if rec is not None:
        guard = require_object(t, rec, not_found=True)   # cross-tenant -> 404 (don't reveal existence)
        if guard is not None:
            return guard
        turns = _outbound_transcript_turns(rec)
        return JSONResponse({
            "call_id": rec.get("id", "") or call_id,
            "direction": "outbound",
            "phone": rec.get("phone", "") or "",
            "name": rec.get("name", "") or "",
            "turns": turns,
            "total": len(turns),
        })
    # INBOUND: treat call_id as an ai_manager_sessions session_id (RLS-scoped to this tenant).
    tid = t["tenant_id"]
    inbound = await asyncio.to_thread(lambda: _inbound_transcript_turns(tid, call_id))
    if inbound:
        # fetch the header phone for the chat title (best-effort).
        phone = ""
        try:
            from ai_manager import store as _store
            if _store.available():
                full = await asyncio.to_thread(lambda: _store.get_session(tid, call_id))
                phone = (full or {}).get("caller_phone", "") or ""
        except Exception:  # noqa: BLE001
            phone = ""
        return JSONResponse({
            "call_id": call_id,
            "direction": "inbound",
            "phone": phone,
            "name": "",
            "turns": inbound,
            "total": len(inbound),
        })
    # nothing matched for this tenant.
    return JSONResponse({"error": "not found"}, status_code=404)


# ── WORD-ACCURATE timed transcript (synced "Spotify" playback highlight) ──────────
# Re-transcribes the call RECORDING once (cached forever per call) with ElevenLabs
# Scribe — batch STT that returns per-WORD start/end + speaker diarization — so the
# panel can highlight word-by-word in sync with the audio. Lazy (only on request),
# single-flight + cached so it's cheap, and it uses a SEPARATE STT key
# (TRANSCRIPT_STT_API_KEY) when set so the live voice agent's quota is never touched.
# Import-guarded + dormant-safe: any failure returns {timed:false} and the panel
# falls back to its estimate.
try:
    import transcript_timed as _tt_mod  # noqa: E402
except Exception:  # noqa: BLE001
    _tt_mod = None
FEATURE_TRANSCRIPT_TIMED = (cfg_get("FEATURE_TRANSCRIPT_TIMED", "1") or "1").strip().lower() in ("1", "true", "yes", "on")
_TT_DIR = VAR / "transcripts_timed"
_TT_STT_KEY = (cfg_get("TRANSCRIPT_STT_API_KEY", "") or cfg_get("ELEVEN_API_KEY", "") or "").strip()


@app.get("/calls/{call_id}/transcript/timed")
async def call_transcript_timed(request: Request, call_id: str):
    """Word-accurate, audio-aligned transcript for the synced highlight. OUTBOUND calls
    only (inbound falls back to the estimate). Returns
    {timed:true, turns:[{role,text,t0,t1,words:[{w,s,e}]}], duration} or {timed:false}."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rec = next((c for c in CALLS if c.get("id") == call_id or c.get("room") == call_id), None)
    if rec is None:
        return JSONResponse({"timed": False})
    guard = require_object(t, rec, not_found=True)   # cross-tenant -> 404 (also guards the cache)
    if guard is not None:
        return guard
    if not (FEATURE_TRANSCRIPT_TIMED and _tt_mod is not None and _TT_STT_KEY):
        return JSONResponse({"timed": False})
    cid = rec.get("id", "") or call_id
    cache_path = _TT_DIR / f"{cid}.json"
    cached = _read_raw(cache_path, None)
    if isinstance(cached, dict) and cached.get("turns"):
        return JSONResponse({"timed": True, **cached})
    url = (_outbound_rec_item(rec).get("recording_presigned_url", "") or "")
    if not url:
        return JSONResponse({"timed": False})
    try:
        result = await _tt_mod.align(url, api_key=_TT_STT_KEY, duration=rec.get("duration_s"))
    except Exception:  # noqa: BLE001
        result = None
    if not result or not result.get("turns"):
        return JSONResponse({"timed": False})
    try:
        _atomic_write_json(cache_path, result)
    except Exception:  # noqa: BLE001
        pass
    return JSONResponse({"timed": True, **result})


@app.get("/stats")
async def stats(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = calls_for(t)
    # W14-WIRE: OPTIONAL date-range filtering (the panel may send from_/to_ or from/to; YYYY-MM-DD).
    # Absent -> lifetime totals (back-compat with the current bare /stats call). started_at is now
    # tz-labelled UTC; the [:10] / >= comparisons are date-prefix based so they remain correct.
    qp = request.query_params
    _frm = (qp.get("from_") or qp.get("from") or "").strip()
    _to = (qp.get("to_") or qp.get("to") or "").strip()
    if _frm:
        rows = [c for c in rows if (c.get("started_at") or "") >= _frm]
    if _to:
        rows = [c for c in rows if (c.get("started_at") or "") <= (_to + "T23:59:59+00:00")]
    total = len(rows)

    def _is_answered(c):
        if "answered" in c:
            return c.get("answered") is True
        return c.get("duration_s", 0) >= 8   # legacy rows without the flag

    answered = len([c for c in rows if _is_answered(c)])
    in_prog = len([c for c in rows if c.get("status") == "calling"])
    voicemail = len([c for c in rows if c.get("outcome") == "voicemail"])
    no_answer = len([c for c in rows if c.get("outcome") == "no_answer"])
    # last-14-days series by date
    from collections import Counter
    cnt = Counter((c.get("started_at", "")[:10]) for c in rows if c.get("started_at"))
    series = [{"name": d[5:], "amt": n} for d, n in sorted(cnt.items())[-14:]]
    return JSONResponse({"total": total, "answered": answered, "in_progress": in_prog,
                         "voicemail": voicemail, "no_answer": no_answer,
                         "campaigns": len(list_campaigns(t)), "series": series})


# ---------- P1.B analytics endpoint ----------
@app.get("/analytics")
async def analytics(request: Request, campaign_id: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    from_ = request.query_params.get("from", "")
    to_ = request.query_params.get("to", "")
    rows = calls_for(t)
    if campaign_id:
        rows = [c for c in rows if c.get("campaign_id") == campaign_id]
    # optional date range filter (YYYY-MM-DD)
    if from_:
        rows = [c for c in rows if (c.get("started_at") or "") >= from_]
    if to_:
        rows = [c for c in rows if (c.get("started_at") or "") <= (to_ + "T23:59:59")]
    dialed = len(rows)
    connected = len([c for c in rows if c.get("status") not in ("queued", "suppressed")])
    answered = len([c for c in rows if c.get("answered") is True or
                    (c.get("answered") is None and c.get("duration_s", 0) >= 8)])
    interested = len([c for c in rows if c.get("outcome") == "interested"])
    callback_cnt = len([c for c in rows if c.get("outcome") == "callback"])
    qualified = len([c for c in rows if (c.get("interest") or 0) >= 70])
    opted_out = len([c for c in rows if c.get("outcome") == "opt_out"])
    voicemail = len([c for c in rows if c.get("outcome") == "voicemail"])
    no_answer = len([c for c in rows if c.get("outcome") == "no_answer"])
    funnel = [
        {"stage": "dialed", "count": dialed},
        {"stage": "connected", "count": connected},
        {"stage": "answered", "count": answered},
        {"stage": "interested", "count": interested},
        {"stage": "callback", "count": callback_cnt},
        {"stage": "qualified", "count": qualified},
    ]
    return JSONResponse({
        "dialed": dialed, "connected": connected, "answered": answered,
        "interested": interested, "callback": callback_cnt, "qualified": qualified,
        "opted_out": opted_out, "voicemail": voicemail, "no_answer": no_answer,
        "funnel": funnel,
    })


# ---------- tenants (admin only) ----------
@app.get("/tenants")
async def list_tenants(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not t.get("is_admin"):
        return JSONResponse({"error": "admin only"}, status_code=403)
    out = [{"tenant_id": x["tenant_id"], "email": x.get("email", ""), "name": x.get("name", ""),
            "is_admin": bool(x.get("is_admin")), "role": _role_of(x),
            "created_at": x.get("created_at", "")}
           for x in _read_tenants()]
    return JSONResponse({"tenants": out})


@app.post("/tenants")
async def create_tenant(request: Request, email: str = Form(""), password: str = Form(""),
                        name: str = Form(""), role: str = Form("manager")):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not t.get("is_admin"):
        return JSONResponse({"error": "admin only"}, status_code=403)
    email = (email or "").strip().lower()
    if not email or not password:
        return JSONResponse({"error": "email and password are required"}, status_code=400)
    role = (role or "manager").strip().lower()
    if role not in ROLES:
        role = "manager"
    tenants = _read_tenants()
    if any((x.get("email") or "").lower() == email for x in tenants):
        return JSONResponse({"error": "email already exists"}, status_code=409)
    salt = secrets.token_hex(8)
    rec = {"tenant_id": uuid.uuid4().hex[:12], "email": email, "salt": salt,
           "pass_hash": _hash_pw(password, salt), "name": (name or email.split("@")[0]).strip(),
           "is_admin": (role == "admin"), "role": role,
           "created_at": datetime.now().isoformat(timespec="seconds")}
    tenants.append(rec)
    _write_tenants(tenants)
    _audit(request, t, "tenant.create", "tenant", rec["tenant_id"],
           meta={"email": rec["email"], "role": rec["role"]})
    return JSONResponse({"tenant_id": rec["tenant_id"], "email": rec["email"],
                         "name": rec["name"], "role": rec["role"]})


# ---------- Client management (Super Admin) — file-based, no Postgres needed ----------
def _client_mgmt_info(t: dict) -> dict:
    """Full management view of a client (tenant) for the Super Admin Clients page."""
    info = {"tenant_id": t["tenant_id"], "email": t.get("email", ""),
            "name": t.get("name", ""), "role": _role_of(t),
            "is_admin": bool(t.get("is_admin")), "status": (t.get("status") or "active"),
            "created_at": t.get("created_at", ""),
            "restricted": list(t.get("restricted") or []), "demo": bool(t.get("demo"))}
    if t.get("demo"):
        info["demo_minutes"] = int(t.get("demo_minutes") or 0)
        info["demo_started_at"] = t.get("demo_started_at") or t.get("created_at") or ""
        rem = int(_demo_remaining_s(t) or 0)
        info["demo_remaining_s"] = rem
        info["demo_expired"] = rem <= 0
    return info


def _parse_restricted(raw) -> list[str]:
    """Accept a JSON array OR a comma-separated string of restricted feature keys."""
    if not raw:
        return []
    try:
        v = json.loads(raw)
        if isinstance(v, list):
            return [str(x) for x in v if str(x).strip()]
    except (ValueError, TypeError):
        pass
    return [s.strip() for s in str(raw).split(",") if s.strip()]


def _require_admin(request: Request):
    """Returns (tenant, None) for an admin, else (None, error_response)."""
    t = resolve_tenant(request)
    if not t:
        return None, need_auth()
    if not t.get("is_admin"):
        return None, JSONResponse({"error": "admin only"}, status_code=403)
    return t, None


@app.get("/admin/clients")
async def admin_list_clients(request: Request):
    t, err = _require_admin(request)
    if err:
        return err
    clients = [_client_mgmt_info(x) for x in _read_tenants() if not x.get("is_admin")]
    return JSONResponse({"clients": clients, "total": len(clients)})


@app.post("/admin/clients")
async def admin_create_client(request: Request, email: str = Form(""), password: str = Form(""),
                              name: str = Form(""), role: str = Form("manager"),
                              demo: str = Form(""), demo_minutes: str = Form("0"),
                              restricted: str = Form("")):
    t, err = _require_admin(request)
    if err:
        return err
    email = (email or "").strip().lower()
    if not email or not password:
        return JSONResponse({"error": "email and password are required"}, status_code=400)
    role = (role or "manager").strip().lower()
    if role not in ROLES or role == "admin":
        role = "manager"          # the admin account is seeded; clients are never admins
    tenants = _read_tenants()
    if any((x.get("email") or "").lower() == email for x in tenants):
        return JSONResponse({"error": "email already exists"}, status_code=409)
    is_demo = str(demo).strip().lower() in ("1", "true", "yes", "on")
    try:
        dmins = max(0, int(demo_minutes or 0))
    except (TypeError, ValueError):
        dmins = 0
    salt = secrets.token_hex(8)
    now = datetime.now().isoformat(timespec="seconds")
    rec = {"tenant_id": uuid.uuid4().hex[:12], "email": email, "salt": salt,
           "pass_hash": _hash_pw(password, salt),
           "name": (name or email.split("@")[0]).strip(),
           "is_admin": False, "role": role, "created_at": now,
           "status": "active", "restricted": _parse_restricted(restricted), "demo": is_demo}
    if is_demo:
        rec["demo_minutes"] = dmins
        rec["demo_started_at"] = now
    tenants.append(rec)
    _write_tenants(tenants)
    _audit(request, t, "client.create", "tenant", rec["tenant_id"],
           meta={"email": email, "role": role, "demo": is_demo, "demo_minutes": dmins})
    return JSONResponse({"client": _client_mgmt_info(rec)})


@app.put("/admin/clients/{tid}")
async def admin_edit_client(request: Request, tid: str, name: str = Form(None),
                            email: str = Form(None), role: str = Form(None),
                            status: str = Form(None), demo: str = Form(None),
                            demo_minutes: str = Form(None), demo_reset: str = Form(None),
                            restricted: str = Form(None)):
    t, err = _require_admin(request)
    if err:
        return err
    tenants = _read_tenants()
    rec = next((x for x in tenants if x.get("tenant_id") == tid), None)
    if rec is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    if rec.get("is_admin"):
        return JSONResponse({"error": "cannot edit the admin account here"}, status_code=400)
    if name is not None and name.strip():
        rec["name"] = name.strip()
    if email is not None and email.strip():
        new_email = email.strip().lower()
        if any((x.get("email") or "").lower() == new_email for x in tenants if x.get("tenant_id") != tid):
            return JSONResponse({"error": "email already exists"}, status_code=409)
        rec["email"] = new_email
    if role is not None:
        r = role.strip().lower()
        if r in ROLES and r != "admin":
            rec["role"] = r
            rec["is_admin"] = False
    if status is not None and status.strip().lower() in ("active", "suspended"):
        rec["status"] = status.strip().lower()
    if restricted is not None:
        rec["restricted"] = _parse_restricted(restricted)
    if demo is not None:
        is_demo = str(demo).strip().lower() in ("1", "true", "yes", "on")
        rec["demo"] = is_demo
        if is_demo and not rec.get("demo_started_at"):
            rec["demo_started_at"] = datetime.now().isoformat(timespec="seconds")
    if demo_minutes is not None:
        try:
            rec["demo_minutes"] = max(0, int(demo_minutes or 0))
        except (TypeError, ValueError):
            pass
    if str(demo_reset or "").strip().lower() in ("1", "true", "yes", "on"):
        rec["demo_started_at"] = datetime.now().isoformat(timespec="seconds")
    _write_tenants(tenants)
    _audit(request, t, "client.edit", "tenant", tid, meta={"status": rec.get("status")})
    return JSONResponse({"client": _client_mgmt_info(rec)})


@app.post("/admin/clients/{tid}/password")
async def admin_reset_client_password(request: Request, tid: str, password: str = Form("")):
    t, err = _require_admin(request)
    if err:
        return err
    if not password or len(password) < 4:
        return JSONResponse({"error": "password must be at least 4 characters"}, status_code=400)
    tenants = _read_tenants()
    rec = next((x for x in tenants if x.get("tenant_id") == tid), None)
    if rec is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    salt = secrets.token_hex(8)
    rec["salt"] = salt
    rec["pass_hash"] = _hash_pw(password, salt)
    _write_tenants(tenants)
    _audit(request, t, "client.reset_password", "tenant", tid)
    return JSONResponse({"ok": True})


@app.post("/admin/clients/{tid}/status")
async def admin_set_client_status(request: Request, tid: str, status: str = Form("")):
    t, err = _require_admin(request)
    if err:
        return err
    s = (status or "").strip().lower()
    if s not in ("active", "suspended"):
        return JSONResponse({"error": "status must be active or suspended"}, status_code=400)
    tenants = _read_tenants()
    rec = next((x for x in tenants if x.get("tenant_id") == tid), None)
    if rec is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    if rec.get("is_admin"):
        return JSONResponse({"error": "cannot suspend the admin account"}, status_code=400)
    rec["status"] = s
    _write_tenants(tenants)
    try:
        if s == "suspended" and _auth_mod is not None:
            _auth_mod.revoke_all(tid)     # instant kill for any JWT sessions
    except Exception:  # noqa: BLE001
        pass
    _audit(request, t, "client.status", "tenant", tid, meta={"status": s})
    return JSONResponse({"ok": True, "status": s})


@app.delete("/admin/clients/{tid}")
async def admin_delete_client(request: Request, tid: str):
    t, err = _require_admin(request)
    if err:
        return err
    tenants = _read_tenants()
    rec = next((x for x in tenants if x.get("tenant_id") == tid), None)
    if rec is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    if rec.get("is_admin"):
        return JSONResponse({"error": "cannot delete the admin account"}, status_code=400)
    _write_tenants([x for x in tenants if x.get("tenant_id") != tid])
    purged: dict = {"tenant": 1}
    # purge every tenant-scoped file store (rows carry tenant_id)
    async with _STORE_LOCK:
        for label, fp in (("leads", LEADS_FILE), ("calls", CALLS_FILE),
                          ("suppression", SUPPRESSION_FILE), ("billing", BILLING_FILE)):
            try:
                rows = _read(fp, [])
                if isinstance(rows, list):
                    kept = [x for x in rows if not (isinstance(x, dict) and x.get("tenant_id") == tid)]
                    purged[label] = len(rows) - len(kept)
                    _write(fp, kept)
            except Exception:  # noqa: BLE001
                pass
    try:
        CALLS[:] = [c for c in CALLS if c.get("tenant_id") != tid]   # in-memory mirror
    except Exception:  # noqa: BLE001
        pass
    try:
        cn = 0
        for p in CAMPAIGN_DIR.glob("*.json"):
            d = _read(p, None)
            if isinstance(d, dict) and d.get("tenant_id") == tid:
                p.unlink()
                cn += 1
        purged["campaigns"] = cn
    except Exception:  # noqa: BLE001
        pass
    try:
        navp = VAR / "nav_configs" / f"{tid}.json"
        if navp.exists():
            navp.unlink()
            purged["nav_config"] = 1
    except Exception:  # noqa: BLE001
        pass
    _audit(request, t, "client.delete", "tenant", tid, meta=purged)
    return JSONResponse({"ok": True, "purged": purged})


# ---------- Public signup: email + 4-digit OTP (via the Axcrio business SMTP) ----------
_SIGNUP_OTP: dict = {}        # email -> {otp, exp, name, salt, pass_hash, tries}
_SIGNUP_OTP_TTL = 600         # 10 minutes


def _signup_default_role() -> str:
    """Admin-configurable default role for self-signups. Default 'agent' (limited credits)."""
    try:
        d = _read(VAR / "signup_settings.json", {})
        r = str((d or {}).get("default_role", "agent")).strip().lower()
        return r if r in ("agent", "manager") else "agent"
    except Exception:  # noqa: BLE001
        return "agent"


def _send_otp_email(to_email: str, otp: str) -> bool:
    """Deliver the 4-digit code. Tries an HTTPS email API (Resend) FIRST — DigitalOcean
    blocks outbound SMTP on droplets, so direct SMTP from this box times out. Falls back to
    Hostinger SMTP (works only if DO has unblocked SMTP for the account). Never raises."""
    _subj = "Your Haptica AI verification code"
    _text = (f"Your Haptica AI verification code is: {otp}\n\nIt expires in 10 minutes. "
             "If you didn't request this, ignore this email.\n\n— Haptica AI (by Famit)")
    _html = (
        '<div style="font-family:system-ui,Segoe UI,Arial,sans-serif;max-width:440px;margin:auto;padding:28px;color:#0b0b0f">'
        '<div style="font-size:13px;letter-spacing:.08em;text-transform:uppercase;color:#8a8a8a">Haptica AI · by Famit</div>'
        '<h2 style="margin:14px 0 6px;font-weight:600">Verify your email</h2>'
        '<p style="color:#555;margin:0 0 20px">Enter this code to finish creating your account:</p>'
        f'<div style="font-size:34px;font-weight:700;letter-spacing:14px;background:#f4f5f7;border-radius:14px;padding:18px 0;text-align:center">{otp}</div>'
        '<p style="color:#999;font-size:13px;margin-top:18px">This code expires in 10 minutes. If you didn\'t request it, ignore this email.</p></div>')
    # 1) HTTPS email API (Resend) — runs over 443, which DO allows.
    _rk = (cfg_get("RESEND_API_KEY", "") or "").strip()
    if _rk:
        try:
            _frm = (cfg_get("EMAIL_FROM", "") or cfg_get("SMTP_FROM", "") or "onboarding@resend.dev").strip()
            _r = httpx.post(
                "https://api.resend.com/emails",
                headers={"Authorization": "Bearer " + _rk, "Content-Type": "application/json"},
                json={"from": (f"Haptica AI <{_frm}>" if "<" not in _frm else _frm),
                      "to": [to_email], "subject": _subj, "html": _html, "text": _text},
                timeout=20)
            if _r.status_code in (200, 201):
                return True
            import logging as _lg0
            _lg0.getLogger("famit-caller").warning("Resend OTP %s: %s", _r.status_code, _r.text[:200])
        except Exception as _re:  # noqa: BLE001
            try:
                import logging as _lg0
                _lg0.getLogger("famit-caller").warning("Resend OTP error: %r", _re)
            except Exception:  # noqa: BLE001
                pass
    # 2) SMTP fallback (only works if DO unblocked outbound SMTP for the account).
    import smtplib
    import ssl as _ssl
    from email.message import EmailMessage as _EmailMessage
    host = (cfg_get("SMTP_HOST", "smtp.hostinger.com") or "").strip()
    port = int(cfg_get("SMTP_PORT", "465") or 465)
    user = (cfg_get("SMTP_USER", "") or "").strip()
    pw = (cfg_get("SMTP_PASS", "") or "").strip()
    frm = (cfg_get("SMTP_FROM", "") or user).strip()
    if not (host and user and pw):
        return False
    msg = _EmailMessage()
    msg["Subject"] = "Your Haptica AI verification code"
    msg["From"] = f"Haptica AI <{frm}>" if "<" not in frm else frm
    msg["To"] = to_email
    msg.set_content(
        f"Your Haptica AI verification code is: {otp}\n\n"
        "It expires in 10 minutes. If you didn't request this, you can ignore this email.\n\n— Haptica AI (by Famit)")
    msg.add_alternative(
        f"""<div style="font-family:system-ui,Segoe UI,Arial,sans-serif;max-width:440px;margin:auto;padding:28px;color:#0b0b0f">
  <div style="font-size:13px;letter-spacing:.08em;text-transform:uppercase;color:#8a8a8a">Haptica AI · by Famit</div>
  <h2 style="margin:14px 0 6px;font-weight:600">Verify your email</h2>
  <p style="color:#555;margin:0 0 20px">Enter this code to finish creating your account:</p>
  <div style="font-size:34px;font-weight:700;letter-spacing:14px;background:#f4f5f7;border-radius:14px;padding:18px 0;text-align:center">{otp}</div>
  <p style="color:#999;font-size:13px;margin-top:18px">This code expires in 10 minutes. If you didn't request it, ignore this email.</p>
</div>""", subtype="html")
    try:
        ctx = _ssl.create_default_context()
        with smtplib.SMTP_SSL(host, port, context=ctx, timeout=20) as s:
            s.login(user, pw)
            s.send_message(msg)
        return True
    except Exception as exc:  # noqa: BLE001
        try:
            import logging as _lg
            _lg.getLogger("famit-caller").warning("OTP email send failed: %r", exc)
        except Exception:  # noqa: BLE001
            pass
        return False


@app.post("/signup/start")
async def signup_start(request: Request, email: str = Form(""), password: str = Form(""),
                       name: str = Form("")):
    email = (email or "").strip().lower()
    if not email or "@" not in email or "." not in email.split("@")[-1]:
        return JSONResponse({"error": "Please enter a valid email address."}, status_code=400)
    if len(password or "") < 6:
        return JSONResponse({"error": "Password must be at least 6 characters."}, status_code=400)
    if any((x.get("email") or "").lower() == email for x in _read_tenants()):
        return JSONResponse({"error": "An account with this email already exists. Try logging in."}, status_code=409)
    otp = f"{secrets.randbelow(10000):04d}"
    salt = secrets.token_hex(8)
    _SIGNUP_OTP[email] = {"otp": otp, "exp": time.time() + _SIGNUP_OTP_TTL,
                          "name": (name or email.split("@")[0]).strip(),
                          "salt": salt, "pass_hash": _hash_pw(password, salt), "tries": 0}
    sent = await asyncio.get_event_loop().run_in_executor(None, _send_otp_email, email, otp)
    if not sent:
        return JSONResponse({"error": "Couldn't send the verification email right now. Please try again."}, status_code=502)
    at = email.index("@")
    masked = (email[0] + "•••" + email[max(1, at - 1):]) if at > 1 else email
    return JSONResponse({"ok": True, "sent_to": masked})


@app.post("/signup/verify")
async def signup_verify(request: Request, email: str = Form(""), otp: str = Form("")):
    email = (email or "").strip().lower()
    otp = (otp or "").strip()
    rec = _SIGNUP_OTP.get(email)
    if not rec:
        return JSONResponse({"error": "No pending signup found. Please start again."}, status_code=400)
    if time.time() > rec["exp"]:
        _SIGNUP_OTP.pop(email, None)
        return JSONResponse({"error": "That code has expired. Please request a new one."}, status_code=400)
    rec["tries"] = int(rec.get("tries", 0)) + 1
    if rec["tries"] > 6:
        _SIGNUP_OTP.pop(email, None)
        return JSONResponse({"error": "Too many attempts. Please start again."}, status_code=429)
    if otp != rec["otp"]:
        return JSONResponse({"error": "Incorrect code. Please check and try again."}, status_code=400)
    tenants = _read_tenants()
    if any((x.get("email") or "").lower() == email for x in tenants):
        _SIGNUP_OTP.pop(email, None)
        return JSONResponse({"error": "An account with this email already exists."}, status_code=409)
    role = _signup_default_role()
    now = datetime.now().isoformat(timespec="seconds")
    new = {"tenant_id": uuid.uuid4().hex[:12], "email": email, "salt": rec["salt"],
           "pass_hash": rec["pass_hash"], "name": rec["name"], "is_admin": False,
           "role": role, "status": "active", "created_at": now, "restricted": [],
           "demo": False, "self_signup": True}
    tenants.append(new)
    _write_tenants(tenants)
    _SIGNUP_OTP.pop(email, None)
    _audit(request, new, "signup.complete", "tenant", new["tenant_id"], meta={"role": role})
    return JSONResponse({"ok": True, "token": _make_token(new["tenant_id"]),
                         "tenant_id": new["tenant_id"], "name": new["name"],
                         "is_admin": False, "role": role})


@app.get("/admin/signup-settings")
async def admin_get_signup_settings(request: Request):
    t, err = _require_admin(request)
    if err:
        return err
    return JSONResponse({"default_role": _signup_default_role()})


@app.put("/admin/signup-settings")
async def admin_set_signup_settings(request: Request, default_role: str = Form("")):
    t, err = _require_admin(request)
    if err:
        return err
    r = (default_role or "").strip().lower()
    if r not in ("agent", "manager"):
        return JSONResponse({"error": "default_role must be agent or manager"}, status_code=400)
    _write(VAR / "signup_settings.json", {"default_role": r})
    _audit(request, t, "signup.default_role", "settings", "", meta={"default_role": r})
    return JSONResponse({"ok": True, "default_role": r})


# ════════════════════════════════════════════════════════════════════════════
# ADVANCED MONITORING — per-tenant session/location/device capture (file-based)
# ════════════════════════════════════════════════════════════════════════════
# Self-contained on two JSON files (no Postgres). Captures, per authenticated
# tenant: client IP (via the reverse proxy's forwarded headers), IP-geolocation
# (country/region/city/isp/timezone), parsed device (browser/os/type), and any
# browser-provided signals (precise GPS w/ consent, timezone, locale, screen).
# Powers the user Profile page and the Super-Admin client-monitoring panel.
SESSIONS_FILE = VAR / "sessions.json"        # {tenant_id: {last_session, sessions[], ...}}
GEO_CACHE_FILE = VAR / "geo_cache.json"      # {ip: {..geo.., _ts}}  (24h TTL)
SESSION_HISTORY_CAP = 25                      # rows kept per tenant
_GEO_TTL = 86400                              # 24h IP-geo cache


def _client_ip(request: Request) -> str:
    """Real client IP behind Caddy/Cloudflare. Trusts forwarded headers (the box
    is only reachable through the proxy), falling back to the socket peer."""
    for h in ("cf-connecting-ip", "x-real-ip"):
        v = (request.headers.get(h) or "").strip()
        if v:
            return v
    xff = (request.headers.get("x-forwarded-for") or "").strip()
    if xff:
        return xff.split(",")[0].strip()
    try:
        return request.client.host if request.client else ""
    except Exception:  # noqa: BLE001
        return ""


def _is_private_ip(ip: str) -> bool:
    """True for loopback/private/link-local addresses that can't be geolocated."""
    if not ip:
        return True
    try:
        import ipaddress
        addr = ipaddress.ip_address(ip)
        return addr.is_private or addr.is_loopback or addr.is_link_local or addr.is_reserved
    except ValueError:
        return True


def _parse_ua(ua: str) -> dict:
    """Lightweight User-Agent parse (no external deps): browser, os, device type."""
    ua = ua or ""
    low = ua.lower()
    # OS
    if "windows nt 10" in low or "windows nt 11" in low:
        os_name = "Windows 10/11"
    elif "windows" in low:
        os_name = "Windows"
    elif "iphone" in low:
        os_name = "iOS"
    elif "ipad" in low:
        os_name = "iPadOS"
    elif "mac os x" in low or "macintosh" in low:
        os_name = "macOS"
    elif "android" in low:
        os_name = "Android"
    elif "cros" in low:
        os_name = "ChromeOS"
    elif "linux" in low:
        os_name = "Linux"
    else:
        os_name = "Unknown"
    # Browser (order matters — Edge/Opera spoof Chrome; iOS browsers spoof Safari)
    if "edg/" in low or "edga/" in low or "edgios/" in low:
        browser = "Edge"
    elif "opr/" in low or "opera" in low:
        browser = "Opera"
    elif "samsungbrowser" in low:
        browser = "Samsung Internet"
    elif "crios/" in low:
        browser = "Chrome"
    elif "fxios/" in low or "firefox" in low:
        browser = "Firefox"
    elif "chrome" in low and "chromium" not in low:
        browser = "Chrome"
    elif "safari" in low:
        browser = "Safari"
    else:
        browser = "Unknown"
    # Device type
    if "ipad" in low or "tablet" in low or ("android" in low and "mobile" not in low):
        device = "Tablet"
    elif "mobi" in low or "iphone" in low or "android" in low:
        device = "Mobile"
    else:
        device = "Desktop"
    return {"browser": browser, "os": os_name, "device": device, "ua": ua[:400]}


def _geolocate_ip(ip: str) -> dict:
    """Resolve an IP to {country, country_code, region, city, lat, lon, isp,
    timezone} via free HTTPS geo providers, cached on disk for 24h. Returns {}
    for private IPs or on total failure. BLOCKING — call via run_in_executor."""
    if not ip or _is_private_ip(ip):
        return {}
    cache = _read(GEO_CACHE_FILE, {})
    hit = cache.get(ip)
    if hit and (time.time() - float(hit.get("_ts", 0)) < _GEO_TTL):
        return {k: v for k, v in hit.items() if k != "_ts"}
    geo: dict = {}
    # Provider 1: ipwho.is (HTTPS, no key)
    try:
        r = httpx.get(f"https://ipwho.is/{ip}", timeout=8)
        if r.status_code == 200:
            d = r.json()
            if d.get("success", True):
                tz = d.get("timezone")
                geo = {
                    "country": d.get("country") or "",
                    "country_code": d.get("country_code") or "",
                    "region": d.get("region") or "",
                    "city": d.get("city") or "",
                    "lat": d.get("latitude"),
                    "lon": d.get("longitude"),
                    "isp": ((d.get("connection") or {}) or {}).get("isp") or d.get("isp") or "",
                    "timezone": (tz.get("id") if isinstance(tz, dict) else tz) or "",
                }
    except Exception:  # noqa: BLE001
        geo = {}
    # Provider 2: ip-api.com (HTTP fallback; DO allows outbound HTTP/HTTPS, only SMTP is blocked)
    if not geo.get("country"):
        try:
            r = httpx.get(
                f"http://ip-api.com/json/{ip}"
                "?fields=status,country,countryCode,regionName,city,lat,lon,isp,timezone",
                timeout=8)
            if r.status_code == 200:
                d = r.json()
                if d.get("status") == "success":
                    geo = {
                        "country": d.get("country") or "", "country_code": d.get("countryCode") or "",
                        "region": d.get("regionName") or "", "city": d.get("city") or "",
                        "lat": d.get("lat"), "lon": d.get("lon"),
                        "isp": d.get("isp") or "", "timezone": d.get("timezone") or "",
                    }
        except Exception:  # noqa: BLE001
            pass
    if geo.get("country"):
        cache[ip] = {**geo, "_ts": time.time()}
        if len(cache) > 800:                      # keep the cache bounded (drop oldest)
            items = sorted(cache.items(), key=lambda kv: float(kv[1].get("_ts", 0)))
            cache = dict(items[-800:])
        try:
            _write(GEO_CACHE_FILE, cache)
        except Exception:  # noqa: BLE001
            pass
    return geo


def _build_session(ip: str, ua: str, cp: dict | None) -> dict:
    """Assemble one session row from IP + UA + browser-provided client payload."""
    dev = _parse_ua(ua)
    geo = _geolocate_ip(ip)
    cp = cp or {}

    def _num(v):
        try:
            return float(v) if v is not None and v != "" else None
        except (TypeError, ValueError):
            return None

    return {
        "ts": _utc_iso(),
        "ip": ip,
        "browser": dev["browser"], "os": dev["os"], "device": dev["device"], "ua": dev["ua"],
        "country": geo.get("country", ""), "country_code": geo.get("country_code", ""),
        "region": geo.get("region", ""), "city": geo.get("city", ""),
        "lat": geo.get("lat"), "lon": geo.get("lon"),
        "isp": geo.get("isp", ""), "ip_timezone": geo.get("timezone", ""),
        # browser-provided (advanced capture; precise GPS only with user consent)
        "tz": str(cp.get("tz") or "")[:64], "locale": str(cp.get("locale") or "")[:32],
        "screen": str(cp.get("screen") or "")[:24], "platform": str(cp.get("platform") or "")[:48],
        "geo_lat": _num(cp.get("geo_lat")), "geo_lon": _num(cp.get("geo_lon")),
        "geo_acc": _num(cp.get("geo_acc")),
    }


def _record_session(tenant_id: str, ip: str, ua: str, cp: dict | None) -> dict:
    """Persist a session for a tenant. De-dupes rapid reloads: if the newest row
    has the same IP+device within 5 minutes, it's refreshed in place (and any new
    precise GPS merged) instead of appending a near-duplicate. BLOCKING."""
    sess = _build_session(ip, ua, cp)
    store = _read(SESSIONS_FILE, {})
    if not isinstance(store, dict):
        store = {}
    rec = store.get(tenant_id) or {}
    hist = list(rec.get("sessions") or [])
    last = hist[0] if hist else None
    dedupe = False
    if last and last.get("ip") == sess["ip"] and last.get("ua") == sess["ua"]:
        try:
            prev = datetime.fromisoformat(last.get("ts"))
            cur = datetime.fromisoformat(sess["ts"])
            dedupe = abs((cur - prev).total_seconds()) < 300
        except (ValueError, TypeError):
            dedupe = False
    if dedupe:
        # refresh timestamp + carry forward precise GPS if newly provided
        last["ts"] = sess["ts"]
        for k in ("geo_lat", "geo_lon", "geo_acc", "tz", "locale", "screen", "platform"):
            if sess.get(k) not in (None, ""):
                last[k] = sess[k]
        hist[0] = last
        sess = last
    else:
        hist.insert(0, sess)
    rec["sessions"] = hist[:SESSION_HISTORY_CAP]
    rec["last_session"] = sess
    rec["first_seen"] = rec.get("first_seen") or sess["ts"]
    if not dedupe:
        rec["sessions_count"] = int(rec.get("sessions_count") or 0) + 1
    store[tenant_id] = rec
    try:
        _write(SESSIONS_FILE, store)
    except Exception:  # noqa: BLE001
        pass
    return sess


def _flag_emoji(cc: str) -> str:
    """ISO-2 country code -> 🇮🇳 flag emoji (empty for unknown)."""
    cc = (cc or "").strip().upper()
    if len(cc) != 2 or not cc.isalpha():
        return ""
    return "".join(chr(0x1F1E6 + ord(c) - ord("A")) for c in cc)


def _public_session(s: dict | None) -> dict:
    """Session row enriched with a derived 'location' label + flag for the UI."""
    if not s:
        return {}
    loc_bits = [b for b in (s.get("city"), s.get("region"), s.get("country")) if b]
    # de-dup consecutive (city == region happens for city-states)
    seen, loc = set(), []
    for b in loc_bits:
        if b not in seen:
            seen.add(b)
            loc.append(b)
    out = dict(s)
    out["location"] = ", ".join(loc)
    out["flag"] = _flag_emoji(s.get("country_code", ""))
    return out


@app.post("/session/beacon")
async def session_beacon(request: Request):
    """Record the caller's current session (IP geo + device + optional browser
    signals). Called by the panel on each authenticated app load."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    cp = body if isinstance(body, dict) else {}
    ip = _client_ip(request)
    ua = request.headers.get("user-agent", "")
    sess = await asyncio.get_event_loop().run_in_executor(
        None, _record_session, t["tenant_id"], ip, ua, cp)
    return JSONResponse({"ok": True, "session": _public_session(sess)})


@app.get("/profile")
async def profile_get(request: Request):
    """The caller's own profile + their location/device monitoring summary."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    store = _read(SESSIONS_FILE, {})
    rec = (store.get(t["tenant_id"]) if isinstance(store, dict) else {}) or {}
    out = {
        "tenant_id": t["tenant_id"], "email": t.get("email", ""), "name": t.get("name", ""),
        "role": _role_of(t), "is_admin": bool(t.get("is_admin")),
        "status": (t.get("status") or "active"), "created_at": t.get("created_at", ""),
        "self_signup": bool(t.get("self_signup")), "demo": bool(t.get("demo")),
        "first_seen": rec.get("first_seen", ""),
        "sessions_count": int(rec.get("sessions_count") or 0),
        "last_session": _public_session(rec.get("last_session") or {}),
        "recent_sessions": [_public_session(s) for s in (rec.get("sessions") or [])[:10]],
    }
    if t.get("demo"):
        out["demo_minutes"] = int(t.get("demo_minutes") or 0)
        out["demo_remaining_s"] = int(_demo_remaining_s(t) or 0)
    return JSONResponse(out)


@app.get("/admin/clients/{tid}/profile")
async def admin_client_profile(request: Request, tid: str):
    """Super-Admin: full profile + session/location/device monitoring for a client."""
    t, err = _require_admin(request)
    if err:
        return err
    rec_t = _tenant_by_id(tid)
    if rec_t is None:
        return JSONResponse({"error": "not found"}, status_code=404)
    store = _read(SESSIONS_FILE, {})
    rec = (store.get(tid) if isinstance(store, dict) else {}) or {}
    info = _client_mgmt_info(rec_t)
    info.update({
        "self_signup": bool(rec_t.get("self_signup")),
        "first_seen": rec.get("first_seen", ""),
        "sessions_count": int(rec.get("sessions_count") or 0),
        "last_session": _public_session(rec.get("last_session") or {}),
        "sessions": [_public_session(s) for s in (rec.get("sessions") or [])[:SESSION_HISTORY_CAP]],
    })
    return JSONResponse({"profile": info})


# ---------- P0.2 suppression endpoints ----------
@app.get("/suppression")
async def get_suppression(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = [x for x in _read(SUPPRESSION_FILE, []) if x.get("tenant_id") == t["tenant_id"]]
    out = [{"phone": x["phone"], "reason": x.get("reason", ""), "source": x.get("source", ""),
            "added_at": x.get("added_at", "")} for x in rows]
    return JSONResponse({"numbers": out, "total": len(out)})


@app.post("/suppression")
async def add_suppression(request: Request, numbers: str = Form(""),
                          csv: UploadFile | None = File(None)):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot edit suppression list")
    csv_bytes = await csv.read() if csv is not None else None
    parsed = parse_leads(numbers, csv_bytes)
    added = 0
    before = _suppressed_set(t["tenant_id"])
    for x in parsed:
        if x["num"] not in before:
            await _add_suppression(t["tenant_id"], x["num"], "upload")
            before.add(x["num"]); added += 1
    total = len(_suppressed_set(t["tenant_id"]))
    _audit(request, t, "suppression.add", "suppression", "", meta={"added": added})
    return JSONResponse({"added": added, "total": total})


@app.delete("/suppression/{phone}")
async def delete_suppression(request: Request, phone: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot edit suppression list")
    p = norm(phone)
    async with _STORE_LOCK:
        store = _read(SUPPRESSION_FILE, [])
        new = [x for x in store if not (x.get("tenant_id") == t["tenant_id"] and x.get("phone") == p)]
        _write(SUPPRESSION_FILE, new)
    _audit(request, t, "suppression.delete", "suppression", p)
    return JSONResponse({"deleted": p})


@app.post("/optout")
async def optout(request: Request, phone: str = Form(""), campaign_id: str = Form(""),
                 source: str = Form("manual")):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot opt out numbers")
    reason = "opt_out_call" if source == "call" else ("manual" if source == "manual" else "api")
    await _add_suppression(t["tenant_id"], phone, reason, source=campaign_id or source)
    await _flip_lead_status(t["tenant_id"], phone, "opted_out")
    await _emit_webhook(t["tenant_id"], "lead.opted_out",
                        {"phone": norm(phone), "campaign_id": campaign_id})
    _audit(request, t, "lead.optout", "lead", norm(phone),
           channel=("call" if source == "call" else "api"),
           meta={"campaign_id": campaign_id, "source": source})
    return JSONResponse({"ok": True})


# ---------- P0.5 callbacks / retry queue endpoints ----------
@app.get("/callbacks")
async def get_callbacks(request: Request, all: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = [r for r in _read(RETRY_FILE, [])
            if t.get("is_admin") or r.get("tenant_id") == t["tenant_id"]]
    if not all:
        rows = [r for r in rows if r.get("reason") == "callback"]
    rows.sort(key=lambda r: r.get("next_attempt_at", ""))
    return JSONResponse({"items": rows})


@app.delete("/callbacks/{rid}")
async def cancel_callback(request: Request, rid: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot cancel callbacks")
    target = next((r for r in _read(RETRY_FILE, []) if r.get("id") == rid), None)
    guard = require_object(t, target)  # 404 if missing or not owned (BOLA)
    if guard is not None:
        return guard
    await _remove_retry(rid)
    _audit(request, t, "callback.cancel", "callback", rid)
    return JSONResponse({"cancelled": rid})


@app.post("/callbacks")
async def add_callback(request: Request, phone: str = Form(""), campaign_id: str = Form(""),
                       when: str = Form(""), name: str = Form("")):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot add callbacks")
    if not get_campaign_for(campaign_id, t):
        return JSONResponse({"error": "campaign not found for this account"}, status_code=404)
    if not norm(phone):
        return JSONResponse({"error": "valid phone required"}, status_code=400)
    await _enqueue_retry(t["tenant_id"], campaign_id, name, phone, 0, 3,
                         when or now_ist().isoformat(), "callback")
    _audit(request, t, "callback.add", "callback", norm(phone),
           meta={"campaign_id": campaign_id, "when": when})
    return JSONResponse({"ok": True})


# ---------- P0.7 usage endpoints ----------
@app.get("/usage")
async def usage(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    tid = t["tenant_id"]
    tr = _tenant_by_id(tid) or {}
    return JSONResponse({
        "today": _tenant_usage(tid, _today_iso()),
        "month": _tenant_usage(tid, _month_iso()),
        "limits": {"max_concurrency": int(tr.get("max_concurrency", 3)),
                   "daily_call_cap": int(tr.get("daily_call_cap", 500)),
                   "monthly_minutes_cap": int(tr.get("monthly_minutes_cap", 5000))},
        "active_now": ACTIVE_CALLS.get(tid, 0),
    })


@app.get("/usage/all")
async def usage_all(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not t.get("is_admin"):
        return JSONResponse({"error": "admin only"}, status_code=403)
    out = []
    for x in _read_tenants():
        tid = x["tenant_id"]
        out.append({"tenant_id": tid, "name": x.get("name", ""), "email": x.get("email", ""),
                    "today": _tenant_usage(tid, _today_iso()),
                    "month": _tenant_usage(tid, _month_iso()),
                    "active_now": ACTIVE_CALLS.get(tid, 0),
                    "limits": {"max_concurrency": int(x.get("max_concurrency", 3)),
                               "daily_call_cap": int(x.get("daily_call_cap", 500)),
                               "monthly_minutes_cap": int(x.get("monthly_minutes_cap", 5000))}})
    return JSONResponse({"tenants": out})


@app.post("/tenants/{tid}/limits")
async def set_tenant_limits(request: Request, tid: str, max_concurrency: int = Form(None),
                            daily_call_cap: int = Form(None), monthly_minutes_cap: int = Form(None)):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not t.get("is_admin"):
        return JSONResponse({"error": "admin only"}, status_code=403)
    async with _STORE_LOCK:
        tenants = _read_tenants()
        target = next((x for x in tenants if x.get("tenant_id") == tid), None)
        if not target:
            return JSONResponse({"error": "not found"}, status_code=404)
        if max_concurrency is not None:
            target["max_concurrency"] = max(1, int(max_concurrency))
        if daily_call_cap is not None:
            target["daily_call_cap"] = max(0, int(daily_call_cap))
        if monthly_minutes_cap is not None:
            target["monthly_minutes_cap"] = max(0, int(monthly_minutes_cap))
        _write_tenants(tenants)
    return JSONResponse({"ok": True, "tenant_id": tid,
                         "max_concurrency": target.get("max_concurrency"),
                         "daily_call_cap": target.get("daily_call_cap"),
                         "monthly_minutes_cap": target.get("monthly_minutes_cap")})


# ════════════════════════════════════════════════════════════════════════════
# CONTROL LAYER — /admin/* super-admin API + /me/entitlements + act-as (CL-B3 / plan C2+C4+C11)
# ════════════════════════════════════════════════════════════════════════════
# EVERY /admin/* route calls require_super_admin (is_admin AND non-legacy-pw — control-security §1.2);
# the TARGET tenant is the path {id}, NEVER a body field (security invariant #3); every write is audited
# to the IMMUTABLE PG `events` leg via _audit(channel="control") with before/after, AND mirrored into the
# entitlement_audit read-copy by the engine. Money/destructive writes (credits, status->disabled,
# impersonate) additionally clear the firewall step-up. Behind CONTROL_ENABLED only for ENFORCEMENT (the
# middleware) — the /admin/* routes themselves are always role-gated (an admin manages entitlements even
# while the master flag is off, so the founder can configure BEFORE flipping enforcement on).

def _control_unavailable() -> JSONResponse:
    return JSONResponse({"error": "control layer unavailable"}, status_code=503)


def _control_audit(request: Request, admin: dict, action: str, *, target_tenant: str | None = None,
                   feature_key: str | None = None, old_value=None, new_value=None,
                   reason: str = "") -> None:
    """Audit a super-admin control action to BOTH the immutable PG `events` leg (channel='control',
    the source of truth) AND the entitlement_audit read-mirror. before/after are MANDATORY (§4.2)."""
    real_admin = (admin or {}).get("real_admin") or (admin or {}).get("tenant_id", "")
    actor = (admin or {}).get("tenant_id", "")
    meta = {"target_tenant": target_tenant, "feature_key": feature_key,
            "old_value": (str(old_value) if old_value is not None else None),
            "new_value": (str(new_value) if new_value is not None else None),
            "reason": reason, "real_admin": real_admin,
            "act_as": (admin or {}).get("act_as"), "auth_method": _auth_method(request)}
    _audit(request, admin, action, "control", target_tenant or feature_key or "",
           channel="control", meta=meta)
    if _ent_mod is not None:
        try:
            _ent_mod._mirror_audit(
                real_admin, actor, action, target_tenant=target_tenant, feature_key=feature_key,
                old_value=meta["old_value"], new_value=meta["new_value"],
                reason=reason, ip=_client_ip(request))
        except Exception:  # noqa: BLE001
            pass


def _vendor_health(tid: str) -> dict:
    """Executive 'last activity' health for a vendor (cheap, reuses existing in-RAM/JSON stores)."""
    try:
        last_call = ""
        for c in reversed(CALLS):
            if c.get("tenant_id") == tid:
                last_call = c.get("at") or c.get("started_at") or ""
                break
        return {"active_now": ACTIVE_CALLS.get(tid, 0), "last_call": last_call}
    except Exception:  # noqa: BLE001
        return {"active_now": 0, "last_call": ""}


# ---------- registry / global flags ----------
@app.get("/admin/features")
async def admin_features(request: Request):
    """The full feature_registry catalog tree (for the Feature-Flags + per-vendor UI)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    return JSONResponse({"features": _ent_mod.registry_tree()})


@app.get("/admin/flags")
async def admin_flags_get(request: Request):
    """Global default_mode per feature (the baseline for every vendor)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    flags = {r["key"]: r.get("default_mode", "on") for r in _ent_mod.registry_tree()}
    return JSONResponse({"flags": flags})


@app.put("/admin/flags/{feature_key}")
async def admin_flags_set(request: Request, feature_key: str, mode: str = Form(...)):
    """Set the GLOBAL baseline mode for a feature (affects ALL vendors -> drops every resolve cache)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    res = _ent_mod.set_global_flag(feature_key, (mode or "").strip(), set_by=t["tenant_id"])
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    _control_audit(request, t, "control.flag.set", feature_key=feature_key,
                   old_value=res.get("before"), new_value=res.get("after"))
    return JSONResponse({"ok": True, "feature_key": feature_key, **res})


# ---------- per-tenant SIDEBAR / NAV config (Super-Admin Sidebar Builder) ----------
# Lets a super-admin control, for ANY tenant, what sidebar items show, in what order,
# and under what label. Stored per-tenant as {order:[], hidden:[], labels:{}, childOrder:{}}
# keyed by a stable nav key (href, or "group:<title>"). PURELY COSMETIC + additive: the
# client applies it on TOP of the static nav + entitlements; the backend 404/402 choke-
# point is still the real boundary. Empty config = the default nav (no behaviour change).
NAV_CONFIG_DIR = VAR / "nav_configs"


def _nav_config_path(tenant_id: str) -> Path:
    safe = "".join(ch for ch in (tenant_id or "") if ch.isalnum() or ch in "-_")
    return NAV_CONFIG_DIR / f"{safe}.json"


def _read_nav_config(tenant_id: str) -> dict:
    if not (tenant_id or "").strip():
        return {}
    return _read(_nav_config_path(tenant_id), {}) or {}


def _sanitize_nav_config(cfg: dict) -> dict:
    """Keep only the known keys with bounded sizes (never trust the client blob).
    Supports: order/hidden/labels/childOrder + parentOf (move a child to another
    category) + custom (admin-created links/sections)."""
    cfg = cfg if isinstance(cfg, dict) else {}
    order = [str(x) for x in (cfg.get("order") or []) if isinstance(x, (str, int))][:300]
    hidden = [str(x) for x in (cfg.get("hidden") or []) if isinstance(x, (str, int))][:300]
    labels_raw = cfg.get("labels") if isinstance(cfg.get("labels"), dict) else {}
    labels = {str(k): str(v)[:60] for k, v in labels_raw.items() if str(v).strip()}
    co_raw = cfg.get("childOrder") if isinstance(cfg.get("childOrder"), dict) else {}
    child_order = {str(k): [str(x) for x in (v or []) if isinstance(x, (str, int))][:100]
                   for k, v in co_raw.items()}
    # parentOf: childKey -> new parent (section) key. Moves a sub-page to another category.
    po_raw = cfg.get("parentOf") if isinstance(cfg.get("parentOf"), dict) else {}
    parent_of = {str(k): str(v)[:120] for k, v in po_raw.items() if str(v).strip()}
    # custom: admin-created items. A section: {key,label,isSection:true}. A link:
    # {key,label,href,parent,icon?}. href is a path/URL the client renders as a link.
    cu_raw = cfg.get("custom") if isinstance(cfg.get("custom"), list) else []
    custom = []
    for it in cu_raw[:100]:
        if not isinstance(it, dict):
            continue
        key = str(it.get("key") or "").strip()[:120]
        if not key:
            continue
        entry = {"key": key, "label": str(it.get("label") or "")[:60]}
        if it.get("isSection"):
            entry["isSection"] = True
        else:
            entry["href"] = str(it.get("href") or "")[:300]
            entry["parent"] = str(it.get("parent") or "")[:120]
        if it.get("icon"):
            entry["icon"] = str(it.get("icon"))[:40]
        custom.append(entry)
    return {"order": order, "hidden": hidden, "labels": labels, "childOrder": child_order,
            "parentOf": parent_of, "custom": custom}


@app.get("/me/nav-config")
async def me_nav_config(request: Request):
    """The logged-in tenant's sidebar config — the sidebar applies it client-side.
    Any authed user reads their OWN config (never 403). Empty {} -> default nav."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    return JSONResponse({"config": _read_nav_config(t["tenant_id"])})


@app.get("/admin/nav-config")
async def admin_nav_config_get(request: Request, tenant_id: str = ""):
    """Super-admin: read a tenant's saved sidebar config (for the Sidebar Builder)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    tid = (tenant_id or "").strip()
    if not tid:
        return JSONResponse({"error": "tenant_id required"}, status_code=400)
    return JSONResponse({"tenant_id": tid, "config": _read_nav_config(tid)})


@app.post("/admin/nav-config")
async def admin_nav_config_set(request: Request, tenant_id: str = Form(...),
                               config: str = Form(...)):
    """Super-admin: save a tenant's sidebar config. `config` = JSON string
    {order:[], hidden:[], labels:{}, childOrder:{}}. Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    tid = (tenant_id or "").strip()
    if not tid:
        return JSONResponse({"error": "tenant_id required"}, status_code=400)
    try:
        raw = json.loads(config or "{}")
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"invalid config json: {exc}"}, status_code=400)
    clean = _sanitize_nav_config(raw)
    try:
        NAV_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        _atomic_write_json(_nav_config_path(tid), clean)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": f"save failed: {exc}"}, status_code=500)
    try:
        _control_audit(request, t, "control.nav.set", feature_key=tid)
    except Exception:  # noqa: BLE001
        pass
    return JSONResponse({"ok": True, "tenant_id": tid, "config": clean})


# ---------- plans ----------
@app.get("/admin/plans")
async def admin_plans_get(request: Request):
    """List plans + their entitlements + limits."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    return JSONResponse({"plans": _ent_mod.plans_detail()})


@app.post("/admin/plans")
async def admin_plans_create(request: Request, plan_id: str = Form(...), name: str = Form(""),
                             description: str = Form("")):
    """Create a plan (catalog seed + PG). Entitlements/limits are set via PUT. Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    pid = (plan_id or "").strip().lower()
    if not pid:
        return JSONResponse({"error": "plan_id required"}, status_code=400)
    if pid in _ent_mod.load_plans():
        return JSONResponse({"error": "plan already exists"}, status_code=409)
    if _ent_mod.available():
        try:
            _ent_mod._exec_admin(
                "INSERT INTO plans (plan_id, name, description) VALUES (:p,:n,:d) "
                "ON CONFLICT (plan_id) DO NOTHING",
                {"p": pid, "n": name or pid, "d": description})
        except Exception:  # noqa: BLE001
            return _control_unavailable()
    # reflect in the in-proc catalog so it's immediately listable.
    _ent_mod.load_plans()[pid] = {"plan_id": pid, "name": name or pid, "is_default": False,
                                  "entitlements": {}, "limits": {}}
    _control_audit(request, t, "control.plan.create", target_tenant=None,
                   old_value=None, new_value=pid)
    return JSONResponse({"ok": True, "plan_id": pid, "name": name or pid})


@app.put("/admin/plans/{plan_id}")
async def admin_plans_edit(request: Request, plan_id: str):
    """Edit a plan bundle: JSON body {entitlements:{key:mode}, limits:{key:int}}. Replaces both sets +
    bumps every tenant ON that plan (global cache drop). Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    pid = (plan_id or "").strip().lower()
    if pid not in _ent_mod.load_plans():
        return JSONResponse({"error": "unknown plan"}, status_code=404)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    ents = {k: v for k, v in (body.get("entitlements") or {}).items()
            if v in _ent_mod.MODES and k in _ent_mod.load_registry()}
    limits = {}
    for k, v in (body.get("limits") or {}).items():
        try:
            limits[k] = int(v)
        except Exception:  # noqa: BLE001
            pass
    before = _ent_mod.plans_detail()
    if _ent_mod.available():
        try:
            _ent_mod._exec_admin("DELETE FROM plan_entitlements WHERE plan_id = :p", {"p": pid})
            for k, m in ents.items():
                _ent_mod._exec_admin(
                    "INSERT INTO plan_entitlements (plan_id, feature_key, mode) VALUES (:p,:k,:m)",
                    {"p": pid, "k": k, "m": m})
            _ent_mod._exec_admin("DELETE FROM plan_limits WHERE plan_id = :p", {"p": pid})
            for k, v in limits.items():
                _ent_mod._exec_admin(
                    "INSERT INTO plan_limits (plan_id, limit_key, value) VALUES (:p,:k,:v)",
                    {"p": pid, "k": k, "v": v})
        except Exception:  # noqa: BLE001
            return _control_unavailable()
    p = _ent_mod.load_plans().get(pid, {})
    p["entitlements"] = ents
    p["limits"] = limits
    _ent_mod.invalidate(None)  # plan change can affect every tenant on it
    _control_audit(request, t, "control.plan.edit", target_tenant=None,
                   old_value=str(len(before)), new_value=f"ents={len(ents)},limits={len(limits)}")
    return JSONResponse({"ok": True, "plan_id": pid, "entitlements": ents, "limits": limits})


# ---------- vendors (list + workspace) ----------
@app.get("/admin/vendors")
async def admin_vendors(request: Request):
    """Vendor list: {tenant_id,name,email,plan,status,created_at,usage_summary,health}. Joins the tenant
    store + control status + usage + last-activity (executive view)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    out = []
    for x in _read_tenants():
        if x.get("is_admin"):
            continue  # the admin tenant is not a managed vendor
        tid = x["tenant_id"]
        st = _ent_mod.load_status(tid) if _ent_mod else {"status": "active", "plan_id": None}
        out.append({
            "tenant_id": tid, "name": x.get("name", ""), "email": x.get("email", ""),
            "role": _role_of(x), "created_at": x.get("created_at", ""),
            "status": st.get("status", "active"), "plan": st.get("plan_id"),
            "usage": {"today": _tenant_usage(tid, _today_iso()),
                      "month": _tenant_usage(tid, _month_iso())},
            "health": _vendor_health(tid),
        })
    return JSONResponse({"vendors": out})


@app.get("/admin/vendors/{vid}")
async def admin_vendor_detail(request: Request, vid: str):
    """Full vendor profile + resolved entitlement map (effective mode + provenance) + usage + health +
    wallet balance. The Permissions/Overview/Usage/Billing tabs all read from here."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    rec = _tenant_by_id(vid)
    if not rec:
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    detail = _ent_mod.vendor_detail(vid) if _ent_mod else {}
    wallet = None
    if _wallet_mod is not None and _wallet_mod.available():
        try:
            bal = await asyncio.to_thread(_wallet_mod.balance, vid, "INR", True)
            if bal:
                wallet = {"available": _minor_to_major(bal["available_minor"]),
                          "held": _minor_to_major(bal["held_minor"])}
        except Exception:  # noqa: BLE001
            wallet = None
    return JSONResponse({
        "tenant_id": vid, "name": rec.get("name", ""), "email": rec.get("email", ""),
        "role": _role_of(rec), "created_at": rec.get("created_at", ""),
        "usage": {"today": _tenant_usage(vid, _today_iso()), "month": _tenant_usage(vid, _month_iso())},
        "health": _vendor_health(vid), "wallet": wallet, **detail,
    })


# ---------- per-vendor entitlement overrides (HIDE/LOCK/ON) ----------
@app.put("/admin/vendors/{vid}/entitlements/{feature_key}")
async def admin_set_override(request: Request, vid: str, feature_key: str,
                             mode: str = Form(...), reason: str = Form("")):
    """Per-vendor override (HIDE/LOCK/ON). Bumps that tenant's ent_version -> real-time. Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    if not _tenant_by_id(vid):
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    res = _ent_mod.set_override(vid, feature_key, (mode or "").strip(),
                                set_by=t["tenant_id"], reason=reason)
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    _control_audit(request, t, "control.override.set", target_tenant=vid, feature_key=feature_key,
                   old_value=res.get("before"), new_value=res.get("after"), reason=reason)
    return JSONResponse({"ok": True, "tenant_id": vid, "feature_key": feature_key, **res})


@app.delete("/admin/vendors/{vid}/entitlements/{feature_key}")
async def admin_clear_override(request: Request, vid: str, feature_key: str):
    """Clear a per-vendor override -> revert to plan/global. Bumps ent_version. Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    if not _tenant_by_id(vid):
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    res = _ent_mod.clear_override(vid, feature_key, set_by=t["tenant_id"])
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    _control_audit(request, t, "control.override.clear", target_tenant=vid, feature_key=feature_key,
                   old_value=res.get("before"), new_value=None)
    return JSONResponse({"ok": True, "tenant_id": vid, "feature_key": feature_key, **res})


@app.put("/admin/vendors/{vid}/plan")
async def admin_set_plan(request: Request, vid: str, plan_id: str = Form(...)):
    """Assign a plan to a vendor. Bumps ent_version. Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    if not _tenant_by_id(vid):
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    res = _ent_mod.set_plan(vid, (plan_id or "").strip(), updated_by=t["tenant_id"])
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    _control_audit(request, t, "control.plan.assign", target_tenant=vid,
                   old_value=res.get("before"), new_value=res.get("after"))
    return JSONResponse({"ok": True, "tenant_id": vid, **res})


@app.put("/admin/vendors/{vid}/status")
async def admin_set_status(request: Request, vid: str, status: str = Form(...), reason: str = Form("")):
    """Vendor lifecycle: active/trial/suspended/disabled/expired. suspended/disabled REVOKES all the
    vendor's tokens (auth.revoke_all -> next call 401) and the status floor hides every non-core feature;
    DATA IS PRESERVED (a flag flip, never a delete). disabled additionally requires firewall step-up.
    Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _ent_mod is None:
        return _control_unavailable()
    if not _tenant_by_id(vid):
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    new_status = (status or "").strip().lower()
    # disabled is the harshest lifecycle change -> spend/destructive step-up on the ACTING admin.
    if new_status == "disabled":
        denied = _step_up_guard(request, "destructive", t)
        if denied is not None:
            _control_audit(request, t, "control.status.stepup_denied", target_tenant=vid,
                           old_value=None, new_value=new_status, reason=reason)
            return denied
    res = _ent_mod.set_status(vid, new_status, reason=reason, updated_by=t["tenant_id"])
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    # INSTANT KILL: revoke the vendor's refresh tokens so a token minted seconds ago dies now (T15).
    revoked = 0
    if new_status in ("suspended", "disabled", "expired") and _auth_mod is not None:
        try:
            revoked = _auth_mod.revoke_all(vid)
        except Exception:  # noqa: BLE001
            revoked = 0
    _control_audit(request, t, "control.status.set", target_tenant=vid,
                   old_value=res.get("before"), new_value=res.get("after"), reason=reason)
    return JSONResponse({"ok": True, "tenant_id": vid, "tokens_revoked": revoked, **res})


@app.post("/admin/vendors/{vid}/credits")
async def admin_credits(request: Request, vid: str, amount: float = Form(...), reason: str = Form("")):
    """Wallet top-up / freeze for a vendor (rupees). Requires firewall step-up (spend scope) on the
    ACTING admin. Rides the F4 wallet ledger (idempotent). Audited."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if not _tenant_by_id(vid):
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    denied = _step_up_guard(request, "spend", t)
    if denied is not None:
        _control_audit(request, t, "control.credit.stepup_denied", target_tenant=vid,
                       old_value=None, new_value=str(amount), reason=reason)
        return denied
    if _wallet_mod is None or not _wallet_mod.available():
        return JSONResponse(_wallet_unavailable_body(), status_code=503)
    minor = int(round(float(amount) * 100))
    if minor <= 0:
        return JSONResponse({"ok": False, "reason": "amount must be positive"}, status_code=400)
    idem = f"admin_credit:{vid}:{secrets.token_hex(8)}"
    res = await asyncio.to_thread(_wallet_mod.topup, vid, minor, t["tenant_id"], idem, "INR", True, None)
    _control_audit(request, t, "control.credit.topup", target_tenant=vid,
                   old_value=None, new_value=_minor_to_major(minor), reason=reason)
    if not res.get("ok"):
        return JSONResponse(res, status_code=400)
    return JSONResponse({"ok": True, "tenant_id": vid, "credited": _minor_to_major(minor),
                         "available": _minor_to_major(res.get("available_minor", 0))})


# ---------- impersonation / act-as (the sharpest knife — gated like root) ----------
@app.post("/admin/vendors/{vid}/impersonate")
async def admin_impersonate(request: Request, vid: str, scope: str = Form("read_only")):
    """Enter act-as: mint a short-TTL (<=10min) act-as token (sub=vendor, real_admin=admin, read-only by
    default). Requires firewall step-up (the sub-bound F3 token). CANNOT target another admin (no lateral
    takeover). Enter is audited; exit is POST /admin/act-as/exit (also audited)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    rec = _tenant_by_id(vid)
    if not rec:
        return JSONResponse({"error": "vendor not found"}, status_code=404)
    # T12: cannot act-as another admin (admin-on-admin impersonation refused).
    if rec.get("is_admin"):
        return _forbidden("cannot impersonate an admin tenant")
    # T9: firewall step-up to ENTER (sub-bound to the acting admin).
    denied = _step_up_guard(request, "destructive", t)
    if denied is not None:
        _control_audit(request, t, "control.impersonate.stepup_denied", target_tenant=vid)
        return denied
    if _auth_mod is None or not _auth_mod.available():
        return JSONResponse({"error": "auth (JWT) unavailable for act-as"}, status_code=503)
    sc = (scope or "read_only").strip().lower()
    if sc not in ("read_only", "read_write"):
        sc = "read_only"
    tok = _auth_mod.make_act_as(vid, t["tenant_id"], sc)
    if not tok:
        return JSONResponse({"error": "could not mint act-as token"}, status_code=503)
    _control_audit(request, t, "control.impersonate.start", target_tenant=vid,
                   old_value=None, new_value=sc)
    resp = JSONResponse({"ok": True, "act_as": vid, "real_admin": t["tenant_id"], "scope": sc,
                         "access_token": tok, "token_type": "Bearer",
                         "expires_in": _auth_mod.ACT_AS_TTL_SECONDS})
    resp.headers["X-Act-As"] = vid  # so the FE banner can't be suppressed via localStorage tamper
    return resp


@app.post("/admin/act-as/exit")
async def admin_act_as_exit(request: Request):
    """Exit act-as. The caller presents the act-as token; we audit the exit (actor=real_admin) and revoke
    the vendor's act-as session. The token's own short TTL bounds it regardless."""
    cred = _extract_cred(request)
    claims = _auth_mod.act_as_claims(cred) if _auth_mod is not None else None
    if not claims:
        return JSONResponse({"error": "not an act-as session"}, status_code=400)
    real_admin = claims.get("real_admin")
    vendor = claims.get("act_as")
    admin_t = _tenant_by_id(real_admin) or {"tenant_id": real_admin}
    _control_audit(request, admin_t, "control.impersonate.stop", target_tenant=vendor,
                   old_value=claims.get("scope"), new_value=None)
    return JSONResponse({"ok": True, "exited": vendor})


# ---------- vendor-facing entitlement read (the panel + AI Copilot consume THIS) ----------
@app.get("/me/entitlements")
async def me_entitlements(request: Request):
    """Versioned resolved entitlement map for the logged-in tenant: {modes, status, plan, version}.
    ETag = the tenant's ent_version; If-None-Match short-circuits to 304 (cheap poll). Core route —
    bypasses the enforcement choke-point (anti-lockout) so it always answers even when suspended."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _ent_mod is None:
        # degrade: all-on (resting). Still a valid, versionless map so the FE renders full nav.
        return JSONResponse({"modes": {}, "status": "active", "plan": None, "version": 1})
    tid = t["tenant_id"]
    payload = _ent_mod.entitlements_payload(tid)
    etag = f'W/"ent-{tid}-{payload.get("version", 1)}"'
    inm = request.headers.get("if-none-match", "")
    if inm and inm == etag:
        return Response(status_code=304, headers={"ETag": etag,
                                                  "Cache-Control": "private, no-cache"})
    return JSONResponse(payload, headers={"ETag": etag, "Cache-Control": "private, no-cache"})


# ---------- WAVE3 Unit4: billing endpoints ----------
def _month_cost(tenant_id: str) -> float:
    month = _month_iso()
    return round(sum(e.get("cost", 0) or 0 for e in _read_ledger(tenant_id)
                     if (e.get("at") or "")[:7] == month), 4)


@app.get("/billing")
async def billing(request: Request):
    """Current plan + month-to-date usage/cost + remaining balance for the caller's tenant."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    tid = t["tenant_id"]
    b = _billing_for(tid)
    mtd = _tenant_usage(tid, _month_iso())
    cost = _month_cost(tid)
    return JSONResponse({
        "tenant_id": tid,
        "plan": b.get("plan", "postpaid"),
        "currency": b.get("currency", "INR"),
        "rate_per_min": float(b.get("rate_per_min", 0) or 0),
        "rate_per_call": float(b.get("rate_per_call", 0) or 0),
        "balance": float(b.get("balance", 0) or 0),
        "included_minutes": int(b.get("included_minutes", 0) or 0),
        "month_to_date": {"calls": mtd.get("calls", 0), "minutes": mtd.get("minutes", 0),
                          "cost": cost},
    })


@app.get("/billing/ledger")
async def billing_ledger(request: Request, limit: int = 100):
    """Itemized recent charges for the caller's tenant (most recent first)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = _read_ledger(t["tenant_id"])[:max(1, min(limit, 1000))]
    return JSONResponse({"ledger": rows, "total": len(_read_ledger(t["tenant_id"]))})


# ---------- WAVE A Unit4: real vendor-cost billing endpoints (additive) ----------
def _vendor_status(vid: str) -> str:
    try:
        if vid == "elevenlabs" and v_elevenlabs:
            return v_elevenlabs.status()
        if vid == "vobiz" and v_vobiz:
            return v_vobiz.status()
        if vid == "groq" and v_groq:
            return v_groq.status()
        if vid == "sarvam" and v_sarvam:
            return v_sarvam.status()
        if vid == "livekit":
            return "configured"
    except Exception:  # noqa: BLE001
        return "error"
    return "not_configured"


def _per_vendor_totals(rows: list[dict]) -> dict:
    """vendor_id -> {cost, qty_by_service} from cost_ledger rows."""
    out: dict = {}
    for r in rows:
        v = r.get("vendor") or "unknown"
        out.setdefault(v, {"cost": 0.0})
        out[v]["cost"] = round(out[v]["cost"] + float(r.get("cost", 0) or 0), 6)
    return out


def _currency_for(tenant: dict) -> str:
    try:
        return _billing_for(tenant["tenant_id"]).get("currency", "INR")
    except Exception:  # noqa: BLE001
        return "INR"


@app.get("/billing/overview")
async def billing_overview(request: Request):
    """Grand total + per-vendor totals + month-to-date for the caller's scope
    (admin = all tenants; manager/agent = own tenant)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = _cost_rows_for(t)
    month = _month_iso()
    grand = round(sum(float(r.get("cost", 0) or 0) for r in rows), 4)
    mtd = round(sum(float(r.get("cost", 0) or 0) for r in rows
                    if (r.get("ts") or "")[:7] == month), 4)
    totals = _per_vendor_totals(rows)
    per_vendor = [{"vendor": vid, "display_name": vendor_display_name(vid),
                   "cost": round(totals.get(vid, {}).get("cost", 0.0), 4),
                   "status": _vendor_status(vid)} for vid in VENDOR_IDS]
    snaps = _read(VENDOR_SNAPSHOTS_FILE, {}) or {}
    updated = ""
    if isinstance(snaps, dict):
        updated = max((v.get("synced_at", "") for v in snaps.values()
                       if isinstance(v, dict)), default="")
    return JSONResponse({"currency": _currency_for(t), "grand_total": grand,
                         "month_to_date": mtd, "per_vendor": per_vendor,
                         "updated_at": updated})


@app.get("/billing/vendors")
async def billing_vendors(request: Request):
    """List every vendor with status + totals + display name (real names now; the
    display-name map makes switching to abstract labels trivial)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = _cost_rows_for(t)
    totals = _per_vendor_totals(rows)
    snaps = _read(VENDOR_SNAPSHOTS_FILE, {}) or {}
    vendors = []
    for vid in VENDOR_IDS:
        snap = snaps.get(vid, {}) if isinstance(snaps, dict) else {}
        vendors.append({"vendor": vid, "display_name": vendor_display_name(vid),
                        "status": _vendor_status(vid),
                        "cost": round(totals.get(vid, {}).get("cost", 0.0), 4),
                        "synced_at": snap.get("synced_at", ""),
                        "stale": bool(snap.get("stale", False)),
                        "estimated": bool(snap.get("estimated", vid in ("groq", "sarvam")))})
    return JSONResponse({"vendors": vendors, "currency": _currency_for(t)})


@app.get("/billing/vendor/{vid}")
async def billing_vendor_detail(request: Request, vid: str):
    """One vendor: status, total cost, and a daily timeseries (from cost_ledger rows)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if vid not in VENDOR_IDS:
        return JSONResponse({"error": "unknown vendor"}, status_code=404)
    rows = [r for r in _cost_rows_for(t) if r.get("vendor") == vid]
    by_day: dict = {}
    for r in rows:
        day = (r.get("ts") or "")[:10] or "unknown"
        by_day[day] = round(by_day.get(day, 0.0) + float(r.get("cost", 0) or 0), 6)
    series = [{"date": d, "cost": c} for d, c in sorted(by_day.items())]
    snap = (_read(VENDOR_SNAPSHOTS_FILE, {}) or {}).get(vid, {})
    return JSONResponse({"vendor": vid, "display_name": vendor_display_name(vid),
                         "status": _vendor_status(vid),
                         "total_cost": round(sum(float(r.get("cost", 0) or 0) for r in rows), 4),
                         "currency": _currency_for(t),
                         "timeseries": series, "rows": len(rows),
                         "synced_at": snap.get("synced_at", ""),
                         "stale": bool(snap.get("stale", False))})


@app.get("/billing/explorer")
async def billing_explorer(request: Request):
    """Per call/lead/campaign cost rows from cost_ledger, filterable by from/to/campaign_id.
    Aggregated by call_id (one row per call, vendor breakdown nested). `from` is a Python
    keyword so query params are read off the request directly."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    qp = request.query_params
    frm = qp.get("from", "")
    to = qp.get("to", "")
    camp = qp.get("campaign_id", "")
    rows = _cost_rows_for(t)
    agg: dict = {}
    for r in rows:
        day = (r.get("ts") or "")[:10]
        if frm and day and day < frm:
            continue
        if to and day and day > to:
            continue
        if camp and r.get("campaign_id") != camp:
            continue
        cid = r.get("call_id") or f"room:{r.get('room','')}"
        a = agg.setdefault(cid, {"call_id": r.get("call_id", ""), "room": r.get("room", ""),
                                 "tenant_id": r.get("tenant_id", ""),
                                 "campaign_id": r.get("campaign_id", ""),
                                 "ts": r.get("ts", ""), "total_cost": 0.0, "by_vendor": {}})
        a["total_cost"] = round(a["total_cost"] + float(r.get("cost", 0) or 0), 6)
        a["by_vendor"][r.get("vendor", "")] = round(
            a["by_vendor"].get(r.get("vendor", ""), 0.0) + float(r.get("cost", 0) or 0), 6)
        if r.get("ts") and (not a["ts"] or r["ts"] < a["ts"]):
            a["ts"] = r["ts"]
    out = sorted(agg.values(), key=lambda x: x.get("ts", ""), reverse=True)
    # Attach lead name + outcome + campaign name by joining the call record.
    for o in out:
        c = _call_by_room(o.get("room", "")) or {}
        o["name"] = c.get("name", "")
        o["phone"] = c.get("phone", "")
        o["campaign_name"] = c.get("campaign_name", "")
        o["outcome"] = c.get("outcome", "")
        o["duration_s"] = c.get("duration_s", 0)
    return JSONResponse({"rows": out[:2000], "total": len(out),
                         "currency": _currency_for(t),
                         "filters": {"from": frm, "to": to, "campaign_id": camp}})


@app.get("/billing/audit")
async def billing_audit(request: Request):
    """Sync status per vendor, stale flags, and vendor-reported vs internal-ledger compare."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    snaps = _read(VENDOR_SNAPSHOTS_FILE, {}) or {}
    rows = _cost_rows_for(t)
    internal = _per_vendor_totals(rows)
    vendors = []
    for vid in VENDOR_IDS:
        snap = snaps.get(vid, {}) if isinstance(snaps, dict) else {}
        # vendor-reported figure (only meaningful for keyed vendors).
        reported = None
        if vid == "vobiz" and isinstance(snap, dict):
            recs = snap.get("records") or []
            reported = round(sum(float(r.get("total_cost", 0) or 0) for r in recs), 4) if recs else None
        elif vid == "elevenlabs" and isinstance(snap, dict):
            u = (snap.get("usage") or {})
            reported = u.get("total_credits") if isinstance(u, dict) else None
        vendors.append({
            "vendor": vid, "display_name": vendor_display_name(vid),
            "status": _vendor_status(vid),
            "synced_at": snap.get("synced_at", ""),
            "stale": bool(snap.get("stale", False)),
            "error": snap.get("error", ""),
            "internal_ledger_cost": round(internal.get(vid, {}).get("cost", 0.0), 4),
            "vendor_reported": reported,
        })
    return JSONResponse({"vendors": vendors, "currency": _currency_for(t),
                         "note": "vendor_reported is workspace/account-level (ElevenLabs, Vobiz balance) "
                                 "and may exceed per-tenant internal ledger; groq/sarvam are estimated."})


@app.post("/billing/sync")
async def billing_sync(request: Request):
    """ADMIN: trigger an immediate vendor-sync pass; return the fresh snapshot.
    (Registered BEFORE /billing/{tenant_id} so 'sync' isn't captured as a tenant_id.)"""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "manage_tenants"):
        return _forbidden("admin only")
    global _LAST_VENDOR_SYNC
    _LAST_VENDOR_SYNC = time.time()
    snaps = await vendor_sync()
    summary = {vid: {"status": (snaps.get(vid, {}) or {}).get("status", _vendor_status(vid)),
                     "synced_at": (snaps.get(vid, {}) or {}).get("synced_at", ""),
                     "stale": bool((snaps.get(vid, {}) or {}).get("stale", False))}
               for vid in VENDOR_IDS}
    return JSONResponse({"ok": True, "synced_at": now_ist().isoformat(timespec="seconds"),
                         "vendors": summary})


@app.get("/admin/store-status")
async def admin_store_status(request: Request):
    """ADMIN (P1 U8): per-store storage-seam visibility — each registered store's effective MODE +
    max_safe + live PG/JSON row counts + mirror ok/fail counters + last error. Read-only; additive.
    JSON count = len(authoritative JSON file); PG count = admin-GUC SELECT count(*). When they match the
    store is converged (an effectively-live drift indicator; the spec's last_shadow_diff stays None
    because shadow_diff.py runs out-of-process and can't write the live spec object). Best-effort
    everywhere: any per-store error is captured into that store's `error` field, never 500s the route."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "manage_tenants"):
        return _forbidden("admin only")

    # base: store.py's own status (modes, max_safe, pg ok/fail, worker liveness, last_error)
    try:
        base = _store.status() if _store is not None else {"ready": False, "stores": {}, "db": {}}
    except Exception as exc:  # noqa: BLE001
        base = {"ready": False, "stores": {}, "db": {}, "status_error": repr(exc)[:200]}

    stores = base.get("stores", {}) or {}

    # registered-store name -> authoritative JSON file path (for the live JSON count)
    _file_for = {
        "leads.json": LEADS_FILE, "calls.json": CALLS_FILE, "suppression.json": SUPPRESSION_FILE,
        "retry_queue.json": RETRY_FILE, "webhooks.json": WEBHOOK_FILE, "billing.json": BILLING_FILE,
        "wa_log.json": WA_LOG_FILE, "usage_events.json": USAGE_EVENTS_FILE,
        "cost_ledger.json": COST_LEDGER_FILE,
    }
    _table_for = {
        "leads.json": "leads", "calls.json": "calls", "suppression.json": "suppression",
        "retry_queue.json": "retry_queue", "webhooks.json": "webhooks", "billing.json": "billing",
        "wa_log.json": "wa_log", "usage_events.json": "usage_events", "cost_ledger.json": "cost_ledger",
        "ledger": "ledger",  # multi_file: per-tenant files var/ledger/<stem>.json
    }

    # live JSON counts (cheap, off the authoritative files)
    for name, info in stores.items():
        try:
            if name == "ledger":
                # multi_file: sum rows across all per-tenant files var/ledger/<stem>.json
                total = 0
                if LEDGER_DIR.exists():
                    for lf in LEDGER_DIR.glob("*.json"):
                        obj = _read_raw(lf, [])
                        total += len(obj) if isinstance(obj, list) else 0
                info["json_count"] = total
                continue
            f = _file_for.get(name)
            if f is not None:
                obj = _read_raw(f, [])
                info["json_count"] = len(obj) if isinstance(obj, (list, dict)) else None
            else:
                info["json_count"] = None
        except Exception as exc:  # noqa: BLE001
            info["json_count"] = None
            info["error"] = repr(exc)[:200]

    # live PG counts (admin GUC; only if the db layer is up). Best-effort, swallow per-store.
    try:
        from db import engine as _eng  # local import: import-safe, never required for the route
        if _eng.available():
            from sqlalchemy import text as _text
            with _eng.session("", is_admin=True) as _s:
                for name, info in stores.items():
                    tbl = _table_for.get(name)
                    if not tbl:
                        info["pg_count"] = None
                        continue
                    try:
                        info["pg_count"] = _s.execute(
                            _text(f"SELECT count(*) FROM {tbl}")).scalar() or 0
                    except Exception as exc:  # noqa: BLE001
                        info["pg_count"] = None
                        info["error"] = repr(exc)[:200]
                # identity mirror (orgs/users/memberships) — not a seam store but useful here
                try:
                    base["identity"] = {
                        "orgs": _s.execute(_text("SELECT count(*) FROM orgs")).scalar() or 0,
                        "users": _s.execute(_text("SELECT count(*) FROM users")).scalar() or 0,
                        "memberships": _s.execute(
                            _text("SELECT count(*) FROM memberships")).scalar() or 0,
                    }
                except Exception:  # noqa: BLE001
                    pass
        else:
            for info in stores.values():
                info.setdefault("pg_count", None)
    except Exception as exc:  # noqa: BLE001
        base["pg_count_error"] = repr(exc)[:200]

    base["at"] = now_ist().isoformat(timespec="seconds")
    return JSONResponse(base)


@app.post("/billing/{tenant_id}")
async def set_billing(request: Request, tenant_id: str, plan: str = Form(None),
                      rate_per_min: float = Form(None), rate_per_call: float = Form(None),
                      currency: str = Form(None), balance: float = Form(None),
                      included_minutes: int = Form(None), topup: float = Form(None)):
    """ADMIN: set a tenant's plan/rates/balance, or top up. `balance` sets absolute;
    `topup` adds to the current balance. Returns the updated billing record."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "manage_tenants"):
        return _forbidden("admin only")
    if not _tenant_by_id(tenant_id):
        return JSONResponse({"error": "tenant not found"}, status_code=404)
    async with _STORE_LOCK:
        store = _read_billing()
        rec = store.get(tenant_id) or _default_billing(_tenant_by_id(tenant_id))
        if plan is not None:
            rec["plan"] = "prepaid" if str(plan).lower() == "prepaid" else "postpaid"
        if rate_per_min is not None:
            rec["rate_per_min"] = max(0.0, float(rate_per_min))
        if rate_per_call is not None:
            rec["rate_per_call"] = max(0.0, float(rate_per_call))
        if currency is not None and str(currency).strip():
            rec["currency"] = str(currency).strip().upper()[:4]
        if included_minutes is not None:
            rec["included_minutes"] = max(0, int(included_minutes))
        if balance is not None:
            rec["balance"] = round(float(balance), 4)
        if topup is not None:
            rec["balance"] = round(float(rec.get("balance", 0) or 0) + float(topup), 4)
        store[tenant_id] = rec
        _write(BILLING_FILE, store)
    _audit(request, t, "billing.config", "billing", tenant_id,
           meta={"plan": rec.get("plan"), "topup": topup})
    return JSONResponse({"tenant_id": tenant_id, **rec})


# ---------- WAVE3 Unit5: WhatsApp manual send ----------
@app.post("/whatsapp/send")
async def whatsapp_send(request: Request, to: str = Form(""), template: str = Form(""),
                        text: str = Form(""), params: str = Form("")):
    """Manager+ manual WhatsApp send. `template` = template name (BSP) OR pass `text`
    for a raw-text generic send. `params` = comma-separated template variables.
    No-ops gracefully (status 'skipped_no_config') if WA_* env not set."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot send WhatsApp")
    to_n = norm(to)
    if not to_n:
        return JSONResponse({"error": "valid 'to' phone required"}, status_code=400)
    tpl = (template or text or "").strip()
    if not tpl:
        return JSONResponse({"error": "template or text is required"}, status_code=400)
    is_text = bool(text.strip() and not template.strip())
    plist = [p.strip() for p in re.split(r"[,\|]", params) if p.strip()] if params else []
    result = await _wa_send(t["tenant_id"], to_n, tpl, plist, kind="manual", is_text=is_text)
    code = 200 if result.get("status") != "skipped_no_config" else 200
    _audit(request, t, "whatsapp.send", "whatsapp", to_n, channel="whatsapp",
           meta={"status": result.get("status")})
    return JSONResponse({"ok": bool(result.get("ok")), "status": result.get("status"),
                         "to": to_n, "configured": bool(wa_mod and wa_mod.is_configured()),
                         # WAFX: surface Meta's REAL error (141006 payment, template-not-registered,
                         # etc.) so the panel shows the true reason instead of "try again".
                         "error": result.get("meta_error")},
                        status_code=code)


# ---------- P1 visibility endpoints ----------
@app.get("/whatsapp/log")
async def whatsapp_log(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = [x for x in _read(WA_LOG_FILE, [])
            if t.get("is_admin") or x.get("tenant_id") == t["tenant_id"]]
    return JSONResponse({"log": rows[:500]})


# ---------- WAVE A2: inbound WhatsApp webhook (Meta) + conversation threads ----------
@app.get("/whatsapp/inbound")
async def whatsapp_inbound_verify(request: Request):
    """Meta webhook verification handshake (NO auth — Meta calls this).
    Echoes hub.challenge as plain text when hub.verify_token == META_WA_VERIFY_TOKEN."""
    q = request.query_params
    mode = q.get("hub.mode", "")
    token = q.get("hub.verify_token", "")
    challenge = q.get("hub.challenge", "")
    expected = os.getenv("META_WA_VERIFY_TOKEN", "").strip()
    if mode == "subscribe" and expected and token == expected:
        return Response(content=challenge, media_type="text/plain", status_code=200)
    return Response(content="verification failed", status_code=403)


def _verify_meta_signature(raw: bytes, header_sig: str) -> bool:
    """Verify X-Hub-Signature-256 (HMAC-SHA256 of the raw body w/ META_WA_APP_SECRET).
    If no app secret is configured (dormant), accept (nothing to verify against yet)."""
    secret = os.getenv("META_WA_APP_SECRET", "").strip()
    if not secret:
        return True  # dormant: can't verify, accept so the pipeline is exercisable
    if not header_sig or "=" not in header_sig:
        return False
    algo, _, sent = header_sig.partition("=")
    if algo != "sha256":
        return False
    digest = hmac.new(secret.encode("utf-8"), raw, hashlib.sha256).hexdigest()
    return hmac.compare_digest(digest, sent)


def _parse_meta_inbound(payload: dict) -> list[dict]:
    """Extract [{phone, text}] from a Meta messages webhook payload. Tolerant."""
    out = []
    try:
        for entry in payload.get("entry", []) or []:
            for ch in entry.get("changes", []) or []:
                val = ch.get("value", {}) or {}
                for m in val.get("messages", []) or []:
                    phone = m.get("from", "")
                    text = ""
                    if m.get("type") == "text":
                        text = (m.get("text", {}) or {}).get("body", "")
                    elif m.get("type") == "button":
                        text = (m.get("button", {}) or {}).get("text", "")
                    elif m.get("type") == "interactive":
                        inter = m.get("interactive", {}) or {}
                        text = ((inter.get("button_reply", {}) or {}).get("title")
                                or (inter.get("list_reply", {}) or {}).get("title") or "")
                    if phone:
                        out.append({"phone": phone, "text": text or ""})
    except Exception:  # noqa: BLE001
        pass
    return out


@app.post("/whatsapp/inbound")
async def whatsapp_inbound(request: Request):
    """Receive Meta message webhooks. Verifies X-Hub-Signature-256, parses sender+text,
    drives the per-contact conversation thread. ALWAYS returns 200 quickly (Meta requires
    a fast 2xx). Dormant-safe: with no WA env it stores inbound + no-ops the reply."""
    raw = await request.body()
    sig = request.headers.get("X-Hub-Signature-256", "")
    if not _verify_meta_signature(raw, sig):
        return JSONResponse({"error": "bad signature"}, status_code=403)
    try:
        payload = json.loads(raw.decode("utf-8") or "{}")
    except Exception:  # noqa: BLE001
        payload = {}
    msgs = _parse_meta_inbound(payload)
    results = []
    for m in msgs:
        try:
            results.append(await _wa_handle_inbound(m["phone"], m["text"]))
        except Exception:  # noqa: BLE001
            pass
    return JSONResponse({"ok": True, "handled": len(results)})


@app.get("/whatsapp/threads")
async def whatsapp_threads(request: Request):
    """List WhatsApp conversation threads (tenant-scoped; manager+/admin)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    out = []
    try:
        if WA_THREADS_DIR.exists():
            # P0-LEAK: threads now live under per-tenant subdirs ({tenant}/{phone}.json)
            # plus the legacy flat files -> glob RECURSIVELY so tenant-scoped threads are
            # listed. Skip atomic-write temp siblings. The tenant filter below is still the
            # authoritative isolation gate (a thread with no tenant_id is shown to admin only).
            for p in sorted(WA_THREADS_DIR.glob("**/*.json"),
                            key=lambda x: x.stat().st_mtime, reverse=True):
                if p.name.startswith("."):
                    continue
                th = _read(p, {}) or {}
                tid = th.get("tenant_id", ADMIN_ID)
                if not t.get("is_admin") and tid != t["tenant_id"]:
                    continue
                out.append({"phone": th.get("phone", ""), "name": th.get("name", ""),
                            "tenant_id": tid, "campaign_id": th.get("campaign_id", ""),
                            "campaign_name": th.get("campaign_name", ""),
                            "status": th.get("status", "active"),
                            "turns": len(th.get("turns") or []),
                            "updated_at": th.get("updated_at", th.get("created_at", ""))})
    except Exception:  # noqa: BLE001
        pass
    return JSONResponse({"threads": out, "total": len(out)})


@app.get("/whatsapp/threads/{phone}")
async def whatsapp_thread_detail(phone: str, request: Request):
    """Full conversation history for one contact (tenant-scoped)."""
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    phone_q = norm(phone) or phone
    if t.get("is_admin"):
        # Admin may view any tenant's thread: locate the file across tenant subdirs
        # (+ the legacy flat path) without tenant-scoping the read.
        th = _wa_thread_find_any(phone_q)
    else:
        th = _wa_thread_read(phone_q, t.get("tenant_id"))
    guard = require_object(t, th)  # 404 if missing or not owned (BOLA)
    if guard is not None:
        return guard
    return JSONResponse({"thread": th})


@app.get("/webhooks")
async def get_webhooks(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = [w for w in _read(WEBHOOK_FILE, []) if w.get("tenant_id") == t["tenant_id"]]
    return JSONResponse({"webhooks": rows})


@app.post("/webhooks")
async def add_webhook(request: Request, url: str = Form(""), secret: str = Form(""),
                      events: str = Form("")):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot add webhooks")
    if not (url or "").strip():
        return JSONResponse({"error": "url is required"}, status_code=400)
    ev = [e.strip() for e in re.split(r"[,\s]+", events) if e.strip()]
    rec = {"id": uuid.uuid4().hex[:10], "tenant_id": t["tenant_id"], "url": url.strip(),
           "secret": secret or secrets.token_hex(16), "events": ev, "active": True,
           "created_at": datetime.now().isoformat(timespec="seconds")}
    async with _STORE_LOCK:
        store = _read(WEBHOOK_FILE, [])
        store.append(rec)
        _write(WEBHOOK_FILE, store)
    _audit(request, t, "webhook.create", "webhook", rec["id"], meta={"url": rec["url"]})
    return JSONResponse({"id": rec["id"], "url": rec["url"], "secret": rec["secret"]})


@app.delete("/webhooks/{wid}")
async def delete_webhook(request: Request, wid: str):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if not can(t, "write"):
        return _forbidden("read-only role cannot delete webhooks")
    async with _STORE_LOCK:
        store = _read(WEBHOOK_FILE, [])
        target = next((w for w in store if w.get("id") == wid), None)
        if not target or (not t.get("is_admin") and target.get("tenant_id") != t["tenant_id"]):
            return JSONResponse({"error": "not found"}, status_code=404)
        store = [w for w in store if w.get("id") != wid]
        _write(WEBHOOK_FILE, store)
    _audit(request, t, "webhook.delete", "webhook", wid)
    return JSONResponse({"deleted": wid})


# ---------- WAVE A Unit3: cost ledger + rollups + vendor sync ----------
def _telephony_rate_per_min(tenant_id: str) -> float:
    """Admin-set rate_per_min from the existing billing config — used as the telephony
    cost fallback when Vobiz is dormant (so the meter is never 0)."""
    try:
        return float(_billing_for(tenant_id).get("rate_per_min", 0) or 0)
    except Exception:  # noqa: BLE001
        return 0.0


def _read_cost_ledger() -> list[dict]:
    rows = _read(COST_LEDGER_FILE, [])
    return rows if isinstance(rows, list) else []


def rebuild_cost_ledger() -> dict:
    """Join usage_events + Vobiz CDR snapshot into a normalized var/cost_ledger.json,
    one row per (call, vendor). Idempotent: rebuilt fully each pass from source stores
    (usage_events + vendor_snapshots), so re-running can't double-count. Also writes
    var/daily_rollups.json. Returns a small summary. Synchronous; callers hold no lock
    (it reads sources + writes two derived files)."""
    events = _read_usage_events()
    snaps = _read(VENDOR_SNAPSHOTS_FILE, {}) or {}
    vobiz_snap = (snaps.get("vobiz") or {}) if isinstance(snaps, dict) else {}
    # Map sip_call_id -> cdr record (deduped by sip_call_id).
    cdr_by_sip = {}
    for rec in (vobiz_snap.get("records") or []):
        sid = rec.get("sip_call_id")
        if sid:
            cdr_by_sip[sid] = rec
    # Secondary index: normalised phone -> list of cdr records (for calls lacking sip_call_id).
    def _norm_phone(p: str) -> str:
        p = (p or "").strip().lstrip("+")
        return p[-10:] if len(p) >= 10 else p
    cdr_by_phone: dict = {}
    for rec in (vobiz_snap.get("records") or []):
        ph = _norm_phone(rec.get("to_number", ""))
        if ph:
            cdr_by_phone.setdefault(ph, []).append(rec)
    # Index calls by room + by phone for telephony attribution.
    call_by_room = {c.get("room"): c for c in CALLS if c.get("room")}

    rows: list[dict] = []
    # 1) Per-vendor internal usage rows (groq/sarvam/elevenlabs/livekit).
    for ev in events:
        rows.append({
            "source": "internal",
            "ts": ev.get("ts", ""),
            "call_id": ev.get("call_id", ""),
            "room": ev.get("room", ""),
            "tenant_id": ev.get("tenant_id", ""),
            "campaign_id": ev.get("campaign_id", ""),
            "vendor": ev.get("vendor", ""),
            "service_type": ev.get("service_type", ""),
            "qty": ev.get("qty", 0),
            "unit": ev.get("unit", ""),
            "cost": round(float(ev.get("est_cost_inr", 0) or 0), 6),
            "currency": "INR",
            "actual_or_estimated": ev.get("actual_or_estimated", "estimated"),
        })
    # 2) Telephony (Vobiz) rows: one per call. Join strategy (in order):
    #    a) sip_call_id exact match (new calls — stamped by run_job since Wave-A activation)
    #    b) phone + 15-min window match (legacy calls lacking sip_call_id; best-effort)
    #    c) fallback: duration_s/60 × admin rate (when Vobiz dormant or call unmatched)
    #    After matching calls, any remaining unmatched CDR records are emitted as orphan rows
    #    (real cost, no call attribution) so Vobiz total is never silently dropped.
    used_sids = set()
    # Build a consumable list of cdr recs keyed by phone for the secondary join.
    # Sort each phone's CDR list by 'at' for deterministic matching.
    cdr_candidates: dict = {ph: sorted(lst, key=lambda x: x.get("at", ""))
                             for ph, lst in cdr_by_phone.items()}
    used_phone_sids: set = set()  # sip_call_ids consumed by secondary join

    for c in CALLS:
        if c.get("status") not in ("done",) and not c.get("duration_s"):
            continue
        dur = int(c.get("duration_s", 0) or 0)
        tid = c.get("tenant_id", "")
        sid = c.get("sip_call_id", "")
        cdr = cdr_by_sip.get(sid) if sid else None
        if cdr and sid not in used_sids:
            # (a) Primary: exact sip_call_id match.
            used_sids.add(sid)
            used_phone_sids.add(sid)
            rows.append({
                "source": "vobiz_cdr", "ts": cdr.get("at", c.get("ended_at", "")),
                "call_id": c.get("id", ""), "room": c.get("room", ""), "tenant_id": tid,
                "campaign_id": c.get("campaign_id", ""), "vendor": "vobiz",
                "service_type": "telephony",
                "qty": cdr.get("billable_seconds", dur), "unit": "seconds",
                "cost": round(float(cdr.get("total_cost", 0) or 0), 6),
                "currency": vobiz_snap.get("currency", "INR"),
                "actual_or_estimated": "actual",
            })
        else:
            # (b) Secondary: phone + 15-min window match (for legacy calls lacking sip_call_id).
            ph = _norm_phone(c.get("phone", ""))
            matched_cdr = None
            if ph and not sid:
                call_ts = c.get("started_at", "") or c.get("ended_at", "")
                for candidate in cdr_candidates.get(ph, []):
                    csid = candidate.get("sip_call_id", "")
                    if csid in used_phone_sids:
                        continue
                    cdr_ts = candidate.get("at", "")
                    # Accept if the CDR end_time is within 15 minutes of call start_at.
                    try:
                        # Normalise ISO strings: strip trailing Z, handle +HH:MM offsets.
                        def _parse_iso(s: str):
                            s = s.replace("Z", "+00:00")
                            return datetime.fromisoformat(s)
                        dt_call = _parse_iso(call_ts) if call_ts else None
                        dt_cdr = _parse_iso(cdr_ts) if cdr_ts else None
                        if dt_call and dt_cdr:
                            # Make both offset-naive for comparison (strip tz).
                            tc = dt_call.replace(tzinfo=None)
                            tcd = dt_cdr.replace(tzinfo=None)
                            if abs((tcd - tc).total_seconds()) <= 900:
                                matched_cdr = candidate
                                used_phone_sids.add(csid)
                                break
                    except Exception:  # noqa: BLE001
                        pass
            if matched_cdr:
                msid = matched_cdr.get("sip_call_id", "")
                used_sids.add(msid)
                rows.append({
                    "source": "vobiz_cdr_phone", "ts": matched_cdr.get("at", c.get("ended_at", "")),
                    "call_id": c.get("id", ""), "room": c.get("room", ""), "tenant_id": tid,
                    "campaign_id": c.get("campaign_id", ""), "vendor": "vobiz",
                    "service_type": "telephony",
                    "qty": matched_cdr.get("billable_seconds", dur), "unit": "seconds",
                    "cost": round(float(matched_cdr.get("total_cost", 0) or 0), 6),
                    "currency": vobiz_snap.get("currency", "INR"),
                    "actual_or_estimated": "actual",
                })
            elif dur > 0:
                # (c) Fallback: estimated from duration × rate.
                rate = _telephony_rate_per_min(tid)
                rows.append({
                    "source": "telephony_fallback", "ts": c.get("ended_at", ""),
                    "call_id": c.get("id", ""), "room": c.get("room", ""), "tenant_id": tid,
                    "campaign_id": c.get("campaign_id", ""), "vendor": "vobiz",
                    "service_type": "telephony",
                    "qty": dur, "unit": "seconds",
                    "cost": round(dur / 60.0 * rate, 6), "currency": "INR",
                    "actual_or_estimated": "estimated",
                })
    # (d) Orphan CDR rows: Vobiz records not matched to any call (e.g. manual test dials,
    #     calls placed before this system existed). Real cost — include under admin tenant.
    all_used_sids = used_sids | used_phone_sids
    admin_tid = "admin"
    for rec in (vobiz_snap.get("records") or []):
        rsid = rec.get("sip_call_id", "")
        if rsid and rsid not in all_used_sids:
            rows.append({
                "source": "vobiz_cdr_orphan", "ts": rec.get("at", ""),
                "call_id": "", "room": "", "tenant_id": admin_tid,
                "campaign_id": "", "vendor": "vobiz",
                "service_type": "telephony",
                "qty": rec.get("billable_seconds", 0), "unit": "seconds",
                "cost": round(float(rec.get("total_cost", 0) or 0), 6),
                "currency": vobiz_snap.get("currency", "INR"),
                "actual_or_estimated": "actual",
            })
    _write(COST_LEDGER_FILE, rows)
    _rebuild_daily_rollups(rows)
    return {"rows": len(rows), "vobiz_cdr_matched": len(used_sids)}


def _rebuild_daily_rollups(rows: list[dict]) -> None:
    """Precompute var/daily_rollups.json: {date: {tenant_id: {vendor: cost}, _total}}."""
    roll: dict = {}
    for r in rows:
        day = (r.get("ts") or "")[:10] or "unknown"
        tid = r.get("tenant_id") or "unknown"
        vendor = r.get("vendor") or "unknown"
        cost = float(r.get("cost", 0) or 0)
        roll.setdefault(day, {})
        roll[day].setdefault(tid, {})
        roll[day][tid][vendor] = round(roll[day][tid].get(vendor, 0.0) + cost, 6)
        roll[day].setdefault("_total", 0.0)
        roll[day]["_total"] = round(roll[day]["_total"] + cost, 6)
    _write(DAILY_ROLLUPS_FILE, roll)


def _cost_rows_for(tenant: dict) -> list[dict]:
    """Tenant-scoped cost_ledger rows (admin sees all)."""
    rows = _read_cost_ledger()
    if tenant.get("is_admin"):
        return rows
    tid = tenant.get("tenant_id")
    return [r for r in rows if r.get("tenant_id") == tid]


async def vendor_sync(now: datetime | None = None) -> dict:
    """~30-min vendor-sync pass. Pulls ElevenLabs analytics + (when keyed) Vobiz CDR for
    the last 24-48h, writes var/vendor_snapshots.json with per-vendor status/timestamp/
    stale flag. On failure: keep last snapshot, mark stale, continue. Then rebuilds the
    cost ledger + rollups. Best-effort; never raises."""
    now = now or now_ist()
    ts = now.isoformat(timespec="seconds")
    try:
        snaps = _read(VENDOR_SNAPSHOTS_FILE, {}) or {}
        if not isinstance(snaps, dict):
            snaps = {}

        # ElevenLabs: subscription + usage-over-time (workspace level).
        if v_elevenlabs is not None:
            try:
                st = v_elevenlabs.status()
                if st == "configured":
                    import time as _t
                    end_ms = int(_t.time() * 1000)
                    start_ms = end_ms - 48 * 3600 * 1000
                    sub = v_elevenlabs.get_subscription()
                    usage = v_elevenlabs.get_usage_over_time(start_ms, end_ms)
                    snaps["elevenlabs"] = {"status": "configured", "synced_at": ts, "stale": False,
                                           "subscription": sub, "usage": usage}
                else:
                    prev = snaps.get("elevenlabs", {})
                    snaps["elevenlabs"] = {**prev, "status": st, "synced_at": prev.get("synced_at", ""),
                                           "stale": st != "configured"}
            except Exception as exc:  # noqa: BLE001
                prev = snaps.get("elevenlabs", {})
                snaps["elevenlabs"] = {**prev, "status": "error", "error": repr(exc)[:120],
                                       "stale": True}

        # Vobiz: CDR + balance (dormant until creds).
        if v_vobiz is not None:
            try:
                st = v_vobiz.status()
                if st == "configured":
                    start_d = (now - timedelta(hours=48)).strftime("%Y-%m-%d")
                    end_d = now.strftime("%Y-%m-%d")
                    cdr = v_vobiz.get_cdr(start_d, end_d)
                    bal = v_vobiz.get_balance()
                    ok = cdr.get("status") == "configured"
                    snaps["vobiz"] = {"status": "configured" if ok else cdr.get("status", "error"),
                                      "synced_at": ts, "stale": not ok,
                                      "records": cdr.get("records", []) if ok else snaps.get("vobiz", {}).get("records", []),
                                      "currency": cdr.get("currency", "INR"),
                                      "balance": bal.get("balance") if bal.get("status") == "configured" else None}
                else:
                    prev = snaps.get("vobiz", {})
                    snaps["vobiz"] = {**prev, "status": st, "stale": True,
                                      "synced_at": prev.get("synced_at", ""),
                                      "records": prev.get("records", [])}
            except Exception as exc:  # noqa: BLE001
                prev = snaps.get("vobiz", {})
                snaps["vobiz"] = {**prev, "status": "error", "error": repr(exc)[:120], "stale": True}

        # Groq / Sarvam: internal meters (no external API) — mark configured + estimated.
        snaps["groq"] = {"status": "configured", "synced_at": ts, "stale": False, "estimated": True}
        snaps["sarvam"] = {"status": "configured", "synced_at": ts, "stale": False, "estimated": True}
        snaps["livekit"] = {"status": "configured", "synced_at": ts, "stale": False, "note": "self-hosted=free"}

        async with _STORE_LOCK:
            _write(VENDOR_SNAPSHOTS_FILE, snaps)
        # Rebuild derived ledger/rollups from the fresh snapshot + usage events.
        rebuild_cost_ledger()
        return snaps
    except Exception as exc:  # noqa: BLE001
        try:
            import logging as _lg
            _lg.getLogger("famit-caller").warning("vendor_sync failed: %r", exc)
        except Exception:  # noqa: BLE001
            pass
        return {}


# ---------- P0.5 scheduler: ONE 60s loop (retry/callback dispatch + opt-out safety sweep) ----------
def _spawn_retry_job(r: dict) -> str:
    """Build a single-lead JOBS entry mirroring /run and kick run_job. No new dial code path."""
    tenant = _tenant_by_id(r["tenant_id"]) or {}
    jid = uuid.uuid4().hex[:10]
    JOBS[jid] = {
        "state": "queued", "campaign_id": r["campaign_id"], "tenant_id": r["tenant_id"],
        "concurrency": 1, "hourly_cap": 200, "daily_cap": int(tenant.get("daily_call_cap", 500)),
        "leads": [{"name": r.get("name", ""), "num": r["phone"], "status": "queued",
                   "room": "", "launched_at": 0.0, "attempt": int(r.get("attempts", 0))}],
        "retry_of": r.get("id"),
    }
    asyncio.create_task(run_job(jid))
    return jid


async def scheduler_loop():
    while True:
        try:
            await asyncio.sleep(60)
            now_iso = now_ist().isoformat()
            # WAVE A Unit1: ingest the agent's per-room usage files into usage_events.json.
            await _drain_usage_raw()
            # WAVE A Unit3: vendor-sync pass every ~30 min (ElevenLabs + Vobiz-when-keyed),
            # then rebuild cost_ledger + rollups. Cheap rebuild also runs each tick so
            # internally-metered cost shows up promptly after a call.
            global _LAST_VENDOR_SYNC
            _now_t = time.time()
            if _now_t - _LAST_VENDOR_SYNC >= 1800:
                _LAST_VENDOR_SYNC = _now_t
                await vendor_sync()
            else:
                try:
                    rebuild_cost_ledger()
                except Exception:  # noqa: BLE001
                    pass
            # 🚨 KILL-SWITCH: only DIAL due retries/callbacks when explicitly enabled.
            # Default OFF → zero redials go out while the retry engine is rebuilt. The
            # reconciliation/classify sweep below still runs (outcomes/scores stay correct).
            # ── W10 smart cadence dial (CALLBACK_CADENCE_ENABLED + RETRY_SCHEDULER_ENABLED): fire_due
            #    owns due-check + max-retries cap + terminal skip + per-lead lock + DND + priority
            #    order (anti-runaway: a pickup -> CALLED -> never redials; attempts monotonic; hard
            #    EXPIRE after max_retries). When armed it REPLACES the flat-file dial below.
            if RETRY_SCHEDULER_ENABLED and _CB_STORE is not None and _cb_fire_due is not None:
                try:
                    for job in await _cb_fire_due(store=_CB_STORE, config=_CB_CFG, bus=_get_event_bus()):
                        camp_fields = (get_campaign(job.campaign_id) or {}).get("fields", {}) or {}
                        if not _in_window(camp_fields)[0]:
                            continue                               # respect calling window
                        if norm(job.phone) in _suppressed_set(job.tenant_id):
                            try:
                                await _cb_release(_CB_STORE, job.tenant_id, job.phone)
                            except Exception:  # noqa: BLE001
                                pass
                            continue                               # opted out since enqueue
                        # NCPR / DND scrub-before-dial (flag-gated, fail-closed): a hit OR an
                        # unscrubbed (cache-miss) number is NOT dialed — released for re-scrub.
                        if _NCPR_SCRUBBER is not None:
                            try:
                                _scrub = _NCPR_SCRUBBER.scrub(job.tenant_id, job.phone)
                            except Exception:  # noqa: BLE001 — fail-closed on a Tier-A check error
                                _scrub = None
                            if _scrub is None or _scrub.block:
                                try:
                                    await _cb_release(_CB_STORE, job.tenant_id, job.phone)
                                except Exception:  # noqa: BLE001
                                    pass
                                continue                           # NCPR-listed / unscrubbed -> skip
                        _spawn_retry_job({
                            "id": f"{job.tenant_id}:{job.phone}",
                            "tenant_id": job.tenant_id, "campaign_id": job.campaign_id,
                            "phone": job.phone, "name": "", "attempt": getattr(job, "attempt", 0),
                            "reason": getattr(job, "reason", "callback"),
                            "recap": getattr(job, "last_summary", ""),
                        })
                except Exception as _cbexc:  # noqa: BLE001 — cadence dial can NEVER kill the loop
                    try:
                        import logging as _lg_cb
                        _lg_cb.getLogger("w10.cadence").warning("fire_due tick failed: %r", _cbexc)
                    except Exception:  # noqa: BLE001
                        pass
            elif RETRY_SCHEDULER_ENABLED:
                # LEGACY flat-file dial (cadence OFF). Unchanged behavior.
                due = [r for r in _read(RETRY_FILE, []) if r.get("next_attempt_at", "") <= now_iso]
                for r in due:
                    camp_fields = (get_campaign(r["campaign_id"]) or {}).get("fields", {}) or {}
                    if not _in_window(camp_fields)[0]:
                        continue                                   # respect calling window
                    if norm(r["phone"]) in _suppressed_set(r["tenant_id"]):
                        await _remove_retry(r["id"]); continue     # opted out since enqueue
                    _spawn_retry_job(r)
                    await _remove_retry(r["id"])
            # Reconciliation sweep: the agent writes the transcript on shutdown, which can
            # LAG run_job's finalize (which then read an empty transcript -> misclassified as
            # no_human/0). Re-reconcile any done call whose transcript now has real data but
            # whose rec wasn't enriched yet (no _reconciled flag). Fixes classify + score +
            # retry/callback + opt-out for late transcripts.
            calls_dirty = False
            for c in list(CALLS):
                room = c.get("room", "")
                if (not room or c.get("status") != "done" or c.get("outcome") == "suppressed"
                        or c.get("_reconciled")):
                    continue
                tr = _read(TRANSCRIPT_DIR / f"{room}.json", {})
                if not tr:
                    continue   # transcript still not written; try again next tick
                tid = c.get("tenant_id", ADMIN_ID)
                phone = c.get("phone", "")
                cid = c.get("campaign_id", "")
                outcome = _classify_outcome(c, tr)
                c["outcome"] = outcome
                c["interest"] = tr.get("interest", 0)
                c["answered"] = outcome in _REAL_CONVO
                c["_reconciled"] = True
                calls_dirty = True
                await _update_lead_after_call(tid, phone, tr.get("interest", 0), outcome,
                                              call_at=c.get("started_at", ""))
                if (tr.get("opt_out") or tr.get("outcome") == "opt_out") and _caller_opted_out(tr):
                    c["outcome"] = "opt_out"; c["answered"] = True
                    if norm(phone) not in _suppressed_set(tid):
                        await _add_suppression(tid, phone, "opt_out_call", source=room)
                        await _flip_lead_status(tid, phone, "opted_out")
                else:
                    camp_fields = (get_campaign(cid) or {}).get("fields", {}) or {}
                    # ── W10 recon enqueue (CALLBACK_CADENCE_ENABLED): idempotent + monotonic — a recon
                    #    tick can NEVER reset attempts or re-enqueue an answered/opted-out lead.
                    _cb_recon_owned = False
                    if _CB_STORE is not None and _cb_enqueue_smart is not None:
                        try:
                            await _cb_enqueue_smart(
                                tid, cid, c, tr, outcome, 0, camp_fields,
                                store=_CB_STORE, config=_CB_CFG, bus=_get_event_bus(),
                                from_reconcile=True,
                            )
                            _cb_recon_owned = True
                        except Exception:  # noqa: BLE001
                            _cb_recon_owned = False
                    if not _cb_recon_owned:
                        maxa = int(camp_fields.get("retry_max_attempts", 3))
                        backoff = camp_fields.get("retry_backoff_mins") or [120, 360, 1440]
                        cb = tr.get("callback_at")
                        already = any(r.get("phone") == norm(phone) and r.get("campaign_id") == cid
                                      for r in _read(RETRY_FILE, []))
                        if cb and not already:
                            await _enqueue_retry(tid, cid, c.get("name", ""), phone, 0, maxa, cb, "callback")
                            await _emit_webhook(tid, "callback.scheduled",
                                                {"phone": phone, "campaign_id": cid,
                                                 "name": c.get("name", ""), "when": cb,
                                                 "callback_raw": tr.get("callback_raw", "")})
                        elif outcome in ("no_answer", "voicemail", "busy") and not already:
                            next_at = _clamp_to_window(now_ist() + timedelta(minutes=backoff[0]), camp_fields)
                            await _enqueue_retry(tid, cid, c.get("name", ""), phone, 1, maxa,
                                                 next_at.isoformat(), outcome)
                # WAVE3 Unit2: emit completion + qualified for late-reconciled calls.
                # Only if _finalize_call hadn't already emitted (it emits with an empty
                # transcript; here we re-emit ONCE with the real summary/score).
                if not c.get("_wh_completed"):
                    _sc = c.get("interest", 0) or 0
                    c["_wh_completed"] = True
                    await _emit_webhook(tid, "call.completed",
                                        {"call_id": c.get("id"), "phone": phone, "name": c.get("name", ""),
                                         "campaign_id": cid, "outcome": c["outcome"], "interest": _sc,
                                         "score": _sc, "summary": tr.get("summary", ""),
                                         "duration_s": c.get("duration_s", 0), "room": room})
                    if c["outcome"] != "opt_out" and _sc >= 70:
                        await _emit_webhook(tid, "lead.qualified",
                                            {"call_id": c.get("id"), "phone": phone,
                                             "name": c.get("name", ""), "campaign_id": cid,
                                             "score": _sc, "outcome": c["outcome"],
                                             "summary": tr.get("summary", "")})
            if calls_dirty:
                async with _STORE_LOCK:
                    _write(CALLS_FILE, CALLS)
        except Exception as exc:  # noqa: BLE001
            try:
                import logging as _lg
                _lg.getLogger("famit-caller").warning("scheduler_loop tick failed: %r", exc)
            except Exception:  # noqa: BLE001
                pass


@app.on_event("startup")
async def _start_scheduler():
    asyncio.create_task(scheduler_loop())
    # CRM CORE: single-source the phone canonicalizer + lazily ensure the projection schema.
    # Best-effort — a crm failure must NEVER affect the scheduler / run-path.
    if _crm_mod is not None:
        try:
            _crm_mod.init(norm=norm)
        except Exception:  # noqa: BLE001
            pass


# ---------- P0.6 hot-leads convenience ----------
@app.get("/leads/hot")
async def leads_hot(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    rows = [x for x in _leads_for(t) if (x.get("score", 0) or 0) >= 70]
    rows.sort(key=lambda x: x.get("score", 0), reverse=True)
    return JSONResponse({"leads": rows})


@app.get("/")
async def index(request: Request):
    if not authed(request):
        return need_auth()
    return HTMLResponse("<h3>Famit backend API is running. Use the panel.</h3>")


# ==============================================================================================
# MODULE MOUNT — ads-engine (autonomous paid-ads engine, prefix /ads). ADDITIVE, FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# The ads_engine router is BARE-OK / TOKEN-DERIVED: its handlers look up
# caller.resolve_tenant / need_auth / can LAZILY (ads_engine.endpoints._auth_helpers) and ALWAYS
# take org_id from the token (t["tenant_id"]) — NEVER from the request body/header. So a plain
# include_router is safe (no cross-tenant hole; no build_router needed — unlike booking/media-gen/
# funnels which read tenant from body and require a token-deriving surface first).
#
# IMPORT-GUARD (house pattern): a missing/broken ads_engine package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_ADS!=1/true => router NOT mounted => byte-identical behavior.
# No scheduler poll_and_enforce tick is wired here (DEFERRED per the build state).
try:
    from ads_engine.endpoints import router as _ads_router  # noqa: E402
except Exception:  # noqa: BLE001
    _ads_router = None

FEATURE_ADS = (cfg_get("FEATURE_ADS", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_ADS and _ads_router is not None:
    try:
        app.include_router(_ads_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_ads
        _lg_ads.getLogger("famit-caller").warning("ads_engine router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — Haptica Grow (Revenue-Truth Signal Loop: L5 scoring + L7 CAPI, prefix /grow).
# ----------------------------------------------------------------------------------------------
# Grow ships ONLY a token-deriving AUTHENTICATED surface (grow.endpoints.build_router): tenant_id
# is ALWAYS resolve_tenant(request)["tenant_id"] — NEVER from the request body (no cross-tenant
# hole). The package is ALSO imported as `_grow_mod` so the _finalize_call hook (above) can feed
# each completed call's outcome into the Signal Loop (scored lead + CAPI Lead/QualifiedLead).
#
# IMPORT-GUARD (house pattern): a missing/broken grow package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_GROW!=1/true => router NOT mounted + the finalize hook is
# inert => byte-identical behavior. Even when ON, the L7 signals stay in SHADOW (log the
# would-send CAPI payload, POST nothing) until Meta CAPI creds + GROW_SIGNALS_LIVE=1 are set
# (founder-gated) — so mounting carries zero live-spend / bad-upload risk.
try:
    import grow as _grow_mod  # noqa: E402  (also referenced by the _finalize_call hook)
    from grow.endpoints import build_router as _build_grow_router  # noqa: E402
except Exception:  # noqa: BLE001
    _grow_mod = None
    _build_grow_router = None

FEATURE_GROW = (cfg_get("FEATURE_GROW", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_GROW and _build_grow_router is not None:
    try:
        _grow_router = _build_grow_router(resolve_tenant, can, need_auth, _forbidden,
                                          firewall=_firewall_mod)
        if _grow_router is not None:
            app.include_router(_grow_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_grow
        _lg_grow.getLogger("famit-caller").warning("grow router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — media-gen (provider-agnostic media-generation ENGINE, prefix /media). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: media-gen is a BODY-TENANT module — its bare module-level `router`
# (`media_gen.router.router`) reads tenant_id from the request body/query and MUST NEVER be mounted
# (mounting it is a cross-tenant hole: any caller could pass tenant_id=<victim>). We mount the
# token-deriving AUTHENTICATED surface `build_router(resolve_tenant, can, need_auth, forbidden,
# firewall)` instead — tenant_id is ALWAYS resolve_tenant(request)["tenant_id"] (token-derived),
# writes enforce can(t,"write"), and by-job_id routes verify rec["tenant_id"]==token (ownership).
# Mirrors the workflow-studio / forms-surveys settled pattern (and unlike ads-engine, which was
# bare-OK / token-derived and needed no build_router).
#
# The video webhook (/media/video/webhook) is intentionally UNAUTHENTICATED on this surface —
# it is provider-signed (providers.verify_webhook, fail-closed) and matched by external_id.
#
# IMPORT-GUARD (house pattern): a missing/broken media_gen package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_MEDIA!=1/true => router NOT mounted => byte-identical behavior.
# firewall is passed for forward-compat (reserved for future step-up on approve; not wired yet).
try:
    from media_gen.router import build_router as _build_media_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_media_router = None

FEATURE_MEDIA = (cfg_get("FEATURE_MEDIA", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_MEDIA and _build_media_router is not None:
    try:
        _media_router = _build_media_router(
            resolve_tenant, can, need_auth, _forbidden, firewall=_firewall_mod
        )
        if _media_router is not None:
            app.include_router(_media_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_media
        _lg_media.getLogger("famit-caller").warning("media_gen router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — pmodel (2D -> 3D Property Studio, prefix /pmodel). FLAG-GATED, default OFF.
# ----------------------------------------------------------------------------------------------
# Turns a 2D floor plan (or a text brief / built-in sample) into an interactive 3D property model
# customers can explore via a public share link — a sales asset the voice agent can hand off.
# build_router injects resolve_tenant/can/need_auth/_forbidden and derives the tenant from the
# TOKEN ONLY (never body/query). The single public surface is the share-by-token read. Import-
# guarded + flag-gated => a missing module or disabled flag is byte-identical to before.
try:
    from pmodel.router import build_router as _build_pmodel_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_pmodel_router = None

FEATURE_PMODEL = (cfg_get("FEATURE_PMODEL", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_PMODEL and _build_pmodel_router is not None:
    try:
        try:
            _pmodel_s3 = _aim_s3()
        except Exception:  # noqa: BLE001 — Spaces optional; uploads degrade to in-record only
            _pmodel_s3 = None
        _pmodel_router = _build_pmodel_router(
            resolve_tenant, can, need_auth, _forbidden,
            s3_client=_pmodel_s3,
            spaces_bucket=(cfg_get("AIM_SPACES_BUCKET", "") or "").strip(),
            presign=_rec_presign,
            audit=_audit,
        )
        if _pmodel_router is not None:
            app.include_router(_pmodel_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_pmodel
        _lg_pmodel.getLogger("famit-caller").warning("pmodel router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — Haptica Flywheel (the RLHF/RLAIF self-improvement engine, prefix /flywheel). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# Every call becomes fuel: a side-pipeline that captures the (state, move, reward) trajectory of each
# finalized call (via the _finalize_call hook above + the dispatch arm-stamp in run_job), scores a
# fused & provenance-stamped reward, distributes it across turns (credit assignment → "which move is
# +/-"), mines a proprietary (chosen,rejected) preference MOAT, and proposes GATED challengers that a
# HUMAN approves in the super-admin console before promotion. The /flywheel/* router is the console
# read/approve surface. `_flywheel_mod` (droplet glue: finalize hook + dispatch arm) is imported
# UNCONDITIONALLY so the run_job/finalize hooks can reference it, but every hook is itself gated by
# FLYWHEEL_ENABLED — so with the flag OFF the resting behaviour is byte-identical (no capture, no
# router, no dispatch change). build_router injects resolve_tenant/can/need_auth/_forbidden +
# require_super_admin + _audit, derives the tenant from the TOKEN, and self-returns None when FastAPI
# is absent. The heavy offline loop runs in a SEPARATE process (flywheel_worker.py), never here.
try:
    import flywheel_app as _flywheel_mod  # noqa: E402 — droplet glue (finalize hook + dispatch arm)
except Exception:  # noqa: BLE001
    _flywheel_mod = None
try:
    from voice_ops.flywheel.router import build_router as _build_flywheel_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_flywheel_router = None

FLYWHEEL_ENABLED = (cfg_get("FLYWHEEL_ENABLED", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FLYWHEEL_ENABLED and _build_flywheel_router is not None:
    try:
        _flywheel_router = _build_flywheel_router(
            resolve_tenant, can, need_auth, _forbidden,
            require_super_admin=require_super_admin, audit=_audit,
        )
        if _flywheel_router is not None:
            app.include_router(_flywheel_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_flywheel
        _lg_flywheel.getLogger("famit-caller").warning("flywheel router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — provider-registry (the universal AI/connector registry, prefix /providers). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: provider-registry ships ONLY a token-deriving AUTHENTICATED surface
# `build_router(resolve_tenant, can, need_auth, forbidden, require_super_admin=, firewall=, audit=)`
# — there is NO body-tenant bare router at all. tenant_id is ALWAYS resolve_tenant(request)["tenant_id"]
# (token-derived), writes enforce can(t,"write"), the /providers/admin/* surface is gated by
# require_super_admin (which EXCLUDES the legacy static password — control-security #1), and a
# provider-key REVEAL requires a firewall.consume_reveal_step_up single-use PIN step-up (the W3 reveal
# scope). A self-hosted base_url is SSRF-validated before a def can be created / tested. The
# request/response field-maps of a custom_field_map def are JSONPath-validated (no-eval) at write-time.
#
# IMPORT-GUARD (house pattern): a missing/broken provider_registry package can NEVER break startup —
# build_router degrades to None when FastAPI/crypto are absent, and any import error is swallowed.
# FEATURE FLAG default OFF: PROVIDER_REGISTRY_ENABLED!=1/true => router NOT mounted => byte-identical
# behavior. Even if mounted, EVERY route self-404s while the flag is OFF (defense in depth) — so the
# resting state is byte-identical whether the flag-check happens here at mount or per-route.
# This is the W4 caller.py mount; serialized against RAG/Vault/Video (only ONE edits caller.py at a time).
try:
    from provider_registry.endpoints import build_router as _build_provreg_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_provreg_router = None

PROVIDER_REGISTRY_ENABLED = (cfg_get("PROVIDER_REGISTRY_ENABLED", "0") or "0").strip().lower() \
    in ("1", "true", "yes", "on")

if PROVIDER_REGISTRY_ENABLED and _build_provreg_router is not None:
    try:
        _provreg_router = _build_provreg_router(
            resolve_tenant, can, need_auth, _forbidden,
            require_super_admin=require_super_admin, firewall=_firewall_mod, audit=_audit,
        )
        if _provreg_router is not None:
            app.include_router(_provreg_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_provreg
        _lg_provreg.getLogger("famit-caller").warning("provider_registry router mount failed",
                                                       exc_info=True)


# ==============================================================================================
# MODULE MOUNT — trunk-registry (own/flexible telephony SIP-trunk registry, prefix /trunk-registry).
# FLAG-GATED (TRUNK_REGISTRY_ENABLED, default OFF => byte-identical resting). TELEPHONY-INDEPENDENCE
# -PLAN §5 T3. This rides caller.py (famit-caller), NEVER agent.py (the earner). At T3 the routes are
# mounted but the strangler dial-loop cut (T5) is NOT wired — so flag-OFF the live `TRUNK` env dial
# path is byte-identical, and flag-ON the routes only manage trunk ROWS (no campaign auto-dial here).
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: trunk-registry ships ONLY a token-deriving AUTHENTICATED surface
# `build_router(resolve_tenant, can, need_auth, forbidden, require_super_admin=, firewall=, audit=)`
# — a TWIN of provider-registry. tenant_id is ALWAYS resolve_tenant(request)["tenant_id"]
# (token-derived), writes enforce can(t,"write"), the /trunk-registry/admin/* surface is gated by
# require_super_admin (which EXCLUDES the legacy static password — control-security #1). A BYO
# sip_host is SSRF-validated before a trunk can be created. A SIP-password REVEAL requires a
# firewall.consume_reveal_step_up single-use PIN step-up. RED-TEAM D: DELETE default = soft-disable;
# a hard-delete REFUSES a `_global`/env-protected trunk + is PIN-gated. RED-TEAM F: /test-call is
# rate-limited (<=3/hr/trunk) + founder-typed destination (the ONLY non-campaign originate; NEVER an
# auto-dial — at T3 it returns a dial-intent only). RED-TEAM E: /quarantine-did is the kill switch.
#
# IMPORT-GUARD (house pattern): a missing/broken trunk_registry package can NEVER break startup —
# build_router degrades to None when FastAPI is absent, and any import error is swallowed.
# FEATURE FLAG default OFF: TRUNK_REGISTRY_ENABLED!=1/true => router NOT mounted => byte-identical
# behavior. Even if mounted, EVERY route self-404s while the flag is OFF (defense in depth).
# Serialized against RAG/Vault/Video/provider-registry (only ONE edits caller.py at a time).
try:
    from trunk_registry.endpoints import build_router as _build_trunkreg_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_trunkreg_router = None

TRUNK_REGISTRY_ENABLED = (cfg_get("TRUNK_REGISTRY_ENABLED", "0") or "0").strip().lower() \
    in ("1", "true", "yes", "on")

if TRUNK_REGISTRY_ENABLED and _build_trunkreg_router is not None:
    try:
        _trunkreg_router = _build_trunkreg_router(
            resolve_tenant, can, need_auth, _forbidden,
            require_super_admin=require_super_admin, firewall=_firewall_mod, audit=_audit,
        )
        if _trunkreg_router is not None:
            app.include_router(_trunkreg_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_trunkreg
        _lg_trunkreg.getLogger("famit-caller").warning("trunk_registry router mount failed",
                                                        exc_info=True)


# ==============================================================================================
# MODULE MOUNT — VIDEO STUDIO (campaign->AI-script->BATCH->Asset Library, prefix /creative/video).
# FLAG-GATED (FEATURE_VIDEO_STUDIO, default OFF => byte-identical resting). VIDEO-STUDIO-MASTER-PLAN
# §8/U4/W8. Video is ASYNC by construction -> adds ZERO to the voice loop; this rides caller.py, NEVER
# agent.py (the earner). The composite tier renders on the Hatchet worker, never in caller.py.
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: video_studio ships ONLY a token-deriving AUTHENTICATED surface
# `build_router(resolve_tenant, can, need_auth, forbidden, list_campaigns=, bridge=)` — the old
# body-`tenant` scaffold surface (a cross-tenant hole) is gone. tenant_id is ALWAYS
# resolve_tenant(request)["tenant_id"] (token-derived); writes enforce can(t,"write"); batch
# ownership is verified per-id (a cross-tenant batch_id -> error:no_such_batch). The submit_gate
# (H1/H2) forces the 1-paid-test choke + per-tenant VIDEO_DAILY_CAP_USD inside propose_batch.
#
# LIBRARY BRIDGE (§5): collect_batch lands finished videos in the ai_asset_* PG library (where
# images already live) via `_video_library_bridge` -> POST the ai_asset internal register-video over
# the VPC loopback, authed with AIASSET_SERVICE_TOKEN. Best-effort: a bridge miss leaves the asset in
# the JSON fallback + logs, never breaks a batch. Dormant until AIASSET_SERVICE_TOKEN + the loopback.
#
# IMPORT-GUARD: a missing/broken video_studio package can NEVER break startup.
def _video_library_bridge(payload: dict) -> dict:
    """POST the ai_asset internal register-video over the VPC loopback (§5). Never raises -> {} on miss.
    Dormant until AIASSET_SERVICE_TOKEN is set (then the studio's finished videos land in the live
    ai_asset_* library). Earner-safe: a short-timeout HTTP call off the request path, never agent.py."""
    try:
        base = (cfg_get("AIASSET_LOOPBACK_BASE", "http://10.122.0.4:8310") or "").rstrip("/")
        token = (cfg_get("AIASSET_SERVICE_TOKEN", "") or "").strip()
        if not base or not token:
            return {}
        with httpx.Client(timeout=8.0) as _c:
            r = _c.post(f"{base}/assets/_internal/register-video",
                        json=payload, headers={"Authorization": f"Bearer {token}"})
            if r.status_code == 200:
                return r.json() if r.headers.get("content-type", "").startswith("application/json") else {}
    except Exception:  # noqa: BLE001 — bridge is best-effort; never break a collect
        pass
    return {}


try:
    from creative.video_studio.endpoints import build_router as _build_video_studio_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_video_studio_router = None

FEATURE_VIDEO_STUDIO = (cfg_get("FEATURE_VIDEO_STUDIO", "0") or "0").strip().lower() \
    in ("1", "true", "yes", "on")

if FEATURE_VIDEO_STUDIO and _build_video_studio_router is not None:
    try:
        _video_studio_router = _build_video_studio_router(
            resolve_tenant, can, need_auth, _forbidden,
            list_campaigns=list_campaigns, bridge=_video_library_bridge,
        )
        if _video_studio_router is not None:
            app.include_router(_video_studio_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_vstudio
        _lg_vstudio.getLogger("famit-caller").warning("video_studio router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — booking (appointments / site-visits / reminders ENGINE, prefix /booking). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: booking is a HEADER-TENANT module on its bare surface — its module-level
# `router` + default `get_ctx` trust the `X-Tenant-Id` header (spoofable) and MUST NEVER be mounted.
# We mount the token-deriving AUTHENTICATED surface `build_router(resolve_tenant, can, need_auth,
# forbidden, firewall)` instead — tenant_id is ALWAYS resolve_tenant(request)["tenant_id"]
# (token-derived), writes enforce can(t,"write") / reads enforce can(t,"read"), and is_admin is
# HARDCODED False into every core.* call (is_admin feeds db.engine.session(tenant_id, is_admin);
# is_admin=1 BYPASSES RLS — so it must never be body/header/claim-derived). This supersedes the
# old "override get_ctx" instruction in REMAINING_MODULES_BUILD_STATE.md — a clean token-deriving
# surface now exists. Mirrors the workflow-studio / forms-surveys / media-gen settled pattern.
#
# DORMANT-SAFE: booking core returns {"status":"not_configured"} when Postgres is down/unconfigured,
# so every endpoint is inert until F1 lands + the deferred Alembic 0003_booking migration + rls.sql
# apply (NOT part of this mount). The risky /tick reminder spend flows through core.tick's own
# firewall(PIN, fail-closed) + wallet gates with the body-supplied pin (only tenant/is_admin move to
# the token; pin legitimately stays in the body).
#
# IMPORT-GUARD (house pattern): a missing/broken booking package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_BOOKING!=1/true => router NOT mounted => byte-identical behavior.
# firewall is passed for signature-uniformity (reserved seam; risky tick spend self-gates in core).
try:
    from booking.router import build_router as _build_booking_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_booking_router = None

FEATURE_BOOKING = (cfg_get("FEATURE_BOOKING", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_BOOKING and _build_booking_router is not None:
    try:
        _booking_router = _build_booking_router(
            resolve_tenant, can, need_auth, _forbidden, firewall=_firewall_mod
        )
        if _booking_router is not None:
            app.include_router(_booking_router)
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_booking
        _lg_booking.getLogger("famit-caller").warning("booking router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — payments (vendor->customer collections: links/invoices/receipts, prefix /payments).
# FLAG-GATED. THIS IS THE LIVE EARNER — mount is dormant-by-flag, byte-identical when OFF.
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: payments ships a CLEAN token-deriving surface — there is no body/header-tenant
# bare router to avoid (unlike media-gen/booking). The module-level `payments.router.router` resolves
# tenant ONLY from the token via the injected resolve_tenant, enforces can(t,"write") on mutating
# routes (create-link / mark-paid / refund), and ALWAYS uses org_id = the resolved tenant_id (never a
# spoofable body/query param). Spend-sensitive routes additionally pass through firewall.require_step_up
# (PASS-THROUGH when FIREWALL disabled / no PIN — non-breaking; 403 challenge when active). The
# /payments/webhooks/{provider} route is intentionally UNAUTHENTICATED — it is provider-signature-
# verified inside core.ingest_webhook (a machine call) and always returns 200 (so no provider retry-storm).
#
# WIRE-THEN-INCLUDE: unlike build_router modules, payments injects its auth helpers via wire(...) (which
# mutates module globals) and then mounts the module-level `router` with prefix="/payments". wire() takes
# KEYWORD-ONLY args. The router has NO internal prefix, so the prefix is applied here at include time.
#
# DORMANT-SAFE: with no PG, core.* returns {"status":"unavailable"}; with no provider creds,
# {"status":"not_configured"}. ensure_schema() is LAZY (first-use, never raises) — payments.init() is
# NOT called at startup ON PURPOSE: calling it would touch PG / apply DDL even with the flag OFF and
# break the byte-identical-when-OFF guarantee. init() + drain_followups() (the scheduler dunning tick)
# are DEFERRED (recorded in the build log), exactly as booking deferred its reminder tick.
#
# IMPORT-GUARD (house pattern): a missing/broken payments package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_PAYMENTS!=1/true => router NOT mounted => byte-identical behavior.
try:
    from payments.router import router as _payments_router, wire as _payments_wire  # noqa: E402
except Exception:  # noqa: BLE001
    _payments_router = None
    _payments_wire = None

FEATURE_PAYMENTS = (cfg_get("FEATURE_PAYMENTS", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_PAYMENTS and _payments_router is not None and _payments_wire is not None:
    try:
        _payments_wire(
            resolve_tenant=resolve_tenant, can=can, need_auth=need_auth,
            forbidden=_forbidden, firewall=_firewall_mod,
        )
        app.include_router(_payments_router, prefix="/payments")
    except Exception:  # noqa: BLE001 — a mount failure must never crash the live spine
        import logging as _lg_payments
        _lg_payments.getLogger("famit-caller").warning("payments router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- credits (credit-wallet + buy-credits + service costing matrix, prefix /credits).
# FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# THE CREDIT LAYER: one customer-facing unit (the credit) on top of the existing wallet/billing.json/
# cost_ledger primitives, via a PLUGGABLE BillingEngine (LocalCreditEngine default; FlexpriceEngine
# when BILLING_ENGINE=flexprice). Razorpay + Stripe top-up rails are DORMANT-UNTIL-KEYS. Tenant is
# ALWAYS token-derived (credits.router._tenant -> resolve_tenant); admin writes pass require_super_admin
# + the Action-Firewall step-up. A credits failure can never break a call (everything best-effort).
#
# IMPORT: the package lives at droplet_work/credits/ (a plain top-level package, like ai_manager/comm)
# so `from credits.router import router` resolves with NO sys.path hack on the box and in the repo.
#
# IMPORT-GUARD (house pattern): a missing/broken credits package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_CREDITS!=1/true => router NOT mounted => byte-identical behavior.
try:
    from credits.router import router as _credits_router, wire as _credits_wire  # noqa: E402
except Exception:  # noqa: BLE001
    _credits_router = None
    _credits_wire = None

FEATURE_CREDITS = (cfg_get("FEATURE_CREDITS", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_CREDITS and _credits_router is not None and _credits_wire is not None:
    try:
        _credits_wire(
            resolve_tenant=resolve_tenant, can=can, need_auth=need_auth,
            forbidden=_forbidden, firewall=_firewall_mod,
            require_super_admin=require_super_admin, audit=_audit,
            step_up_guard=_step_up_guard,
        )
        app.include_router(_credits_router, prefix="/credits")
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_credits
        _lg_credits.getLogger("famit-caller").warning("credits router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- support (AI Customer Support: omnichannel ticketing + KB-grounded AI replies,
# prefix /support). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# TENANT ISOLATION: support ships a CLEAN token-deriving surface (same shape as payments -- there is
# NO body/header-tenant bare router to avoid, unlike media-gen/booking). Every tenant route resolves
# tenant ONLY from the token via the injected resolve_tenant; org_id is ALWAYS the resolved tenant_id
# (NEVER a spoofable body/query param). Mutating routes (inbound / draft / reply / escalate / claim /
# resolve) enforce can(t,"write"). support is NOT a money path, so there is no blanket spend step-up;
# the ONLY step-up-gated route is /support/tickets/{id}/resolve (a risky human force-close, scope
# "support_override") -- PASS-THROUGH when FIREWALL disabled / no PIN (non-breaking), 403 when active.
# The /support/webhooks/{channel} route is intentionally UNAUTHENTICATED (a machine call): today it is
# a dormant no-op returning {"status":"not_configured"} (200), so providers never retry-storm; the
# channel adapter + tenant binding + signature verify are DEFERRED (Omnichannel Inbox unit).
#
# WIRE-THEN-INCLUDE (identical to payments): support injects its auth helpers via wire(...) (mutates
# module globals) and then mounts the module-level `router` with prefix="/support". wire() takes
# KEYWORD-ONLY args. The router has NO internal prefix, so the prefix is applied here at include time.
#
# DORMANT-SAFE: with no PG, core.* returns {"status":"unavailable"}; with an empty KB / no LLM creds,
# the deterministic extractive KB draft fires (grounded-or-escalate -- never hallucinates).
# ensure_schema() is LAZY (first-use, never raises) -- support.init() is NOT called at startup ON
# PURPOSE: calling it would touch PG / apply DDL even with the flag OFF and break the
# byte-identical-when-OFF guarantee. init() is DEFERRED (gate it INSIDE the flag-on block when
# activated), exactly as payments/booking deferred theirs.
#
# IMPORT-GUARD (house pattern): a missing/broken support package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_SUPPORT!=1/true => router NOT mounted => byte-identical behavior.
try:
    from support.router import router as _support_router, wire as _support_wire  # noqa: E402
except Exception:  # noqa: BLE001
    _support_router = None
    _support_wire = None

FEATURE_SUPPORT = (cfg_get("FEATURE_SUPPORT", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_SUPPORT and _support_router is not None and _support_wire is not None:
    try:
        _support_wire(
            resolve_tenant=resolve_tenant, can=can, need_auth=need_auth,
            forbidden=_forbidden, firewall=_firewall_mod,
        )
        app.include_router(_support_router, prefix="/support")
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_support
        _lg_support.getLogger("famit-caller").warning("support router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- forms-surveys (Form/Lead-Capture builder + Survey/Feedback engine). FLAG-GATED.
# Authed CRUD at /forms*, PUBLIC token routes at /f/{token}. No router prefix (paths are absolute).
# ----------------------------------------------------------------------------------------------
# TENANT ISOLATION: forms-surveys ships a CLEAN token-deriving surface (the build_router pattern --
# same shape as media-gen/booking, NOT the wire() shape). Authed CRUD routes resolve tenant ONLY
# from the token via the injected resolve_tenant; org_id is ALWAYS the resolved tenant_id (NEVER a
# spoofable body/query param), writes enforce can(t,"write") / reads can(t,"read"), and is_admin is
# token-derived (feeds db.engine.session -> RLS). The PUBLIC routes /f/{public_token} (render) and
# /f/{public_token}/submit are intentionally UNAUTHENTICATED by design: there is no authenticated
# tenant on the public path, so org_id is SERVER-DERIVED from the form record resolved by the
# unguessable public_token (secrets.token_urlsafe) -- never a request param. Anti-abuse on that
# unauth endpoint is mandatory and lives in the router/core (per-(token,IP) rate limit, raw-body
# size cap pre-parse, honeypot silent-drop, allow-list schema validation, sha256(ip+token)
# forensics, tenant-scoped audit). The two existing @app.middleware("http") handlers are rate-limit
# + metrics ONLY (no global auth wall), so the public /f/ routes are reachable as designed.
#
# SIGNATURE: build_router(resolve_tenant, can, need_auth, forbidden, *, ratelimit=None, audit=None).
# NOTE -- UNLIKE media-gen/booking, forms' build_router has NO `firewall` param (forms are FREE: no
# spend, no wallet hold). Passing firewall= would TypeError (swallowed by the except -> silently
# mounts nothing). We pass only the 4 positional auth helpers; ratelimit/audit fall back to the
# router's own import-guarded `import ratelimit`/`import audit`.
#
# HYPHENATED PACKAGE: the dir is `forms-surveys` (not a legal Python identifier), so a plain
# `import forms_surveys` cannot find it. We register the package under the alias `forms_surveys`
# in sys.modules via importlib spec_from_file_location (submodule_search_locations=[pkgdir]) so the
# inner relative imports resolve -- self-contained, no dependency on the package's own _bootstrap.
#
# init() DELIBERATELY DEFERRED (protects byte-identical-when-OFF, same as payments/support/booking):
# forms.core.init() calls ensure_schema() which applies DDL when PG is reachable -- calling it would
# touch the live PG even with the flag OFF. The deferred leads-write + workflow-trigger emit hooks
# (init(emit_lead=, emit_workflow=)) are likewise DEFERRED (recorded in the build log). build_router
# stands alone: routes call core.* directly; ensure_schema() is LAZY (first-use, never raises) so
# schema applies only on the first authed call AFTER the flag is turned on.
#
# IMPORT-GUARD (house pattern): a missing/broken forms-surveys package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_FORMS!=1/true => router NOT mounted => byte-identical behavior.
try:
    import importlib.util as _fs_ilu  # noqa: E402
    import os as _fs_os  # noqa: E402
    import sys as _fs_sys  # noqa: E402
    _fs_pkgdir = _fs_os.path.join(_fs_os.path.dirname(_fs_os.path.abspath(__file__)), "forms-surveys")
    if "forms_surveys" in _fs_sys.modules:
        _fs_pkg = _fs_sys.modules["forms_surveys"]
    else:
        _fs_spec = _fs_ilu.spec_from_file_location(
            "forms_surveys", _fs_os.path.join(_fs_pkgdir, "__init__.py"),
            submodule_search_locations=[_fs_pkgdir],
        )
        _fs_pkg = _fs_ilu.module_from_spec(_fs_spec)
        _fs_sys.modules["forms_surveys"] = _fs_pkg
        _fs_spec.loader.exec_module(_fs_pkg)
    _build_forms_router = _fs_pkg.build_router
except Exception:  # noqa: BLE001
    _build_forms_router = None

FEATURE_FORMS = (cfg_get("FEATURE_FORMS", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_FORMS and _build_forms_router is not None:
    try:
        _forms_router = _build_forms_router(resolve_tenant, can, need_auth, _forbidden)
        if _forms_router is not None:
            app.include_router(_forms_router)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_forms
        _lg_forms.getLogger("famit-caller").warning("forms-surveys router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- workflow-studio (durable visual-automation engine, prefix /workflows). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# TENANT ISOLATION: workflow-studio ships a CLEAN token-deriving surface (the build_router pattern --
# same shape as media-gen/booking/forms). We mount ONLY build_router(resolve_tenant, can, need_auth,
# forbidden, firewall) -- tenant_id is ALWAYS resolve_tenant(request)["tenant_id"] (token-derived),
# NEVER a body/query field. Writes enforce can(t,"write"); /workflows/killswitch is admin-only
# (can(t,"manage_tenants")); /workflows/runs/{id}/approve verifies the firewall step-up token bound to
# the authed tenant (sub==tenant_id) inside the approval node. ⚠ We NEVER mount the bare module-level
# `workflow.endpoints.router` (kept for the offline test only): it reads tenant_id from the request
# body so the package can run decoupled, which is the exact cross-tenant hole build_router closes
# (RLS cannot save it -- the store would be invoked with the attacker-supplied tenant).
#
# EVENT BRIDGE: attach_event_bridge(app) is called inside the flag-on block. Today it is a DEFINED-not-
# wired no-op descriptor (zero side effects -- it does NOT touch caller.py's emit points): the real
# spine emit wiring (_emit_webhook / _finalize_call -> dispatch_event) is DEFERRED (Lifecycle-Trigger
# unit), so calling it is safe and satisfies the checklist without changing behavior.
#
# NO init()/ensure_schema TO DEFER (unlike payments/support/forms/booking): workflow-studio has none.
# make_store() defaults to the IN-MEMORY backend unless WORKFLOW_STORE=pg, so even with the flag ON the
# live PG is NOT touched. The Hatchet engine is dormant-until-creds (hatchet_sdk is LAZY-imported inside
# get_hatchet(), never at module top), so the same single interpreter runs in-process when dormant.
#
# IMPORT: the inner package was deployed to /opt/famit-agent/workflow/ (a plain top-level package, like
# ads_engine/media_gen) so `from workflow.endpoints import build_router` resolves with NO sys.path hack;
# this also lets funnels' later `import workflow` resolve for free. Mount workflow BEFORE funnels.
#
# IMPORT-GUARD (house pattern): a missing/broken workflow package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_WORKFLOWS!=1/true => router NOT mounted => byte-identical behavior.
try:
    from workflow.endpoints import build_router as _build_workflow_router  # noqa: E402
    from workflow.events import attach_event_bridge as _wf_attach_event_bridge  # noqa: E402
except Exception:  # noqa: BLE001
    _build_workflow_router = None
    _wf_attach_event_bridge = None

FEATURE_WORKFLOWS = (cfg_get("FEATURE_WORKFLOWS", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_WORKFLOWS and _build_workflow_router is not None:
    try:
        _workflow_router = _build_workflow_router(
            resolve_tenant, can, need_auth, _forbidden, firewall=_firewall_mod
        )
        if _workflow_router is not None:
            app.include_router(_workflow_router)
        if _wf_attach_event_bridge is not None:
            _wf_attach_event_bridge(app)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_workflow
        _lg_workflow.getLogger("famit-caller").warning("workflow-studio router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- ai-manager (voice/chat command-center management API, prefix /ai-manager). FLAG-GATED.
# ----------------------------------------------------------------------------------------------
# TENANT ISOLATION: ai-manager is a BARE-OK / TOKEN-DERIVED module (same class as ads-engine, NOT the
# body/header-tenant class of media-gen/booking/funnels). ai_manager.endpoints derives tenant ONLY from
# the authenticated request via a lazy `import caller; caller.resolve_tenant(request)` -- tenant_id is
# ALWAYS t["tenant_id"] (token), NEVER a body/query field. So a plain app.include_router(router) is safe
# (no cross-tenant hole; no build_router needed). Reads enforce can(t,"read"), writes can(t,"write"),
# grant/revoke require can(t,"manage_tenants") + a firewall step-up. The two SERVICE-TOKEN endpoints
# (/numbers/lookup, POST /sessions) are DORMANT until AIM_SERVICE_TOKEN is set (always 401 otherwise).
#
# NO init()/ensure_schema TO DEFER: ai-manager persists sessions as JSONL on the control plane (no PG /
# no DDL), so there is nothing schema-side to gate -- the live PG is never touched, flag ON or OFF.
# The LiveKit voice front (inbound_agent.py) + SIP dispatch are a SEPARATE later wire (do NOT pass
# through caller.py); this mount is the management/HTTP surface only.
#
# IMPORT: the package was deployed to /opt/famit-agent/ai_manager/ (a plain top-level package, like
# ads_engine/media_gen/workflow) so `from ai_manager.endpoints import router` resolves with NO sys.path
# hack. router is None when FastAPI is absent (offline test env) -- guarded below.
#
# IMPORT-GUARD (house pattern): a missing/broken ai_manager package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_AI_MANAGER!=1/true => router NOT mounted => byte-identical behavior.
try:
    from ai_manager.endpoints import router as _ai_manager_router  # noqa: E402
except Exception:  # noqa: BLE001
    _ai_manager_router = None

FEATURE_AI_MANAGER = (cfg_get("FEATURE_AI_MANAGER", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_AI_MANAGER and _ai_manager_router is not None:
    try:
        app.include_router(_ai_manager_router)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_aimanager
        _lg_aimanager.getLogger("famit-caller").warning("ai-manager router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- funnels (ad->landing->lead->call->WA->booking->payment funnel builder, prefix
# /funnels). FLAG-GATED. ⚠ SECURITY-CRITICAL: was the "BLOCKED" checklist row #9.
# ----------------------------------------------------------------------------------------------
# TENANT ISOLATION (the build-state must-fix, RESOLVED): funnels ships BOTH a bare module-level `router`
# (funnels.endpoints.router) that reads tenant_id FROM THE BODY (`payload.get("tenant_id")`) -- the exact
# cross-tenant hole -- AND a CLEAN token-deriving `build_router(resolve_tenant, can, need_auth, forbidden,
# firewall)` (funnels/endpoints.py L132, the same shape as workflow-studio/forms-surveys, ADDED by the
# 2026-06-10 security fix). We mount ONLY `build_router`: every route does
# `t = resolve_tenant(request)` (token), `if not t: need_auth()`, tenant_id is ALWAYS t["tenant_id"]
# (verified by Read), writes enforce `can(t,"write")`. Because publish/run DELEGATE to the workflow engine,
# deriving tenant from the TOKEN here is what stops an attacker body-tenant from flowing into
# workflow.publish/run. We NEVER mount the bare `router`, and we DO NOT apply funnel_wiring.diff (the
# shipped diff mounts the bare body-tenant router = the hole; it is inert text in the package, ignored).
#
# ORDER: mounted AFTER workflow-studio (above) -- funnels lazy-`import workflow` to delegate publish/run,
# and the workflow package is already on the box (/opt/famit-agent/workflow/), so this resolves for free.
#
# NO init()/ensure_schema TO DEFER: make_store() defaults to the IN-MEMORY backend unless FUNNELS_STORE=pg,
# so even with the flag ON the live PG is NOT touched. No DDL. config.killswitch() (FUNNELS_KILLSWITCH) is a
# separate runtime break-glass, NOT this mount flag.
#
# IMPORT-GUARD (house pattern): a missing/broken funnels package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_FUNNELS!=1/true => router NOT mounted => byte-identical behavior.
try:
    from funnels.endpoints import build_router as _build_funnels_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_funnels_router = None

FEATURE_FUNNELS = (cfg_get("FEATURE_FUNNELS", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_FUNNELS and _build_funnels_router is not None:
    try:
        _funnels_router = _build_funnels_router(
            resolve_tenant, can, need_auth, _forbidden, firewall=_firewall_mod
        )
        if _funnels_router is not None:
            app.include_router(_funnels_router)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_funnels
        _lg_funnels.getLogger("famit-caller").warning("funnels router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- whatsapp-builder (AI WhatsApp template-generation BRAIN, prefix
# /whatsapp/campaign). FLAG-GATED, default OFF, dormant-until-creds.
# ----------------------------------------------------------------------------------------------
# WHAT: the upstream brain for the WhatsApp Campaign Builder -- pick a campaign -> the LLM
# (reused Groq->OpenRouter seam) PROPOSES Meta-compliant template suggestions + variations + CTAs
# + personalization tokens + media recs + campaign structure; a DETERMINISTIC Meta-compliance
# validator is the AUTHORITY (grammar, category auto-classify, NO-INVENT scrub). Spec:
# design/wa-template-ai-backend.md. Package deployed to /opt/famit-agent/whatsapp_builder/.
#
# DISTINCT from the live whatsapp.py (send/receive/webhook) -- this NEVER edits it. Generation
# spend rides wallet.py (resource_type="wa_template_gen", idempotent F4 reserve/settle/release);
# the eventual per-message SEND cost stays the whatsapp.py meter's job (no double-charge here).
#
# TENANT ISOLATION: we mount ONLY the token-deriving build_router (funnels/workflow shape). Every
# route does `t = resolve_tenant(request)` (token), tenant_id is ALWAYS t["tenant_id"], NEVER a
# body/query field. Writes enforce can(t,"write"). There is NO bare body-tenant router to avoid.
#
# NO init()/ensure_schema TO DEFER AT IMPORT: the ai_wa_* DDL (whatsapp_builder/db/ddl_ai_wa.sql)
# is applied STANDALONE via psql as famit_app (off the Alembic chain, same as ddl_wallet.sql).
# With no DB reachable the module falls back to var/whatsapp_builder/*.jsonl, so the live PG is
# untouched whether the flag is ON or OFF until the schema is explicitly applied + ENSURE run.
#
# IMPORT-GUARD (house pattern): a missing/broken whatsapp_builder package can NEVER break startup.
# FEATURE FLAG default OFF: FEATURE_WHATSAPP_BUILDER!=1/true => router NOT mounted => byte-identical.
try:
    from whatsapp_builder.router import build_router as _build_wab_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_wab_router = None

FEATURE_WHATSAPP_BUILDER = (cfg_get("FEATURE_WHATSAPP_BUILDER", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_WHATSAPP_BUILDER and _build_wab_router is not None:
    try:
        _wab_router = _build_wab_router(
            resolve_tenant, can, need_auth, _forbidden, firewall=_firewall_mod
        )
        if _wab_router is not None:
            app.include_router(_wab_router)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_wab
        _lg_wab.getLogger("famit-caller").warning("whatsapp-builder router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — communication (the omnichannel comms tab: Telegram now; Email/SMS later, prefix /comm).
# FLAG-GATED (COMM_ENABLED, default OFF => byte-identical resting). COMMUNICATION-MASTER-PLAN §8 WAVE 1.
# ----------------------------------------------------------------------------------------------
# ⚠ TENANT ISOLATION: comm ships the token-deriving AUTHENTICATED surface
# `build_router(resolve_tenant, can, need_auth, forbidden, require_super_admin=, firewall=, audit=)`
# — same shape as whatsapp_builder/provider_registry. Every AUTHENTICATED route derives tenant_id from
# resolve_tenant(request)["tenant_id"] (token), NEVER a body/query field; writes enforce can(t,"write").
#
# THE ONE UNAUTHENTICATED route is POST /comm/webhook/telegram/{tenant_id} — Telegram (a machine) calls
# it. It is FAIL-CLOSED (COMMUNICATION-MASTER-PLAN §4 S2): the {tenant_id} is UNTRUSTED until
# comm.webhook.handle() constant-time-verifies the per-tenant secret_token (bound to the PATH tenant +
# that tenant's bot provider_def); the RLS GUC is set ONLY AFTER that verify. No/wrong/other-tenant
# secret -> 403; a dormant tenant -> 403 (NOT fail-open like the legacy Meta webhook). It NEVER blocks.
#
# EARNER LAW: comm rides caller.py (a separate process). It NEVER imports agent.py (the live earner).
# Every contact-facing send (the post-call hook, a later phase) is asyncio.create_task'd + bounded by a
# per-channel asyncio.wait_for inside the engine — NEVER awaited on the dial loop.
#
# IMPORT-GUARD (house pattern): a missing/broken comm package can NEVER break startup — build_router
# degrades to None when FastAPI is absent, and any import error is swallowed. FEATURE FLAG default OFF:
# COMM_ENABLED!=1/true => router NOT mounted => byte-identical. Even if mounted, EVERY route self-404s
# while the flag is OFF (defense in depth) — so the resting state is byte-identical either way.
try:
    from comm.router import build_router as _build_comm_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_comm_router = None

COMM_ENABLED = (cfg_get("COMM_ENABLED", "0") or "0").strip().lower() in ("1", "true", "yes", "on")

if COMM_ENABLED and _build_comm_router is not None:
    try:
        _comm_router = _build_comm_router(
            resolve_tenant, can, need_auth, _forbidden,
            require_super_admin=require_super_admin, firewall=_firewall_mod, audit=_audit,
        )
        if _comm_router is not None:
            app.include_router(_comm_router)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_comm
        _lg_comm.getLogger("famit-caller").warning("communication router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT — twenty-crm (deep Twenty CRM integration, prefix /twenty). TWENTY-CRM-INTEGRATION.
# ----------------------------------------------------------------------------------------------
# WHAT: a server-side proxy + value-bridge over a tenant's Twenty CRM (https://twenty.com). The panel
# renders NATIVE Haptica UI (Pipeline kanban / Companies / People) over the normalized /twenty/* contract
# this router serves — NO iframe, NO cross-origin. The workspace API key (a JWT) stays SERVER-SIDE: the
# browser never sees it (status reads hand back only a masked tail).
#
# TENANT ISOLATION: build_router(resolve_tenant, can, need_auth, forbidden) — same token-deriving shape as
# forms/workflow/comm. The Twenty connection (URL+key) is resolved PER TENANT from
# resolve_tenant(request)["tenant_id"] via a small JSON store under VAR (never a body/query field). Writes
# enforce can(t,"write"). A genuine 401 still bounces via need_auth.
#
# DORMANT-SAFE BY DESIGN: with no per-tenant connection AND no env fallback, every READ returns
# {connected:false}+empty (200) so the panel shows a calm "Connect your Twenty CRM" state, and the module
# makes ZERO external calls / touches NOTHING. So even with the flag ON the resting state for an
# unconnected tenant is inert. The flag defaults ON (this is an explicitly-shipped feature that must be
# connectable out of the box); set FEATURE_TWENTY_CRM=0 to unmount entirely. Optional single-tenant/dev
# fallback creds: TWENTY_API_URL + TWENTY_API_KEY.
#
# IMPORT-GUARD (house pattern): a missing/broken twenty_crm package or absent FastAPI can NEVER break
# startup — the import is swallowed and the mount is wrapped.
try:
    from twenty_crm import build_router as _build_twenty_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_twenty_router = None

FEATURE_TWENTY_CRM = (cfg_get("FEATURE_TWENTY_CRM", "1") or "1").strip().lower() in ("1", "true", "yes", "on")

if FEATURE_TWENTY_CRM and _build_twenty_router is not None:
    try:
        _twenty_self_host = (cfg_get("TWENTY_SELF_HOST", "0") or "0").strip().lower() in ("1", "true", "yes", "on")
        _twenty_router = _build_twenty_router(
            resolve_tenant, can, need_auth, _forbidden,
            var_dir=VAR,
            env_url=cfg_get("TWENTY_API_URL", "") or "",
            env_key=cfg_get("TWENTY_API_KEY", "") or "",
            # Self-hosted (zero-touch) mode: Haptica auto-provisions an isolated Twenty
            # workspace per tenant against the internal Twenty (no API key from the user).
            self_host=_twenty_self_host,
            internal_url=cfg_get("TWENTY_INTERNAL_URL", "") or "",
            provision_domain=cfg_get("TWENTY_PROVISION_DOMAIN", "crm.haptica.local") or "crm.haptica.local",
            provision_secret=cfg_get("TWENTY_PROVISION_SECRET", "") or cfg_get("TWENTY_APP_SECRET", "") or "",
        )
        if _twenty_router is not None:
            app.include_router(_twenty_router)
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_twenty
        _lg_twenty.getLogger("famit-caller").warning("twenty-crm router mount failed", exc_info=True)


# ==============================================================================================
# MODULE MOUNT -- auto-lead (real-time multi-source lead ingestion, prefix /auto-lead). AUTO-LEAD.
# ----------------------------------------------------------------------------------------------
# WHAT: connect sources (website/custom webhooks, Zapier, Meta/Google lead ads, WhatsApp, email,
# Apollo) -> ingest leads in REAL TIME -> validate/dedupe -> route into the LEADS store (so Riya can
# call them) + a live activity feed. The hero is the PUBLIC webhook POST /auto-lead/ingest/{token}:
# unauthenticated by design (tenant derived from the unguessable per-source token, never a request
# field), with a 64KB size cap + honeypot on top of caller.py's global IP rate-limit (forms' model).
#
# LEADS INTEGRATION: we inject ONE async callable `_al_add_lead(tenant_id, lead)` that does the exact
# lock-guarded, phone-normalised, per-tenant-deduped write to leads.json the /leads endpoint does -- so
# ingested leads are first-class Haptica leads (single source of truth; no schema drift).
#
# PULL SOURCES (email IMAP / Apollo) are polled by the router's poll_once(), driven from a startup task
# below (a separate @app.on_event so it can't disturb scheduler_loop). Default interval 120s.
#
# IMPORT-GUARD (house pattern): a missing/broken auto_lead package can NEVER break startup.
# FEATURE FLAG default ON (additive + dormant-safe; set FEATURE_AUTO_LEAD=0 to unmount).
try:
    from auto_lead import build_router as _build_auto_lead_router  # noqa: E402
except Exception:  # noqa: BLE001
    _build_auto_lead_router = None

FEATURE_AUTO_LEAD = (cfg_get("FEATURE_AUTO_LEAD", "1") or "1").strip().lower() in ("1", "true", "yes", "on")
_auto_lead_router = None

if FEATURE_AUTO_LEAD and _build_auto_lead_router is not None:
    async def _al_add_lead(tenant_id: str, lead: dict) -> dict:
        """Add ONE ingested lead to leads.json (same dedup/normalise/tenant-scope as POST /leads).
        Returns {added: bool, lead_id?, reason?}. Lock-guarded against the run-path."""
        phone = norm(lead.get("phone", "") or "")
        if not phone:
            return {"added": False, "reason": "invalid_phone"}
        async with _STORE_LOCK:
            store = _read_raw(LEADS_FILE, [])
            existing = None
            for x in store:
                if x.get("tenant_id", ADMIN_ID) == tenant_id and x.get("phone") == phone:
                    existing = x
                    break
            if existing is not None:
                return {"added": False, "reason": "duplicate", "lead_id": existing.get("id")}
            lid = uuid.uuid4().hex[:8]
            rec = {
                "id": lid, "tenant_id": tenant_id,
                "name": (lead.get("name") or "").strip(), "phone": phone,
                "status": (lead.get("status") or "new"),
                "added_at": datetime.now().isoformat(timespec="seconds"),
                "source": lead.get("source") or "auto_lead",
            }
            if lead.get("email"):
                rec["email"] = str(lead["email"]).strip()
            if lead.get("tags"):
                rec["tags"] = [str(x) for x in lead["tags"] if str(x).strip()]
            if lead.get("hot"):
                rec["hot"] = True
                rec["score"] = 70
            store.append(rec)
            _write_raw(LEADS_FILE, store)
        return {"added": True, "lead_id": lid}

    try:
        _auto_lead_router = _build_auto_lead_router(
            resolve_tenant, can, need_auth, _forbidden,
            var_dir=VAR, add_lead=_al_add_lead, norm=norm, client_ip=_client_ip,
        )
        if _auto_lead_router is not None:
            app.include_router(_auto_lead_router)

        _AL_POLL_INTERVAL = int(cfg_get("AUTO_LEAD_POLL_INTERVAL", "120") or "120")

        @app.on_event("startup")
        async def _start_auto_lead_pollers():  # noqa: ANN202
            async def _loop():
                while True:
                    await asyncio.sleep(max(30, _AL_POLL_INTERVAL))
                    try:
                        if _auto_lead_router is not None:
                            await _auto_lead_router.poll_once()
                    except Exception:  # noqa: BLE001 -- a poll hiccup never stops the loop
                        pass
            asyncio.create_task(_loop())
    except Exception:  # noqa: BLE001 -- a mount failure must never crash the live spine
        import logging as _lg_al
        _lg_al.getLogger("famit-caller").warning("auto-lead router mount failed", exc_info=True)


# ═══════════════════════════════════════════════════════════════════════════════
# LPR — PLATFORM PROVIDER KEY-STORE (super-admin only). Hot-reloadable: a key added
# here reaches the live AIM rotation on the next pick() with NO redeploy/restart.
# These are PLATFORM keys (Groq/Sarvam/SambaNova/OpenRouter), NOT per-tenant. Gated
# by require_super_admin (legacy_pw EXCLUDED). Raw key NEVER returned — only masked.
# Import-guarded: a missing llm_router can never crash startup or any other route.
# ═══════════════════════════════════════════════════════════════════════════════
try:
    from llm_router import key_store as _pk_store, get_pool as _pk_get_pool  # noqa: E402
except Exception:  # noqa: BLE001
    _pk_store = None  # type: ignore
    _pk_get_pool = None  # type: ignore

_PK_PROVIDERS = ("groq", "sarvam", "sambanova", "openrouter")


def _pk_unavailable():
    return JSONResponse({"error": "provider key-store unavailable"}, status_code=503)


@app.get("/admin/provider-keys")
async def admin_provider_keys_list(request: Request):
    """List platform provider keys (masked) grouped by provider. Raw key NEVER returned."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _pk_store is None:
        return _pk_unavailable()
    return JSONResponse({"providers": _pk_store.list_all_masked()})


@app.post("/admin/provider-keys")
async def admin_provider_keys_add(request: Request, provider: str = Form(...),
                                  key: str = Form(...), label: str = Form("")):
    """Add a provider key. Enters the live rotation on the next pick() (hot-reload)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _pk_store is None:
        return _pk_unavailable()
    p = (provider or "").strip().lower()
    if p not in _PK_PROVIDERS:
        return JSONResponse({"error": "unknown provider", "allowed": list(_PK_PROVIDERS)}, status_code=400)
    if not (key or "").strip():
        return JSONResponse({"error": "empty key"}, status_code=400)
    try:
        res = _pk_store.add_key(p, key, label=label)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": "add failed", "detail": type(exc).__name__}, status_code=400)
    _audit(request, t, "provider_key.add", "provider_key", res.get("id", ""),
           channel="control", meta={"provider": p, "masked": res.get("masked", ""),
                                    "deduped": res.get("deduped", False)})
    return JSONResponse({"ok": True, **res})


@app.put("/admin/provider-keys/{key_id}")
async def admin_provider_keys_update(request: Request, key_id: str,
                                     enabled: str = Form(""), label: str = Form("")):
    """Toggle enabled / relabel by id. Never edits the secret. Hot-reloads live."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _pk_store is None:
        return _pk_unavailable()
    en = None
    if str(enabled).strip() != "":
        en = str(enabled).strip().lower() in ("1", "true", "yes", "on", "enabled")
    lbl = label if str(label).strip() != "" else None
    res = _pk_store.update_key(key_id, enabled=en, label=lbl)
    if not res.get("ok"):
        return JSONResponse({"error": "key not found", "id": key_id}, status_code=404)
    _audit(request, t, "provider_key.update", "provider_key", key_id,
           channel="control", meta={"enabled": en, "relabel": lbl is not None})
    return JSONResponse(res)


@app.delete("/admin/provider-keys/{key_id}")
async def admin_provider_keys_delete(request: Request, key_id: str):
    """Delete a provider key by id. Removed from rotation on the next pick()."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _pk_store is None:
        return _pk_unavailable()
    res = _pk_store.delete_key(key_id)
    if not res.get("deleted"):
        return JSONResponse({"error": "key not found", "id": key_id}, status_code=404)
    _audit(request, t, "provider_key.delete", "provider_key", key_id, channel="control")
    return JSONResponse(res)


@app.get("/admin/provider-keys/status")
async def admin_provider_keys_status(request: Request):
    """Live pool view per key: cooling_until / pick_count / last_429_at / available.
    Reads each pool's snapshot() (masked; no raw key)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _pk_get_pool is None:
        return _pk_unavailable()
    out = {}
    for p in _PK_PROVIDERS:
        try:
            pool = _pk_get_pool(p)
            out[p] = pool.snapshot() if pool is not None else []
        except Exception:  # noqa: BLE001
            out[p] = []
    return JSONResponse({"status": out})


# ════════════════════════════════════════════════════════════════════════════════════════════════
# P2: MANAGED PROVIDER LAYER — encrypted ProviderKeyStore + health-scored KeyRouter (the W13 engine).
# Additive + IMPORT-GUARDED + DORMANT-SAFE: every route self-503s until voice_ops.config is importable
# AND (for writes) a keystore master secret is set. This is the canonical store going forward; the
# legacy /admin/provider-keys above stays until the panel + agent are fully repointed. No agent path
# is touched here — the live call is unaffected. Backs the P3 Service Control Center.
# ════════════════════════════════════════════════════════════════════════════════════════════════
try:
    from voice_ops.config.router_bridge import get_key_router as _vk_get_router  # noqa: E402
    from voice_ops.config.keys import ProviderKeyStore as _VkKeyStore  # noqa: E402
except Exception:  # noqa: BLE001
    _vk_get_router = None  # type: ignore
    _VkKeyStore = None  # type: ignore

# Platform-scoped key set the earner shares (per-tenant BYO is a later unit). is_admin=True so the
# managed router/store operate cross-tenant for the super-admin console.
_PLATFORM_TENANT = (os.getenv("PLATFORM_TENANT_ID") or "_platform").strip()


def _vk_unavailable():
    return JSONResponse({"error": "managed provider layer unavailable (voice_ops.config not importable)"},
                        status_code=503)


@app.get("/admin/provider-pool/health")
async def admin_provider_pool_health(request: Request, provider: str = ""):
    """Per-key health (score / circuit / latency / success-rate / status) for the managed providers.
    No secrets — fingerprints only. Degrades to an empty shape if the key store isn't reachable."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _vk_get_router is None:
        return _vk_unavailable()
    try:
        return JSONResponse({"health": _vk_get_router(_PLATFORM_TENANT, is_admin=True).health(provider or None),
                             "platform_tenant": _PLATFORM_TENANT})
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"health": {}, "error": type(exc).__name__})


@app.get("/admin/provider-pool/usage")
async def admin_provider_pool_usage(request: Request, minutes: int = 1440):
    """Per-key provider usage. `durable` is the CROSS-PROCESS truth (ClickHouse, flushed by the
    agent workers — success/failure/429/latency/score/status per key); `live` is THIS process's
    in-memory snapshot (usually empty in the API process). The P3 UI reads `durable`."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _vk_get_router is None:
        return _vk_unavailable()
    live = {"providers": {}}
    try:
        live = _vk_get_router(_PLATFORM_TENANT, is_admin=True).analytics_snapshot()
    except Exception:  # noqa: BLE001
        pass
    durable, durable_err = [], None
    if _obs_q is not None:
        try:
            r = await _obs_q.provider_key_usage(minutes)
            durable = r.get("rows", [])
            durable_err = r.get("error")
        except Exception as exc:  # noqa: BLE001
            durable_err = type(exc).__name__
    return JSONResponse({"live": live, "durable": durable,
                         **({"durable_error": durable_err} if durable_err else {})})


@app.post("/admin/provider-pool/keys")
async def admin_provider_pool_add_key(request: Request, provider: str = Form(...),
                                      key: str = Form(...), label: str = Form("")):
    """Add an ENCRYPTED provider key (AES-256-GCM at rest; plaintext only at call time). Joins the
    health pool live. Requires a keystore master secret (else 400 with a clear hint)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _VkKeyStore is None:
        return _vk_unavailable()
    p = (provider or "").strip().lower()
    if not p:
        return JSONResponse({"error": "provider required"}, status_code=400)
    if not (key or "").strip():
        return JSONResponse({"error": "empty key"}, status_code=400)
    actor = t.get("tenant_id", "") if isinstance(t, dict) else ""
    try:
        rec = _VkKeyStore().add_key(_PLATFORM_TENANT, p, key, label=label, added_by=actor, is_admin=True)
    except Exception as exc:  # noqa: BLE001 — most often VaultError (no master secret) or store down
        return JSONResponse({"error": "add failed (is a keystore master secret + config store configured?)",
                             "detail": type(exc).__name__}, status_code=400)
    _audit(request, t, "provider_pool.key.add", "provider_key", rec.get("fingerprint", ""),
           channel="control", meta={"provider": p})
    return JSONResponse({"ok": True, **rec})


@app.put("/admin/provider-pool/keys/{provider}/{fingerprint}")
async def admin_provider_pool_update_key(request: Request, provider: str, fingerprint: str,
                                         enabled: str = Form("")):
    """Enable/disable a managed key by fingerprint (never edits the secret). Hot-reloads live."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _VkKeyStore is None:
        return _vk_unavailable()
    p = (provider or "").strip().lower()
    en = str(enabled).strip().lower() in ("1", "true", "yes", "on", "enabled")
    actor = t.get("tenant_id", "") if isinstance(t, dict) else ""
    try:
        store = _VkKeyStore()
        res = (store.enable_key(_PLATFORM_TENANT, p, fingerprint, actor=actor, is_admin=True) if en
               else store.disable_key(_PLATFORM_TENANT, p, fingerprint, actor=actor, is_admin=True))
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": "update failed", "detail": type(exc).__name__}, status_code=400)
    if res.get("status") == "not_found":
        return JSONResponse({"error": "key not found", "fingerprint": fingerprint}, status_code=404)
    _audit(request, t, "provider_pool.key.update", "provider_key", fingerprint,
           channel="control", meta={"provider": p, "enabled": en})
    return JSONResponse(res)


@app.delete("/admin/provider-pool/keys/{provider}/{fingerprint}")
async def admin_provider_pool_delete_key(request: Request, provider: str, fingerprint: str):
    """Delete a managed key by fingerprint. Removed from rotation on the next resolve."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _VkKeyStore is None:
        return _vk_unavailable()
    p = (provider or "").strip().lower()
    actor = t.get("tenant_id", "") if isinstance(t, dict) else ""
    try:
        res = _VkKeyStore().remove_key(_PLATFORM_TENANT, p, fingerprint, actor=actor, is_admin=True)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": "delete failed", "detail": type(exc).__name__}, status_code=400)
    _audit(request, t, "provider_pool.key.delete", "provider_key", fingerprint,
           channel="control", meta={"provider": p})
    return JSONResponse({"ok": True, **res})


# === PVS PHASE-1: CUSTOM PROVIDER CRUD (super-admin gated; isolated store) ===
# Separate from the live provider key-store (which feeds the earner pool) — registering a custom
# provider here NEVER changes the earner pipeline. Routing an outbound call through a custom
# provider is PHASE-2 / OB-PROV (gated). Phase 1 = persist + list + delete + surface in /providers.
def _cp_store():
    try:
        from llm_router import custom_providers as _cp
        return _cp
    except Exception:  # noqa: BLE001
        return None


@app.get("/admin/custom-providers")
async def admin_custom_providers_list(request: Request):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    cp = _cp_store()
    if cp is None:
        return JSONResponse({"error": "custom-provider store unavailable"}, status_code=503)
    return JSONResponse({"custom_providers": cp.list_masked()})


@app.post("/admin/custom-providers")
async def admin_custom_providers_add(request: Request, name: str = Form(...), kind: str = Form(...),
                                     base_url: str = Form(...), model: str = Form(""),
                                     key: str = Form(""), logo_url: str = Form("")):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    cp = _cp_store()
    if cp is None:
        return JSONResponse({"error": "custom-provider store unavailable"}, status_code=503)
    try:
        res = cp.add(name, kind, base_url, model, key=key, logo_url=logo_url)
    except ValueError as exc:
        return JSONResponse({"error": str(exc)}, status_code=400)
    except Exception as exc:  # noqa: BLE001
        return JSONResponse({"error": "add failed", "detail": type(exc).__name__}, status_code=400)
    _audit(request, t, "custom_provider.add", "custom_provider", res.get("id", ""),
           channel="control", meta={"name": res.get("name"), "kind": res.get("kind"),
                                    "model": res.get("model")})
    return JSONResponse({"ok": True, **res})


@app.put("/admin/custom-providers/{cid}")
async def admin_custom_providers_update(request: Request, cid: str, enabled: str = Form(""),
                                        name: str = Form(""), base_url: str = Form(""),
                                        model: str = Form(""), key: str = Form("")):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    cp = _cp_store()
    if cp is None:
        return JSONResponse({"error": "custom-provider store unavailable"}, status_code=503)
    en = None
    if str(enabled).strip() != "":
        en = str(enabled).strip().lower() in ("1", "true", "yes", "on", "enabled")
    res = cp.update(cid, enabled=en,
                    label=(name if str(name).strip() else None),
                    base_url=(base_url if str(base_url).strip() else None),
                    model=(model if str(model).strip() else None),
                    key=(key if str(key).strip() else None))
    if not res.get("ok"):
        return JSONResponse({"error": "custom provider not found", "id": cid}, status_code=404)
    _audit(request, t, "custom_provider.update", "custom_provider", cid, channel="control")
    return JSONResponse(res)


@app.delete("/admin/custom-providers/{cid}")
async def admin_custom_providers_delete(request: Request, cid: str):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    cp = _cp_store()
    if cp is None:
        return JSONResponse({"error": "custom-provider store unavailable"}, status_code=503)
    res = cp.delete(cid)
    if not res.get("deleted"):
        return JSONResponse({"error": "custom provider not found", "id": cid}, status_code=404)
    _audit(request, t, "custom_provider.delete", "custom_provider", cid, channel="control")
    return JSONResponse(res)
# === /PVS PHASE-1 custom-provider CRUD ===


# ════════════════════════════════════════════════════════════════════════════════════════════════
# TOLEX — agent tooling & capability system (CONTROL PLANE). Import-guarded (a missing tolex module
# returns 503, never a startup error) + every route require_super_admin. The agent RUNTIME hook is a
# separate flag (TOLEX_ENABLED) in agent.py; this control plane works regardless so you can configure
# grants before turning the runtime on.
# ════════════════════════════════════════════════════════════════════════════════════════════════
try:
    import tolex as _tolex_mod  # noqa: E402
except Exception:  # noqa: BLE001
    _tolex_mod = None


@app.get("/admin/tolex/catalog")
async def admin_tolex_catalog(request: Request):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    return JSONResponse({"catalog": _tolex_mod.catalog(), "runtime_enabled": _tolex_mod.enabled()})


@app.get("/admin/tolex/grants")
async def admin_tolex_grants_get(request: Request, campaign_id: str = ""):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    return JSONResponse({"grants": _tolex_mod.get_grants("", campaign_id)})


@app.put("/admin/tolex/grants")
async def admin_tolex_grants_put(request: Request):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    cid = str((body or {}).get("campaign_id", "") or "")
    en = bool((body or {}).get("enabled", False))
    tools = (body or {}).get("tools", {}) or {}
    res = _tolex_mod.set_grants("", cid, en, tools)
    _audit(request, t, "tolex.grants.set", "tolex", cid or "_default", channel="control",
           meta={"enabled": en, "tools": list(tools.keys()) if isinstance(tools, dict) else []})
    return JSONResponse({"ok": True, "grants": res})


@app.post("/admin/tolex/grants/enable-recommended")
async def admin_tolex_enable_recommended(request: Request, campaign_id: str = ""):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    res = _tolex_mod.enable_recommended("", campaign_id)
    _audit(request, t, "tolex.grants.enable_recommended", "tolex", campaign_id or "_default", channel="control")
    return JSONResponse({"ok": True, "grants": res})


@app.get("/admin/tolex/ops")
async def admin_tolex_ops(request: Request, campaign_id: str = "", limit: int = 200):
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    return JSONResponse({"ops": _tolex_mod.recent_ops(campaign_id, limit)})


# ── Tolex TENANT surface (/tolex/*) — a tenant manages ONLY its own agent's tooling. The scope is the
# AUTH-derived tenant id (never a client-supplied one), so a tenant can only ever read/write its own
# node + see its own operations: hard isolation, no cross-tenant access. Mirrors the /bookings auth.
@app.get("/tolex/catalog")
async def tolex_catalog(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    return JSONResponse({"catalog": _tolex_mod.catalog(), "runtime_enabled": _tolex_mod.enabled()})


@app.get("/tolex/grants")
async def tolex_grants_get(request: Request, campaign_id: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    return JSONResponse({"grants": _tolex_mod.get_grants(t.get("tenant_id", "") or "", campaign_id)})


@app.put("/tolex/grants")
async def tolex_grants_put(request: Request):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    try:
        body = await request.json()
    except Exception:  # noqa: BLE001
        body = {}
    cid = str((body or {}).get("campaign_id", "") or "")
    en = bool((body or {}).get("enabled", False))
    tools = (body or {}).get("tools", {}) or {}
    res = _tolex_mod.set_grants(t.get("tenant_id", "") or "", cid, en, tools)
    return JSONResponse({"ok": True, "grants": res})


@app.post("/tolex/grants/enable-recommended")
async def tolex_enable_recommended(request: Request, campaign_id: str = ""):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    res = _tolex_mod.enable_recommended(t.get("tenant_id", "") or "", campaign_id)
    return JSONResponse({"ok": True, "grants": res})


@app.get("/tolex/ops")
async def tolex_ops_tenant(request: Request, campaign_id: str = "", limit: int = 200):
    t = resolve_tenant(request)
    if not t:
        return need_auth()
    if _tolex_mod is None:
        return JSONResponse({"error": "tolex unavailable"}, status_code=503)
    return JSONResponse({"ops": _tolex_mod.recent_ops(campaign_id, limit, t.get("tenant_id", "") or "")})
# === /TOLEX control plane ===


# ════════════════════════════════════════════════════════════════════════════════════════════════
# VOICE-CONFIG STORE — super-admin reads/writes VAR/voice_keys.json, the SAME file the agent reads.
# Raw keys NEVER leave the server: GET masks every secret to its last-4 chars. POST is a MERGE — only
# the fields the caller actually sends are touched (an empty string CLEARS that one field; absent
# fields are left untouched so you can't accidentally wipe sibling keys). Super-admin gated.
# ════════════════════════════════════════════════════════════════════════════════════════════════
VOICE_KEYS_FILE = VAR / "voice_keys.json"   # canonical voice STT/TTS provider keys (agent reads this)
_VOICE_KEY_FIELDS = ("deepgram_api_key", "sarvam_api_key", "elevenlabs_api_key", "groq_api_key")
_VOICE_PROVIDERS = ("deepgram", "sarvam", "elevenlabs", "groq")


def _mask_secret(val: str) -> str:
    """Mask a secret to a fixed dot prefix + its last 4 chars. Empty/short -> ''."""
    s = str(val or "")
    if len(s) < 4:
        return ""
    return "••••" + s[-4:]


def _voice_config_masked(cfg: dict) -> dict:
    """Build the public (masked) voice-config view. Never echoes a raw key."""
    cfg = cfg if isinstance(cfg, dict) else {}
    providers = {}
    for prov in _VOICE_PROVIDERS:
        raw = str(cfg.get(f"{prov}_api_key", "") or "")
        # emit BOTH name sets so any reader works (FE reads configured/key_masked; has_key/masked kept).
        providers[prov] = {"has_key": bool(raw), "configured": bool(raw),
                           "masked": _mask_secret(raw), "key_masked": _mask_secret(raw)}
    return {
        "stt_provider": str(cfg.get("stt_provider", "") or "") or "sarvam",
        "providers": providers,
    }


@app.get("/admin/voice-config")
async def admin_voice_config_get(request: Request):
    """Return the current voice STT/TTS provider config (MASKED — raw keys never returned).
    Reads VAR/voice_keys.json (the same file the agent reads); missing file -> defaults."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    cfg = _read_raw(VOICE_KEYS_FILE, {})
    if not isinstance(cfg, dict):
        cfg = {}
    return JSONResponse(_voice_config_masked(cfg))


@app.post("/admin/voice-config")
async def admin_voice_config_set(request: Request):
    """MERGE a partial voice-config into VAR/voice_keys.json. Only the fields the caller sends
    are touched: an empty string CLEARS that field; an absent field is left untouched (so other
    keys are never wiped). Writes atomically. Returns the masked config (same shape as GET)."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    try:
        body = await request.json()
        if not isinstance(body, dict):
            return JSONResponse({"error": "body must be a JSON object"}, status_code=400)
    except Exception:  # noqa: BLE001
        return JSONResponse({"error": "invalid JSON body"}, status_code=400)
    cfg = _read_raw(VOICE_KEYS_FILE, {})
    if not isinstance(cfg, dict):
        cfg = {}
    changed = []
    if "stt_provider" in body:
        sp = str(body.get("stt_provider", "") or "").strip().lower()
        if sp:
            cfg["stt_provider"] = sp
            changed.append("stt_provider")
    for field in _VOICE_KEY_FIELDS:
        if field in body:
            val = body.get(field)
            if val is None:
                continue   # explicit null == "leave untouched", same as absent
            cfg[field] = str(val)   # "" clears the field; non-empty sets it
            changed.append(field)
    _atomic_write_json(VOICE_KEYS_FILE, cfg)
    _audit(request, t, "voice_config.update", "voice_config", "voice_keys.json",
           channel="control", meta={"fields": changed})
    return JSONResponse(_voice_config_masked(cfg))


@app.get("/admin/recording-proxy")
async def recording_proxy(request: Request, url: str = ""):
    """Same-origin proxy for a presigned recording: the browser fetches the WHOLE file from us (one
    fast hop) and plays it from memory, instead of streaming cross-region (choppy on a Singapore→India
    path). SSRF-guarded to either a *.digitaloceanspaces.com presigned URL (X-Amz-Signature) OR a
    *.blob.core.windows.net Azure SAS URL (sig) — the signature IS the capability in both cases.
    Super-admin gated; NEVER raises into the request."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    try:
        p = urllib.parse.urlparse(url or "")
        host = (p.hostname or "").lower()
        qs = urllib.parse.parse_qs(p.query or "")
        # Accept a DO Spaces presigned URL (X-Amz-Signature) OR an Azure Blob SAS URL (sig). Both
        # carry their own capability in the signature; anything else is rejected as a forbidden host.
        ok_spaces = host.endswith(".digitaloceanspaces.com") and "X-Amz-Signature" in qs
        ok_azure = host.endswith(".blob.core.windows.net") and "sig" in qs
        if not (ok_spaces or ok_azure):
            return JSONResponse({"error": "forbidden host"}, status_code=400)
        async with httpx.AsyncClient(timeout=60.0, follow_redirects=True) as cli:
            r = await cli.get(url)
        if r.status_code != 200:
            return JSONResponse({"error": "upstream", "status": r.status_code}, status_code=502)
        return Response(content=r.content,
                        media_type=r.headers.get("content-type", "audio/mpeg"),
                        headers={"Cache-Control": "private, max-age=600"})
    except Exception:  # noqa: BLE001
        return JSONResponse({"error": "proxy failed"}, status_code=502)


# ════════════════════════════════════════════════════════════════════════════════════════════════
# COMPANY LOGO FETCH — best-effort logo resolver for a company website. NEVER raises: on any failure
# it falls back to the always-works Google s2 favicon. Tries Clearbit, then the site's own
# <link rel=icon>/apple-touch-icon/og:image, then the Google favicon. Super-admin gated; httpx async.
# ════════════════════════════════════════════════════════════════════════════════════════════════
def _logo_domain(url: str) -> str:
    """Normalize an arbitrary website string to a bare registrable host (no scheme/path/www)."""
    s = str(url or "").strip()
    if not s:
        return ""
    if "://" not in s:
        s = "http://" + s
    try:
        host = urllib.parse.urlparse(s).netloc or ""
    except Exception:  # noqa: BLE001
        host = ""
    host = host.split("@")[-1].split(":")[0].strip().lower()
    if host.startswith("www."):
        host = host[4:]
    return host


def _google_favicon(domain: str) -> str:
    return f"https://www.google.com/s2/favicons?domain={urllib.parse.quote(domain)}&sz=128"


@app.get("/admin/fetch-logo")
async def admin_fetch_logo(request: Request, url: str = ""):
    """Best-effort company-logo resolver for a website URL. Returns {logo_url, source}. NEVER
    raises — on any error it returns the always-works Google favicon. Strategy (first that works):
    (1) Clearbit logo by domain, (2) the site's own <link rel=icon>/apple-touch-icon/og:image,
    (3) Google s2 favicon fallback. Super-admin gated."""
    t = require_super_admin(request)
    if isinstance(t, JSONResponse):
        return t
    domain = _logo_domain(url)
    if not domain:
        return JSONResponse({"error": "missing or invalid url"}, status_code=400)
    fallback = _google_favicon(domain)
    try:
        async with httpx.AsyncClient(timeout=4.0, follow_redirects=True,
                                     headers={"User-Agent": "Mozilla/5.0 (famit-logo-fetch)"}) as cli:
            # (1) Clearbit — a clean, high-quality brand logo when they have it.
            try:
                r = await cli.head(f"https://logo.clearbit.com/{domain}")
                if r.status_code == 200:
                    return JSONResponse({"logo_url": f"https://logo.clearbit.com/{domain}",
                                         "source": "clearbit"})
            except Exception:  # noqa: BLE001
                pass
            # (2) Parse the site's own HTML for an icon / og:image, resolve to absolute, HEAD-check.
            try:
                page = await cli.get(f"https://{domain}/")
                html = page.text or ""
                base = str(page.url) or f"https://{domain}/"
                cand = None
                src = "favicon"
                m = re.search(
                    r'<link[^>]+rel=["\'][^"\']*(?:apple-touch-icon|shortcut icon|icon)[^"\']*["\'][^>]*>',
                    html, re.IGNORECASE)
                if m:
                    hm = re.search(r'href=["\']([^"\']+)["\']', m.group(0), re.IGNORECASE)
                    if hm:
                        cand = hm.group(1)
                        src = "favicon"
                if not cand:
                    om = re.search(
                        r'<meta[^>]+property=["\']og:image["\'][^>]*content=["\']([^"\']+)["\']',
                        html, re.IGNORECASE)
                    if not om:
                        om = re.search(
                            r'<meta[^>]+content=["\']([^"\']+)["\'][^>]*property=["\']og:image["\']',
                            html, re.IGNORECASE)
                    if om:
                        cand = om.group(1)
                        src = "og"
                if cand:
                    abs_url = urllib.parse.urljoin(base, cand.strip())
                    try:
                        hr = await cli.head(abs_url)
                        if hr.status_code == 200:
                            return JSONResponse({"logo_url": abs_url, "source": src})
                    except Exception:  # noqa: BLE001
                        pass
            except Exception:  # noqa: BLE001
                pass
    except Exception:  # noqa: BLE001
        pass
    # (3) Always-works fallback.
    return JSONResponse({"logo_url": fallback, "source": "favicon"})
